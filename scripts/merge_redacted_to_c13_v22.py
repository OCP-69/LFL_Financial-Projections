"""
Merge 260315_LFL_BM_Vorlage_normal_redacted_final.xlsx
     → C13_Template_financial_projections_neu.xlsx
Output: LFL_BM_C13_Normal_redacted_v22_20260315.xlsx

Mapping:
  Source M5 (Pre-Seed start, Aug 2026) = Target Month 1
  Source M5–M52 (48 months) → Target Monthly rows 2–49

Column mapping (Monthly sheet):
  A  = Year (calendar year)
  B  = Month number (1–48, counting from Pre-Seed start)
  C  = Revenue           ← 6_P&L R8
  D  = Cash in Revenue   ← 6_P&L R8  (simplified: same as revenue)
  E  = Other Cash In     ← 7_BS_CF R9 (Equity Funding)
  F  = CoR Cash Out      ← 6_P&L R14 (COGS)
  G  = OpEx Cash Out     ← 6_P&L R27 (Total Operating Expenses)
  H  = Interest          ← 0 (not modeled)
  I  = Tax               ← 6_P&L R35
  J  = Founder Bonus     ← 0
  K  = Opening Cash      ← 7_BS_CF R14
  L  = Closing Cash      ← 7_BS_CF R15
  M  = Debt Outstanding  ← 0
  N  = Debtors           ← 6_P&L R8  (1-month revenue)
  O  = Creditors         ← (R14+R27) * 10%

Annual sheet: Aggregated by calendar year (2026–2030)
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from datetime import datetime
from copy import copy
import shutil
import os

# ── Paths ──────────────────────────────────────────────────────────────────
SRC   = '260315_LFL_BM_Vorlage_normal_redacted_final.xlsx'
TPL   = 'C13_Template_financial_projections_neu.xlsx'
OUT   = 'LFL_BM_C13_Normal_redacted_v22_20260315.xlsx'

# ── Source row constants (6_P&L) ──────────────────────────────────────────
PL_REVENUE  = 8   # TOTAL REVENUE
PL_COGS     = 14  # TOTAL COGS
PL_OPEX     = 27  # TOTAL OPERATING EXPENSES
PL_TAX      = 35  # Income Tax
PL_NETINC   = 37  # NET INCOME

# ── Source row constants (7_BS_CF) ────────────────────────────────────────
CF_EQUITY   = 9   # Equity Funding Received
CF_BEG      = 14  # Beginning Cash Balance
CF_END      = 15  # ENDING CASH BALANCE

# ── Month mapping ─────────────────────────────────────────────────────────
# Source M5 → Target Month 1
# Source M52 → Target Month 48
SRC_START = 5   # first source month to include
SRC_END   = 52  # last source month (inclusive)
N_MONTHS  = SRC_END - SRC_START + 1  # = 48

# Calendar year structure (source month → calendar year)
# M1 = Apr 2026 → M5 = Aug 2026
def src_to_year(m):
    """Return calendar year for source month m."""
    # M1=Apr 2026, month offset = (m-1)
    # Apr 2026 = month index 3 (0-based Jan=0)
    month_of_year_0based = (3 + (m - 1)) % 12  # 0=Jan ... 11=Dec
    year = 2026 + (3 + (m - 1)) // 12
    return year

# Validate:
# M5=Aug 2026: (3+4)%12=7 → Aug ✓, 2026+(3+4)//12=2026+0=2026 ✓
# M10=Jan 2027: (3+9)%12=0 → Jan ✓, 2026+(3+9)//12=2026+1=2027 ✓
# M52=Jul 2030: (3+51)%12=54%12=6 → Jul ✓, 2026+(3+51)//12=2026+4=2030 ✓

def src_col(m):
    """Source column index for month m (M1=col2, M5=col6)."""
    return m + 1

# ── Load source (data_only) ────────────────────────────────────────────────
print(f'Loading source: {SRC}')
wb_src = openpyxl.load_workbook(SRC, data_only=True)
ws_pl  = wb_src['6_P&L']
ws_cf  = wb_src['7_BS_CF']

def get(ws, row, m):
    """Get numeric value from worksheet at (row, src_col(m))."""
    v = ws.cell(row, src_col(m)).value
    return float(v) if v is not None else 0.0

# ── Collect all 48 months of data ─────────────────────────────────────────
months_data = []
for m in range(SRC_START, SRC_END + 1):
    target_month = m - SRC_START + 1   # 1-based
    year         = src_to_year(m)
    rev          = get(ws_pl, PL_REVENUE, m)
    cogs         = get(ws_pl, PL_COGS,    m)
    opex         = get(ws_pl, PL_OPEX,    m)
    tax          = get(ws_pl, PL_TAX,     m)
    net_inc      = get(ws_pl, PL_NETINC,  m)
    equity       = get(ws_cf, CF_EQUITY,  m)
    beg_cash     = get(ws_cf, CF_BEG,     m)
    end_cash     = get(ws_cf, CF_END,     m)
    creditors    = round((cogs + opex) * 0.10, 2)
    months_data.append({
        'src_m':       m,
        'tgt_m':       target_month,
        'year':        year,
        'revenue':     rev,
        'cor':         cogs,
        'opex':        opex,
        'tax':         tax,
        'net_inc':     net_inc,
        'equity':      equity,
        'beg_cash':    beg_cash,
        'end_cash':    end_cash,
        'creditors':   creditors,
    })

# ── Copy template ──────────────────────────────────────────────────────────
print(f'Copying template: {TPL} → {OUT}')
shutil.copy2(TPL, OUT)
wb_out = openpyxl.load_workbook(OUT)

# ── Styles ─────────────────────────────────────────────────────────────────
DARK_BLUE   = '1F3864'
MID_BLUE    = '2E75B6'
LIGHT_BLUE  = 'D6E4F0'
LIGHT_GREY  = 'F2F2F2'
WHITE       = 'FFFFFF'
GREEN       = '70AD47'
ORANGE      = 'ED7D31'

def hdr_font(bold=True, white=True, sz=10):
    return Font(name='Calibri', bold=bold, color=WHITE if white else '000000', size=sz)

def num_font(bold=False):
    return Font(name='Calibri', bold=bold, size=9)

def fill(hex_color):
    return PatternFill('solid', fgColor=hex_color)

def border_thin():
    s = Side(style='thin', color='BFBFBF')
    return Border(left=s, right=s, top=s, bottom=s)

FMT_EUR  = '#,##0.00'
FMT_INT  = '#,##0'
FMT_PCT  = '0.0%'

# ── Fill Monthly sheet ─────────────────────────────────────────────────────
print('Filling Monthly sheet...')
ws_mo = wb_out['Monthly']

# Header row R1 – style it
hdr_cols = {
    1: 'Year',
    2: 'Month',
    3: 'Revenue',
    4: 'Cash in from Revenue',
    5: 'Other Cash In (equity, debt, grants)',
    6: 'CoR Cash Out',
    7: 'Operating Cash Out (incl R&D)',
    8: 'Interest',
    9: 'Tax',
    10: 'Founder Bonus',
    11: 'Opening Cash',
    12: 'Closing Cash',
    13: 'Debt Outstanding',
    14: 'Debtors',
    15: 'Creditors',
}
for c, label in hdr_cols.items():
    cell = ws_mo.cell(1, c)
    cell.value = label
    cell.font  = Font(name='Calibri', bold=True, color=WHITE, size=9)
    cell.fill  = fill(DARK_BLUE)
    cell.alignment = Alignment(horizontal='center', wrap_text=True)
    cell.border = border_thin()

# Column widths
ws_mo.column_dimensions['A'].width = 8
ws_mo.column_dimensions['B'].width = 8
for c in range(3, 16):
    ws_mo.column_dimensions[get_column_letter(c)].width = 16

# Data rows: 48 months → rows 2–49
for i, d in enumerate(months_data):
    row = i + 2
    bg = LIGHT_BLUE if (i % 2 == 0) else WHITE

    # Determine phase label for col A
    m = d['src_m']
    if m <= 16:
        phase_label = 'Pre-Seed'
    elif m <= 28:
        phase_label = 'Seed'
    elif m <= 40:
        phase_label = 'Series A'
    else:
        phase_label = 'Series B'

    data_row = [
        d['year'],           # A
        d['tgt_m'],          # B  Month number (1-48)
        d['revenue'],        # C  Revenue
        d['revenue'],        # D  Cash in from Revenue (= revenue, simplified)
        d['equity'],         # E  Other Cash In (equity)
        d['cor'],            # F  CoR
        d['opex'],           # G  OpEx
        0.0,                 # H  Interest
        d['tax'],            # I  Tax
        0.0,                 # J  Founder Bonus
        d['beg_cash'],       # K  Opening Cash
        d['end_cash'],       # L  Closing Cash
        0.0,                 # M  Debt
        d['revenue'],        # N  Debtors (1-month revenue)
        d['creditors'],      # O  Creditors (10% of costs)
    ]

    for c_idx, val in enumerate(data_row, 1):
        cell = ws_mo.cell(row, c_idx)
        cell.value  = val
        cell.fill   = fill(bg)
        cell.border = border_thin()
        cell.font   = num_font()
        if c_idx == 1:   # Year
            cell.alignment = Alignment(horizontal='center')
            cell.number_format = '0'
        elif c_idx == 2: # Month
            cell.alignment = Alignment(horizontal='center')
            cell.number_format = '0'
            # Add phase color hint in col A
            ws_mo.cell(row, 1).value = d['year']
        else:
            cell.alignment = Alignment(horizontal='right')
            cell.number_format = FMT_EUR

    # Phase coloring for col A
    phase_colors = {
        'Pre-Seed': '9DC3E6',
        'Seed':     'A9D18E',
        'Series A': 'FFD966',
        'Series B': 'F4B183',
    }
    ws_mo.cell(row, 1).fill = fill(phase_colors.get(phase_label, LIGHT_BLUE))
    ws_mo.cell(row, 1).font = Font(name='Calibri', bold=True, size=9, color='000000')

# ── Phase legend below data ────────────────────────────────────────────────
legend_row = 51
ws_mo.cell(legend_row, 1).value = 'Phase legend:'
ws_mo.cell(legend_row, 1).font  = Font(bold=True, size=9)
legend_items = [
    ('Pre-Seed', '9DC3E6', 'M5–M16 (Aug 2026–Jul 2027)'),
    ('Seed',     'A9D18E', 'M17–M28 (Aug 2027–Jul 2028)'),
    ('Series A', 'FFD966', 'M29–M40 (Aug 2028–Jul 2029)'),
    ('Series B', 'F4B183', 'M41–M52 (Aug 2029–Jul 2030)'),
]
for li, (name, color, desc) in enumerate(legend_items):
    r = legend_row + 1 + li
    ws_mo.cell(r, 1).value = name
    ws_mo.cell(r, 1).fill  = fill(color)
    ws_mo.cell(r, 1).font  = Font(bold=True, size=9)
    ws_mo.cell(r, 2).value = desc
    ws_mo.cell(r, 2).font  = Font(size=9)

# ── Fill Annual sheet ──────────────────────────────────────────────────────
print('Filling Annual sheet...')
ws_an = wb_out['Annual']

# Annual aggregation by calendar year
year_groups = {
    2026: [d for d in months_data if d['year'] == 2026],
    2027: [d for d in months_data if d['year'] == 2027],
    2028: [d for d in months_data if d['year'] == 2028],
    2029: [d for d in months_data if d['year'] == 2029],
    2030: [d for d in months_data if d['year'] == 2030],
}

# Style Annual header R1
an_headers = {
    1:  'Year',
    2:  'Revenue',
    3:  'Cost of Revenue',
    4:  'Gross Profit',
    5:  'Total Opex',
    6:  'EBIT',
    7:  'Interest',
    8:  'Tax',
    9:  'Founder Bonus',
    10: 'Grants/Equity In',
    11: 'Net Profit',
    12: 'Min Cash',
    13: 'Avg Monthly Gross Burn',
    14: 'Avg Monthly EBIT Burn',
    15: 'Cash (year-end)',
    16: 'Debtors',
    17: 'R&D Intangible',
    18: 'Creditors',
    19: 'Debt',
    20: 'Net Assets',
    21: 'Total Equity Investment',
    22: 'Cumulative Net Profit',
    23: 'Quick Ratio',
}
for c, label in an_headers.items():
    cell = ws_an.cell(1, c)
    cell.value = label
    cell.font  = Font(name='Calibri', bold=True, color=WHITE, size=9)
    cell.fill  = fill(DARK_BLUE)
    cell.alignment = Alignment(horizontal='center', wrap_text=True)
    cell.border = border_thin()
    ws_an.column_dimensions[get_column_letter(c)].width = 16
ws_an.column_dimensions['A'].width = 8

cumulative_net = 0.0
for yr_idx, (year, group) in enumerate(sorted(year_groups.items())):
    if not group:
        continue
    row = yr_idx + 2

    rev     = sum(d['revenue']  for d in group)
    cor     = sum(d['cor']      for d in group)
    opex    = sum(d['opex']     for d in group)
    tax     = sum(d['tax']      for d in group)
    equity  = sum(d['equity']   for d in group)
    net_inc = rev - cor - opex - tax
    gp      = rev - cor
    ebit    = gp - opex
    min_cash      = min(d['end_cash'] for d in group)
    end_cash_yr   = group[-1]['end_cash']
    debtors_yr    = group[-1]['revenue']
    creditors_yr  = group[-1]['creditors']
    n = len(group)
    avg_burn       = (cor + opex) / n if n > 0 else 0
    avg_ebit_burn  = min(0, ebit / n) if n > 0 else 0
    net_assets     = end_cash_yr + debtors_yr - creditors_yr
    quick_ratio    = (end_cash_yr / creditors_yr) if creditors_yr > 0 else 0
    cumulative_net += net_inc

    bg = LIGHT_BLUE if yr_idx % 2 == 0 else WHITE
    annual_row = [
        year, rev, cor, gp, opex, ebit, 0.0, tax, 0.0,
        equity, net_inc, min_cash, avg_burn, avg_ebit_burn,
        end_cash_yr, debtors_yr, 0.0, creditors_yr, 0.0,
        net_assets, equity, cumulative_net, quick_ratio,
    ]
    for c_idx, val in enumerate(annual_row, 1):
        cell = ws_an.cell(row, c_idx)
        cell.value  = val
        cell.fill   = fill(bg)
        cell.border = border_thin()
        cell.font   = num_font()
        if c_idx == 1:
            cell.alignment = Alignment(horizontal='center')
            cell.number_format = '0'
            cell.font = Font(name='Calibri', bold=True, size=9)
        elif c_idx == 23:  # Quick Ratio
            cell.number_format = '0.0x'
            cell.alignment = Alignment(horizontal='right')
        else:
            cell.alignment = Alignment(horizontal='right')
            cell.number_format = FMT_EUR

    # Highlight positive/negative net profit
    np_cell = ws_an.cell(row, 11)
    if net_inc < 0:
        np_cell.font = Font(name='Calibri', size=9, color='C00000', bold=True)
    else:
        np_cell.font = Font(name='Calibri', size=9, color='375623', bold=True)

    # Note months in year
    ws_an.cell(row, 1).value = f'{year}\n({n}m)'
    ws_an.cell(row, 1).alignment = Alignment(horizontal='center', wrap_text=True)
    ws_an.row_dimensions[row].height = 30

# ── Add info note at bottom of Annual ────────────────────────────────────
note_row = 9
ws_an.cell(note_row, 1).value = (
    'Note: Counting starts at Pre-Seed Phase (Source M5 = Target Month 1). '
    'Source: 260315_LFL_BM_Vorlage_normal_redacted_final.xlsx'
)
ws_an.cell(note_row, 1).font = Font(name='Calibri', size=8, italic=True, color='595959')
ws_an.merge_cells(f'A{note_row}:W{note_row}')

# ── Add summary tab "Merge_Info" ───────────────────────────────────────────
print('Creating Merge_Info sheet...')
if 'Merge_Info' in wb_out.sheetnames:
    del wb_out['Merge_Info']
ws_info = wb_out.create_sheet('Merge_Info')

info_rows = [
    ('LFL Financial Projections – Merge Info', None),
    ('', None),
    ('Erstellt:', datetime.now().strftime('%Y-%m-%d %H:%M')),
    ('Quelle:', '260315_LFL_BM_Vorlage_normal_redacted_final.xlsx'),
    ('Template:', 'C13_Template_financial_projections_neu.xlsx'),
    ('Output:', OUT),
    ('', None),
    ('Mapping-Regel:', 'Quelle M5 (Pre-Seed-Start, Aug 2026) = Ziel Monat 1'),
    ('Quelle M5–M52:', '48 Monate → Ziel Monthly Rows 2–49'),
    ('', None),
    ('Phasen-Abgrenzung (Quelle):', None),
    ('  Ideation:',  'M1–M4  (Apr–Jul 2026) – NICHT im Merge enthalten'),
    ('  Pre-Seed:',  'M5–M16 (Aug 2026–Jul 2027) → Ziel Month 1–12'),
    ('  Seed:',      'M17–M28 (Aug 2027–Jul 2028) → Ziel Month 13–24'),
    ('  Series A:',  'M29–M40 (Aug 2028–Jul 2029) → Ziel Month 25–36'),
    ('  Series B:',  'M41–M52 (Aug 2029–Jul 2030) → Ziel Month 37–48'),
    ('', None),
    ('Spalten-Mapping (Monthly):', None),
    ('  C – Revenue:',          '6_P&L R8  (TOTAL REVENUE)'),
    ('  D – Cash in Revenue:',  '6_P&L R8  (vereinfacht = Revenue)'),
    ('  E – Other Cash In:',    '7_BS_CF R9 (Equity Funding Received)'),
    ('  F – CoR Cash Out:',     '6_P&L R14 (TOTAL COGS)'),
    ('  G – OpEx Cash Out:',    '6_P&L R27 (TOTAL OPERATING EXPENSES)'),
    ('  H – Interest:',         '0 (nicht modelliert)'),
    ('  I – Tax:',              '6_P&L R35 (Income Tax)'),
    ('  J – Founder Bonus:',    '0'),
    ('  K – Opening Cash:',     '7_BS_CF R14 (Beginning Cash Balance)'),
    ('  L – Closing Cash:',     '7_BS_CF R15 (ENDING CASH BALANCE)'),
    ('  M – Debt:',             '0 (nicht modelliert)'),
    ('  N – Debtors:',          '6_P&L R8  (1 Monat Revenue)'),
    ('  O – Creditors:',        '(COGS + OpEx) × 10%'),
]

ws_info.column_dimensions['A'].width = 30
ws_info.column_dimensions['B'].width = 50

for r_idx, (label, val) in enumerate(info_rows, 1):
    ws_info.cell(r_idx, 1).value = label
    if val is not None:
        ws_info.cell(r_idx, 2).value = val
    if r_idx == 1:
        ws_info.cell(r_idx, 1).font = Font(name='Calibri', bold=True, size=12, color=DARK_BLUE)
    elif label.endswith(':') and not label.startswith('  '):
        ws_info.cell(r_idx, 1).font = Font(name='Calibri', bold=True, size=10)
    else:
        ws_info.cell(r_idx, 1).font = Font(name='Calibri', size=9)
        ws_info.cell(r_idx, 2).font = Font(name='Calibri', size=9)

# ── Save ───────────────────────────────────────────────────────────────────
print(f'Saving: {OUT}')
wb_out.save(OUT)

file_size = os.path.getsize(OUT) / 1024
print(f'✓ Fertig: {OUT}  ({file_size:.0f} KB)')
print(f'  Monthly: {N_MONTHS} Monate (Rows 2–{N_MONTHS+1})')
print(f'  Annual:  {len([g for g in year_groups.values() if g])} Jahre (2026–2030)')
print()
print('=== Quick-Check (erste 5 + letzte 3 Monate) ===')
for d in months_data[:5] + months_data[-3:]:
    print(f'  Src M{d["src_m"]:2d} → Tgt M{d["tgt_m"]:2d} ({d["year"]}): '
          f'Rev={d["revenue"]:>10,.0f}  OpEx={d["opex"]:>10,.0f}  EndCash={d["end_cash"]:>12,.0f}')
