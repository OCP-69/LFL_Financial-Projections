"""
Merge 260315_LFL_BM_Szen_normal_final.xlsx → C13_Template_financial_projections_ext_2026.xlsx
Output: C13_Template_financial_projections_ext_2026_v02.xlsx

Mapping rule:
  Source M5 (Pre-Seed start) = Target Month 1 (Year 1, Col C)
  Source M5–M16  → Year 1  Col C–N  (12 months)
  Source M17–M28 → Year 2  Col B–M  (12 months)
  Source M29–M40 → Year 3  Col B–M  (12 months)
  Source M41–M52 → Year 4  Col B–M  (12 months)
  Year 5: no source data → zeroed out

Strategy:
  - Clear Startup & Pre-Opening Costs sheet (zero out all amounts)
  - Override all monthly data cells (formulas & values) with actual source data
  - Preserve aggregate formula rows: R14, R20, R38, R44, R46, R49, R53, R55, R57, R59
  - Opening Cash chain (R9 formulas) and Closing Cash (R46) preserved throughout

Source rows used:
  6_P&L  R8:  Total Revenue
  6_P&L  R14: Total COGS
  5_Costs R13: Total Personnel (→ Staff R30)
  5_Costs R18: Total Technology (→ R&D Maintenance R29)
  5_Costs R24: Total Office/Büro (→ Rent R27)
  5_Costs R30: Total Professional Services (→ Consultants R25)
  5_Costs R35: Total Insurance & Bank (→ Insurance R26)
  5_Costs R40: Total Marketing & Sales (→ Marketing R37)
  5_Costs R45+R48: Total Other + Payment (→ Travel/Other R35)
  7_BS_CF R9:  Equity Funding (→ Investment Received R16)
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import shutil, os

SRC  = '260315_LFL_BM_Szen_normal_final.xlsx'
TPL  = 'C13_Template_financial_projections_ext_2026.xlsx'
OUT  = 'C13_Template_financial_projections_ext_2026_v02.xlsx'

# ── Load source ────────────────────────────────────────────────────────────────
print('Loading source data...')
wb_src = openpyxl.load_workbook(SRC, data_only=True)
ws_pl  = wb_src['6_P&L']
ws_cf  = wb_src['7_BS_CF']
ws_co  = wb_src['5_Costs']

def src_val(ws, row, src_month):
    """Get numeric value from source worksheet at given row and source month."""
    v = ws.cell(row, src_month + 1).value  # M1=col2, M5=col6
    return round(float(v), 2) if v else 0.0

# Collect all 48 source months (M5–M52) into a list of dicts
src_data = []
for m in range(5, 53):
    rev    = src_val(ws_pl, 8,  m)
    cogs   = src_val(ws_pl, 14, m)
    equity = src_val(ws_cf, 9,  m)
    staff  = src_val(ws_co, 13, m)   # Total Personnel
    tech   = src_val(ws_co, 18, m)   # Total Technology  → R&D Maintenance
    office = src_val(ws_co, 24, m)   # Total Office      → Rent
    prof   = src_val(ws_co, 30, m)   # Total Professional → Consultants
    ins    = src_val(ws_co, 35, m)   # Total Insurance   → Insurance
    mktg   = src_val(ws_co, 40, m)   # Total Marketing   → Marketing
    other  = src_val(ws_co, 45, m)   # Total Sonstige    → Travel
    pay    = src_val(ws_co, 48, m)   # Payment Processing → Travel (combined)
    src_data.append({
        'src_m':  m,
        'tgt_m':  m - 4,   # 1-based target month
        'rev':    rev,
        'cogs':   cogs,
        'equity': equity,
        'staff':  staff,
        'tech':   tech,
        'office': office,
        'prof':   prof,
        'ins':    ins,
        'mktg':   mktg,
        'other':  other + pay,  # combined into Travel/Other row
    })

print(f'  Loaded {len(src_data)} months (M5–M52)')
print(f'  M5  (tgt M1):  Rev={src_data[0]["rev"]:>10,.0f}  COGS={src_data[0]["cogs"]:>8,.0f}  Equity={src_data[0]["equity"]:>9,.0f}  Staff={src_data[0]["staff"]:>8,.0f}')
print(f'  M52 (tgt M48): Rev={src_data[47]["rev"]:>10,.0f}  COGS={src_data[47]["cogs"]:>8,.0f}  Equity={src_data[47]["equity"]:>9,.0f}  Staff={src_data[47]["staff"]:>8,.0f}')

# ── Copy template ──────────────────────────────────────────────────────────────
print(f'Copying {TPL} → {OUT}')
shutil.copy2(TPL, OUT)
wb = openpyxl.load_workbook(OUT)  # load WITH formulas (data_only=False)

# ── Helper: set a cell to a plain value (overrides any formula) ───────────────
def set_val(ws, row, col, value):
    """Override cell with a plain numeric value (0 = clear)."""
    ws.cell(row, col).value = value if value != 0 else 0

def is_formula(ws, row, col):
    v = ws.cell(row, col).value
    return v is not None and str(v).startswith('=')

# ── Step 1: Zero out Startup & Pre-Opening Costs sheet ───────────────────────
print('Clearing Startup & Pre-Opening Costs...')
ws_sp = wb['Startup & Pre-Opening Costs']
for r in range(2, ws_sp.max_row + 1):
    for c in range(2, ws_sp.max_column + 1):
        cell = ws_sp.cell(r, c)
        v = cell.value
        if v is not None and not str(v).startswith('='):
            cell.value = 0

# ── Step 2: Define year-sheet column mapping ─────────────────────────────────
# Year 1: Pre-Opening=colB(2), Month1=colC(3)...Month12=colN(14), Total=colO(15)
# Year 2: Month13=colB(2)...Month24=colM(13), Total=colN(14)
# Year 3: Month25=colB(2)...Month36=colM(13), Total=colN(14)
# Year 4: Month37=colB(2)...Month48=colM(13), Total=colN(14)
# Year 5: Month49=colB(2)...Month60=colM(13), Total=colN(14)

SHEET_MAP = {
    'Year 1 Cash Budget': {'start_tgt': 1,  'end_tgt': 12, 'col_offset': 2},  # Month1=col3
    'Year 2 Cash Budget': {'start_tgt': 13, 'end_tgt': 24, 'col_offset': 1},  # Month13=col2
    'Year 3 Cash Budget': {'start_tgt': 25, 'end_tgt': 36, 'col_offset': 1},
    'Year 4 Cash Budget': {'start_tgt': 37, 'end_tgt': 48, 'col_offset': 1},
    'Year 5 Cash Budget': {'start_tgt': 49, 'end_tgt': 60, 'col_offset': 1},
}

def tgt_to_col(tgt_month, col_offset):
    """Convert 1-based target month to column index within its year sheet."""
    m_in_yr = ((tgt_month - 1) % 12) + 1   # 1-12
    return m_in_yr + col_offset

# Row numbers in the year sheets (both Year 1 and Year 2-5 use same structure)
ROWS = {
    'revenue':     7,   # Revenue (= AvgSale×Customers formula, we override)
    'rev_cash':    10,  # Revenue (this month) cash in
    'rev_wk1':     11,  # Revenue 1 Week Credit this month
    'rev_wk1_lag': 12,  # Revenue 1 Week Credit last month
    'rev_mo':      13,  # Revenue 1 Month Credit
    'investment':  16,  # Investment Received
    'cor_now':     18,  # CoR (this month)
    'cor_cr':      19,  # CoR (1 month supplier credit)
    'r22':         22,  # Contracts/GPDR
    'consultants': 25,  # Consultants → Professional Services
    'insurance':   26,  # Insurance → Insurance & Bank
    'rent':        27,  # Rent → Office/Büro
    'rd_design':   28,  # R&D Design
    'rd_maint':    29,  # R&D Maintenance → Technology
    'staff':       30,  # Staff → Personnel
    'comms':       31,  # Communications
    'incorporation':32, # Incorporation Fees
    'web_design':  33,  # Website Design
    'web_maint':   34,  # Website Maintenance
    'travel':      35,  # Travel → Other/Sonstige + Payment
    'vehicle':     36,  # Vehicle Rental
    'marketing':   37,  # Marketing
    'equip':       40,  # Computer Equipment
    'comms_asset': 41,  # Communications (asset)
    'furniture':   42,  # Furniture
    'buildings':   43,  # Buildings
    'cor_pl':      51,  # Cost of Revenue (P&L section) → override formula
}

# ── Step 3: Fill all year sheets ─────────────────────────────────────────────
print('Filling year sheets...')

for sheet_name, cfg in SHEET_MAP.items():
    ws = wb[sheet_name]
    start, end, offset = cfg['start_tgt'], cfg['end_tgt'], cfg['col_offset']
    print(f'  {sheet_name}: target months {start}–{end}')

    # --- Zero out Pre-Opening column (col B=2) for Year 1 ---
    if sheet_name == 'Year 1 Cash Budget':
        pre_open_col = 2
        for r in list(ROWS.values()) + [3, 4, 5, 6]:
            cell = ws.cell(r, pre_open_col)
            if not is_formula(ws, r, pre_open_col):
                cell.value = 0
        # Pre-Opening Opening Cash
        ws.cell(9, 2).value = 0  # B9 = 0 (start with no cash)

    # --- Fill each month ---
    for tgt_m in range(start, end + 1):
        src_idx = tgt_m - 1   # 0-based index into src_data (M5=index0)
        col     = tgt_to_col(tgt_m, offset)

        if src_idx < len(src_data):
            d = src_data[src_idx]
        else:
            # Year 5 or beyond: zero everything
            d = {k: 0 for k in ['rev','cogs','equity','staff','tech','office','prof','ins','mktg','other']}

        # Revenue (R7): override formula with actual value
        ws.cell(ROWS['revenue'],     col).value = d['rev']
        # Revenue cash (R10): actual revenue = 100% cash collected this month
        ws.cell(ROWS['rev_cash'],    col).value = d['rev']
        # Credit rows: 0 (no credit terms in our model)
        ws.cell(ROWS['rev_wk1'],     col).value = 0
        ws.cell(ROWS['rev_wk1_lag'], col).value = 0
        ws.cell(ROWS['rev_mo'],      col).value = 0
        # Investment received
        ws.cell(ROWS['investment'],  col).value = d['equity']
        # Cost of Revenue: cash out this month (no supplier credit)
        ws.cell(ROWS['cor_now'],     col).value = d['cogs']
        ws.cell(ROWS['cor_cr'],      col).value = 0
        # Expense rows: override all with actual values
        ws.cell(ROWS['r22'],         col).value = 0
        ws.cell(ROWS['consultants'], col).value = d['prof']
        ws.cell(ROWS['insurance'],   col).value = d['ins']
        ws.cell(ROWS['rent'],        col).value = d['office']
        ws.cell(ROWS['rd_design'],   col).value = 0
        ws.cell(ROWS['rd_maint'],    col).value = d['tech']
        ws.cell(ROWS['staff'],       col).value = d['staff']
        ws.cell(ROWS['comms'],       col).value = 0
        ws.cell(ROWS['incorporation'],col).value = 0
        ws.cell(ROWS['web_design'],  col).value = 0
        ws.cell(ROWS['web_maint'],   col).value = 0
        ws.cell(ROWS['travel'],      col).value = d['other']
        ws.cell(ROWS['vehicle'],     col).value = 0
        ws.cell(ROWS['marketing'],   col).value = d['mktg']
        # Assets: 0 (hardware is renting = OpEx, already in staff/tech)
        ws.cell(ROWS['equip'],       col).value = 0
        ws.cell(ROWS['comms_asset'], col).value = 0
        ws.cell(ROWS['furniture'],   col).value = 0
        ws.cell(ROWS['buildings'],   col).value = 0
        # P&L Cost of Revenue: override formula (R49=R7 already correct)
        ws.cell(ROWS['cor_pl'],      col).value = d['cogs']

# ── Step 4: Add phase labels in Row 1 for Year 1 (Pre-Opening = Ideation note) ─
ws1 = wb['Year 1 Cash Budget']
# Update Pre-Opening label to note Ideation
ws1.cell(1, 2).value = 'Pre-Opening\n(Ideation\nM1–M4: 0)'

# ── Step 5: Quick-verify closing cash Year 1 Month 1 (should = 683,717) ──────
# C46 = C9 + C14 - C20 - C38 - C44 + C16
# C9=B46=0, C14=C10=0 (no revenue M5), C20=300 (cogs), C38=sum expenses, C44=0, C16=720000
d0 = src_data[0]
closing_manual = (0
                  + d0['rev']
                  - d0['cogs']
                  - (d0['staff'] + d0['tech'] + d0['office'] + d0['prof'] +
                     d0['ins'] + d0['mktg'] + d0['other'])
                  + d0['equity'])
print(f'  Verify M5→M1 closing: Revenue={d0["rev"]:.0f}  COGS={d0["cogs"]:.0f}  '
      f'Total Opex={d0["staff"]+d0["tech"]+d0["office"]+d0["prof"]+d0["ins"]+d0["mktg"]+d0["other"]:.0f}  '
      f'Equity={d0["equity"]:.0f}  → Closing≈{closing_manual:,.0f} (source={683717})')

# ── Save ───────────────────────────────────────────────────────────────────────
print(f'Saving {OUT}...')
wb.save(OUT)
kb = os.path.getsize(OUT) / 1024
print(f'✓ Fertig: {OUT}  ({kb:.0f} KB)')

print()
print('=== Phase Summary ===')
phases = [
    ('Pre-Seed',  1,  12),
    ('Seed',     13,  24),
    ('Series A', 25,  36),
    ('Series B', 37,  48),
]
for ph, s, e in phases:
    seg = src_data[s-1:e]
    rev   = sum(d['rev']   for d in seg)
    cogs  = sum(d['cogs']  for d in seg)
    opex  = sum(d['staff']+d['tech']+d['office']+d['prof']+d['ins']+d['mktg']+d['other'] for d in seg)
    eq    = sum(d['equity'] for d in seg)
    print(f'  {ph:10s} (M{s:2d}–M{e:2d}): Rev={rev:>12,.0f}  COGS={cogs:>10,.0f}  OpEx={opex:>12,.0f}  Equity={eq:>12,.0f}')
