"""
Update LFL_BM_C13_Normal_redacted_v25_20260315.xlsx:
  Fill R&D column in all relevant sheets.

R&D definition:
  AI/ML APIs      → Category A (Produktentwicklungs-Tools)
  SaaS Tools      → Category A (Produktentwicklungs-Tools)
  Cloud Hosting   → Category C (Infrastruktur der Plattform)
  Sicherheit      → Category C (GDPR, IT-Security)

Treatment:
  - Monthly: add col P "R&D Cash Out" (monthly spend)
  - Annual col 17 "R&D Intangible": annual R&D investment (gross new spend per year)
  - BalanceSheet col 4 "R&D Intangible": cumulative net book value
       (gross cumulative R&D − accumulated amortization over 36 months)
  - BalanceSheet col 7 "Net Assets": updated to include R&D Intangible
  - Annual col 20 "Net Assets": updated to include R&D Intangible

Source: 260315_LFL_BM_Vorlage_normal_redacted_final.xlsx  (5_Costs rows 15/16/17/34)
Output: LFL_BM_C13_Normal_redacted_v26_20260315.xlsx
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import shutil, os

SRC_COSTS = '260315_LFL_BM_Vorlage_normal_redacted_final.xlsx'
SRC_V25   = 'LFL_BM_C13_Normal_redacted_v25_20260315.xlsx'
OUT       = 'LFL_BM_C13_Normal_redacted_v26_20260315.xlsx'

AMORT_MONTHS = 36   # 3-year straight-line amortization

# ── Styles ────────────────────────────────────────────────────────────────────
DARK_BLUE  = '1F3864'
MID_BLUE   = '2E75B6'
GREEN_DARK = '375623'
GREEN_LIGHT= 'E2EFDA'
BLUE_LIGHT = 'D6E4F0'
ORANGE_LIGHT='FCE4D6'
WHITE      = 'FFFFFF'

def fill_c(hex_c):
    return PatternFill('solid', fgColor=hex_c)

def thin_border():
    s = Side(style='thin', color='BFBFBF')
    return Border(left=s, right=s, top=s, bottom=s)

FMT_EUR = '#,##0.00'
FMT_INT = '#,##0'

# ── Step 1: Extract monthly R&D from source costs ────────────────────────────
print('Loading source costs...')
wb_costs = openpyxl.load_workbook(SRC_COSTS, data_only=True)
ws_costs = wb_costs['5_Costs']

# Source M5..M52 → target months 1..48
# Source col for month m = m+1
rd_monthly = []   # index 0 = target month 1 (source M5)
for m in range(5, 53):
    col    = m + 1
    cloud  = float(ws_costs.cell(15, col).value or 0)  # Cloud Hosting Basis
    aiml   = float(ws_costs.cell(16, col).value or 0)  # AI/ML APIs
    saas   = float(ws_costs.cell(17, col).value or 0)  # SaaS Tools & Lizenzen
    sec    = float(ws_costs.cell(34, col).value or 0)  # Sicherheit & Compliance
    rd     = cloud + aiml + saas + sec
    rd_monthly.append({
        'src_m':   m,
        'tgt_m':   m - 4,          # 1-based target month
        'cloud':   cloud,
        'aiml':    aiml,
        'saas':    saas,
        'sec':     sec,
        'total':   rd,
    })

print(f'  Monthly R&D collected: {len(rd_monthly)} months')
print(f'  Total R&D M5-M52: {sum(d["total"] for d in rd_monthly):,.2f} EUR')

# ── Step 2: Compute cumulative gross & net (after amortization) ───────────────
# gross_cum[i] = cumulative R&D through target month i+1 (0-based)
# amort_cum[i] = accumulated amortization through target month i+1
# net_nbv[i]   = gross_cum[i] - amort_cum[i]
gross_cum = []
amort_cum = []
net_nbv   = []

cum_gross = 0.0
cum_amort = 0.0
for i, d in enumerate(rd_monthly):
    cum_gross += d['total']
    # Amortization from all prior months (including this one):
    # Month j (0-based) contributes (min(i-j, 36) / 36) * rd[j] to amortization by month i
    # More efficiently: add 1/36 of each R&D spend for each month it has been on the books
    # At month i: amortize all previous spend by 1/36 per elapsed month, capped at 36
    # Recompute fully for accuracy:
    total_amort = 0.0
    for j, dj in enumerate(rd_monthly[:i+1]):
        elapsed = i - j       # months since cost was incurred (0 = same month)
        # Amortize starting month after incurrence → elapsed = 0 → 0 months amortized
        amort_months_so_far = min(elapsed, AMORT_MONTHS)
        total_amort += (amort_months_so_far / AMORT_MONTHS) * dj['total']
    gross_cum.append(cum_gross)
    amort_cum.append(total_amort)
    net_nbv.append(cum_gross - total_amort)

# Quick check
print(f'  Month 1 RD: {rd_monthly[0]["total"]:.2f}  NBV: {net_nbv[0]:.2f}')
print(f'  Month 12 RD gross cum: {gross_cum[11]:.2f}  amort: {amort_cum[11]:.2f}  NBV: {net_nbv[11]:.2f}')
print(f'  Month 48 RD gross cum: {gross_cum[47]:.2f}  amort: {amort_cum[47]:.2f}  NBV: {net_nbv[47]:.2f}')

# ── Step 3: Copy v25 → v26 ────────────────────────────────────────────────────
print(f'Copying {SRC_V25} → {OUT}')
shutil.copy2(SRC_V25, OUT)
wb = openpyxl.load_workbook(OUT)

# ── Helper to read Monthly rows ───────────────────────────────────────────────
ws_mo = wb['Monthly']

# Map target-month → row in Monthly sheet (month=col B, rows 2..49)
tgt_m_to_row = {}
for r in range(2, 50):
    m_val = ws_mo.cell(r, 2).value
    if m_val is not None:
        tgt_m_to_row[int(m_val)] = r

# ── Step 4: Add R&D column to Monthly (col P = 16) ───────────────────────────
print('Updating Monthly sheet...')
COL_RD_MO = 16   # column P

# Header
hdr_cell = ws_mo.cell(1, COL_RD_MO)
hdr_cell.value     = 'R&D Cash Out'
hdr_cell.font      = Font(name='Calibri', bold=True, color=WHITE, size=9)
hdr_cell.fill      = fill_c('4472C4')   # distinct blue for R&D
hdr_cell.alignment = Alignment(horizontal='center', wrap_text=True)
hdr_cell.border    = thin_border()
ws_mo.column_dimensions[get_column_letter(COL_RD_MO)].width = 16

# Also add sub-columns for breakdown (cols Q-T)
COL_CLOUD = 17
COL_AIML  = 18
COL_SAAS  = 19
COL_SEC   = 20
sub_headers = {
    COL_CLOUD: 'R&D: Cloud Hosting\n(Cat C)',
    COL_AIML:  'R&D: AI/ML APIs\n(Cat A)',
    COL_SAAS:  'R&D: SaaS Tools\n(Cat A)',
    COL_SEC:   'R&D: Security\n(Cat C)',
}
for col, label in sub_headers.items():
    cell = ws_mo.cell(1, col)
    cell.value     = label
    cell.font      = Font(name='Calibri', bold=True, color=WHITE, size=8)
    cell.fill      = fill_c('8EA9C1')
    cell.alignment = Alignment(horizontal='center', wrap_text=True)
    cell.border    = thin_border()
    ws_mo.column_dimensions[get_column_letter(col)].width = 15

# Fill data rows
for i, d in enumerate(rd_monthly):
    row = tgt_m_to_row.get(d['tgt_m'])
    if row is None:
        continue
    bg = BLUE_LIGHT if (i % 2 == 0) else WHITE

    for col, val in [
        (COL_RD_MO, d['total']),
        (COL_CLOUD, d['cloud']),
        (COL_AIML,  d['aiml']),
        (COL_SAAS,  d['saas']),
        (COL_SEC,   d['sec']),
    ]:
        cell = ws_mo.cell(row, col)
        cell.value         = round(val, 2)
        cell.font          = Font(name='Calibri', size=9)
        cell.fill          = fill_c('EBF3FB') if col == COL_RD_MO else fill_c(bg[1:] if bg.startswith('#') else bg)
        cell.border        = thin_border()
        cell.alignment     = Alignment(horizontal='right')
        cell.number_format = FMT_EUR

# ── Step 5: Fill Annual col 17 "R&D Intangible" ──────────────────────────────
print('Updating Annual sheet...')
ws_an = wb['Annual']

# Year → set of target months
year_to_tgt_months = {}
for i, d in enumerate(rd_monthly):
    # Determine year from Monthly sheet
    row  = tgt_m_to_row.get(d['tgt_m'])
    yr   = ws_mo.cell(row, 1).value if row else None
    if yr is not None:
        year_to_tgt_months.setdefault(yr, []).append(i)

# Annual sheet rows: R2=2026, R3=2027, R4=2028, R5=2029, R6=2030
# Find year → row mapping
year_to_an_row = {}
for r in range(2, 15):
    v = ws_an.cell(r, 1).value
    if v is not None:
        yr_str = str(v).split('\n')[0].strip()
        try:
            yr = int(yr_str)
            year_to_an_row[yr] = r
        except ValueError:
            pass

print(f'  Annual year→row: {year_to_an_row}')

# For each year: annual R&D spend = sum of monthly R&D in that year
# R&D Intangible (annual col 17) = annual gross R&D spend
# Net Assets (col 20) updated

for yr, an_row in sorted(year_to_an_row.items()):
    month_indices = year_to_tgt_months.get(yr, [])
    if not month_indices:
        continue

    # Annual R&D: gross new spend this year
    annual_rd_gross = sum(rd_monthly[i]['total'] for i in month_indices)

    # Year-end net book value: use the net_nbv of the last month in this year
    last_idx    = max(month_indices)
    yr_end_nbv  = net_nbv[last_idx]
    yr_end_gross= gross_cum[last_idx]

    # Fill col 17: annual R&D gross spend (new investment this year)
    cell_17 = ws_an.cell(an_row, 17)
    cell_17.value         = round(annual_rd_gross, 2)
    cell_17.font          = Font(name='Calibri', size=9, bold=True, color=GREEN_DARK)
    cell_17.fill          = fill_c(GREEN_LIGHT)
    cell_17.border        = thin_border()
    cell_17.alignment     = Alignment(horizontal='right')
    cell_17.number_format = FMT_EUR

    # Update col 20 "Net Assets" to include R&D Intangible (net book value)
    old_na = ws_an.cell(an_row, 20).value or 0
    new_na = round(float(old_na) + yr_end_nbv, 2)
    cell_20 = ws_an.cell(an_row, 20)
    cell_20.value         = new_na
    cell_20.font          = Font(name='Calibri', size=9, bold=True)
    cell_20.border        = thin_border()
    cell_20.alignment     = Alignment(horizontal='right')
    cell_20.number_format = FMT_EUR

    print(f'  {yr}: Annual R&D={annual_rd_gross:>10,.2f}  Gross cum={yr_end_gross:>12,.2f}  '
          f'NBV={yr_end_nbv:>12,.2f}  Updated Net Assets={new_na:>14,.2f}')

# ── Step 6: Fill BalanceSheet col 4 "R&D Intangible" ─────────────────────────
print('Updating BalanceSheet...')
ws_bs = wb['BalanceSheet']

# Find year → BS row
year_to_bs_row = {}
for r in range(2, 15):
    v = ws_bs.cell(r, 1).value
    if v is not None:
        try:
            yr = int(str(v).split('\n')[0].strip())
            year_to_bs_row[yr] = r
        except ValueError:
            pass

print(f'  BalanceSheet year→row: {year_to_bs_row}')

for yr, bs_row in sorted(year_to_bs_row.items()):
    month_indices = year_to_tgt_months.get(yr, [])
    if not month_indices:
        continue

    last_idx   = max(month_indices)
    yr_end_nbv = net_nbv[last_idx]

    # Fill col 4: R&D Intangible (net book value at year end)
    cell_4 = ws_bs.cell(bs_row, 4)
    cell_4.value         = round(yr_end_nbv, 2)
    cell_4.font          = Font(name='Calibri', size=9, bold=True, color=GREEN_DARK)
    cell_4.fill          = fill_c(GREEN_LIGHT)
    cell_4.border        = thin_border()
    cell_4.alignment     = Alignment(horizontal='right')
    cell_4.number_format = FMT_EUR

    # Update col 7 "Net Assets": add R&D Intangible
    old_na = ws_bs.cell(bs_row, 7).value or 0
    new_na = round(float(old_na) + yr_end_nbv, 2)
    cell_7 = ws_bs.cell(bs_row, 7)
    cell_7.value         = new_na
    cell_7.font          = Font(name='Calibri', size=9, bold=True)
    cell_7.border        = thin_border()
    cell_7.alignment     = Alignment(horizontal='right')
    cell_7.number_format = FMT_EUR

    print(f'  {yr}: R&D NBV={yr_end_nbv:>12,.2f}  Updated Net Assets={new_na:>14,.2f}')

# ── Step 7: Add R&D summary note to BalanceSheet ─────────────────────────────
last_bs_row = max(year_to_bs_row.values()) + 2
ws_bs.cell(last_bs_row, 1).value = (
    'R&D Intangible = Cumulative net book value (AI/ML APIs + SaaS Tools + '
    'Cloud Hosting Basis + Sicherheit & Compliance), amortized over 36 months. '
    'Cat A: AI/ML+SaaS (Produktentwicklung); Cat C: Cloud+Security (Infrastruktur/GDPR).'
)
ws_bs.cell(last_bs_row, 1).font = Font(name='Calibri', size=8, italic=True, color='595959')
try:
    ws_bs.merge_cells(f'A{last_bs_row}:J{last_bs_row}')
except:
    pass

# ── Step 8: Add R&D breakdown tab ────────────────────────────────────────────
print('Creating R&D_Breakdown sheet...')
if 'R&D_Breakdown' in wb.sheetnames:
    del wb['R&D_Breakdown']
ws_rd = wb.create_sheet('R&D_Breakdown')

# Header banner
ws_rd.merge_cells('A1:H1')
ws_rd['A1'] = 'R&D Kostenallokation — AI/ML · SaaS · Cloud · Security'
ws_rd['A1'].font      = Font(name='Calibri', bold=True, size=12, color=WHITE)
ws_rd['A1'].fill      = fill_c(DARK_BLUE)
ws_rd['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws_rd.row_dimensions[1].height = 26

ws_rd.merge_cells('A2:H2')
ws_rd['A2'] = (
    'Kategorie A: AI/ML APIs + SaaS Tools (Produktentwicklungs-Tools)  |  '
    'Kategorie C: Cloud Hosting Basis + Sicherheit & Compliance (Infrastruktur/GDPR)  |  '
    'Amortisation: 36 Monate'
)
ws_rd['A2'].font      = Font(name='Calibri', size=9, italic=True, color='595959')
ws_rd['A2'].alignment = Alignment(horizontal='center')
ws_rd.row_dimensions[2].height = 14

# Table header R4
rd_headers = [
    'Tgt Monat', 'Src Monat', 'Cloud Basis\n(Cat C)', 'AI/ML APIs\n(Cat A)',
    'SaaS Tools\n(Cat A)', 'Security\n(Cat C)', 'Total R&D', 'Kum. NBV\n(nach Amort.)'
]
for c_idx, hdr in enumerate(rd_headers, 1):
    cell = ws_rd.cell(4, c_idx)
    cell.value     = hdr
    cell.font      = Font(name='Calibri', bold=True, color=WHITE, size=9)
    cell.fill      = fill_c(MID_BLUE)
    cell.alignment = Alignment(horizontal='center', wrap_text=True)
    cell.border    = thin_border()

ws_rd.row_dimensions[4].height = 30
col_widths = [10, 10, 16, 16, 16, 16, 16, 18]
for c_idx, w in enumerate(col_widths, 1):
    ws_rd.column_dimensions[get_column_letter(c_idx)].width = w

# Data rows
PHASE_COLORS = {
    range(1, 13):  '9DC3E6',   # Pre-Seed
    range(13, 25): 'A9D18E',   # Seed
    range(25, 37): 'FFD966',   # Series A
    range(37, 49): 'F4B183',   # Series B
}
def phase_color(tgt_m):
    for rng, col in PHASE_COLORS.items():
        if tgt_m in rng:
            return col
    return 'FFFFFF'

# Phase subtotals
phase_defs = [
    ('Pre-Seed',  1,  12),
    ('Seed',     13,  24),
    ('Series A', 25,  36),
    ('Series B', 37,  48),
]

data_row = 5
for i, d in enumerate(rd_monthly):
    bg = phase_color(d['tgt_m'])
    row_vals = [
        d['tgt_m'], d['src_m'],
        d['cloud'], d['aiml'], d['saas'], d['sec'],
        d['total'], round(net_nbv[i], 2),
    ]
    for c_idx, val in enumerate(row_vals, 1):
        cell = ws_rd.cell(data_row, c_idx)
        cell.value  = val
        cell.fill   = fill_c(bg)
        cell.border = thin_border()
        if c_idx <= 2:
            cell.font      = Font(name='Calibri', size=9, bold=(c_idx == 1))
            cell.alignment = Alignment(horizontal='center')
            cell.number_format = '0'
        else:
            cell.font          = Font(name='Calibri', size=9)
            cell.alignment     = Alignment(horizontal='right')
            cell.number_format = FMT_EUR
    data_row += 1

# Phase subtotal rows
data_row += 1
ws_rd.cell(data_row, 1).value = 'PHASEN-ZUSAMMENFASSUNG'
ws_rd.cell(data_row, 1).font  = Font(name='Calibri', bold=True, size=10, color=DARK_BLUE)
ws_rd.merge_cells(f'A{data_row}:H{data_row}')
data_row += 1

ph_hdr = ['Phase', 'Monate', 'Cloud Basis (C)', 'AI/ML (A)', 'SaaS (A)', 'Security (C)', 'Total R&D', 'Phase-End NBV']
for c_idx, hdr in enumerate(ph_hdr, 1):
    cell = ws_rd.cell(data_row, c_idx)
    cell.value     = hdr
    cell.font      = Font(name='Calibri', bold=True, color=WHITE, size=9)
    cell.fill      = fill_c(DARK_BLUE)
    cell.border    = thin_border()
    cell.alignment = Alignment(horizontal='center')
data_row += 1

cat_a_total = cat_c_total = grand_total = 0.0
for ph_name, m_s, m_e in phase_defs:
    indices = [i for i, d in enumerate(rd_monthly) if m_s <= d['tgt_m'] <= m_e]
    ph_cloud = sum(rd_monthly[i]['cloud'] for i in indices)
    ph_aiml  = sum(rd_monthly[i]['aiml']  for i in indices)
    ph_saas  = sum(rd_monthly[i]['saas']  for i in indices)
    ph_sec   = sum(rd_monthly[i]['sec']   for i in indices)
    ph_total = sum(rd_monthly[i]['total'] for i in indices)
    ph_nbv   = net_nbv[max(indices)]
    cat_a    = ph_aiml + ph_saas
    cat_c    = ph_cloud + ph_sec
    cat_a_total += cat_a
    cat_c_total += cat_c
    grand_total += ph_total
    bg = phase_color(m_s)
    row_vals = [ph_name, f'M{m_s}–M{m_e}', ph_cloud, ph_aiml, ph_saas, ph_sec, ph_total, ph_nbv]
    for c_idx, val in enumerate(row_vals, 1):
        cell = ws_rd.cell(data_row, c_idx)
        cell.value  = val
        cell.fill   = fill_c(bg)
        cell.border = thin_border()
        if c_idx <= 2:
            cell.font      = Font(name='Calibri', size=9, bold=True)
            cell.alignment = Alignment(horizontal='center')
        else:
            cell.font          = Font(name='Calibri', size=9, bold=True)
            cell.alignment     = Alignment(horizontal='right')
            cell.number_format = FMT_EUR
    data_row += 1

# Grand total
row_vals = ['GESAMT', 'M1–M48', '', '', '', '', grand_total, net_nbv[-1]]
for c_idx, val in enumerate(row_vals, 1):
    cell = ws_rd.cell(data_row, c_idx)
    cell.value  = val if val != '' else None
    cell.fill   = fill_c(DARK_BLUE)
    cell.font   = Font(name='Calibri', bold=True, color=WHITE, size=10)
    cell.border = thin_border()
    cell.alignment = Alignment(horizontal='center' if c_idx <= 2 else 'right')
    if c_idx > 6 and val != '':
        cell.number_format = FMT_EUR
data_row += 2

# Category totals
ws_rd.cell(data_row, 1).value = 'Kategorie A (Produktentwicklung):'
ws_rd.cell(data_row, 1).font  = Font(bold=True, size=9)
ws_rd.cell(data_row, 2).value = round(cat_a_total, 2)
ws_rd.cell(data_row, 2).font  = Font(bold=True, size=9, color='1F497D')
ws_rd.cell(data_row, 2).number_format = FMT_EUR
ws_rd.cell(data_row, 3).value = '(AI/ML APIs + SaaS Tools)'
ws_rd.cell(data_row, 3).font  = Font(size=9, italic=True, color='595959')
data_row += 1
ws_rd.cell(data_row, 1).value = 'Kategorie C (Infrastruktur/GDPR):'
ws_rd.cell(data_row, 1).font  = Font(bold=True, size=9)
ws_rd.cell(data_row, 2).value = round(cat_c_total, 2)
ws_rd.cell(data_row, 2).font  = Font(bold=True, size=9, color='1F497D')
ws_rd.cell(data_row, 2).number_format = FMT_EUR
ws_rd.cell(data_row, 3).value = '(Cloud Hosting Basis + Sicherheit & Compliance)'
ws_rd.cell(data_row, 3).font  = Font(size=9, italic=True, color='595959')

# ── Save ──────────────────────────────────────────────────────────────────────
print(f'Saving: {OUT}')
wb.save(OUT)
size_kb = os.path.getsize(OUT) / 1024
print(f'✓ Fertig: {OUT}  ({size_kb:.0f} KB)')
print()
print('=== R&D Summary ===')
print(f'  Total R&D M5-M52:  {grand_total:>12,.2f} EUR')
print(f'  Final NBV (M52):   {net_nbv[-1]:>12,.2f} EUR')
print(f'  Category A:        {cat_a_total:>12,.2f} EUR  ({cat_a_total/grand_total*100:.1f}%)')
print(f'  Category C:        {cat_c_total:>12,.2f} EUR  ({cat_c_total/grand_total*100:.1f}%)')
