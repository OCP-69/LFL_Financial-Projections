"""
create_annual_income_statement.py
Generates C13_Fin Proj_Ann_Income_Statement.xlsx from the LFL Financial Model inputs.

Reads: Input/260304_LFL_SaaS_Startup_Financial_Model_v0.4.xlsx
Writes: scenarios/C13_Fin Proj_Ann_Income_Statement.xlsx

The script fully replicates the Excel formula logic in Python, then aggregates
the 52 monthly periods into annual columns (Year 1–4, partial Year 5).
"""

import math
import datetime
import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from pathlib import Path

# ---------------------------------------------------------------------------
# 1. LOAD INPUTS
# ---------------------------------------------------------------------------
BASE_DIR = Path(__file__).parent.parent
INPUT_FILE = BASE_DIR / "Input" / "260304_LFL_SaaS_Startup_Financial_Model_v0.4.xlsx"
OUTPUT_FILE = BASE_DIR / "scenarios" / "C13_Fin Proj_Ann_Income_Statement.xlsx"

wb_in = openpyxl.load_workbook(INPUT_FILE, data_only=True)
ws_inp = wb_in["Inputs"]
ws_sbx = wb_in["00_Input_Sandbox"]

def inp(row):
    return ws_inp.cell(row=row, column=2).value

def sbx_col(scenario="gering"):
    """Return the sandbox column index (4=gering, 5=normal, 6=stark)."""
    for col in range(4, 7):
        if str(ws_sbx.cell(row=1, column=col).value).lower() == scenario.lower():
            return col
    return 4  # default gering

def sbx(row, scenario="gering"):
    return ws_sbx.cell(row=row, column=sbx_col(scenario)).value

# Active scenario
ACTIVE_SCENARIO = str(ws_sbx.cell(row=1, column=2).value).strip()

# General
TAX_RATE         = float(inp(7))  # 0.30
START_DATE       = inp(5)         # 2026-04-01
if isinstance(START_DATE, str):
    START_DATE = datetime.datetime.strptime(START_DATE, "%Y-%m-%d")

# Financing
IDEATION_AMT     = float(inp(10));  IDEATION_MO    = int(inp(11))
PRESEED_AMT      = float(inp(12));  PRESEED_MO     = int(inp(13))
SEED_AMT         = float(inp(14));  SEED_MO        = int(inp(15))
SERIESA_AMT      = float(inp(16));  SERIESA_MO     = int(inp(17))

# Revenue – Sandbox overrides Inputs rows 20 & 22
SEAT_PRICE_YR    = float(sbx(6, ACTIVE_SCENARIO)) * 12   # €/month * 12 -> €/year
PRICE_INCREASE   = float(inp(21))    # 0.08 per year
FIRST_CUST_MO    = int(sbx(5, ACTIVE_SCENARIO))           # first paying month
INIT_SEATS       = float(inp(23))    # 5
SEAT_GROWTH_MO   = float(inp(24))    # 0.05 per month
ENT_START_MO     = int(inp(25))      # 24
ENT_AVG_ACV      = float(inp(26))    # 150 000
ENT_DEALS_QTR    = float(inp(27))    # 1 deal per quarter
CHURN_ANNUAL     = float(inp(28))    # 0.08
NRR              = float(inp(29))    # 1.18

# Salaries
SAL = {
    "ceo":      float(inp(32)),
    "cto":      float(inp(33)),
    "cco":      float(inp(34)),
    "senior":   float(inp(35)),
    "junior":   float(inp(36)),
    "ml":       float(inp(37)),
    "pm":       float(inp(38)),
    "sales":    float(inp(39)),
    "mktg":     float(inp(40)),
    "cs":       float(inp(41)),
    "office":   float(inp(42)),
    "finance":  float(inp(43)),
}
SAL_INCREASE     = float(inp(44))   # 0.05
LOHNNEBENKOSTEN  = float(inp(45))   # 0.21

# Hiring plan – column F = effective entry month (accounts for KI strategy)
def eff_month(row):
    v = ws_inp.cell(row=row, column=6).value
    try:
        return int(float(v))
    except (TypeError, ValueError):
        return 99

HIRE = {
    "senior_1":  eff_month(48),
    "senior_2":  eff_month(49),
    "junior_1":  eff_month(50),
    "ml_1":      eff_month(51),
    "ml_2":      eff_month(52),
    "sales_1":   eff_month(53),
    "sales_2":   eff_month(54),
    "cs_1":      eff_month(55),
    "mktg_1":    eff_month(56),
    "pm_1":      eff_month(57),
    "office_1":  eff_month(58),
    "senior_3":  eff_month(59),
    "senior_4":  eff_month(60),
    "junior_2":  eff_month(61),
    "ml_3":      eff_month(62),
    "sales_3":   eff_month(63),
    "cs_2":      eff_month(64),
    "mktg_2":    eff_month(65),
    "senior_5":  eff_month(66),
    "junior_3":  eff_month(67),
    "ml_4":      eff_month(68),
    "pm_2":      eff_month(69),
    "sales_4":   eff_month(70),
    "cs_3":      eff_month(71),
    "finance_1": eff_month(72),
    "senior_6":  eff_month(73),
    "senior_7":  eff_month(74),
    "junior_4":  eff_month(75),
    "ml_5":      eff_month(76),
    "sales_5":   eff_month(77),
    "cs_4":      eff_month(78),
    "mktg_3":    eff_month(79),
}

# Tech
CLOUD_BASE       = float(inp(82))   # 1200
CLOUD_PER_SEAT   = float(inp(83))   # 50
AIML_BASE        = float(inp(84))   # 1000
AIML_GROWTH_MO   = float(inp(85))   # 0.05
SAAS_BASE        = float(inp(86))   # 400
SAAS_PER_HC      = float(inp(87))   # 100
SW_LICENSES      = float(inp(88))   # 5000
SECURITY         = float(inp(89))   # 3500

# Hardware
LAPTOP           = float(inp(92))   # 2500
MONITOR          = float(inp(93))   # 800
REPLACE_CYCLE    = float(inp(94))   # 36
IT_OTHER_YR      = float(inp(95))   # 2000

# Office
RENT1            = float(inp(98))   # 1500
UPGRADE_MO       = int(inp(99))     # 18
RENT2            = float(inp(100))  # 4500
UTILITIES        = float(inp(101))  # 300
INTERNET         = float(inp(102))  # 150
OFFICE_SUPPLIES_PER_HC = float(inp(104))  # 30

# Professional Services
LEGAL_BASE       = float(inp(107))  # 8000 per year -> /12 monthly
LEGAL_FINANCING  = float(inp(108))  # 15000 per financing event
TAX_ADVISOR      = float(inp(109))  # 800/month
AUDITOR_YR       = float(inp(110))  # 12000
AUDITOR_START    = int(inp(111))    # 17
CONSULTANT_YR    = float(inp(112))  # 5000

# Insurance & Bank
DO_INS_YR        = float(inp(115))  # 3000
LIAB_INS_YR      = float(inp(116))  # 1500
CYBER_INS_YR     = float(inp(117))  # 2000
BANK_FEES        = float(inp(118))  # 50
PAYMENT_PROC_PCT = float(inp(119))  # 0.025

# Marketing
ADS_INIT         = float(inp(122))  # 500
ADS_GROWTH       = float(inp(123))  # 0.05
CONTENT_SEO      = float(inp(124))  # 1500
EVENTS_YR        = float(inp(125))  # 25000
SALES_TOOLS      = float(inp(126))  # 300
SALES_COMM_PCT   = float(inp(127))  # 0.10
SALES_TRAVEL_PER_MO = float(inp(128))  # 500 per sales rep/month

# Other
TRAVEL_PER_HC_YR = float(inp(131))  # 2000
TRAINING_PER_HC_YR = float(inp(132))  # 1500
TEAM_EVENTS_PER_HC_YR = float(inp(133))  # 1000
CONTINGENCY_PCT  = float(inp(134))  # 0.05
DEPRECIATION_YRS = float(inp(135))  # 3

# Financing events lookup
FINANCING_EVENTS = {
    IDEATION_MO: IDEATION_AMT,
    PRESEED_MO:  PRESEED_AMT,
    SEED_MO:     SEED_AMT,
    SERIESA_MO:  SERIESA_AMT,
}

# ---------------------------------------------------------------------------
# 2. COMPUTE MONTHLY MODEL (52 months)
# ---------------------------------------------------------------------------
N_MONTHS = 52

def month_date(m):
    """Return (year, month) for model month m (1-based)."""
    d = START_DATE + datetime.timedelta(days=30.44 * (m - 1))
    return d.year, d.month

def year_label(m):
    y, mo = month_date(m)
    return y

# --- Headcount per month ---
def headcount_per_month():
    """Returns dict: month -> {role: count}. Founders always present from M1."""
    hc = []
    for m in range(1, N_MONTHS + 1):
        h = {
            "ceo": 1, "cto": 1, "cco": 1,
            "senior": 0, "junior": 0, "ml": 0,
            "pm": 0, "sales": 0, "mktg": 0, "cs": 0,
            "office": 0, "finance": 0,
        }
        senior_roles = [k for k in HIRE if k.startswith("senior_")]
        for k in senior_roles:
            if HIRE[k] <= m:
                h["senior"] += 1
        for k in [k for k in HIRE if k.startswith("junior_")]:
            if HIRE[k] <= m:
                h["junior"] += 1
        for k in [k for k in HIRE if k.startswith("ml_")]:
            if HIRE[k] <= m:
                h["ml"] += 1
        for k in [k for k in HIRE if k.startswith("pm_")]:
            if HIRE[k] <= m:
                h["pm"] += 1
        for k in [k for k in HIRE if k.startswith("sales_")]:
            if HIRE[k] <= m:
                h["sales"] += 1
        for k in [k for k in HIRE if k.startswith("mktg_")]:
            if HIRE[k] <= m:
                h["mktg"] += 1
        for k in [k for k in HIRE if k.startswith("cs_")]:
            if HIRE[k] <= m:
                h["cs"] += 1
        if HIRE.get("office_1", 99) <= m:
            h["office"] += 1
        if HIRE.get("finance_1", 99) <= m:
            h["finance"] += 1
        hc.append(h)
    return hc

def total_hc(h):
    return sum(h.values())

# Salary for a given role considering annual increases (applied at start of each year)
def annual_salary(role, year_in_model):
    """year_in_model starts at 1."""
    base = SAL[role]
    return base * ((1 + SAL_INCREASE) ** (year_in_model - 1))

def monthly_gross(role, m, hc_count):
    """Monthly gross salary for headcount."""
    year_num = math.ceil(m / 12)
    return (annual_salary(role, year_num) / 12) * hc_count

def compute_monthly():
    hc_by_month = headcount_per_month()

    # Track state
    active_seats = [0.0] * (N_MONTHS + 1)   # index 0 unused
    sub_rev      = [0.0] * (N_MONTHS + 1)
    ent_rev      = [0.0] * (N_MONTHS + 1)
    total_rev    = [0.0] * (N_MONTHS + 1)

    active_ent   = 0.0

    churn_mo     = CHURN_ANNUAL / 12

    seat_price_mo = SEAT_PRICE_YR / 12   # monthly price per seat

    for m in range(1, N_MONTHS + 1):
        year_num = math.ceil(m / 12)
        # Price increases annually
        price_factor = (1 + PRICE_INCREASE) ** (year_num - 1)
        price_mo = seat_price_mo * price_factor

        if m < FIRST_CUST_MO:
            active_seats[m] = 0.0
        elif m == FIRST_CUST_MO:
            active_seats[m] = INIT_SEATS
        else:
            prev = active_seats[m - 1]
            new_seats = prev * SEAT_GROWTH_MO
            churned   = prev * churn_mo
            active_seats[m] = prev + new_seats - churned

        sub_rev[m] = active_seats[m] * price_mo

        # Enterprise revenue
        if m >= ENT_START_MO:
            # New deals each quarter
            if (m - ENT_START_MO) % 3 == 0:
                active_ent += ENT_DEALS_QTR
            ent_rev[m] = active_ent * (ENT_AVG_ACV / 12)
        else:
            ent_rev[m] = 0.0

        total_rev[m] = sub_rev[m] + ent_rev[m]

    # Personnel costs
    personnel = [0.0] * (N_MONTHS + 1)
    for m in range(1, N_MONTHS + 1):
        h = hc_by_month[m - 1]
        year_num = math.ceil(m / 12)
        gross = sum(monthly_gross(role, m, cnt) for role, cnt in h.items())
        personnel[m] = gross * (1 + LOHNNEBENKOSTEN)

    # Technology costs (total monthly opex tech)
    tech = [0.0] * (N_MONTHS + 1)
    for m in range(1, N_MONTHS + 1):
        hc = total_hc(hc_by_month[m - 1])
        cloud = CLOUD_BASE + CLOUD_PER_SEAT * active_seats[m]
        aiml  = AIML_BASE * ((1 + AIML_GROWTH_MO) ** (m - 1))
        saas  = SAAS_BASE + SAAS_PER_HC * hc
        tech[m] = cloud + aiml + saas + SW_LICENSES + SECURITY

    # Office costs
    office_cost = [0.0] * (N_MONTHS + 1)
    for m in range(1, N_MONTHS + 1):
        hc = total_hc(hc_by_month[m - 1])
        rent = RENT2 if m >= UPGRADE_MO else RENT1
        supplies = OFFICE_SUPPLIES_PER_HC * hc
        office_cost[m] = rent + UTILITIES + INTERNET + supplies

    # Professional Services
    prof = [0.0] * (N_MONTHS + 1)
    fin_months = set(FINANCING_EVENTS.keys())
    for m in range(1, N_MONTHS + 1):
        legal = LEGAL_BASE / 12
        if m in fin_months:
            legal += LEGAL_FINANCING
        audit = (AUDITOR_YR / 12) if m >= AUDITOR_START else 0.0
        prof[m] = legal + TAX_ADVISOR + audit + (CONSULTANT_YR / 12)

    # Insurance & Bank
    insbank = [0.0] * (N_MONTHS + 1)
    for m in range(1, N_MONTHS + 1):
        insbank[m] = (DO_INS_YR + LIAB_INS_YR + CYBER_INS_YR) / 12 + BANK_FEES

    # Marketing & Sales
    mktg_cost = [0.0] * (N_MONTHS + 1)
    for m in range(1, N_MONTHS + 1):
        h = hc_by_month[m - 1]
        ads = ADS_INIT * ((1 + ADS_GROWTH) ** (m - 1))
        events = EVENTS_YR / 12
        commission = SALES_COMM_PCT * total_rev[m]
        travel = SALES_TRAVEL_PER_MO * h["sales"]
        mktg_cost[m] = ads + CONTENT_SEO + events + SALES_TOOLS + commission + travel

    # Other OpEx
    other_opex = [0.0] * (N_MONTHS + 1)
    for m in range(1, N_MONTHS + 1):
        hc = total_hc(hc_by_month[m - 1])
        travel   = (TRAVEL_PER_HC_YR / 12) * hc
        training = (TRAINING_PER_HC_YR / 12) * hc
        events   = (TEAM_EVENTS_PER_HC_YR / 12) * hc
        hw       = (LAPTOP + MONITOR) * _new_hires(m, hc_by_month) + IT_OTHER_YR / 12
        other_opex[m] = travel + training + events + hw

    # COGS
    cogs_cloud = [0.0] * (N_MONTHS + 1)
    cogs_aiml  = [0.0] * (N_MONTHS + 1)
    cogs_pay   = [0.0] * (N_MONTHS + 1)
    total_cogs = [0.0] * (N_MONTHS + 1)
    for m in range(1, N_MONTHS + 1):
        cogs_cloud[m] = CLOUD_PER_SEAT * active_seats[m]
        cogs_aiml[m]  = AIML_BASE * ((1 + AIML_GROWTH_MO) ** (m - 1))
        cogs_pay[m]   = PAYMENT_PROC_PCT * total_rev[m]
        total_cogs[m] = cogs_cloud[m] + cogs_aiml[m] + cogs_pay[m]

    # Gross Profit
    gross_profit = [0.0] * (N_MONTHS + 1)
    for m in range(1, N_MONTHS + 1):
        gross_profit[m] = total_rev[m] - total_cogs[m]

    # Contingency buffer on total opex (excl COGS)
    total_opex_pre = [0.0] * (N_MONTHS + 1)
    contingency    = [0.0] * (N_MONTHS + 1)
    total_opex     = [0.0] * (N_MONTHS + 1)
    for m in range(1, N_MONTHS + 1):
        base = personnel[m] + tech[m] + office_cost[m] + prof[m] + insbank[m] + mktg_cost[m] + other_opex[m]
        contingency[m]    = base * CONTINGENCY_PCT
        total_opex[m]     = base + contingency[m]
        total_opex_pre[m] = base

    # EBITDA
    ebitda = [0.0] * (N_MONTHS + 1)
    for m in range(1, N_MONTHS + 1):
        ebitda[m] = gross_profit[m] - total_opex[m]

    # Depreciation (hardware / IT over 3 years)
    depr = [0.0] * (N_MONTHS + 1)
    for m in range(1, N_MONTHS + 1):
        # Hardware for new hires at time of joining, depreciated over 36 months
        new_hw_cost = (LAPTOP + MONITOR) * _new_hires(m, hc_by_month)
        # Spread depreciation from all past purchases
        depr_total = 0.0
        for past_m in range(1, m + 1):
            nh = _new_hires(past_m, hc_by_month)
            hw = (LAPTOP + MONITOR) * nh + IT_OTHER_YR / 12
            months_old = m - past_m + 1
            if months_old <= DEPRECIATION_YRS * 12:
                depr_total += hw / (DEPRECIATION_YRS * 12)
        depr[m] = depr_total

    # EBIT, EBT, Tax, Net Income
    ebit       = [0.0] * (N_MONTHS + 1)
    ebt        = [0.0] * (N_MONTHS + 1)
    tax        = [0.0] * (N_MONTHS + 1)
    net_income = [0.0] * (N_MONTHS + 1)
    for m in range(1, N_MONTHS + 1):
        ebit[m] = ebitda[m] - depr[m]
        ebt[m]  = ebit[m]   # simplified: no interest
        tax[m]  = max(0.0, ebt[m] * TAX_RATE)
        net_income[m] = ebt[m] - tax[m]

    return {
        "active_seats":  active_seats,
        "sub_rev":       sub_rev,
        "ent_rev":       ent_rev,
        "total_rev":     total_rev,
        "cogs_cloud":    cogs_cloud,
        "cogs_aiml":     cogs_aiml,
        "cogs_pay":      cogs_pay,
        "total_cogs":    total_cogs,
        "gross_profit":  gross_profit,
        "personnel":     personnel,
        "tech":          tech,
        "office_cost":   office_cost,
        "prof":          prof,
        "insbank":       insbank,
        "mktg_cost":     mktg_cost,
        "other_opex":    other_opex,
        "contingency":   contingency,
        "total_opex":    total_opex,
        "ebitda":        ebitda,
        "depr":          depr,
        "ebit":          ebit,
        "ebt":           ebt,
        "tax":           tax,
        "net_income":    net_income,
        "hc_by_month":   hc_by_month,
    }


def _new_hires(m, hc_by_month):
    """Count employees hired in month m (headcount delta)."""
    if m == 1:
        return total_hc(hc_by_month[0])
    return max(0, total_hc(hc_by_month[m - 1]) - total_hc(hc_by_month[m - 2]))


def aggregate_annual(data):
    """
    Aggregate monthly arrays into annual columns.
    Returns list of dicts with keys matching data, one per year period.
    Percentage lines (margin %) are computed from summed values.
    """
    # Determine year groupings
    # Year 1: M1-M12, Year 2: M13-M24, Year 3: M25-M36, Year 4: M37-M48, Year 5: M49-52
    periods = [
        ("Year 1 (FY2026/27)", range(1, 13)),
        ("Year 2 (FY2027/28)", range(13, 25)),
        ("Year 3 (FY2028/29)", range(25, 37)),
        ("Year 4 (FY2029/30)", range(37, 49)),
        ("Year 5 (4 mo. FY2030)", range(49, 53)),
    ]

    # Also compute total 52-month and 4-year full columns
    sum_keys = [
        "sub_rev", "ent_rev", "total_rev",
        "cogs_cloud", "cogs_aiml", "cogs_pay", "total_cogs",
        "gross_profit",
        "personnel", "tech", "office_cost", "prof", "insbank",
        "mktg_cost", "other_opex", "contingency", "total_opex",
        "ebitda", "depr", "ebit", "ebt", "tax", "net_income",
    ]

    results = []
    for label, months in periods:
        row = {"label": label}
        for k in sum_keys:
            row[k] = sum(data[k][m] for m in months)
        # Derived ratios
        rev = row["total_rev"]
        row["gross_margin_pct"]  = row["gross_profit"]  / rev if rev else 0
        row["ebitda_margin_pct"] = row["ebitda"]        / rev if rev else 0
        row["net_margin_pct"]    = row["net_income"]    / rev if rev else 0
        # Headcount at end of period
        last_m = list(months)[-1]
        row["headcount_eop"] = total_hc(data["hc_by_month"][last_m - 1])
        results.append(row)

    # Totals column (full 52 months)
    total = {"label": "Total (52 mo.)"}
    for k in sum_keys:
        total[k] = sum(data[k][m] for m in range(1, N_MONTHS + 1))
    rev = total["total_rev"]
    total["gross_margin_pct"]  = total["gross_profit"]  / rev if rev else 0
    total["ebitda_margin_pct"] = total["ebitda"]        / rev if rev else 0
    total["net_margin_pct"]    = total["net_income"]    / rev if rev else 0
    total["headcount_eop"]     = total_hc(data["hc_by_month"][N_MONTHS - 1])
    results.append(total)

    return results


# ---------------------------------------------------------------------------
# 3. WRITE EXCEL OUTPUT
# ---------------------------------------------------------------------------
SCENARIO_LABEL_MAP = {
    "gering": "Konservativ (Auto-Fokus)",
    "normal": "Hybrid",
    "stark":  "Aggressiv (Packaging-Fokus)",
}

# Colour palette (LFL branding)
COL_HEADER_BG  = "1F3864"   # dark navy
COL_HEADER_FG  = "FFFFFF"
COL_SECTION_BG = "2E75B6"   # medium blue
COL_SECTION_FG = "FFFFFF"
COL_SUBHEAD_BG = "D6E4F0"   # light blue
COL_ALT_BG     = "F2F7FC"   # very light blue
COL_TOTAL_BG   = "1F3864"   # dark navy (total column)
COL_TOTAL_FG   = "FFFFFF"
COL_POS        = "1E6B42"   # green for positive margins
COL_NEG        = "C00000"   # red for negative

def make_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def make_font(bold=False, color="000000", size=10, italic=False):
    return Font(bold=bold, color=color, size=size, italic=italic)

def thin_border():
    s = Side(style="thin", color="BDD7EE")
    return Border(left=s, right=s, top=s, bottom=s)

def bottom_border():
    s_bot = Side(style="medium", color="2E75B6")
    s_thin = Side(style="thin", color="BDD7EE")
    return Border(left=s_thin, right=s_thin, bottom=s_bot)


def write_excel(annual, monthly_data):
    wb = openpyxl.Workbook()

    # ---- Sheet 1: Annual Income Statement ----
    ws = wb.active
    ws.title = "Annual Income Statement"

    # Column widths
    ws.column_dimensions["A"].width = 38
    for i, ltr in enumerate(["B", "C", "D", "E", "F", "G"], start=1):
        ws.column_dimensions[ltr].width = 18

    # ---- TITLE BLOCK ----
    ws.merge_cells("A1:G1")
    ws["A1"] = "LoopForgeLab GmbH – Annual Income Statement"
    ws["A1"].font = Font(bold=True, size=14, color=COL_HEADER_FG)
    ws["A1"].fill = make_fill(COL_HEADER_BG)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:G2")
    scenario_name = SCENARIO_LABEL_MAP.get(ACTIVE_SCENARIO, ACTIVE_SCENARIO)
    ws["A2"] = f"Szenario: {scenario_name}  |  Modell-Start: April 2026  |  Erstellt: {datetime.date.today().strftime('%d.%m.%Y')}"
    ws["A2"].font = Font(italic=True, size=9, color="AAAAAA")
    ws["A2"].fill = make_fill("0D1F3C")
    ws["A2"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 16

    ws.row_dimensions[3].height = 6  # spacer

    # ---- COLUMN HEADERS (row 4) ----
    ws.row_dimensions[4].height = 30
    headers = ["Position"] + [p["label"] for p in annual]
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col_idx, value=h)
        cell.fill = make_fill(COL_HEADER_BG)
        cell.font = Font(bold=True, color=COL_HEADER_FG, size=9)
        cell.alignment = Alignment(horizontal="center" if col_idx > 1 else "left",
                                   vertical="center", wrap_text=True)
        cell.border = thin_border()

    # Helper: write a row
    def write_row(row_num, label, values, style="normal", fmt="€",
                  is_pct=False, indent=0, row_height=16):
        ws.row_dimensions[row_num].height = row_height

        # Label cell
        lc = ws.cell(row=row_num, column=1, value=(" " * indent) + label)
        if style == "section":
            lc.fill = make_fill(COL_SECTION_BG)
            lc.font = Font(bold=True, color=COL_SECTION_FG, size=10)
        elif style == "subhead":
            lc.fill = make_fill(COL_SUBHEAD_BG)
            lc.font = Font(bold=True, size=9)
        elif style == "total":
            lc.fill = make_fill(COL_HEADER_BG)
            lc.font = Font(bold=True, color=COL_TOTAL_FG, size=9)
        elif style == "bold":
            lc.fill = make_fill("E8F0FA")
            lc.font = Font(bold=True, size=9)
        else:
            lc.font = Font(size=9)
            if row_num % 2 == 0:
                lc.fill = make_fill(COL_ALT_BG)
        lc.alignment = Alignment(indent=indent // 2 if indent else 0)
        lc.border = thin_border()

        # Value cells
        n_data = len(values)
        for col_idx, v in enumerate(values, start=2):
            c = ws.cell(row=row_num, column=col_idx, value=v)
            is_last_col = (col_idx == n_data + 1)

            if style == "section":
                c.fill = make_fill(COL_SECTION_BG)
                c.font = Font(bold=True, color=COL_SECTION_FG, size=10)
            elif style == "total":
                c.fill = make_fill(COL_HEADER_BG)
                c.font = Font(bold=True, color=COL_TOTAL_FG, size=9)
            elif style == "subhead":
                c.fill = make_fill(COL_SUBHEAD_BG)
                c.font = Font(bold=True, size=9)
            elif style == "bold":
                c.fill = make_fill("E8F0FA")
                c.font = Font(bold=True, size=9)
                if isinstance(v, (int, float)) and v < 0:
                    c.font = Font(bold=True, color=COL_NEG, size=9)
            else:
                c.font = Font(size=9)
                if row_num % 2 == 0:
                    c.fill = make_fill(COL_ALT_BG)
                if isinstance(v, (int, float)) and v < 0 and not is_pct:
                    c.font = Font(color=COL_NEG, size=9)

            if is_last_col:
                c.fill = make_fill("D6E4F0")
                c.font = Font(bold=True, size=9,
                              color=COL_TOTAL_FG if style == "total" else "000000")

            if is_pct:
                c.number_format = '0.0%'
            elif fmt == "€":
                c.number_format = '#,##0 "€";[Red]-#,##0 "€"'
            elif fmt == "k€":
                if isinstance(v, (int, float)):
                    ws.cell(row=row_num, column=col_idx, value=round(v / 1000, 1))
                c.number_format = '#,##0.0 "k€";[Red]-#,##0.0 "k€"'
            elif fmt == "int":
                c.number_format = '#,##0'
            c.alignment = Alignment(horizontal="right")
            c.border = thin_border()

    # Row pointer
    r = 5

    def vals(key):
        return [p[key] for p in annual]

    def spacer(h=6):
        nonlocal r
        ws.row_dimensions[r].height = h
        r += 1

    # ---- REVENUE SECTION ----
    write_row(r, "REVENUE", vals("total_rev"), style="section"); r += 1
    write_row(r, "Subscription Revenue",  vals("sub_rev"),  indent=4); r += 1
    write_row(r, "Enterprise Revenue",    vals("ent_rev"),  indent=4); r += 1
    write_row(r, "TOTAL REVENUE",         vals("total_rev"), style="bold"); r += 1

    spacer(4)

    # ---- COGS ----
    write_row(r, "COST OF GOODS SOLD (COGS)", vals("total_cogs"), style="section"); r += 1
    write_row(r, "Cloud Hosting (variable)",  vals("cogs_cloud"), indent=4); r += 1
    write_row(r, "AI/ML API Kosten",          vals("cogs_aiml"),  indent=4); r += 1
    write_row(r, "Payment Processing",        vals("cogs_pay"),   indent=4); r += 1
    write_row(r, "TOTAL COGS",               vals("total_cogs"), style="bold"); r += 1

    spacer(4)

    # ---- GROSS PROFIT ----
    write_row(r, "GROSS PROFIT",         vals("gross_profit"),      style="subhead", row_height=18); r += 1
    write_row(r, "Gross Margin %",       vals("gross_margin_pct"),  is_pct=True, indent=4); r += 1

    spacer(4)

    # ---- OPERATING EXPENSES ----
    write_row(r, "OPERATING EXPENSES", vals("total_opex"), style="section"); r += 1
    write_row(r, "Personnel (incl. social charges)", vals("personnel"),   indent=4); r += 1
    write_row(r, "Technology & Cloud (OpEx)",        vals("tech"),        indent=4); r += 1
    write_row(r, "Office & Facilities",              vals("office_cost"), indent=4); r += 1
    write_row(r, "Professional Services",            vals("prof"),        indent=4); r += 1
    write_row(r, "Insurance & Bank Fees",            vals("insbank"),     indent=4); r += 1
    write_row(r, "Marketing & Sales",               vals("mktg_cost"),   indent=4); r += 1
    write_row(r, "Other OpEx",                      vals("other_opex"),  indent=4); r += 1
    write_row(r, "Contingency / Buffer (5%)",       vals("contingency"), indent=4); r += 1
    write_row(r, "TOTAL OPERATING EXPENSES",        vals("total_opex"),  style="bold"); r += 1

    spacer(4)

    # ---- EBITDA ----
    write_row(r, "EBITDA",          vals("ebitda"),          style="subhead", row_height=18); r += 1
    write_row(r, "EBITDA Margin %", vals("ebitda_margin_pct"), is_pct=True, indent=4); r += 1

    spacer(4)

    write_row(r, "Depreciation & Amortization", vals("depr"), indent=4); r += 1

    spacer(4)

    write_row(r, "EBIT (Operating Income)", vals("ebit"), style="bold"); r += 1
    write_row(r, "EBT (Earnings Before Tax)", vals("ebt")); r += 1
    write_row(r, "Income Tax (30%)",          vals("tax")); r += 1

    spacer(4)

    write_row(r, "NET INCOME",    vals("net_income"),    style="total", row_height=20); r += 1
    write_row(r, "Net Margin %",  vals("net_margin_pct"), is_pct=True, indent=4); r += 1

    spacer(6)

    # ---- MEMO ITEMS ----
    write_row(r, "MEMO ITEMS", [""] * len(annual), style="section"); r += 1
    write_row(r, "Total Headcount (EoP)", [p["headcount_eop"] for p in annual], fmt="int"); r += 1

    # ---- Sheet 2: Monthly Detail ----
    ws2 = wb.create_sheet("Monthly Detail")
    ws2.column_dimensions["A"].width = 32
    for col in range(2, N_MONTHS + 2):
        ws2.column_dimensions[get_column_letter(col)].width = 10

    # Header row
    ws2.cell(row=1, column=1, value="Position").font = Font(bold=True)
    for m in range(1, N_MONTHS + 1):
        y, mo = month_date(m)
        label_m = f"{y}-{mo:02d}"
        c = ws2.cell(row=1, column=m + 1, value=label_m)
        c.font = Font(bold=True, size=8)
        c.alignment = Alignment(horizontal="center")

    monthly_rows = [
        ("TOTAL REVENUE",           "total_rev"),
        ("  Subscription Revenue",  "sub_rev"),
        ("  Enterprise Revenue",    "ent_rev"),
        ("TOTAL COGS",              "total_cogs"),
        ("GROSS PROFIT",            "gross_profit"),
        ("TOTAL OPEX",              "total_opex"),
        ("  Personnel",             "personnel"),
        ("  Technology",            "tech"),
        ("  Office",                "office_cost"),
        ("  Professional Services", "prof"),
        ("  Insurance & Bank",      "insbank"),
        ("  Marketing & Sales",     "mktg_cost"),
        ("  Other OpEx",            "other_opex"),
        ("  Contingency",           "contingency"),
        ("EBITDA",                  "ebitda"),
        ("Depreciation",            "depr"),
        ("EBIT",                    "ebit"),
        ("NET INCOME",              "net_income"),
    ]

    for ri, (label, key) in enumerate(monthly_rows, start=2):
        ws2.cell(row=ri, column=1, value=label).font = Font(
            bold=key in ("total_rev", "gross_profit", "ebitda", "net_income"), size=8)
        for m in range(1, N_MONTHS + 1):
            c = ws2.cell(row=ri, column=m + 1, value=round(monthly_data[key][m]))
            c.number_format = '#,##0;[Red]-#,##0'
            c.font = Font(size=7)

    wb.save(OUTPUT_FILE)
    print(f"Saved: {OUTPUT_FILE}")


# ---------------------------------------------------------------------------
# 4. MAIN
# ---------------------------------------------------------------------------
if __name__ == "__main__":
    print(f"Computing model for scenario: '{ACTIVE_SCENARIO}'...")
    data = compute_monthly()
    annual = aggregate_annual(data)

    print("\n=== ANNUAL INCOME STATEMENT PREVIEW ===")
    print(f"{'Position':<38}", end="")
    for p in annual:
        print(f"  {p['label'][:16]:>16}", end="")
    print()
    print("-" * (38 + 18 * len(annual)))

    preview_rows = [
        ("Total Revenue",        "total_rev",       False),
        ("Total COGS",           "total_cogs",       False),
        ("Gross Profit",         "gross_profit",     False),
        ("  Gross Margin %",     "gross_margin_pct", True),
        ("Total OpEx",           "total_opex",       False),
        ("EBITDA",               "ebitda",           False),
        ("  EBITDA Margin %",    "ebitda_margin_pct",True),
        ("Net Income",           "net_income",       False),
        ("  Net Margin %",       "net_margin_pct",   True),
        ("Headcount EoP",        "headcount_eop",    False),
    ]

    for label, key, is_pct in preview_rows:
        print(f"{label:<38}", end="")
        for p in annual:
            v = p[key]
            if is_pct:
                print(f"  {v:>15.1%}", end="")
            else:
                print(f"  {v:>14,.0f} ", end="")
        print()

    write_excel(annual, data)
    print("\nDone.")
