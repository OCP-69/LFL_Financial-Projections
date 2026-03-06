"""
build_v6.py  –  LFL_BM_Konservativ_v6_TreiberMatrix.xlsx
══════════════════════════════════════════════════════════
Integriert 260305_LFL_Treiber_Matrix_v1.xlsx in das Business-Modell v5.

ÄNDERUNGEN NACH SHEET:
───────────────────────
1. Treiber_Matrix  (NEU)      Import der vollständigen Matrix als Referenz-Sheet
2. Szenarien_Analyse (UPDATE) + NRR, Churn, Consulting-Startmonat als Szenario-Zeilen
3. 00_Input_Sandbox (UPDATE)  + NRR, Churn, Consulting-Startmonat (→ Szenarien_Analyse)
4. Inputs           (UPDATE)  NRR/Churn → VLOOKUP; + Consulting-Startmonat
                              Kostenkorrekturen aus Treiber-Matrix
5. Revenue          (UPDATE)  Consulting-Revenue-Formel + Startmonat-Check
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SRC_BM  = "/home/user/LFL_Financial-Projections/scenarios/LFL_BM_Konservativ_v5.xlsx"
SRC_TM  = "/home/user/LFL_Financial-Projections/260305_LFL_Treiber_Matrix_v1.xlsx"
DEST    = "/home/user/LFL_Financial-Projections/scenarios/LFL_BM_Konservativ_v6_TreiberMatrix.xlsx"

# ── Neue Sandbox-Zeilen ───────────────────────────────────────────────────────
SB_NRR       = 13   # Net Revenue Retention  → Szenarien_Analyse!C9/D9/E9
SB_CHURN     = 14   # Churn Rate             → Szenarien_Analyse!C10/D10/E10
SB_CSTART    = 15   # Consulting-Startmonat  → Szenarien_Analyse!C11/D11/E11

# Neue Szenarien_Analyse-Zeilen
SZ_NRR       = 9
SZ_CHURN     = 10
SZ_CSTART    = 11

# Neue Inputs-Zeilen (nach bestehender letzte Z140)
IN_CSTART    = 141  # Consulting-Startmonat

# Farben
BLUE_DARK   = "1F3864"
BLUE_MID    = "2E75B6"
BLUE_LIGHT  = "BDD7EE"
GREEN_DARK  = "375623"
GREEN_LIGHT = "E2EFDA"
AMBER       = "FFF2CC"
TEAL        = "1F6B75"
TEAL_LIGHT  = "E9F7F8"
ORANGE      = "FCE4D6"
ORANGE_DARK = "C55A11"
GREY        = "F2F2F2"
WHITE       = "FFFFFF"
RED_LIGHT   = "FFE0E0"
PURPLE      = "7030A0"
PURPLE_LIGHT= "E8D5F5"


def _thin():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

def _medium():
    s = Side(style="medium", color="2E75B6")
    return Border(left=s, right=s, top=s, bottom=s)

def apply(cell, value=None, bold=False, italic=False, size=10,
          color="000000", bg=None, halign="left", valign="center",
          wrap=True, fmt=None):
    if value is not None:
        cell.value = value
    cell.font = Font(bold=bold, italic=italic, size=size, color=color, name="Calibri")
    cell.alignment = Alignment(horizontal=halign, vertical=valign, wrap_text=wrap)
    cell.border = _thin()
    if bg:
        cell.fill = PatternFill("solid", fgColor=bg)
    if fmt:
        cell.number_format = fmt

def sec_hdr(ws, row, text, ncols, bg=BLUE_MID):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row, 1)
    apply(c, value=text, bold=True, color=WHITE, bg=bg, halign="left", wrap=False)


# ══════════════════════════════════════════════════════════════════════════════
# 1. Treiber_Matrix – neues Sheet importieren
# ══════════════════════════════════════════════════════════════════════════════
def build_treiber_sheet(wb_bm, wb_tm):
    ws_src = wb_tm["Treiber-Matrix"]
    ws_dst = wb_bm.create_sheet("Treiber_Matrix", 1)   # an Position 2 (nach Anleitung)

    # Category colours
    cat_colors = {
        "UMSATZ":   "DDEEFF",
        "KOSTEN":   "E8F5E9",
        "CAPEX":    "FFF3E0",
        "NEU":      PURPLE_LIGHT,
    }
    section_colors = {
        "Revenue 1: SaaS / Lizenzen":         BLUE_LIGHT,
        "Revenue 2: Consulting":              GREEN_LIGHT,
        "Personal & Gehälter":               AMBER,
        "Technologie & Cloud":               TEAL_LIGHT,
        "Marketing & Vertrieb":              ORANGE,
        "Professional Services & Sonstiges": GREY,
        "Hardware & Ausstattung":            "F0F0F0",
        "Büro & Facilities":                "F5F5DC",
        "NEUE / POTENZIELLE SZENARIEN":      PURPLE_LIGHT,
        "Umsatz – Neue Treiber":             PURPLE_LIGHT,
        "Kosten – Neue Treiber":             PURPLE_LIGHT,
        "LEGENDE – Szenario-Abhängigkeit":   GREY,
    }

    # Title
    ws_dst.merge_cells("A1:M1")
    apply(ws_dst["A1"],
          value="LoopForgeLab – Treiber- & Annahmen-Matrix (importiert aus 260305_LFL_Treiber_Matrix_v1.xlsx)",
          bold=True, size=13, color=WHITE, bg=BLUE_DARK, halign="center", wrap=False)
    ws_dst.row_dimensions[1].height = 30

    ws_dst.merge_cells("A2:M2")
    apply(ws_dst["A2"],
          value=("Alle Umsatz- und Kostentreiber | Szenarien: Gering · Normal · Stark | "
                 "Spalten L & M: Eigene Bewertungen und Alternativwerte | "
                 "NEU-markierte Zeilen: noch nicht im Modell — Entwicklungspotenzial"),
          italic=True, size=9, color="595959", bg=TEAL_LIGHT, halign="left")
    ws_dst.row_dimensions[2].height = 20

    # Column headers (row 3)
    headers = ["Kategorie","Unterkategorie","Parameter / Treiber","Aktueller Wert",
               "Einheit","Szenario-Abhängigkeit","Gering","Normal","Stark",
               "Einflussfaktor & Wirkungslogik","Annahmen & Begründung",
               "► Ihre Bewertung","► Neue Annahme"]
    for col, h in enumerate(headers, 1):
        apply(ws_dst.cell(3, col), value=h, bold=True, color=WHITE,
              bg=BLUE_DARK, halign="center", size=9)
    ws_dst.row_dimensions[3].height = 30

    # Copy data rows 4-76
    for src_row in range(4, ws_src.max_row + 1):
        vals = [ws_src.cell(src_row, c).value for c in range(1, 14)]
        if not any(v is not None for v in vals):
            continue

        dst_row = src_row   # keep same row numbers for simplicity
        ws_dst.row_dimensions[dst_row].height = 45

        cat = vals[0] if vals[0] else ""

        # Section header rows (only one non-None value in col A, rest None)
        is_section = (vals[0] is not None and vals[1] is None and
                      str(vals[0]) not in ("UMSATZ","KOSTEN","CAPEX","NEU"))
        if is_section:
            sec_bg = section_colors.get(str(vals[0]), GREY)
            ws_dst.merge_cells(f"A{dst_row}:M{dst_row}")
            apply(ws_dst.cell(dst_row, 1), value=vals[0],
                  bold=True, size=10, color=WHITE if sec_bg == BLUE_MID else "000000",
                  bg=sec_bg, halign="left", wrap=False)
            ws_dst.row_dimensions[dst_row].height = 18
            continue

        # Category header rows (UMSATZ, KOSTEN, CAPEX, NEU)
        if vals[0] in ("UMSATZ","KOSTEN","CAPEX","NEU") and vals[1] is None:
            continue   # skip pure category markers

        row_bg = cat_colors.get(str(cat), WHITE)
        is_neu = str(cat) == "NEU"

        for col_idx in range(1, 14):
            v = vals[col_idx - 1]
            c = ws_dst.cell(dst_row, col_idx)
            # Editable columns L & M
            if col_idx >= 12:
                apply(c, value=v, bg=AMBER, halign="left", size=9, wrap=True)
            elif col_idx in (7, 8, 9):  # Gering/Normal/Stark
                bg = AMBER if is_neu else GREEN_LIGHT
                apply(c, value=v, bold=True, bg=bg, halign="center", size=9, wrap=False)
            elif col_idx == 6:  # Szenario-Abhängigkeit
                dep_colors = {
                    "Szenario":    BLUE_LIGHT,
                    "Unabhängig":  ORANGE,
                    "Berechnet":   TEAL_LIGHT,
                    "Skalierend":  GREEN_LIGHT,
                    "Fest":        GREY,
                    "Zu definieren": PURPLE_LIGHT,
                }
                dep_bg = dep_colors.get(str(v) if v else "", WHITE)
                apply(c, value=v, bg=dep_bg, halign="center", size=9,
                      bold=(str(v) == "Unabhängig"),
                      color=PURPLE if is_neu else "000000")
            elif col_idx in (1, 2):
                apply(c, value=v, bg=row_bg if not is_neu else PURPLE_LIGHT,
                      halign="left", size=9, bold=(col_idx == 1 and is_neu))
            else:
                apply(c, value=v, bg=row_bg if not is_neu else PURPLE_LIGHT,
                      halign="left", size=9, wrap=True)

    # Column widths
    widths = [14, 20, 30, 18, 14, 16, 14, 14, 14, 35, 35, 22, 22]
    for col_idx, w in enumerate(widths, 1):
        ws_dst.column_dimensions[get_column_letter(col_idx)].width = w

    ws_dst.freeze_panes = "C4"

    # Legend at bottom
    legend_row = 78
    ws_dst.merge_cells(f"A{legend_row}:M{legend_row}")
    apply(ws_dst.cell(legend_row, 1),
          value=("LEGENDE: Szenario=szenarioabhängig (Sandbox)  |  "
                 "Unabhängig=eigenständiger Schalter (z.B. Consulting-Quote)  |  "
                 "Berechnet=abgeleiteter Wert  |  Skalierend=fixer Satz, skaliert mit MA/Kunden  |  "
                 "NEU (lila)=noch nicht im Modell"),
          italic=True, size=8, color="595959", bg=GREY)

    print(f"  Treiber_Matrix Sheet: {ws_src.max_row} Zeilen importiert")


# ══════════════════════════════════════════════════════════════════════════════
# 2. Szenarien_Analyse – NRR, Churn, Consulting-Startmonat hinzufügen
# ══════════════════════════════════════════════════════════════════════════════
def patch_szenarien(wb):
    ws = wb["Szenarien_Analyse"]

    # Row 9: NRR (Net Revenue Retention)
    ws.cell(SZ_NRR, 1).value = "Net Revenue Retention (NRR)"
    ws.cell(SZ_NRR, 2).value = 1.15    # Referenz
    ws.cell(SZ_NRR, 3).value = 1.05    # Gering  (105%)
    ws.cell(SZ_NRR, 4).value = 1.15    # Normal  (115%)
    ws.cell(SZ_NRR, 5).value = 1.30    # Stark   (130%)
    ws.cell(SZ_NRR, 6).value = "=E9-D9"
    ws.cell(SZ_NRR, 7).value = (
        "Treiber-Matrix: NRR szenarioabhängig machen (bisher fest 125%).\n"
        "Gering=105%: Automotive, wenig Upselling-Potenzial.\n"
        "Normal=115%: Hybrid, DPP-Module.\n"
        "Stark=130%: Packaging, Analytics-Premium + Multi-Werk-Rollouts."
    )

    # Row 10: Churn Rate
    ws.cell(SZ_CHURN, 1).value = "Jährliche Churn Rate"
    ws.cell(SZ_CHURN, 2).value = 0.06
    ws.cell(SZ_CHURN, 3).value = 0.10  # Gering  (10%)
    ws.cell(SZ_CHURN, 4).value = 0.06  # Normal  (6%)
    ws.cell(SZ_CHURN, 5).value = 0.03  # Stark   (3%)
    ws.cell(SZ_CHURN, 6).value = "=E10-D10"
    ws.cell(SZ_CHURN, 7).value = (
        "Treiber-Matrix: Churn szenarioabhängig (bisher fest 6%).\n"
        "Gering=10%: Automotive-Budgetzyklen, höheres Abwanderungsrisiko.\n"
        "Normal=6%: Standardwert (SaaS-Industrie).\n"
        "Stark=3%: Packaging, tiefe API-Integration → hoher Switching-Cost."
    )

    # Row 11: Consulting-Startmonat
    ws.cell(SZ_CSTART, 1).value = "Consulting-Startmonat"
    ws.cell(SZ_CSTART, 2).value = 4
    ws.cell(SZ_CSTART, 3).value = 3    # Gering: früh (Automotive braucht sofort Begleitung)
    ws.cell(SZ_CSTART, 4).value = 4    # Normal
    ws.cell(SZ_CSTART, 5).value = 6    # Stark: Packaging nutzt länger Self-Service
    ws.cell(SZ_CSTART, 6).value = "=E11-D11"
    ws.cell(SZ_CSTART, 7).value = (
        "Treiber-Matrix Z.21: Consulting startet als Enabler vor dem SaaS-Produkt.\n"
        "Gering=Monat 3: Automotive benötigt sofort IT-Begleitung.\n"
        "Stark=Monat 6: Packaging-Kunden länger Self-Service-fähig."
    )

    # Style new rows
    for row in (SZ_NRR, SZ_CHURN, SZ_CSTART):
        ws.row_dimensions[row].height = 55
        ws.cell(row, 1).font = Font(bold=True, size=10, name="Calibri", color=TEAL)
        ws.cell(row, 1).fill = PatternFill("solid", fgColor=TEAL_LIGHT)
        ws.cell(row, 1).alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(row, 1).border = _thin()
        for col in (3, 4, 5):
            c = ws.cell(row, col)
            c.fill = PatternFill("solid", fgColor=AMBER)
            c.font = Font(bold=True, size=11, name="Calibri")
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = _thin()
            if row in (SZ_NRR, SZ_CHURN):
                c.number_format = "0%"
        for col in (2, 6, 7):
            ws.cell(row, col).border = _thin()
        ws.cell(row, 8).value = "→ Sandbox → Inputs"
        ws.cell(row, 8).font = Font(italic=True, size=8, color=TEAL)

    print(f"  Szenarien_Analyse Z.{SZ_NRR}-{SZ_CSTART}: NRR, Churn, Consulting-Startmonat")


# ══════════════════════════════════════════════════════════════════════════════
# 3. Sandbox – NRR, Churn, Consulting-Startmonat hinzufügen
# ══════════════════════════════════════════════════════════════════════════════
def patch_sandbox(wb):
    ws = wb["00_Input_Sandbox"]

    rows = [
        (SB_NRR,  "Retention",  "Net Revenue Retention (NRR)", "%",
         f"=Szenarien_Analyse!C{SZ_NRR}", f"=Szenarien_Analyse!D{SZ_NRR}", f"=Szenarien_Analyse!E{SZ_NRR}",
         "Szenarioabhängig (Szenarien_Analyse Z.9). Gering=105% | Normal=115% | Stark=130%."),
        (SB_CHURN,"Retention",  "Churn Rate",                  "%/Jahr",
         f"=Szenarien_Analyse!C{SZ_CHURN}", f"=Szenarien_Analyse!D{SZ_CHURN}", f"=Szenarien_Analyse!E{SZ_CHURN}",
         "Szenarioabhängig (Szenarien_Analyse Z.10). Gering=10% | Normal=6% | Stark=3%."),
        (SB_CSTART,"Consulting","Consulting-Startmonat",        "Monat",
         f"=Szenarien_Analyse!C{SZ_CSTART}", f"=Szenarien_Analyse!D{SZ_CSTART}", f"=Szenarien_Analyse!E{SZ_CSTART}",
         "Szenarioabhängig (Szenarien_Analyse Z.11). Gering=3 | Normal=4 | Stark=6."),
    ]

    for (row, kat, var, einheit, gering, normal, stark, quelle) in rows:
        ws.cell(row, 1).value = kat
        ws.cell(row, 2).value = var
        ws.cell(row, 3).value = einheit
        ws.cell(row, 4).value = gering
        ws.cell(row, 5).value = normal
        ws.cell(row, 6).value = stark
        ws.cell(row, 7).value = quelle
        ws.row_dimensions[row].height = 30
        for col in range(1, 8):
            c = ws.cell(row, col)
            c.font = Font(size=10, name="Calibri")
            c.border = _thin()
            c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        for col in (4, 5, 6):
            c = ws.cell(row, col)
            c.font = Font(italic=True, size=10, name="Calibri", color="1F3864")
            c.fill = PatternFill("solid", fgColor=BLUE_LIGHT)
            c.alignment = Alignment(horizontal="right", vertical="center")

    print(f"  Sandbox Z.{SB_NRR}: NRR → =Szenarien_Analyse!C{SZ_NRR}/D{SZ_NRR}/E{SZ_NRR}")
    print(f"  Sandbox Z.{SB_CHURN}: Churn → =Szenarien_Analyse!C{SZ_CHURN}/D{SZ_CHURN}/E{SZ_CHURN}")
    print(f"  Sandbox Z.{SB_CSTART}: Consulting-Startmonat → =Szenarien_Analyse!C{SZ_CSTART}/D{SZ_CSTART}/E{SZ_CSTART}")


# ══════════════════════════════════════════════════════════════════════════════
# 4. Inputs – NRR/Churn → VLOOKUP; + Consulting-Startmonat; Kostenkorrekturen
# ══════════════════════════════════════════════════════════════════════════════
def patch_inputs(wb):
    ws = wb["Inputs"]

    SANDBOX_VLOOKUP = (
        "=VLOOKUP(\"{var}\",'00_Input_Sandbox'!$B$3:$F$20,"
        "'00_Input_Sandbox'!$B$2-1,FALSE)"
    )

    # ── NRR (Z29): hardcoded 1.05 → VLOOKUP ─────────────────────────────────
    ws.cell(29, 2).value = SANDBOX_VLOOKUP.format(var="Net Revenue Retention (NRR)")
    ws.cell(29, 2).number_format = "0%"
    ws.cell(29, 3).value = "Net Revenue Retention (szenarioabhängig): Gering=105% | Normal=115% | Stark=130%"
    ws.cell(29, 4).value = (
        "Treiber-Matrix: NRR jetzt szenarioabhängig (vorher fest 125%). "
        "Expansion-Umsatz minus Churn. NRR>100%=Bestandskunden wachsen."
    )

    # ── Churn (Z28): hardcoded 0.15 → VLOOKUP ───────────────────────────────
    ws.cell(28, 2).value = SANDBOX_VLOOKUP.format(var="Churn Rate")
    ws.cell(28, 2).number_format = "0%"
    ws.cell(28, 3).value = "Jährliche Churn Rate (szenarioabhängig): Gering=10% | Normal=6% | Stark=3%"
    ws.cell(28, 4).value = (
        "Treiber-Matrix: Churn jetzt szenarioabhängig (vorher fest 15%). "
        "Automotive: 10% (Budgetzyklen). Packaging: 3% (hoher Switching-Cost)."
    )

    # ── Z141: Consulting-Startmonat (NEU) ────────────────────────────────────
    ws.cell(IN_CSTART, 1).value = "Consulting-Startmonat"
    ws.cell(IN_CSTART, 2).value = SANDBOX_VLOOKUP.format(var="Consulting-Startmonat")
    ws.cell(IN_CSTART, 3).value = "Monat des ersten Consulting-Umsatzes (aus Sandbox/Szenarien_Analyse)"
    ws.cell(IN_CSTART, 4).value = (
        "Treiber-Matrix Z.21: Consulting startet als Enabler vor dem SaaS-Produkt. "
        "Gering=Monat 3 | Normal=Monat 4 | Stark=Monat 6. "
        "Revenue-Formel prüft: IF(Monat >= Consulting-Startmonat, Umsatz, 0)."
    )
    ws.cell(IN_CSTART, 5).value = "Szenariogesteuert (Sandbox Z.15)"
    from openpyxl.styles import Font as F, PatternFill as PF, Alignment as A
    ws.cell(IN_CSTART, 1).font = F(bold=True, size=10, name="Calibri")
    ws.cell(IN_CSTART, 1).border = _thin()
    ws.cell(IN_CSTART, 2).fill = PF("solid", fgColor=BLUE_LIGHT)
    ws.cell(IN_CSTART, 2).border = _thin()
    ws.cell(IN_CSTART, 2).alignment = A(horizontal="right", vertical="center")
    ws.cell(IN_CSTART, 4).font = F(italic=True, size=9, color="595959", name="Calibri")
    ws.cell(IN_CSTART, 4).alignment = A(wrap_text=True, vertical="top")
    ws.cell(IN_CSTART, 4).border = _thin()
    ws.row_dimensions[IN_CSTART].height = 55

    # ── Update DATENFLUSS Banner ─────────────────────────────────────────────
    ws.cell(136, 1).value = (
        "ℹ DATENFLUSS CONSULTING:  "
        "Szenarien_Analyse (Z.5 Wskt. / Z.8 Tage / Z.11 Startmonat)  →  "
        "Sandbox (Z.11 Tage / Z.12 Wskt. / Z.15 Startmonat)  →  "
        "Inputs (Z.138 Tagessatz / Z.139 Wskt. / Z.140 Tage / Z.141 Startmonat)  →  "
        "Revenue Z.28-29 → P&L"
    )

    # ── Kostenkorrekturen aus Treiber-Matrix ─────────────────────────────────
    cost_updates = {
        82:  ("Cloud/Hosting Basis",        2000,   "800→2.000 €/Monat (Treiber-Matrix)"),
        83:  ("Cloud Skalierung pro Seat",   80,     "50→80 €/Seat/Monat (Treiber-Matrix)"),
        84:  ("AI/ML APIs Basis",            2000,   "600→2.000 €/Monat (Treiber-Matrix)"),
        85:  ("AI Kosten Wachstum/Monat",    0.08,   "3%→8%/Monat (Treiber-Matrix)"),
        122: ("Paid Ads Budget Initial",     1000,   "300→1.000 €/Monat (Treiber-Matrix)"),
        124: ("Content & SEO",               2500,   "1.000→2.500 €/Monat (Treiber-Matrix)"),
        125: ("Events & Messen/Jahr",        50000,  "25.000→50.000 €/Jahr (Treiber-Matrix)"),
        127: ("Sales Provision",             0.12,   "8%→12% vom neuen ARR (Treiber-Matrix)"),
        128: ("Reisekosten Sales/MA/Monat",  700,    "300→700 €/MA/Monat (Treiber-Matrix)"),
    }
    for row, (label, new_val, note) in cost_updates.items():
        ws.cell(row, 2).value = new_val
        # Append note to column D (Bemerkungen)
        existing = ws.cell(row, 4).value or ""
        ws.cell(row, 4).value = (existing + "\n" + note).strip()

    print(f"  Inputs Z.28: Churn → VLOOKUP('Churn Rate', Sandbox)")
    print(f"  Inputs Z.29: NRR   → VLOOKUP('Net Revenue Retention (NRR)', Sandbox)")
    print(f"  Inputs Z.{IN_CSTART}: Consulting-Startmonat (NEU) → VLOOKUP Sandbox")
    print(f"  Inputs: 9 Kostenpositionen aus Treiber-Matrix korrigiert")


# ══════════════════════════════════════════════════════════════════════════════
# 5. Revenue – Consulting-Revenue-Formel + Startmonat-Check
# ══════════════════════════════════════════════════════════════════════════════
def patch_revenue(wb):
    ws    = wb["Revenue"]
    ncols = 53   # col A + M1…M52

    # Z28: Consulting-Tage/Monat — nur ab Consulting-Startmonat (Inputs Z141)
    # Formel: =IF(Monat >= Inputs!$B$141, TotalKunden × Wskt × Tage, 0)
    for col in range(2, ncols + 1):
        monat = col - 1   # col 2 = M1, col 3 = M2, ...
        col_l = get_column_letter(col)
        ws.cell(28, col).value = (
            f"=IF({monat}>=Inputs!$B${IN_CSTART},"
            f"{col_l}27*Inputs!$B$139*Inputs!$B$140,0)"
        )
        ws.cell(28, col).fill = PatternFill("solid", fgColor=AMBER)
        ws.cell(28, col).border = _thin()
        ws.cell(28, col).alignment = Alignment(horizontal="right", vertical="center")

    # Z28 Label update
    ws.cell(28, 1).value = (
        "Consulting-Tage/Monat  [ab Startmonat; Kunden × Wskt. × Tage/Einsatz]"
    )

    print(f"  Revenue Z.28: Formel + Startmonat-Check  IF(Monat>={IN_CSTART},...)")
    print(f"  Revenue Z.29: unverändert (×Tagessatz, greift automatisch)")


# ══════════════════════════════════════════════════════════════════════════════
# 6. Neue Sheet: NEU_Entwicklung – "Zu definieren"-Treiber als Roadmap
# ══════════════════════════════════════════════════════════════════════════════
def build_neu_sheet(wb):
    ws = wb.create_sheet("Entwicklungs_Roadmap")

    ws.merge_cells("A1:F1")
    apply(ws["A1"],
          value="LFL – Entwicklungs-Roadmap: Neue Treiber aus Treiber-Matrix (noch nicht modelliert)",
          bold=True, size=12, color=WHITE, bg=PURPLE, halign="center", wrap=False)
    ws.row_dimensions[1].height = 28

    headers = ["Kategorie", "Parameter / Treiber", "Mögliche Werte (G/N/S)",
               "Auswirkung", "Priorität", "Nächster Schritt"]
    for col, h in enumerate(headers, 1):
        apply(ws.cell(2, col), value=h, bold=True, color=WHITE,
              bg=BLUE_DARK, halign="center", size=9)
    ws.row_dimensions[2].height = 22

    roadmap = [
        ("SaaS-Pricing", "Produkt-Tiers (Basic/Pro/Enterprise)",
         "Basic ~5K€/J | Pro ~12K€/J | Ent ~25K€/J",
         "Höhere Preisflexibilität, bessere Marktabdeckung → ARR ↑",
         "P1 HOCH",
         "Seat-Preis in Sandbox durch Tier-Verteilung ersetzen"),
        ("SaaS-Umsatz", "NRR-Varianten (bereits integriert!)",
         "Gering=105% | Normal=115% | Stark=130%",
         "✓ In v6 implementiert — Sandbox Z.13 / Szenarien Z.9",
         "✓ ERLEDIGT",
         "—"),
        ("SaaS-Umsatz", "Churn-Varianten (bereits integriert!)",
         "Gering=10% | Normal=6% | Stark=3%",
         "✓ In v6 implementiert — Sandbox Z.14 / Szenarien Z.10",
         "✓ ERLEDIGT",
         "—"),
        ("Marktexpansion", "Internationalisierung",
         "Keine | EU Y2 | US Y3",
         "Multiplikator auf Kundenwachstum ab definiertem Monat → ARR ↑↑",
         "P2 MITTEL",
         "Revenue: Multiplikator-Zeile ab Monat X; Inputs: Expansion-Startmonat"),
        ("Personal", "Gehaltsniveau-Faktor",
         "85% | 100% | 115% des Markts",
         "Direkter Hebel auf gesamte Personalkostenbasis → EBITDA ±",
         "P2 MITTEL",
         "Inputs: Faktor auf alle Gehälter in Einstellungsplan"),
        ("Büro/Facilities", "Remote-Work-Strategie",
         "Office-First | Hybrid | Full-Remote",
         "Büromiete 0–4.500€/Monat → Burn Rate ↓ bei Remote",
         "P3 NIEDRIG",
         "Inputs: Büromiete durch Strategie-Faktor (0/0.5/1.0) steuern"),
        ("Technologie", "Technologie-Stack-Strategie",
         "Open-Source -40% | Proprietär | Multi-Provider +15%",
         "AI/ML-API-Kosten direkt → COGS Margin ±",
         "P2 MITTEL",
         "Sandbox: Stack-Faktor auf AI/ML APIs Basis"),
    ]

    for i, row in enumerate(roadmap):
        dst_row = 3 + i
        ws.row_dimensions[dst_row].height = 50
        bgs = [PURPLE_LIGHT, PURPLE_LIGHT, TEAL_LIGHT, TEAL_LIGHT,
               AMBER, GREEN_LIGHT, ORANGE]
        prio_bgs = {
            "P1 HOCH": "FF6B6B",
            "P2 MITTEL": "FFB347",
            "P3 NIEDRIG": "FFD700",
            "✓ ERLEDIGT": "C6EFCE",
        }
        for col_idx, val in enumerate(row, 1):
            c = ws.cell(dst_row, col_idx)
            bg = prio_bgs.get(val, PURPLE_LIGHT if i % 2 == 0 else "F0EAF8")
            if col_idx == 5:
                apply(c, value=val, bold=True, bg=bg,
                      halign="center", size=9, color=WHITE if "FF" in bg else "000000")
            else:
                apply(c, value=val, bg=PURPLE_LIGHT if i % 2 == 0 else "F0EAF8",
                      halign="left", size=9, wrap=True)

    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 25
    ws.column_dimensions["D"].width = 35
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 40
    print("  Entwicklungs_Roadmap Sheet: 7 Einträge (davon 2 bereits umgesetzt)")


# ══════════════════════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════════════════════
def main():
    print(f"Lade Business-Modell: {SRC_BM}")
    wb_bm = openpyxl.load_workbook(SRC_BM, data_only=False)
    print(f"Lade Treiber-Matrix:   {SRC_TM}")
    wb_tm = openpyxl.load_workbook(SRC_TM, data_only=True)

    print("\n1. Treiber_Matrix Sheet importieren …")
    build_treiber_sheet(wb_bm, wb_tm)

    print("\n2. Szenarien_Analyse: NRR / Churn / Consulting-Startmonat …")
    patch_szenarien(wb_bm)

    print("\n3. 00_Input_Sandbox: NRR / Churn / Consulting-Startmonat …")
    patch_sandbox(wb_bm)

    print("\n4. Inputs: VLOOKUPs + Consulting-Startmonat + Kostenkorrekturen …")
    patch_inputs(wb_bm)

    print("\n5. Revenue: Consulting-Startmonat-Check in Formel …")
    patch_revenue(wb_bm)

    print("\n6. Entwicklungs_Roadmap Sheet …")
    build_neu_sheet(wb_bm)

    wb_bm.save(DEST)
    print(f"\n✅ Gespeichert: {DEST}")

    # Verifikation
    wb2 = openpyxl.load_workbook(DEST, data_only=False)
    sz  = wb2["Szenarien_Analyse"]
    sb  = wb2["00_Input_Sandbox"]
    inp = wb2["Inputs"]
    rev = wb2["Revenue"]

    print("\n── Verifikation ──")
    print(f"  Sheets: {wb2.sheetnames}")
    print(f"  Szenarien Z.{SZ_NRR}  (NRR):         G={sz.cell(SZ_NRR,3).value}  N={sz.cell(SZ_NRR,4).value}  S={sz.cell(SZ_NRR,5).value}")
    print(f"  Szenarien Z.{SZ_CHURN}  (Churn):       G={sz.cell(SZ_CHURN,3).value}  N={sz.cell(SZ_CHURN,4).value}  S={sz.cell(SZ_CHURN,5).value}")
    print(f"  Szenarien Z.{SZ_CSTART}  (C-Startm.):  G={sz.cell(SZ_CSTART,3).value}  N={sz.cell(SZ_CSTART,4).value}  S={sz.cell(SZ_CSTART,5).value}")
    print(f"  Sandbox  Z.{SB_NRR} D:  {sb.cell(SB_NRR,4).value}")
    print(f"  Sandbox  Z.{SB_CHURN} D:  {sb.cell(SB_CHURN,4).value}")
    print(f"  Sandbox  Z.{SB_CSTART} D:  {sb.cell(SB_CSTART,4).value}")
    print(f"  Inputs Z.28 (Churn):      {inp.cell(28,2).value[:55]}…")
    print(f"  Inputs Z.29 (NRR):        {inp.cell(29,2).value[:55]}…")
    print(f"  Inputs Z.{IN_CSTART} (C-Startm.):  {inp.cell(IN_CSTART,2).value[:55]}…")
    print(f"  Inputs Z.82 (Cloud):      {inp.cell(82,2).value}  (war 800)")
    print(f"  Inputs Z.84 (AI API):     {inp.cell(84,2).value}  (war 600)")
    print(f"  Inputs Z.127 (Provision): {inp.cell(127,2).value}  (war 0.08)")
    print(f"  Revenue Z.28 M1:          {rev.cell(28,2).value}")
    print(f"  Revenue Z.29 M1:          {rev.cell(29,2).value}")

if __name__ == "__main__":
    main()
