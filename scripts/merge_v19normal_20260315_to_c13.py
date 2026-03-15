"""
Data Merge Script: 260315_LFL_BM_Vorlage_normal_v19.xlsx → C13_Template_financial_projections_neu.xlsx
Mapping-Logik: 260313_LFL_Finm Proj_ Data Merge_C13 TEmplate.txt

Identische Mapping-Logik wie merge_v19_to_c13.py –
angepasst auf neue Quelldatei vom 15.03.2026 (Normal-Szenario).

Mapping-Regeln:
  Revenue            → Monthly.Revenue + Monthly.Cash in from Revenue
  COGS               → Monthly.CoR Cash Out
  Total OpEx         → Monthly.Operating Cash Out
  Equity/Grants      → Monthly.Other Cash In
  Ending Cash        → Monthly.Closing Cash
  Beginning Cash     → Monthly.Opening Cash
  Income Tax         → Monthly.Tax
  Annual Aggregates  → Annual sheet (Jahresaggregation)
  Year-End Balances  → BalanceSheet

Zeitachse:
  Source M1  = April 2026  → C13 Monthly Row 5
  Source M52 = Juli 2030   → C13 Monthly Row 56
  Leere Monate (Jan–Mär 2026): Nullwerte
"""

import openpyxl
from datetime import datetime
import os

# ── Pfade ─────────────────────────────────────────────────────────────────────
BASE_DIR   = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SOURCE     = os.path.join(BASE_DIR, '260315_LFL_BM_Vorlage_normal_v19.xlsx')
TARGET_TPL = os.path.join(BASE_DIR, 'C13_Template_financial_projections_neu.xlsx')
TIMESTAMP  = datetime.now().strftime('%Y%m%d_%H%M')
OUTPUT     = os.path.join(BASE_DIR, f'LFL_BM_C13_Normal_v19_20260315_{TIMESTAMP}.xlsx')

# ── Hilfsfunktionen ───────────────────────────────────────────────────────────
def read_row_values(ws, row_num, num_months=52):
    """Lese 52 Monatswerte aus Spalten B (col=2) bis col=53."""
    data = []
    for col in range(2, 2 + num_months):
        val = ws.cell(row=row_num, column=col).value
        data.append(float(val) if val is not None and isinstance(val, (int, float)) else 0.0)
    return data

def year_months_range(year_offset):
    """
    Gibt (src_start, src_end) Index-Tupel für das jeweilige Kalenderjahr zurück.
    Source M1 = April 2026; Jan–Mär 2026 enthalten keine Quelldaten.

    2026: M1–M9   (April–Dez)  → src idx 0..8
    2027: M10–M21 (Jan–Dez)    → src idx 9..20
    2028: M22–M33              → src idx 21..32
    2029: M34–M45              → src idx 33..44
    2030: M46–M52 (Jan–Jul)    → src idx 45..51
    """
    ranges = [(0, 9), (9, 21), (21, 33), (33, 45), (45, 52)]
    return ranges[year_offset] if year_offset < len(ranges) else (0, 0)

# ── Source laden ──────────────────────────────────────────────────────────────
print(f"Lade Quelldatei: {os.path.basename(SOURCE)} …")
wb_src = openpyxl.load_workbook(SOURCE, data_only=True)

ws_rev   = wb_src['4_Revenue']
ws_pl    = wb_src['6_P&L']
ws_cf    = wb_src['7_BS_CF']

# Schlüsseldaten (52 Monate)
total_revenue  = read_row_values(ws_rev, 32)   # R32: Total Revenue (€/Monat)
mrr            = read_row_values(ws_rev, 25)   # R25: MRR
arr            = read_row_values(ws_rev, 26)   # R26: ARR

total_cogs     = read_row_values(ws_pl,  14)   # R14: TOTAL COGS
total_opex     = read_row_values(ws_pl,  27)   # R27: TOTAL OPERATING EXPENSES
net_income     = read_row_values(ws_pl,  37)   # R37: NET INCOME
income_tax     = read_row_values(ws_pl,  35)   # R35: Income Tax
ebitda         = read_row_values(ws_pl,  29)   # R29: EBITDA
gross_profit   = read_row_values(ws_pl,  16)   # R16: GROSS PROFIT

equity_funding = read_row_values(ws_cf,  9)    # R9:  Equity Funding Received
beginning_cash = read_row_values(ws_cf,  14)   # R14: Beginning Cash Balance
ending_cash    = read_row_values(ws_cf,  15)   # R15: ENDING CASH BALANCE
burn_rate      = read_row_values(ws_cf,  18)   # R18: Monthly Burn Rate

# Berechnete Hilfsgrößen
debtors   = [rev for rev in total_revenue]           # 1 Monat Umsatz als Debitoren
creditors = [(total_cogs[i] + total_opex[i]) * 0.10 # 10 % der Gesamtkosten als Kreditoren
             for i in range(52)]

cum_equity = []
cum_sum = 0.0
for v in equity_funding:
    cum_sum += v
    cum_equity.append(cum_sum)

cum_profit = []
cum_sum = 0.0
for v in net_income:
    cum_sum += v
    cum_profit.append(cum_sum)

print(f"  → Gesamtumsatz M1–M52:        {sum(total_revenue):>15,.0f} €")
print(f"  → Eigenkapital eingeflossen:  {sum(equity_funding):>15,.0f} €")
print(f"  → Letzter Cash-Stand (M52):   {ending_cash[51]:>15,.0f} €")

# ── Target Template laden ─────────────────────────────────────────────────────
print(f"\nLade C13-Template: {os.path.basename(TARGET_TPL)} …")
wb_dst = openpyxl.load_workbook(TARGET_TPL, data_only=False)

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 1: Monthly
# Spalten:
#  A=Year | B=Month | C=Revenue | D=Cash in from Revenue | E=Other Cash In
#  F=CoR Cash Out | G=Operating Cash Out | H=Interest | I=Tax | J=Founder Bonus
#  K=Opening Cash | L=Closing Cash | M=Debt Outstanding | N=Debtors | O=Creditors
# ══════════════════════════════════════════════════════════════════════════════
print("Befülle Monthly-Sheet …")
ws_m = wb_dst['Monthly']

# Monats-Nummern für Zeilen 2–49 (4 Jahre × 12 Monate = 2026–2029)
month_num = 1
for row in range(2, 50):
    ws_m.cell(row=row, column=2).value = month_num
    month_num = (month_num % 12) + 1

# Neue Zeilen für 2030 (Jan–Jul = 7 Monate) → Rows 50–56
for m_idx in range(7):
    row = 50 + m_idx
    ws_m.cell(row=row, column=1).value = 2030
    ws_m.cell(row=row, column=2).value = m_idx + 1  # Jan=1 … Jul=7

def write_monthly_row(row, src_idx):
    """Schreibe alle Finanzdaten für einen Monat in eine C13-Zeile."""
    rev  = total_revenue[src_idx]
    cogs = total_cogs[src_idx]
    opex = total_opex[src_idx]
    eqty = equity_funding[src_idx]
    tax  = income_tax[src_idx]
    beg  = beginning_cash[src_idx]
    end  = ending_cash[src_idx]
    deb  = debtors[src_idx]
    cred = creditors[src_idx]

    ws_m.cell(row=row, column=3).value  = round(rev,  2)   # C: Revenue
    ws_m.cell(row=row, column=4).value  = round(rev,  2)   # D: Cash in from Revenue
    ws_m.cell(row=row, column=5).value  = round(eqty, 2)   # E: Other Cash In (Equity)
    ws_m.cell(row=row, column=6).value  = round(cogs, 2)   # F: CoR Cash Out
    ws_m.cell(row=row, column=7).value  = round(opex, 2)   # G: Operating Cash Out
    ws_m.cell(row=row, column=8).value  = 0.0              # H: Interest (nicht modelliert)
    ws_m.cell(row=row, column=9).value  = round(tax,  2)   # I: Tax
    ws_m.cell(row=row, column=10).value = 0.0              # J: Founder Bonus
    ws_m.cell(row=row, column=11).value = round(beg,  2)   # K: Opening Cash
    ws_m.cell(row=row, column=12).value = round(end,  2)   # L: Closing Cash
    ws_m.cell(row=row, column=13).value = 0.0              # M: Debt Outstanding
    ws_m.cell(row=row, column=14).value = round(deb,  2)   # N: Debtors
    ws_m.cell(row=row, column=15).value = round(cred, 2)   # O: Creditors

# Leerzeilen Jan–Mär 2026 (Rows 2–4, keine Quelldaten)
for row in range(2, 5):
    for col in [3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15]:
        ws_m.cell(row=row, column=col).value = 0.0

# Quelldaten M1–M52 → Rows 5–56 (M1=Apr 2026, M52=Jul 2030)
for m_idx in range(52):
    write_monthly_row(5 + m_idx, m_idx)

print(f"  → 52 Monate eingetragen (Apr 2026 – Jul 2030, Rows 5–56)")

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 2: Annual
# Spalten:
#  A=Year | B=Revenue | C=CoR | D=Gross Profit | E=Total Opex (incl R&D amort)
#  F=EBIT | G=Interest | H=Tax | I=Founder Bonus | J=Grants (non-taxable)
#  K=Net Profit | L=Min Cash | M=Avg Monthly Gross Burn | N=Avg Monthly EBIT Burn
#  O=Cash (year-end) | P=Debtors | Q=R&D Intangible | R=Creditors
#  S=Debt | T=Net Assets | U=Total Equity Investment | V=Cumulative Profit | W=Quick Ratio
# ══════════════════════════════════════════════════════════════════════════════
print("Befülle Annual-Sheet …")
ws_a = wb_dst['Annual']

annual_years = [2026, 2027, 2028, 2029, 2030]

# Kumulativen Equity- und Gewinnstand am Jahresende berechnen
cum_eq_year = []
cum_pr_year = []
for yo, _ in enumerate(annual_years):
    _, e = year_months_range(yo)
    cum_eq_year.append(sum(equity_funding[:e]))
    cum_pr_year.append(sum(net_income[:e]))

print(f"\n  {'Jahr':>6} | {'Umsatz':>14} | {'Gross Profit':>14} | {'EBIT':>14} | {'Net Income':>12} | {'Cash Jahresende':>15}")
print(f"  {'-'*6}-+-{'-'*14}-+-{'-'*14}-+-{'-'*14}-+-{'-'*12}-+-{'-'*15}")

for yo, year in enumerate(annual_years):
    row = 2 + yo
    s, e = year_months_range(yo)
    months_n = e - s

    rev_yr   = sum(total_revenue[s:e])
    cogs_yr  = sum(total_cogs[s:e])
    opex_yr  = sum(total_opex[s:e])
    gp_yr    = rev_yr - cogs_yr
    ebit_yr  = gp_yr - opex_yr
    tax_yr   = sum(income_tax[s:e])
    ni_yr    = sum(net_income[s:e])

    min_cash   = min(ending_cash[s:e]) if months_n > 0 else 0.0
    last_cash  = ending_cash[e - 1] if months_n > 0 else 0.0
    last_deb   = debtors[e - 1]   if months_n > 0 else 0.0
    last_cred  = creditors[e - 1] if months_n > 0 else 0.0
    net_assets = last_cash + last_deb - last_cred
    quick_ratio = (last_cash / last_cred) if last_cred > 0 else 0.0

    # Avg Monthly Gross Burn = (COGS + OpEx) / Monate
    gross_burns    = [total_cogs[i] + total_opex[i] for i in range(s, e)]
    avg_gross_burn = sum(gross_burns) / months_n if months_n > 0 else 0.0

    # Avg Monthly EBIT Burn (nur Verlustmonate)
    ebit_vals      = [ebitda[i] for i in range(s, e) if ebitda[i] < 0]
    avg_ebit_burn  = sum(ebit_vals) / len(ebit_vals) if ebit_vals else 0.0

    ws_a.cell(row=row, column=2).value  = round(rev_yr, 2)
    ws_a.cell(row=row, column=3).value  = round(cogs_yr, 2)
    ws_a.cell(row=row, column=4).value  = round(gp_yr, 2)
    ws_a.cell(row=row, column=5).value  = round(opex_yr, 2)
    ws_a.cell(row=row, column=6).value  = round(ebit_yr, 2)
    ws_a.cell(row=row, column=7).value  = 0.0
    ws_a.cell(row=row, column=8).value  = round(tax_yr, 2)
    ws_a.cell(row=row, column=9).value  = 0.0
    ws_a.cell(row=row, column=10).value = 0.0
    ws_a.cell(row=row, column=11).value = round(ni_yr, 2)
    ws_a.cell(row=row, column=12).value = round(min_cash, 2)
    ws_a.cell(row=row, column=13).value = round(avg_gross_burn, 2)
    ws_a.cell(row=row, column=14).value = round(avg_ebit_burn, 2)
    ws_a.cell(row=row, column=15).value = round(last_cash, 2)
    ws_a.cell(row=row, column=16).value = round(last_deb, 2)
    ws_a.cell(row=row, column=17).value = 0.0   # R&D Intangible (nicht separat aktiviert)
    ws_a.cell(row=row, column=18).value = round(last_cred, 2)
    ws_a.cell(row=row, column=19).value = 0.0   # Debt
    ws_a.cell(row=row, column=20).value = round(net_assets, 2)
    ws_a.cell(row=row, column=21).value = round(cum_eq_year[yo], 2)
    ws_a.cell(row=row, column=22).value = round(cum_pr_year[yo], 2)
    ws_a.cell(row=row, column=23).value = round(quick_ratio, 4)

    print(f"  {year:>6} | {rev_yr:>14,.0f} | {gp_yr:>14,.0f} | {ebit_yr:>14,.0f} | {ni_yr:>12,.0f} | {last_cash:>15,.0f}")

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 3: BalanceSheet
# Spalten:
#  A=Year | B=Cash | C=Debtors | D=R&D Intangible | E=Creditors
#  F=Debt | G=Net Assets | H=Total Equity Investment | I=Cumulative Profit | J=Quick Ratio
# ══════════════════════════════════════════════════════════════════════════════
print("\nBefülle BalanceSheet …")
ws_bs = wb_dst['BalanceSheet']

for yo, year in enumerate(annual_years):
    row = 2 + yo
    _, e = year_months_range(yo)

    cash_end = ending_cash[e - 1] if e > 0 else 0.0
    deb_end  = debtors[e - 1]     if e > 0 else 0.0
    cred_end = creditors[e - 1]   if e > 0 else 0.0
    net_ass  = cash_end + deb_end - cred_end
    qr       = (cash_end / cred_end) if cred_end > 0 else 0.0

    ws_bs.cell(row=row, column=1).value  = year
    ws_bs.cell(row=row, column=2).value  = round(cash_end, 2)
    ws_bs.cell(row=row, column=3).value  = round(deb_end, 2)
    ws_bs.cell(row=row, column=4).value  = 0.0
    ws_bs.cell(row=row, column=5).value  = round(cred_end, 2)
    ws_bs.cell(row=row, column=6).value  = 0.0
    ws_bs.cell(row=row, column=7).value  = round(net_ass, 2)
    ws_bs.cell(row=row, column=8).value  = round(cum_eq_year[yo], 2)
    ws_bs.cell(row=row, column=9).value  = round(cum_pr_year[yo], 2)
    ws_bs.cell(row=row, column=10).value = round(qr, 4)

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 4: Prompt Business Model_Slides – Summary KPIs
# ══════════════════════════════════════════════════════════════════════════════
print("Aktualisiere Prompt-Sheet …")
ws_prompt = wb_dst['Prompt Business Model_Slides']

total_rev_all  = sum(total_revenue)
final_rev      = sum(total_revenue[45:52])     # 2030 Jan–Jul
final_ni       = sum(net_income[45:52])
cum_ni_all     = sum(net_income)
total_eq_all   = sum(equity_funding)
peak_cash      = max(ending_cash)
lowest_cash    = min(ending_cash)
avg_burn       = sum(total_cogs[i] + total_opex[i] for i in range(52)) / 52
neg_ebit_m     = [ebitda[i] for i in range(52) if ebitda[i] < 0]
avg_ebit_burn  = sum(neg_ebit_m) / len(neg_ebit_m) if neg_ebit_m else 0.0
first_neg      = next((i for i, c in enumerate(ending_cash) if c < 0), None)
first_neg_lbl  = f"M{first_neg + 1} (2026)" if first_neg is not None else "nie"

breakeven_year = "nicht erreicht (in 52 Monaten)"
for yo, year in enumerate(annual_years):
    s, e = year_months_range(yo)
    if sum(net_income[s:e]) > 0:
        breakeven_year = str(year)
        break

prompt_updates = {
    7:  "- Währung: EUR",
    8:  "- Unternehmenstyp: B2B SaaS / AI-First (LoopforgeLab GmbH)",
    11: f"- Gesamtumsatz über den Zeitraum (M1–M52): {total_rev_all:,.0f} €",
    12: f"- Umsatz letztes erfasstes Jahr (2030, Jan–Jul): {final_rev:,.0f} €",
    13: f"- Nettoergebnis letztes erfasstes Jahr: {final_ni:,.0f} €",
    14: f"- Kumuliertes Nettoergebnis: {cum_ni_all:,.0f} €",
    15: f"- Gesamtes eingesetztes Eigenkapital: {total_eq_all:,.0f} €",
    16: "- Nicht-verwässernde Förderung: 0 €",
    19: f"- Höchster Cash-Stand: {peak_cash:,.0f} €",
    20: f"- Niedrigster Cash-Stand: {lowest_cash:,.0f} €",
    21: f"- Erster Monat mit negativem Cash: {first_neg_lbl}",
    22: f"- Durchschnittliche monatliche Gesamtausgaben: {avg_burn:,.0f} €",
    23: f"- Durchschnittlicher EBIT-Burn (Verlustmonate): {avg_ebit_burn:,.0f} €",
    27: f"- Break-even Jahr (Nettogewinn): {breakeven_year}",
}
for row_num, text in prompt_updates.items():
    ws_prompt.cell(row=row_num, column=1).value = text

# ── Speichern ─────────────────────────────────────────────────────────────────
wb_dst.save(OUTPUT)
print(f"\n✓ Datei gespeichert: {os.path.basename(OUTPUT)}")

# ══════════════════════════════════════════════════════════════════════════════
# VALIDIERUNG – C13 Stress-Test
# ══════════════════════════════════════════════════════════════════════════════
print("\n═══════════════════════════════════════════════════════")
print("VALIDIERUNG (C13 Stress-Test)")
print("═══════════════════════════════════════════════════════")

# 1. Cash-Check
neg_months = [(i+1, ending_cash[i]) for i in range(52) if ending_cash[i] < 0]
if neg_months:
    print(f"⚠  CASH NEGATIV in {len(neg_months)} Monat(en):")
    for m, c in neg_months[:5]:
        print(f"     M{m}: {c:,.0f} €")
    if len(neg_months) <= 3:
        print("   → Pre-Seed-Phase vor Ideation-Funding (erwartet)")
else:
    print("✓  Cash-Check: Closing Cash in allen Monaten ≥ 0")

# 2. R&D-Amortisation (nicht separat modelliert → Hinweis)
print("ℹ  R&D Intangible: nicht als separater Bilanzposten modelliert")
print("   → Entwicklungskosten fließen direkt in Operating Cash Out")

# 3. Founder Bonus
profitable_months = [i+1 for i in range(52) if net_income[i] > 0]
if profitable_months:
    print(f"✓  Erste profitable Monate: M{profitable_months[0]}")
    print(f"   → Founder Bonus erst ab M{profitable_months[0]} zulässig (aktuell: 0 €)")
else:
    print("ℹ  Keine profitablen Monate in 52-Monats-Zeitraum")

# Zusatz: Break-even
print(f"\n  Break-even Jahr (Net Income > 0 im Jahr): {breakeven_year}")
print(f"  Kumuliertes Net Income M1–M52:            {cum_ni_all:>15,.0f} €")
print(f"  Peak Cash (M52):                          {peak_cash:>15,.0f} €")
print(f"  Lowest Cash (M1–M52):                     {lowest_cash:>15,.0f} €")
print("═══════════════════════════════════════════════════════")
print(f"\n  Quelldatei: {os.path.basename(SOURCE)}")
print(f"  Ausgabedatei: {os.path.basename(OUTPUT)}")
