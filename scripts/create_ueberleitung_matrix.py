"""
Erstellt Überleitung-Matrix Excel:
  Quelle: 260315_LFL_BM_Vorlage_normal_v19.xlsx
  Ziel:   LFL_BM_C13_Normal_v21_20260315.xlsx
  Output: LFL_BM_Ueberleitung_Matrix_v21.xlsx
"""

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side)
import os

BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SRC  = os.path.join(BASE, '260315_LFL_BM_Vorlage_normal_v19.xlsx')
DST  = os.path.join(BASE, 'LFL_BM_C13_Normal_v21_20260315.xlsx')
OUT  = os.path.join(BASE, 'LFL_BM_Ueberleitung_Matrix_v21.xlsx')

# ── Quelldaten lesen ──────────────────────────────────────────────────────────
wb_src = openpyxl.load_workbook(SRC, data_only=True)
ws_rev = wb_src['4_Revenue']
ws_pl  = wb_src['6_P&L']
ws_cf  = wb_src['7_BS_CF']

# M1=Apr 2026, Monatsnamen
MONTH_NAMES = ['Apr 26','Mai 26','Jun 26','Jul 26','Aug 26','Sep 26','Okt 26','Nov 26','Dez 26',
               'Jan 27','Feb 27','Mär 27','Apr 27','Mai 27','Jun 27','Jul 27','Aug 27','Sep 27',
               'Okt 27','Nov 27','Dez 27','Jan 28','Feb 28','Mär 28','Apr 28','Mai 28','Jun 28',
               'Jul 28','Aug 28','Sep 28','Okt 28','Nov 28','Dez 28','Jan 29','Feb 29','Mär 29',
               'Apr 29','Mai 29','Jun 29','Jul 29','Aug 29','Sep 29','Okt 29','Nov 29','Dez 29',
               'Jan 30','Feb 30','Mär 30','Apr 30','Mai 30','Jun 30','Jul 30']

PHASES = [
    {'name':'Ideation',  's':  0, 'e':  4, 'annual_row': 2},
    {'name':'Pre-Seed',  's':  4, 'e': 16, 'annual_row': 3},
    {'name':'Seed',      's': 16, 'e': 28, 'annual_row': 4},
    {'name':'Series A',  's': 28, 'e': 40, 'annual_row': 5},
    {'name':'Series B',  's': 40, 'e': 52, 'annual_row': 6},
]

# Alle Monatswerte lesen (0-based index)
rev_all = [float(ws_rev.cell(row=32, column=2+i).value or 0) for i in range(52)]
ni_all  = [float(ws_pl.cell(row=37, column=2+i).value or 0) for i in range(52)]
cf_all  = [float(ws_cf.cell(row=15, column=2+i).value or 0) for i in range(52)]

# ── Styles ─────────────────────────────────────────────────────────────────────
def fill(hex_):   return PatternFill('solid', fgColor=hex_)
def font(bold=False, sz=10, color='000000', italic=False):
    return Font(bold=bold, size=sz, color=color, italic=italic)
def align(h='center', v='center', wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)
def border_thin():
    s = Side(style='thin', color='CCCCCC')
    return Border(left=s, right=s, top=s, bottom=s)
def border_medium():
    s = Side(style='medium', color='888888')
    return Border(left=s, right=s, top=s, bottom=s)

# Farben
C_DARK_BLUE  = '1F4E79'   # Phase-Header
C_MID_BLUE   = '2E75B6'   # Spalten-Header
C_LIGHT_BLUE = 'D6E4F0'   # Metrik-Header-Zeile
C_REV        = 'EBF5FB'   # Revenue-Zeile
C_NI         = 'FEF9E7'   # Net-Profit-Zeile
C_CF         = 'F9EBEA'   # Cash-Zeile
C_TOTAL      = 'D5F5E3'   # Gesamt-Spalte
C_TARGET     = 'EDE7F6'   # Ziel-Spalten
C_GREY_CELL  = 'F2F2F2'   # Zelle ausserhalb Phase

FMT_EUR  = '#,##0.00 "€"'
FMT_EUR0 = '#,##0 "€"'
FMT_TEXT = '@'

# ── Workbook erstellen ─────────────────────────────────────────────────────────
wb = openpyxl.Workbook()
ws = wb.active
ws.title = 'Überleitung_Matrix'
ws.sheet_view.showGridLines = False
ws.sheet_properties.tabColor = '1F4E79'

# ── Titel ──────────────────────────────────────────────────────────────────────
ws.merge_cells('A1:AJ1')
tc = ws['A1']
tc.value = 'ÜBERLEITUNG: 260315_LFL_BM_Vorlage_normal_v19.xlsx  →  LFL_BM_C13_Normal_v21_20260315.xlsx'
tc.font  = Font(bold=True, size=14, color='FFFFFF')
tc.fill  = fill(C_DARK_BLUE)
tc.alignment = align('left', 'center')
ws.row_dimensions[1].height = 28

ws.merge_cells('A2:AJ2')
sc = ws['A2']
sc.value = ('Kennzahlen: Revenue = SUM(4_Revenue!Zxx je Phase) | Net Profit = SUM(6_P&L!Zxx je Phase) | '
            'Year-end Cash = Einzelzelle 7_BS_CF!Zxx (letzter Phasenmonat)')
sc.font  = Font(size=9, italic=True, color='444444')
sc.fill  = fill('E8EAF6')
sc.alignment = align('left', 'center')
ws.row_dimensions[2].height = 16

# Leerzeile
ws.row_dimensions[3].height = 6

# ── Feste Spalten (Struktur) ───────────────────────────────────────────────────
# A: Kennzahl  B: Quelle-Sheet  C: Quelle-Zeile  D: Quelle-Zellbereich
# E–BA: M1–M52 (Monatswerte)
# BB: GESAMT / Phasenend-Wert
# BC: Ziel-Sheet (v21)   BD: Ziel-Zeile   BE: Ziel-Spalte (v21)

COL_METRIC    = 1   # A
COL_SRC_SH    = 2   # B
COL_SRC_ROW   = 3   # C
COL_SRC_RANGE = 4   # D
COL_M_START   = 5   # E = M1
COL_TOTAL     = 57  # BB  (5 + 52)
COL_DST_SH    = 58  # BC
COL_DST_ROW   = 59  # BD
COL_DST_COL   = 60  # BE

# Spaltenbreiten
ws.column_dimensions[get_column_letter(COL_METRIC)].width    = 14
ws.column_dimensions[get_column_letter(COL_SRC_SH)].width    = 14
ws.column_dimensions[get_column_letter(COL_SRC_ROW)].width   = 10
ws.column_dimensions[get_column_letter(COL_SRC_RANGE)].width = 18
for mi in range(52):
    ws.column_dimensions[get_column_letter(COL_M_START + mi)].width = 13
ws.column_dimensions[get_column_letter(COL_TOTAL)].width   = 16
ws.column_dimensions[get_column_letter(COL_DST_SH)].width  = 12
ws.column_dimensions[get_column_letter(COL_DST_ROW)].width = 11
ws.column_dimensions[get_column_letter(COL_DST_COL)].width = 14

# Spalten-Header (Zeile 4)
HDR_ROW = 4
def hdr(col, val, w=None):
    c = ws.cell(row=HDR_ROW, column=col, value=val)
    c.font      = Font(bold=True, size=9, color='FFFFFF')
    c.fill      = fill(C_MID_BLUE)
    c.alignment = align('center', 'center', wrap=True)
    c.border    = border_thin()
    if w: ws.column_dimensions[get_column_letter(col)].width = w

hdr(COL_METRIC,    'Kennzahl')
hdr(COL_SRC_SH,    'Quelle\nSheet')
hdr(COL_SRC_ROW,   'Quelle\nZeile')
hdr(COL_SRC_RANGE, 'Quelle\nZellbereich')
for mi in range(52):
    c = ws.cell(row=HDR_ROW, column=COL_M_START + mi,
                value=f'M{mi+1}\n{MONTH_NAMES[mi]}')
    c.font      = Font(bold=True, size=8, color='FFFFFF')
    c.fill      = fill(C_MID_BLUE)
    c.alignment = align('center', 'center', wrap=True)
    c.border    = border_thin()
hdr(COL_TOTAL,   'GESAMT /\nPhasenend-Wert')
hdr(COL_DST_SH,  'Ziel\nSheet (v21)')
hdr(COL_DST_ROW, 'Ziel\nZeile (v21)')
hdr(COL_DST_COL, 'Ziel\nSpalte (v21)')
ws.row_dimensions[HDR_ROW].height = 30

# ── Datenzeilen je Phase ───────────────────────────────────────────────────────
cur_row = HDR_ROW + 1

for ph in PHASES:
    pname = ph['name']
    ps, pe = ph['s'], ph['e']
    n = pe - ps
    arow = ph['annual_row']
    col_s_ltr = get_column_letter(ps + 2)    # Quell-Spalte erster Monat
    col_e_ltr = get_column_letter(pe + 1)    # Quell-Spalte letzter Monat

    # ── Phase-Header-Zeile ────────────────────────────────────────────────────
    end_col_ltr = get_column_letter(COL_DST_COL)
    ws.merge_cells(f'{get_column_letter(1)}{cur_row}:{end_col_ltr}{cur_row}')
    ph_cell = ws.cell(row=cur_row, column=1,
        value=(f'  ▶  {pname.upper()}   |   '
               f'M{ps+1}–M{pe}   |   '
               f'{MONTH_NAMES[ps]} – {MONTH_NAMES[pe-1]}   |   '
               f'Quell-Spalten: {col_s_ltr}–{col_e_ltr}   |   '
               f'{n} Monate'))
    ph_cell.font      = Font(bold=True, size=10, color='FFFFFF')
    ph_cell.fill      = fill(C_DARK_BLUE)
    ph_cell.alignment = align('left', 'center')
    ws.row_dimensions[cur_row].height = 20
    cur_row += 1

    # ── 3 Metrik-Zeilen ───────────────────────────────────────────────────────
    metrics = [
        {
            'label':      'Revenue',
            'src_sheet':  '4_Revenue',
            'src_row':    32,
            'src_range':  f'{col_s_ltr}32:{col_e_ltr}32',
            'values':     rev_all,
            'is_sum':     True,
            'row_fill':   C_REV,
            'dst_sheet':  'Annual',
            'dst_row':    arow,
            'dst_col':    'B (Revenue)',
        },
        {
            'label':      'Net Profit',
            'src_sheet':  '6_P&L',
            'src_row':    37,
            'src_range':  f'{col_s_ltr}37:{col_e_ltr}37',
            'values':     ni_all,
            'is_sum':     True,
            'row_fill':   C_NI,
            'dst_sheet':  'Annual',
            'dst_row':    arow,
            'dst_col':    'K (Net Profit)',
        },
        {
            'label':      'Year-end Cash',
            'src_sheet':  '7_BS_CF',
            'src_row':    15,
            'src_range':  f'{col_e_ltr}15',   # Einzelzelle: letzter Phasenmonat
            'values':     cf_all,
            'is_sum':     False,               # Einzelwert, kein SUM
            'row_fill':   C_CF,
            'dst_sheet':  'Annual',
            'dst_row':    arow,
            'dst_col':    'O (Cash year-end)',
        },
    ]

    for mt in metrics:
        ws.row_dimensions[cur_row].height = 18

        # Strukturspalten
        def set_fixed(col, val, bg, bold=False, fmt=None, ha='left'):
            c = ws.cell(row=cur_row, column=col, value=val)
            c.font      = Font(bold=bold, size=9, color='1A1A1A')
            c.fill      = fill(bg)
            c.alignment = align(ha, 'center')
            c.border    = border_thin()
            if fmt: c.number_format = fmt

        set_fixed(COL_METRIC,    mt['label'],     mt['row_fill'], bold=True)
        set_fixed(COL_SRC_SH,    mt['src_sheet'], mt['row_fill'])
        set_fixed(COL_SRC_ROW,   mt['src_row'],   mt['row_fill'], ha='center')
        set_fixed(COL_SRC_RANGE, mt['src_range'], mt['row_fill'])

        # Monatswerte
        phase_total = 0.0
        for mi in range(52):
            col = COL_M_START + mi
            in_phase = (ps <= mi < pe)
            val = mt['values'][mi]

            if not in_phase:
                # Grau: ausserhalb dieser Phase
                c = ws.cell(row=cur_row, column=col, value='')
                c.fill      = fill(C_GREY_CELL)
                c.border    = border_thin()
            else:
                if not mt['is_sum'] and mi != pe - 1:
                    # Year-end Cash: nur letzter Monat relevant
                    c = ws.cell(row=cur_row, column=col, value='–')
                    c.fill      = fill(mt['row_fill'])
                    c.alignment = align('center', 'center')
                    c.font      = Font(size=9, color='999999', italic=True)
                    c.border    = border_thin()
                else:
                    c = ws.cell(row=cur_row, column=col, value=round(val, 2))
                    c.fill          = fill(mt['row_fill'])
                    c.number_format = FMT_EUR0
                    c.alignment     = align('right', 'center')
                    c.font          = Font(size=9)
                    c.border        = border_thin()
                    if mt['is_sum']:
                        phase_total += val
                    else:
                        phase_total = val   # Einzelwert

        # Gesamtspalte
        tc = ws.cell(row=cur_row, column=COL_TOTAL, value=round(phase_total, 2))
        tc.font          = Font(bold=True, size=9, color='1A3A1A')
        tc.fill          = fill(C_TOTAL)
        tc.number_format = FMT_EUR0
        tc.alignment     = align('right', 'center')
        tc.border        = border_medium()

        # Ziel-Spalten
        set_fixed(COL_DST_SH,  mt['dst_sheet'],   C_TARGET, ha='center')
        set_fixed(COL_DST_ROW, f'Zeile {mt["dst_row"]}', C_TARGET, ha='center')
        set_fixed(COL_DST_COL, mt['dst_col'],      C_TARGET)

        cur_row += 1

    # Leerzeile zwischen Phasen
    ws.row_dimensions[cur_row].height = 6
    cur_row += 1

# ── Gesamt-Summen-Zeile ────────────────────────────────────────────────────────
cur_row += 1
totals_label_row = cur_row

# Merge Label
ws.merge_cells(f'A{cur_row}:D{cur_row}')
tc = ws.cell(row=cur_row, column=1, value='GESAMT M1–M52 (alle Phasen)')
tc.font      = Font(bold=True, size=10, color='FFFFFF')
tc.fill      = fill(C_DARK_BLUE)
tc.alignment = align('left', 'center')
ws.row_dimensions[cur_row].height = 22

grand_rev  = sum(rev_all)
grand_ni   = sum(ni_all)
grand_cash = cf_all[51]  # Letzter Monat M52

# Revenue Gesamt
for col in range(COL_M_START, COL_M_START + 52):
    c = ws.cell(row=cur_row, column=col, value=round(rev_all[col - COL_M_START], 2))
    c.font = Font(bold=True, size=8, color='FFFFFF')
    c.fill = fill(C_DARK_BLUE)
    c.number_format = FMT_EUR0
    c.alignment = align('right', 'center')
    c.border = border_thin()

c = ws.cell(row=cur_row, column=COL_TOTAL,
            value=f'Revenue: {grand_rev:,.0f} €  |  NI: {grand_ni:,.0f} €  |  Cash M52: {grand_cash:,.0f} €')
c.font      = Font(bold=True, size=9, color='FFFFFF')
c.fill      = fill(C_DARK_BLUE)
c.alignment = align('center', 'center', wrap=True)
c.border    = border_medium()
ws.merge_cells(f'{get_column_letter(COL_TOTAL)}{cur_row}:{get_column_letter(COL_DST_COL)}{cur_row}')

# ── Zell-Mapping Referenz-Sheet ────────────────────────────────────────────────
wb2 = openpyxl.load_workbook(DST, data_only=True)

# Neues Sheet: Referenz-Tabelle Annual v21
ws_ref = wb.create_sheet('Referenz_Annual_v21')
ws_ref.sheet_properties.tabColor = '2E75B6'
ws_ref.sheet_view.showGridLines = False

# Header
ws_ref.merge_cells('A1:P1')
h = ws_ref['A1']
h.value = 'Annual-Sheet LFL_BM_C13_Normal_v21_20260315.xlsx – Vollständige Werte je Phase'
h.font  = Font(bold=True, size=12, color='FFFFFF')
h.fill  = fill(C_DARK_BLUE)
h.alignment = align('left', 'center')
ws_ref.row_dimensions[1].height = 24

# Annual-Spalten aus v21
annual_headers = ['Phase','Revenue','Cost of Revenue','Gross Profit','Total Opex',
                  'EBIT','Interest','Tax','Founder Bonus','Grants',
                  'Net Profit','Min Cash','Avg Mo. Burn','Avg EBIT Burn',
                  'Cash (Phase-end)','Debtors']
ws_annual_src = wb2['Annual']

for c_idx, h in enumerate(annual_headers, 1):
    cell = ws_ref.cell(row=2, column=c_idx, value=h)
    cell.font      = Font(bold=True, size=9, color='FFFFFF')
    cell.fill      = fill(C_MID_BLUE)
    cell.alignment = align('center', 'center', wrap=True)
    cell.border    = border_thin()
    ws_ref.column_dimensions[get_column_letter(c_idx)].width = 16
ws_ref.row_dimensions[2].height = 28

phase_fills = [C_REV, C_NI, C_CF, C_LIGHT_BLUE, C_TARGET]
phase_names_order = ['Ideation','Pre-Seed','Seed','Series A','Series B']

for i, pn in enumerate(phase_names_order):
    row = 3 + i
    ws_ref.row_dimensions[row].height = 18
    src_row = 2 + i
    bg = phase_fills[i % len(phase_fills)]

    # Phase-Name
    c = ws_ref.cell(row=row, column=1, value=pn)
    c.font = Font(bold=True, size=9); c.fill = fill(bg)
    c.border = border_thin(); c.alignment = align('left','center')

    # Werte aus v21 Annual
    for col_idx in range(2, 17):
        val = ws_annual_src.cell(row=src_row, column=col_idx).value
        c2 = ws_ref.cell(row=row, column=col_idx, value=val)
        c2.font = Font(size=9); c2.fill = fill(bg)
        c2.border = border_thin()
        c2.alignment = align('right','center')
        if isinstance(val, (int, float)):
            c2.number_format = FMT_EUR0

# Zelladressen-Legende
ws_ref.row_dimensions[10].height = 8
legend = [
    ('B', 'Revenue'),('C','Cost of Revenue'),('D','Gross Profit'),('E','Total Opex'),
    ('F','EBIT'),('H','Tax'),('K','Net Profit'),('L','Min Cash'),
    ('M','Avg Monthly Burn'),('N','Avg EBIT Burn'),('O','Cash Phase-end'),('P','Debtors'),
]
ws_ref.cell(row=11, column=1, value='Spalten-Mapping Annual-Sheet (v21):').font = Font(bold=True, size=9)
for i, (col_l, desc) in enumerate(legend):
    c = ws_ref.cell(row=12 + i//4, column=1 + (i%4)*2,
                    value=f'Spalte {col_l}: {desc}')
    c.font = Font(size=9, italic=True, color='333333')

# ── Speichern ─────────────────────────────────────────────────────────────────
wb.save(OUT)
print(f'✓ Gespeichert: {os.path.basename(OUT)}')
print(f'  Sheets: {wb.sheetnames}')
print(f'  Zeilen in Matrix: {cur_row}')
