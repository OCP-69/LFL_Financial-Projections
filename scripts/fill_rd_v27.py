"""
Recalculate R&D for ALL sheets starting from v25.
Output: LFL_BM_C13_Normal_redacted_v27_20260315.xlsx

R&D Components (per Nutzeranforderung):
  1. CTO Gehalt (AG-Brutto)         ← 5_Costs R7 (direkt)
  2. SW Developer (1.+2.) Gehalt    ← berechnet aus 2_Inputs (M6, M21)
  3. Mechanical/Domain Eng. Gehalt  ← berechnet aus 2_Inputs (M11)
  4. Hardware-Renting (Engineering) ← Engineering-HC × 89 EUR/MA/Mo
  5. AI/ML APIs                     ← 5_Costs R16 (direkt)
  6. SaaS Tools & Lizenzen          ← 5_Costs R17 (direkt)

Total R&D M5-M52: 1,575,310 EUR
Amortisation: 36 Monate linear → NBV (M52): berechnet
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import shutil, os, math

SRC_COSTS = '260315_LFL_BM_Vorlage_normal_redacted_final.xlsx'
SRC_V25   = 'LFL_BM_C13_Normal_redacted_v25_20260315.xlsx'
OUT       = 'LFL_BM_C13_Normal_redacted_v27_20260315.xlsx'
AMORT_M   = 36

# ── Styles ────────────────────────────────────────────────────────────────────
DARK_BLUE   = '1F3864'
MID_BLUE    = '2E75B6'
TEAL        = '1F7391'
GREEN_DARK  = '375623'
GREEN_LIGHT = 'E2EFDA'
BLUE_LIGHT  = 'DDEEFF'
ORANGE_LIGHT= 'FCE4D6'
WHITE       = 'FFFFFF'

def fc(h): return PatternFill('solid', fgColor=h)
def tb():
    s = Side(style='thin', color='BFBFBF')
    return Border(left=s, right=s, top=s, bottom=s)

FMT_EUR = '#,##0.00'

# ── Load source data ──────────────────────────────────────────────────────────
print('Loading source data...')
wb_src    = openpyxl.load_workbook(SRC_COSTS, data_only=True)
ws_costs  = wb_src['5_Costs']
ws_inputs = wb_src['2_Inputs']

ag_faktor   = 1 + float(ws_inputs.cell(85, 2).value)   # 1.22
raise_pa    = float(ws_inputs.cell(84, 2).value)        # 0.03
hw_per_ma   = float(ws_inputs.cell(197, 2).value)       # 89.0

sw_salary   = float(ws_inputs.cell(71, 2).value)        # 75,000
mech_salary = float(ws_inputs.cell(74, 2).value)        # 68,000
sw1_month   = int(ws_inputs.cell(177, 2).value)         # 6
sw2_month   = int(ws_inputs.cell(178, 2).value)         # 21
mech_month  = int(ws_inputs.cell(181, 2).value)         # 11

def emp_monthly(hire_m, annual_salary, m):
    """AG-Brutto monthly cost for employee hired at hire_m."""
    if m < hire_m: return 0.0
    return (annual_salary * ag_faktor / 12) * (1 + raise_pa) ** math.floor((m - 1) / 12)

# ── Collect monthly R&D data M5-M52 (→ target M1-M48) ───────────────────────
rd_data = []
for m in range(5, 53):
    col = m + 1
    cto    = float(ws_costs.cell(7,  col).value or 0)
    sw1    = emp_monthly(sw1_month,  sw_salary,   m)
    sw2    = emp_monthly(sw2_month,  sw_salary,   m)
    mech   = emp_monthly(mech_month, mech_salary, m)
    eng_hc = (1 +
              (1 if m >= sw1_month  else 0) +
              (1 if m >= sw2_month  else 0) +
              (1 if m >= mech_month else 0))
    hw_eng = eng_hc * hw_per_ma
    aiml   = float(ws_costs.cell(16, col).value or 0)
    saas   = float(ws_costs.cell(17, col).value or 0)
    total  = cto + sw1 + sw2 + mech + hw_eng + aiml + saas
    rd_data.append({
        'src_m': m, 'tgt_m': m - 4,
        'cto': cto, 'sw1': sw1, 'sw2': sw2, 'mech': mech,
        'hw_eng': hw_eng, 'aiml': aiml, 'saas': saas,
        'total': total,
    })

# ── Amortization: net book value per month ────────────────────────────────────
gross_cum, amort_cum, net_nbv = [], [], []
for i, d in enumerate(rd_data):
    g = sum(x['total'] for x in rd_data[:i+1])
    a = sum(min(i - j, AMORT_M) / AMORT_M * rd_data[j]['total'] for j in range(i+1))
    gross_cum.append(g)
    amort_cum.append(a)
    net_nbv.append(g - a)

total_rd = sum(d['total'] for d in rd_data)
print(f'  Total R&D (M5-M52): {total_rd:,.2f} EUR')
print(f'  Final NBV (M52):    {net_nbv[-1]:,.2f} EUR')

# ── Copy v25 → v27 ────────────────────────────────────────────────────────────
print(f'Copying {SRC_V25} → {OUT}')
shutil.copy2(SRC_V25, OUT)
wb = openpyxl.load_workbook(OUT)

# ── Build target-month → Monthly-row map ─────────────────────────────────────
ws_mo = wb['Monthly']
tgt_to_row = {}
yr_to_tgt  = {}   # year → list of target-month indices (0-based)
for r in range(2, 50):
    m_val = ws_mo.cell(r, 2).value
    yr_val = ws_mo.cell(r, 1).value
    if m_val is not None:
        tgt_to_row[int(m_val)] = r
        yr_to_tgt.setdefault(yr_val, []).append(int(m_val) - 1)  # 0-based index into rd_data

# ── 1. MONTHLY SHEET: add R&D columns ────────────────────────────────────────
print('Filling Monthly sheet...')

# Column layout: after existing O(15), add P onward
COLS = {
    'total':  (16, 'R&D Gesamt',             DARK_BLUE,  WHITE),
    'cto':    (17, 'CTO Gehalt\n(AG-Brutto)', '4472C4',   WHITE),
    'sw1':    (18, 'SW Dev 1\n(AG-Brutto)',   '4472C4',   WHITE),
    'sw2':    (19, 'SW Dev 2\n(AG-Brutto)',   '4472C4',   WHITE),
    'mech':   (20, 'Mech. Eng.\n(AG-Brutto)', '4472C4',   WHITE),
    'hw_eng': (21, 'HW-Renting\n(Eng.)',       '8EA9C1',   WHITE),
    'aiml':   (22, 'AI/ML APIs',               TEAL,       WHITE),
    'saas':   (23, 'SaaS Tools',               TEAL,       WHITE),
}

for key, (col, label, hdr_bg, hdr_fg) in COLS.items():
    cell = ws_mo.cell(1, col)
    cell.value     = label
    cell.font      = Font(name='Calibri', bold=True, color=hdr_fg, size=8)
    cell.fill      = fc(hdr_bg)
    cell.alignment = Alignment(horizontal='center', wrap_text=True)
    cell.border    = tb()
    ws_mo.column_dimensions[get_column_letter(col)].width = 14
ws_mo.row_dimensions[1].height = 32

PHASE_BG = {range(1,13): '9DC3E6', range(13,25): 'A9D18E',
            range(25,37): 'FFD966', range(37,49): 'F4B183'}
def p_bg(tm):
    for rng, c in PHASE_BG.items():
        if tm in rng: return c
    return 'FFFFFF'

for i, d in enumerate(rd_data):
    row = tgt_to_row.get(d['tgt_m'])
    if not row: continue
    for key, (col, *_) in COLS.items():
        cell = ws_mo.cell(row, col)
        cell.value         = round(d[key], 4)
        cell.font          = Font(name='Calibri', size=9,
                                  bold=(key == 'total'))
        cell.fill          = fc('EAF2FF' if key == 'total' else p_bg(d['tgt_m']))
        cell.border        = tb()
        cell.alignment     = Alignment(horizontal='right')
        cell.number_format = FMT_EUR

# ── 2. ANNUAL SHEET ───────────────────────────────────────────────────────────
print('Filling Annual sheet...')
ws_an = wb['Annual']

# Find year → Annual row
yr_to_an = {}
for r in range(2, 15):
    v = ws_an.cell(r, 1).value
    if v:
        try: yr_to_an[int(str(v).split('\n')[0])] = r
        except: pass

for yr, an_row in sorted(yr_to_an.items()):
    indices = yr_to_tgt.get(yr, [])
    if not indices: continue
    annual_rd   = sum(rd_data[i]['total'] for i in indices)
    last_i      = max(indices)
    yr_end_nbv  = net_nbv[last_i]

    # Col 17: annual gross R&D spend (new investment this year)
    c17 = ws_an.cell(an_row, 17)
    c17.value         = round(annual_rd, 2)
    c17.font          = Font(name='Calibri', size=9, bold=True, color=GREEN_DARK)
    c17.fill          = fc(GREEN_LIGHT)
    c17.border        = tb()
    c17.alignment     = Alignment(horizontal='right')
    c17.number_format = FMT_EUR

    # Col 20: Net Assets + R&D NBV
    old_na = float(ws_an.cell(an_row, 20).value or 0)
    new_na = round(old_na + yr_end_nbv, 2)
    c20 = ws_an.cell(an_row, 20)
    c20.value         = new_na
    c20.font          = Font(name='Calibri', size=9, bold=True)
    c20.border        = tb()
    c20.alignment     = Alignment(horizontal='right')
    c20.number_format = FMT_EUR

    print(f'  {yr}: Annual R&D={annual_rd:>12,.2f}  NBV={yr_end_nbv:>12,.2f}  Net Assets={new_na:>14,.2f}')

# ── 3. BALANCE SHEET ─────────────────────────────────────────────────────────
print('Filling BalanceSheet...')
ws_bs = wb['BalanceSheet']

yr_to_bs = {}
for r in range(2, 15):
    v = ws_bs.cell(r, 1).value
    if v:
        try: yr_to_bs[int(str(v).split('\n')[0])] = r
        except: pass

for yr, bs_row in sorted(yr_to_bs.items()):
    indices = yr_to_tgt.get(yr, [])
    if not indices: continue
    last_i     = max(indices)
    yr_end_nbv = net_nbv[last_i]

    # Col 4: R&D Intangible (net book value)
    c4 = ws_bs.cell(bs_row, 4)
    c4.value         = round(yr_end_nbv, 2)
    c4.font          = Font(name='Calibri', size=9, bold=True, color=GREEN_DARK)
    c4.fill          = fc(GREEN_LIGHT)
    c4.border        = tb()
    c4.alignment     = Alignment(horizontal='right')
    c4.number_format = FMT_EUR

    # Col 7: Net Assets updated
    old_na = float(ws_bs.cell(bs_row, 7).value or 0)
    new_na = round(old_na + yr_end_nbv, 2)
    c7 = ws_bs.cell(bs_row, 7)
    c7.value         = new_na
    c7.font          = Font(name='Calibri', size=9, bold=True)
    c7.border        = tb()
    c7.alignment     = Alignment(horizontal='right')
    c7.number_format = FMT_EUR

    print(f'  {yr}: R&D NBV={yr_end_nbv:>12,.2f}  Net Assets={new_na:>14,.2f}')

# Note
last_bs = max(yr_to_bs.values()) + 2
note = ws_bs.cell(last_bs, 1)
note.value = ('R&D Intangible = Kum. Netto-Buchwert nach Amortisation (36 Mo. linear). '
              'Komponenten: CTO-Gehalt + SW Dev 1+2 + Mech. Eng. + HW-Renting (Eng.) + AI/ML APIs + SaaS Tools.')
note.font = Font(name='Calibri', size=8, italic=True, color='595959')
try: ws_bs.merge_cells(f'A{last_bs}:J{last_bs}')
except: pass

# ── 4. R&D_BREAKDOWN SHEET ───────────────────────────────────────────────────
print('Creating R&D_Breakdown sheet...')
if 'R&D_Breakdown' in wb.sheetnames: del wb['R&D_Breakdown']
ws_rd = wb.create_sheet('R&D_Breakdown')

# Title
ws_rd.merge_cells('A1:J1')
ws_rd['A1'] = 'R&D Kostenallokation — Vollständige Aufstellung (M1–M48 = Quelle M5–M52)'
ws_rd['A1'].font      = Font(name='Calibri', bold=True, size=12, color=WHITE)
ws_rd['A1'].fill      = fc(DARK_BLUE)
ws_rd['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws_rd.row_dimensions[1].height = 26

ws_rd.merge_cells('A2:J2')
ws_rd['A2'] = ('Komponenten: CTO (AG-Brutto) · SW Dev 1+2 · Mech. Eng. · HW-Renting Eng. · '
               'AI/ML APIs · SaaS Tools  |  Amortisation: 36 Monate linear')
ws_rd['A2'].font      = Font(name='Calibri', size=9, italic=True, color='595959')
ws_rd['A2'].alignment = Alignment(horizontal='center')
ws_rd.row_dimensions[2].height = 14

# Header row R4
rd_cols = ['Tgt M','Src M','CTO (AG-Br.)', 'SW Dev 1', 'SW Dev 2',
           'Mech. Eng.', 'HW Renting\n(Eng.)', 'AI/ML APIs', 'SaaS Tools',
           'Total R&D', 'Kum. Brutto', 'Kum. Amort.', 'NBV (Netto)']
for ci, h in enumerate(rd_cols, 1):
    cell = ws_rd.cell(4, ci)
    cell.value     = h
    cell.font      = Font(name='Calibri', bold=True, color=WHITE, size=8)
    cell.fill      = fc(MID_BLUE)
    cell.alignment = Alignment(horizontal='center', wrap_text=True)
    cell.border    = tb()
ws_rd.row_dimensions[4].height = 30
for ci, w in enumerate([7,7,14,12,12,12,12,12,12,12,14,14,14], 1):
    ws_rd.column_dimensions[get_column_letter(ci)].width = w

# Data rows
for i, d in enumerate(rd_data):
    row = i + 5
    bg = p_bg(d['tgt_m'])
    vals = [d['tgt_m'], d['src_m'],
            d['cto'], d['sw1'], d['sw2'], d['mech'],
            d['hw_eng'], d['aiml'], d['saas'],
            d['total'], gross_cum[i], amort_cum[i], net_nbv[i]]
    for ci, v in enumerate(vals, 1):
        cell = ws_rd.cell(row, ci)
        cell.value  = round(v, 2) if isinstance(v, float) else v
        cell.fill   = fc(bg)
        cell.border = tb()
        if ci <= 2:
            cell.font      = Font(name='Calibri', size=8, bold=(ci==1))
            cell.alignment = Alignment(horizontal='center')
            cell.number_format = '0'
        else:
            cell.font          = Font(name='Calibri', size=8,
                                      bold=(ci == 10))  # bold Total
            cell.alignment     = Alignment(horizontal='right')
            cell.number_format = FMT_EUR

# Phase summary
PHASES = [('Pre-Seed','M1–M12',0,12),('Seed','M13–M24',12,24),
          ('Series A','M25–M36',24,36),('Series B','M37–M48',36,48)]
SUM_ROW = len(rd_data) + 6

ws_rd.merge_cells(f'A{SUM_ROW}:M{SUM_ROW}')
ws_rd.cell(SUM_ROW, 1).value = 'PHASEN-ZUSAMMENFASSUNG'
ws_rd.cell(SUM_ROW, 1).font  = Font(name='Calibri', bold=True, size=10, color=DARK_BLUE)
SUM_ROW += 1

ph_hdr = ['Phase','Monate','CTO','SW Dev 1','SW Dev 2','Mech. Eng.',
          'HW Renting','AI/ML','SaaS','Total R&D','Phase-End NBV']
for ci, h in enumerate(ph_hdr, 1):
    cell = ws_rd.cell(SUM_ROW, ci)
    cell.value     = h
    cell.font      = Font(name='Calibri', bold=True, color=WHITE, size=9)
    cell.fill      = fc(DARK_BLUE)
    cell.border    = tb()
    cell.alignment = Alignment(horizontal='center', wrap_text=True)
SUM_ROW += 1

grand = {'cto':0,'sw1':0,'sw2':0,'mech':0,'hw_eng':0,'aiml':0,'saas':0,'total':0}
for ph_name, ph_mstr, i_s, i_e in PHASES:
    seg = rd_data[i_s:i_e]
    ph  = {k: sum(d[k] for d in seg) for k in grand}
    nbv = net_nbv[i_e - 1]
    bg  = p_bg(i_s + 1)
    for grand_k in grand: grand[grand_k] += ph[grand_k]
    vals = [ph_name, ph_mstr, ph['cto'], ph['sw1'], ph['sw2'], ph['mech'],
            ph['hw_eng'], ph['aiml'], ph['saas'], ph['total'], nbv]
    for ci, v in enumerate(vals, 1):
        cell = ws_rd.cell(SUM_ROW, ci)
        cell.value  = round(v, 2) if isinstance(v, float) else v
        cell.fill   = fc(bg)
        cell.border = tb()
        cell.font   = Font(name='Calibri', size=9, bold=(ci in (1, 10)))
        cell.alignment = Alignment(horizontal='center' if ci <= 2 else 'right')
        if ci > 2: cell.number_format = FMT_EUR
    SUM_ROW += 1

# Grand total row
gt_vals = ['GESAMT','M1–M48',grand['cto'],grand['sw1'],grand['sw2'],grand['mech'],
           grand['hw_eng'],grand['aiml'],grand['saas'],grand['total'],net_nbv[-1]]
for ci, v in enumerate(gt_vals, 1):
    cell = ws_rd.cell(SUM_ROW, ci)
    cell.value     = round(v, 2) if isinstance(v, float) else v
    cell.fill      = fc(DARK_BLUE)
    cell.font      = Font(name='Calibri', bold=True, color=WHITE, size=10)
    cell.border    = tb()
    cell.alignment = Alignment(horizontal='center' if ci <= 2 else 'right')
    if ci > 2: cell.number_format = FMT_EUR

# Component category totals
SUM_ROW += 2
cat_A = grand['sw1'] + grand['sw2'] + grand['mech'] + grand['aiml'] + grand['saas']
cat_B = grand['cto'] + grand['hw_eng']
rows_cat = [
    ('Personal (Engineering-Team)', cat_A,
     'SW Dev 1+2 + Mech. Eng. + AI/ML APIs + SaaS Tools', 'A9D18E'),
    ('CTO + Hardware-Renting', cat_B,
     'CTO AG-Brutto + Hardware-Renting Engineering', '9DC3E6'),
    ('TOTAL R&D', grand['total'],
     'Alle Komponenten', DARK_BLUE),
]
for label, val, desc, bg in rows_cat:
    ws_rd.cell(SUM_ROW, 1).value  = label
    ws_rd.cell(SUM_ROW, 1).font   = Font(name='Calibri', bold=True, size=9)
    ws_rd.cell(SUM_ROW, 1).fill   = fc(bg)
    ws_rd.cell(SUM_ROW, 1).font   = Font(name='Calibri', bold=True, size=9,
                                          color=WHITE if bg == DARK_BLUE else '000000')
    ws_rd.cell(SUM_ROW, 2).value  = round(val, 2)
    ws_rd.cell(SUM_ROW, 2).number_format = FMT_EUR
    ws_rd.cell(SUM_ROW, 2).font   = Font(name='Calibri', bold=True, size=9)
    ws_rd.cell(SUM_ROW, 3).value  = desc
    ws_rd.cell(SUM_ROW, 3).font   = Font(name='Calibri', size=9, italic=True, color='595959')
    SUM_ROW += 1

# ── Save ──────────────────────────────────────────────────────────────────────
print(f'Saving {OUT}...')
wb.save(OUT)
kb = os.path.getsize(OUT) / 1024
print(f'✓ Fertig: {OUT}  ({kb:.0f} KB)')
print()
print('=== Gesamtergebnis ===')
print(f'  Total R&D M1–M48:  {total_rd:>14,.2f} EUR')
print(f'  NBV Ende M48:      {net_nbv[-1]:>14,.2f} EUR')
for k in ['cto','sw1','sw2','mech','hw_eng','aiml','saas']:
    s = sum(d[k] for d in rd_data)
    print(f'    {k:8s}: {s:>12,.2f} EUR  ({s/total_rd*100:.1f}%)')
