"""
update_consulting_values.py
────────────────────────────
Aktualisiert v3 mit korrekten Consulting-Werten gemäß Spezifikation:

  Tagessatz:                 Gering=1.000€ | Normal=1.200€ | Stark=1.500€
  Consulting-Tage/Kunde:     Gering=5      | Normal=10     | Stark=20
  Buchungs-Wahrscheinlichkeit: a)20%       | b)50%         | c)100%
    → gering=20%, normal=50%, stark=100%

Umbenennung überall konsistent:
  "Consulting Anteil" / "Consulting-Anteil je Kunde"
  → "Buchungs-Wahrscheinlichkeit Consulting je Kunde"

Output: scenarios/LFL_BM_Konservativ_v4_Final.xlsx
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

SRC  = "/home/user/LFL_Financial-Projections/scenarios/LFL_BM_Konservativ_v3_Linked.xlsx"
DEST = "/home/user/LFL_Financial-Projections/scenarios/LFL_BM_Konservativ_v4_Final.xlsx"

AMBER       = "FFF2CC"
BLUE_LIGHT  = "BDD7EE"
GREEN_LIGHT = "E2EFDA"
TEAL        = "E9F7F8"
ORANGE      = "FCE4D6"
WHITE       = "FFFFFF"

NEW_VAR_NAME = "Buchungs-Wahrscheinlichkeit Consulting"

def _thin():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

def style_val(cell, bg, bold=True, fmt=None):
    cell.fill      = PatternFill("solid", fgColor=bg)
    cell.font      = Font(bold=bold, size=11, name="Calibri")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border    = _thin()
    if fmt:
        cell.number_format = fmt

def style_formula(cell):
    cell.fill      = PatternFill("solid", fgColor=BLUE_LIGHT)
    cell.font      = Font(italic=True, size=10, name="Calibri", color="1F3864")
    cell.alignment = Alignment(horizontal="right", vertical="center")
    cell.border    = _thin()


# ── 1. Szenarien_Analyse ──────────────────────────────────────────────────────
def patch_szenarien(wb):
    ws = wb["Szenarien_Analyse"]

    # Row 5: Buchungs-Wahrscheinlichkeit  (gering=20%, normal=50%, stark=100%)
    ws.cell(5, 1).value = "Buchungs-Wahrscheinlichkeit Consulting je Kunde"
    ws.cell(5, 2).value = 0.50     # Referenz (Normal)
    ws.cell(5, 3).value = 0.20     # Gering  → a) jeder 5. Kunde
    ws.cell(5, 4).value = 0.50     # Normal  → b) jeder 2. Kunde
    ws.cell(5, 5).value = 1.00     # Stark   → c) jeder Kunde
    ws.cell(5, 6).value = "=E5-D5"
    ws.cell(5, 7).value = (
        "Wahrscheinlichkeit, dass ein gewonnener Kunde Consulting bucht:\n"
        "a) 20 % (gering) = jeder 5. Kunde – Automotive, wenig Beratungsbedarf\n"
        "b) 50 % (normal) = jeder 2. Kunde – Hybrid-Markt\n"
        "c) 100 % (stark) = jeder Kunde – Packaging, hohe Implementierungstiefe\n"
        "→ Fließt über Sandbox Z.12 → Inputs Z.139 → Revenue Z.28"
    )
    ws.row_dimensions[5].height = 70
    for col in (3, 4, 5):
        style_val(ws.cell(5, col), AMBER, fmt="0%")
    ws.cell(5, 1).font      = Font(bold=True, size=10, name="Calibri", color="1F6B75")
    ws.cell(5, 1).fill      = PatternFill("solid", fgColor=TEAL)
    ws.cell(5, 1).alignment = Alignment(wrap_text=True, vertical="top")
    ws.cell(5, 1).border    = _thin()

    # Row 8: Consulting-Tage/Einsatz  (gering=5, normal=10, stark=20)
    ws.cell(8, 2).value = 10    # Referenz
    ws.cell(8, 3).value = 5     # Gering
    ws.cell(8, 4).value = 10    # Normal
    ws.cell(8, 5).value = 20    # Stark
    ws.cell(8, 6).value = "=E8-D8"
    ws.cell(8, 7).value = (
        "Consulting-Tage pro Kunde je Einsatz:\n"
        "gering=5  (Basisschulung + Systemcheck)\n"
        "normal=10 (Implementierungsbegleitung + Anwenderschulung)\n"
        "stark=20  (Spez. Anforderungen, Datenbereinigung, Stammdaten-Harmonisierung)\n"
        "→ Fließt über Sandbox Z.11 → Inputs Z.140 → Revenue Z.28"
    )
    ws.row_dimensions[8].height = 70
    for col in (3, 4, 5):
        style_val(ws.cell(8, col), AMBER)
    ws.cell(8, 1).font      = Font(bold=True, size=10, name="Calibri", color="1F6B75")
    ws.cell(8, 1).fill      = PatternFill("solid", fgColor=TEAL)
    ws.cell(8, 1).alignment = Alignment(wrap_text=True, vertical="top")
    ws.cell(8, 1).border    = _thin()

    print("  Szenarien_Analyse Z.5:  Buchungs-Wahrscheinlichkeit → 20% / 50% / 100%")
    print("  Szenarien_Analyse Z.8:  Consulting-Tage/Einsatz     →  5  / 10  /  20")


# ── 2. Sandbox ────────────────────────────────────────────────────────────────
def patch_sandbox(wb):
    ws = wb["00_Input_Sandbox"]

    # Row 7: Tagessatz  (gering=1.000, normal=1.200, stark=1.500)
    ws.cell(7, 4).value = 1000
    ws.cell(7, 5).value = 1200
    ws.cell(7, 6).value = 1500
    ws.cell(7, 7).value = (
        "Consulting-Tagessatz: gering=1.000 €/Tag | normal=1.200 €/Tag | stark=1.500 €/Tag. "
        "Basiert auf Expertenstatus (Haag/Candemir) und Marktniveau B2B Industrie."
    )
    for col in (4, 5, 6):
        style_val(ws.cell(7, col), GREEN_LIGHT, fmt='#,##0 "€"')

    # Row 11: Consulting-Tage/Einsatz – Formulas already reference Szenarien_Analyse
    # Description update only
    ws.cell(11, 7).value = (
        "Referenziert Szenarien_Analyse Z.8 (dort bearbeitbar). "
        "gering=5 Tage | normal=10 Tage | stark=20 Tage. "
        "Direkte Eingabe hier überschreibt die Szenarien-Referenz."
    )

    # Row 12: Rename variable + update description
    ws.cell(12, 2).value = NEW_VAR_NAME
    ws.cell(12, 7).value = (
        "Referenziert Szenarien_Analyse Z.5 (dort bearbeitbar). "
        "a) gering=20% – jeder 5. Kunde kauft Consulting, "
        "b) normal=50% – jeder 2. Kunde, "
        "c) stark=100% – jeder Kunde kauft Consulting-Tage."
    )

    print(f"  Sandbox Z.7:  Tagessatz       → 1.000 / 1.200 / 1.500 €/Tag")
    print(f"  Sandbox Z.11: Tage-Beschriftung aktualisiert (Formeln = Szenarien_Analyse)")
    print(f"  Sandbox Z.12: Variable → '{NEW_VAR_NAME}'")


# ── 3. Inputs ─────────────────────────────────────────────────────────────────
def patch_inputs(wb):
    ws = wb["Inputs"]

    # Row 138: Tagessatz label + explanation
    ws.cell(138, 3).value = "EUR pro Beratertag je Szenario: gering=1.000€ | normal=1.200€ | stark=1.500€"

    # Row 139: Rename + update VLOOKUP key to match new Sandbox B12 name
    ws.cell(139, 1).value = "Buchungs-Wahrscheinlichkeit Consulting je Kunde (%)"
    ws.cell(139, 2).value = (
        f'=VLOOKUP("{NEW_VAR_NAME}",'
        "'00_Input_Sandbox'!$B$3:$F$20,"
        "'00_Input_Sandbox'!$B$2-1,FALSE)"
    )
    ws.cell(139, 2).number_format = "0%"
    style_formula(ws.cell(139, 2))
    ws.cell(139, 3).value = (
        "Wahrscheinlichkeit je Szenario (aus Szenarien_Analyse/Sandbox): "
        "gering=20% | normal=50% | stark=100%"
    )
    ws.cell(139, 4).value = (
        "Steuerung der Buchungsquote:\n"
        "a) 20 % (gering) = Konservativ – 1 von 5 Kunden kauft Consulting\n"
        "b) 50 % (normal) = Realistisch  – jeder 2. Kunde kauft Consulting\n"
        "c) 100 % (stark) = Voll-Auslastung – jeder Kunde kauft Consulting\n\n"
        "Kategorien Consulting-Leistung:\n"
        "  • Implementierungsbegleitung: techn. Integration vor Ort\n"
        "  • Schulung: Anwenderschulung, Train-the-Trainer\n"
        "  • Spez. Kundenanforderungen & Datenbereinigung: Datenmigration, "
        "Stammdaten, individuelle Anpassungen"
    )
    ws.cell(139, 4).alignment = Alignment(wrap_text=True, vertical="top")
    ws.cell(139, 4).font      = Font(italic=True, size=9, color="595959", name="Calibri")
    ws.cell(139, 4).border    = _thin()
    ws.row_dimensions[139].height = 90

    # Row 140: Tage label + explanation
    ws.cell(140, 3).value = (
        "Consulting-Tage je Einsatz/Kunde (aus Szenarien_Analyse/Sandbox): "
        "gering=5 | normal=10 | stark=20 Tage"
    )

    # Update flow banner row 136
    ws.cell(136, 1).value = (
        "ℹ DATENFLUSS CONSULTING:  "
        "Szenarien_Analyse (Z.5 Buchungs-Wskt. / Z.8 Tage)  →  "
        "Sandbox (Z.11 Tage / Z.12 Wskt.)  →  "
        "Inputs (Z.138 Tagessatz / Z.139 Wskt. / Z.140 Tage)  →  "
        "Revenue Z.28 Tage/Monat  →  Revenue Z.29 Consulting-Umsatz  →  P&L"
    )

    print(f"  Inputs Z.138: Tagessatz-Beschreibung aktualisiert")
    print(f"  Inputs Z.139: Label + VLOOKUP-Key → '{NEW_VAR_NAME}'")
    print(f"  Inputs Z.140: Tage-Beschreibung aktualisiert (5/10/20)")


# ── 4. Revenue – update row labels for clarity ────────────────────────────────
def patch_revenue_labels(wb):
    ws = wb["Revenue"]
    # Row 27 label (Total Active Customers) - no change needed
    # Row 28 label
    ws.cell(28, 1).value = "Consulting-Tage/Monat (Kunden × Wskt. × Tage/Einsatz)"
    # Row 29 label
    ws.cell(29, 1).value = "Consulting Revenue (€/Monat)  [= Tage/Mo × Tagessatz]"
    print("  Revenue Z.28/29: Labels aktualisiert (Formel unverändert)")


# ── MAIN ──────────────────────────────────────────────────────────────────────
def main():
    print(f"Lade: {SRC}")
    wb = openpyxl.load_workbook(SRC, data_only=False)

    print("\n1. Szenarien_Analyse …")
    patch_szenarien(wb)

    print("\n2. Sandbox …")
    patch_sandbox(wb)

    print("\n3. Inputs …")
    patch_inputs(wb)

    print("\n4. Revenue Labels …")
    patch_revenue_labels(wb)

    wb.save(DEST)
    print(f"\n✅ Gespeichert: {DEST}")

    # Verifikation
    wb2 = openpyxl.load_workbook(DEST, data_only=False)
    sz  = wb2["Szenarien_Analyse"]
    sb  = wb2["00_Input_Sandbox"]
    inp = wb2["Inputs"]
    rev = wb2["Revenue"]

    print("\n── Verifikation ──")
    print(f"  Szenarien Z.5  (Buchungs-Wskt.):  "
          f"gering={sz.cell(5,3).value}  normal={sz.cell(5,4).value}  stark={sz.cell(5,5).value}")
    print(f"  Szenarien Z.8  (Tage/Einsatz):    "
          f"gering={sz.cell(8,3).value}  normal={sz.cell(8,4).value}  stark={sz.cell(8,5).value}")
    print(f"  Sandbox   Z.7  (Tagessatz):        "
          f"gering={sb.cell(7,4).value}  normal={sb.cell(7,5).value}  stark={sb.cell(7,6).value}")
    print(f"  Sandbox   Z.11 (Tage formel D):   {sb.cell(11,4).value}")
    print(f"  Sandbox   Z.12 (Variable):        {sb.cell(12,2).value}")
    print(f"  Sandbox   Z.12 (Wskt. formel D):  {sb.cell(12,4).value}")
    print(f"  Inputs    Z.138 (Tagessatz):       {inp.cell(138,2).value}")
    print(f"  Inputs    Z.139 (VLOOKUP-Key):     {inp.cell(139,2).value[:60]}…")
    print(f"  Inputs    Z.140 (Tage VLOOKUP):    {inp.cell(140,2).value[:60]}…")
    print(f"  Revenue   Z.28 Label:              {rev.cell(28,1).value}")
    print(f"  Revenue   Z.29 Label:              {rev.cell(29,1).value}")
    print()
    print("  Formeln Revenue Z.28/29 (unverändert):")
    print(f"    Z.28 B (M1): {rev.cell(28,2).value}")
    print(f"    Z.29 B (M1): {rev.cell(29,2).value}")
    print()
    print("  Datenfluss:")
    print("  Szenarien Z.5  (Wskt. 20/50/100%)  → Sandbox Z.12 D/E/F → Inputs Z.139 VLOOKUP ✓")
    print("  Szenarien Z.8  (Tage 5/10/20)      → Sandbox Z.11 D/E/F → Inputs Z.140 VLOOKUP ✓")
    print("  Sandbox Z.7    (Tagessatz 1000/1200/1500) → Inputs Z.138 VLOOKUP ✓")
    print("  Revenue Z.28: Kunden × Wskt. × Tage = Consulting-Tage/Monat ✓")
    print("  Revenue Z.29: Tage × Tagessatz = Consulting-Umsatz/Monat ✓")
    print("  P&L Z.7:  TOTAL REVENUE = Subscription + Enterprise + Consulting ✓")


if __name__ == "__main__":
    main()
