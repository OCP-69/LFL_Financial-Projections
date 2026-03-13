"""
Data Merge Script: BM_Vorlage_v19 → C13_Template_financial_projections_neu
Mapping-Logik basierend auf: 260313_LFL_Finm Proj_ Data Merge_C13 TEmplate.txt

Mapping-Regeln:
  Revenue            → Monthly.Revenue + Monthly.Cash in from Revenue
  COGS               → Monthly.CoR Cash Out
  Total OpEx         → Monthly.Operating Cash Out
  Equity/Grants      → Monthly.Other Cash In
  Ending Cash        → Monthly.Closing Cash
  Beginning Cash     → Monthly.Opening Cash
  Income Tax         → Monthly.Tax
  Annual Aggregates  → Annual sheet
  Year-End Balances  → BalanceSheet

Zeitachse:
  Source M1  = April 2026  → C13 Monthly Row 5
  Source M52 = Juli 2030   → C13 Monthly Row 56
  Leere Monate (Jan–Mär 2026): Nullwerte
"""

import openpyxl
from datetime import datetime
import os

# ── Pfade ────────────────────────────────────────────────────────────────────
BASE_DIR   = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SOURCE     = os.path.join(BASE_DIR, '260312_LFL_BM_Vorlage_v19.xlsx')
TARGET_TPL = os.path.join(BASE_DIR, 'C13_Template_financial_projections_neu.xlsx')
TIMESTAMP  = datetime.now().strftime('%Y%m%d_%H%M')
OUTPUT     = os.path.join(BASE_DIR, f'LFL_BM_C13_Merge_{TIMESTAMP}.xlsx')

# ── Hilfsfunktionen ───────────────────────────────────────────────────────────
def read_row_values(ws, row_num, num_months=52):
    """Lese 52 Monatswerte aus Spalten B (col=2) bis col=53."""
    data = []
    for col in range(2, 2 + num_months):
        val = ws.cell(row=row_num, column=col).value
        data.append(float(val) if val is not None and isinstance(val, (int, float)) else 0.0)
    return data

def year_months_range(year_offset):
    """
    Gibt (src_start, src_end) für das Jahr zurück.
    year_offset=0 → 2026, year_offset=1 → 2027 usw.
    Source M1 = April 2026; Jan–Mär 2026 enthalten keine Quelldaten.

    2026: M1–M9   (April–Dez)  → src idx 0..8   (nur 9 Monate!)
    2027: M10–M21 (Jan–Dez)    → src idx 9..20
    2028: M22–M33              → src idx 21..32
    2029: M34–M45              → src idx 33..44
    2030: M46–M52 (Jan–Jul)    → src idx 45..51
    """
    if year_offset == 0:    # 2026: April–Dez = M1–M9
        return (0, 9)
    elif year_offset == 1:  # 2027: M10–M21
        return (9, 21)
    elif year_offset == 2:  # 2028: M22–M33
        return (21, 33)
    elif year_offset == 3:  # 2029: M34–M45
        return (33, 45)
    elif year_offset == 4:  # 2030: M46–M52 (Jan–Jul)
        return (45, 52)
    return (0, 0)

# ── Source laden ──────────────────────────────────────────────────────────────
print("Lade Quelldatei …")
wb_src = openpyxl.load_workbook(SOURCE, data_only=True)

ws_rev   = wb_src['4_Revenue']
ws_pl    = wb_src['6_P&L']
ws_cf    = wb_src['7_BS_CF']
ws_costs = wb_src['5_Costs']

# Schlüsseldaten (52 Monate)
total_revenue   = read_row_values(ws_rev, 32)   # Total Revenue (€/Monat)
mrr             = read_row_values(ws_rev, 25)   # MRR
arr             = read_row_values(ws_rev, 26)   # ARR

total_cogs      = read_row_values(ws_pl,  14)   # TOTAL COGS
total_opex      = read_row_values(ws_pl,  27)   # TOTAL OPERATING EXPENSES
net_income      = read_row_values(ws_pl,  37)   # NET INCOME
income_tax      = read_row_values(ws_pl,  35)   # Income Tax
ebitda          = read_row_values(ws_pl,  29)   # EBITDA
gross_profit    = read_row_values(ws_pl,  16)   # GROSS PROFIT

equity_funding  = read_row_values(ws_cf,  9)    # Equity Funding Received
beginning_cash  = read_row_values(ws_cf,  14)   # Beginning Cash Balance
ending_cash     = read_row_values(ws_cf,  15)   # ENDING CASH BALANCE
burn_rate       = read_row_values(ws_cf,  18)   # Monthly Burn Rate

# Berechnete Hilfsgrößen
debtors    = [rev for rev in total_revenue]            # 1 Monat Umsatz als Debitoren
creditors  = [(total_cogs[i] + total_opex[i]) * 0.10  # 10 % der Gesamtkosten
              for i in range(52)]
cum_equity = []
cum_sum    = 0.0
for v in equity_funding:
    cum_sum += v
    cum_equity.append(cum_sum)

cum_profit = []
cum_sum    = 0.0
for v in net_income:
    cum_sum += v
    cum_profit.append(cum_sum)

print(f"  → {sum(total_revenue):,.0f} € Gesamtumsatz (M1–M52)")
print(f"  → {sum(equity_funding):,.0f} € Eigenkapital eingeflossen")
print(f"  → Letzter Cash-Stand: {ending_cash[51]:,.0f} €")

# ── Target Template laden ─────────────────────────────────────────────────────
print("\nLade C13-Template …")
wb_dst = openpyxl.load_workbook(TARGET_TPL, data_only=False)

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 1: Monthly
# Spalten:
#  1=Year | 2=Month | 3=Revenue | 4=Cash in from Revenue | 5=Other Cash In
#  6=CoR Cash Out | 7=Operating Cash Out | 8=Interest | 9=Tax | 10=Founder Bonus
#  11=Opening Cash | 12=Closing Cash | 13=Debt Outstanding | 14=Debtors | 15=Creditors
# ══════════════════════════════════════════════════════════════════════════════
print("Befülle Monthly-Sheet …")
ws_m = wb_dst['Monthly']

# Monate-Mapping:
# Row 2  = Jan 2026 (kein Source-Datum)
# Row 5  = Apr 2026 = M1  → idx 0
# Row 13 = Dez 2026 = M9  → idx 8
# Row 14 = Jan 2027 = M10 → idx 9
# Row 56 = Jul 2030 = M52 → idx 51

# Zuerst: Monats-Nummern für bereits existierende 48 Zeilen (2026–2029) eintragen
# Jahres-Blöcke: 2026 rows 2–13, 2027 rows 14–25, 2028 rows 26–37, 2029 rows 38–49
month_num = 1
for row in range(2, 50):  # rows 2..49 = 4 Jahre × 12 Monate
    ws_m.cell(row=row, column=2).value = month_num
    month_num = (month_num % 12) + 1

# Neue Zeilen für 2030 (Jan–Jul = 7 Monate) anhängen → rows 50–56
for m_idx in range(7):
    row = 50 + m_idx
    ws_m.cell(row=row, column=1).value = 2030
    ws_m.cell(row=row, column=2).value = m_idx + 1  # Jan=1 … Jul=7

# Helper: Schreibe Monatswerte
def write_monthly_row(row, src_idx):
    """Schreibe alle Finanzdaten für einen Monat in eine Zeile."""
    rev   = total_revenue[src_idx]
    cogs  = total_cogs[src_idx]
    opex  = total_opex[src_idx]
    eqty  = equity_funding[src_idx]
    tax   = income_tax[src_idx]
    beg   = beginning_cash[src_idx]
    end   = ending_cash[src_idx]
    deb   = debtors[src_idx]
    cred  = creditors[src_idx]

    ws_m.cell(row=row, column=3).value  = round(rev, 2)    # Revenue
    ws_m.cell(row=row, column=4).value  = round(rev, 2)    # Cash in from Revenue (= Revenue)
    ws_m.cell(row=row, column=5).value  = round(eqty, 2)   # Other Cash In
    ws_m.cell(row=row, column=6).value  = round(cogs, 2)   # CoR Cash Out
    ws_m.cell(row=row, column=7).value  = round(opex, 2)   # Operating Cash Out
    ws_m.cell(row=row, column=8).value  = 0.0              # Interest (nicht modelliert)
    ws_m.cell(row=row, column=9).value  = round(tax, 2)    # Tax
    ws_m.cell(row=row, column=10).value = 0.0              # Founder Bonus
    ws_m.cell(row=row, column=11).value = round(beg, 2)    # Opening Cash
    ws_m.cell(row=row, column=12).value = round(end, 2)    # Closing Cash
    ws_m.cell(row=row, column=13).value = 0.0              # Debt Outstanding
    ws_m.cell(row=row, column=14).value = round(deb, 2)    # Debtors
    ws_m.cell(row=row, column=15).value = round(cred, 2)   # Creditors

# Leerzeilen für Jan–Mär 2026 (Rows 2–4, keine Quelldaten)
for row in range(2, 5):
    for col in [3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15]:
        ws_m.cell(row=row, column=col).value = 0.0

# Quelldaten M1–M52 eintragen
# M1 (April 2026) → Row 5, M52 (Juli 2030) → Row 56
for m_idx in range(52):
    row = 5 + m_idx  # Row 5..56
    write_monthly_row(row, m_idx)

monthly_rows_filled = 52
print(f"  → {monthly_rows_filled} Monate eingetragen (Apr 2026 – Jul 2030)")

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 2: Annual
# Spalten:
#  1=Year | 2=Revenue | 3=CoR | 4=Gross Profit | 5=Total Opex (incl R&D amort)
#  6=EBIT | 7=Interest | 8=Tax | 9=Founder Bonus | 10=Grants (non-taxable)
#  11=Net Profit | 12=Min Cash | 13=Avg Monthly Gross Burn | 14=Avg Monthly EBIT Burn
#  15=Cash (year-end) | 16=Debtors | 17=R&D Intangible | 18=Creditors
#  19=Debt | 20=Net Assets | 21=Total Equity Investment | 22=Cumulative Profit | 23=Quick Ratio
# ══════════════════════════════════════════════════════════════════════════════
print("Befülle Annual-Sheet …")
ws_a = wb_dst['Annual']

# Jahres-Definitionen
# year_offset: 0=2026, 1=2027, 2=2028, 3=2029, 4=2030
annual_years = [2026, 2027, 2028, 2029, 2030]  # Zeilen 2–6

# Kumulative Equity bis Jahresende
cum_eq_year = [0.0] * 5
cum_pr_year = [0.0] * 5

# Berechnung der kumulativen Equity bis Jahresende
for yo, year in enumerate(annual_years):
    s, e = year_months_range(yo)
    # Kumulativ über alle vergangenen Monate bis Jahresende
    end_month = min(e, 52)  # letzter Quell-Index +1
    ce = sum(equity_funding[:end_month])
    cp = sum(net_income[:end_month])
    cum_eq_year[yo] = ce
    cum_pr_year[yo] = cp

for yo, year in enumerate(annual_years):
    row = 2 + yo  # Row 2–6
    s, e = year_months_range(yo)

    rev_yr    = sum(total_revenue[s:e])
    cogs_yr   = sum(total_cogs[s:e])
    opex_yr   = sum(total_opex[s:e])
    gp_yr     = rev_yr - cogs_yr
    ebit_yr   = gp_yr - opex_yr
    tax_yr    = sum(income_tax[s:e])
    ni_yr     = sum(net_income[s:e])
    eqty_yr   = sum(equity_funding[s:e])

    # Min Cash im Jahr
    min_cash  = min(ending_cash[s:e]) if e > s else 0.0
    # Letzter Monats-Cash (Jahresende)
    last_cash = ending_cash[e - 1] if e > s else 0.0
    # Avg Monthly Gross Burn (totaler Cash-Abfluss / Monate)
    gross_burn_vals = [cogs_yr_m + opex_yr_m
                       for cogs_yr_m, opex_yr_m in
                       zip(total_cogs[s:e], total_opex[s:e])]
    months_in_year  = e - s
    avg_gross_burn  = sum(gross_burn_vals) / months_in_year if months_in_year > 0 else 0.0
    # Avg Monthly EBIT Burn (nur Verlustmonate)
    ebit_months = [ebitda[i] for i in range(s, e) if ebitda[i] < 0]
    avg_ebit_burn = (sum(ebit_months) / len(ebit_months)) if ebit_months else 0.0

    # Debtors / Creditors am Jahresende
    last_deb  = debtors[e - 1] if e > s else 0.0
    last_cred = creditors[e - 1] if e > s else 0.0
    net_assets = last_cash + last_deb - last_cred
    quick_ratio = (last_cash / last_cred) if last_cred > 0 else 0.0

    ws_a.cell(row=row, column=2).value  = round(rev_yr, 2)
    ws_a.cell(row=row, column=3).value  = round(cogs_yr, 2)
    ws_a.cell(row=row, column=4).value  = round(gp_yr, 2)
    ws_a.cell(row=row, column=5).value  = round(opex_yr, 2)
    ws_a.cell(row=row, column=6).value  = round(ebit_yr, 2)
    ws_a.cell(row=row, column=7).value  = 0.0               # Interest
    ws_a.cell(row=row, column=8).value  = round(tax_yr, 2)
    ws_a.cell(row=row, column=9).value  = 0.0               # Founder Bonus
    ws_a.cell(row=row, column=10).value = 0.0               # Grants
    ws_a.cell(row=row, column=11).value = round(ni_yr, 2)
    ws_a.cell(row=row, column=12).value = round(min_cash, 2)
    ws_a.cell(row=row, column=13).value = round(avg_gross_burn, 2)
    ws_a.cell(row=row, column=14).value = round(avg_ebit_burn, 2)
    ws_a.cell(row=row, column=15).value = round(last_cash, 2)
    ws_a.cell(row=row, column=16).value = round(last_deb, 2)
    ws_a.cell(row=row, column=17).value = 0.0               # R&D Intangible
    ws_a.cell(row=row, column=18).value = round(last_cred, 2)
    ws_a.cell(row=row, column=19).value = 0.0               # Debt
    ws_a.cell(row=row, column=20).value = round(net_assets, 2)
    ws_a.cell(row=row, column=21).value = round(cum_eq_year[yo], 2)
    ws_a.cell(row=row, column=22).value = round(cum_pr_year[yo], 2)
    ws_a.cell(row=row, column=23).value = round(quick_ratio, 4)

    print(f"  {year}: Rev={rev_yr:>12,.0f} € | EBIT={ebit_yr:>12,.0f} € | Cash={last_cash:>12,.0f} €")

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 3: BalanceSheet
# Spalten:
#  1=Year | 2=Cash | 3=Debtors | 4=R&D Intangible | 5=Creditors
#  6=Debt | 7=Net Assets | 8=Total Equity Investment | 9=Cumulative Profit | 10=Quick Ratio
# ══════════════════════════════════════════════════════════════════════════════
print("Befülle BalanceSheet …")
ws_bs = wb_dst['BalanceSheet']

for yo, year in enumerate(annual_years):
    row = 2 + yo  # Row 2–6
    _, e = year_months_range(yo)

    cash_end  = ending_cash[e - 1] if e > 0 else 0.0
    deb_end   = debtors[e - 1] if e > 0 else 0.0
    cred_end  = creditors[e - 1] if e > 0 else 0.0
    net_ass   = cash_end + deb_end - cred_end
    qr        = (cash_end / cred_end) if cred_end > 0 else 0.0

    # Sicherstellen, dass Zeile mit Jahr befüllt ist
    ws_bs.cell(row=row, column=1).value  = year
    ws_bs.cell(row=row, column=2).value  = round(cash_end, 2)
    ws_bs.cell(row=row, column=3).value  = round(deb_end, 2)
    ws_bs.cell(row=row, column=4).value  = 0.0              # R&D Intangible
    ws_bs.cell(row=row, column=5).value  = round(cred_end, 2)
    ws_bs.cell(row=row, column=6).value  = 0.0              # Debt
    ws_bs.cell(row=row, column=7).value  = round(net_ass, 2)
    ws_bs.cell(row=row, column=8).value  = round(cum_eq_year[yo], 2)
    ws_bs.cell(row=row, column=9).value  = round(cum_pr_year[yo], 2)
    ws_bs.cell(row=row, column=10).value = round(qr, 4)

# ══════════════════════════════════════════════════════════════════════════════
# Prompt Business Model_Slides – Werte aktualisieren
# ══════════════════════════════════════════════════════════════════════════════
print("Aktualisiere Prompt-Sheet …")
ws_prompt = wb_dst['Prompt Business Model_Slides']

total_rev_all     = sum(total_revenue)
final_year_rev    = sum(total_revenue[45:52])   # 2030 (Jan–Jul)
final_year_ni     = sum(net_income[45:52])
cum_ni_all        = sum(net_income)
total_eq_all      = sum(equity_funding)
peak_cash         = max(ending_cash)
lowest_cash       = min(ending_cash)

# Erster negativer Cash-Monat
first_neg_month   = next((i for i, c in enumerate(ending_cash) if c < 0), None)
first_neg_label   = f"2026, Monat {first_neg_month + 1}" if first_neg_month is not None else "nie"

avg_burn          = sum(total_cogs[i] + total_opex[i] for i in range(52)) / 52
neg_ebit_months   = [ebitda[i] for i in range(52) if ebitda[i] < 0]
avg_ebit_burn     = sum(neg_ebit_months) / len(neg_ebit_months) if neg_ebit_months else 0.0
starting_cash     = ending_cash[0]  # End of M1
monthly_burn_avg  = avg_burn
runway            = (starting_cash / monthly_burn_avg) if monthly_burn_avg > 0 else 0.0

# Break-even Jahr ermitteln
breakeven_year = "nicht erreicht (in 52 Monaten)"
for yo, year in enumerate(annual_years):
    s, e = year_months_range(yo)
    if sum(net_income[s:e]) > 0:
        breakeven_year = str(year)
        break

# Prompt-Zeilen ersetzen
prompt_updates = {
    11: f"- Gesamtumsatz über den Zeitraum: {total_rev_all:,.0f} €",
    12: f"- Umsatz letztes erfasstes Jahr (2030, Jan–Jul): {final_year_rev:,.0f} €",
    13: f"- Nettoergebnis letztes erfasstes Jahr: {final_year_ni:,.0f} €",
    14: f"- Kumuliertes Nettoergebnis: {cum_ni_all:,.0f} €",
    15: f"- Gesamtes eingesetztes Eigenkapital: {total_eq_all:,.0f} €",
    16: "- Nicht-verwässernde Förderung: 0 €",
    19: f"- Höchster Cash-Stand: {peak_cash:,.0f} €",
    20: f"- Niedrigster Cash-Stand: {lowest_cash:,.0f} €",
    21: f"- Erster Monat mit negativem Cash: {first_neg_label}",
    22: f"- Durchschnittliche monatliche Gesamtausgaben: {avg_burn:,.0f} €",
    23: f"- Durchschnittlicher EBIT-Burn (Verlustmonate): {avg_ebit_burn:,.0f} €",
    27: f"- Break-even Jahr (Nettogewinn): {breakeven_year}",
}

for row_num, text in prompt_updates.items():
    ws_prompt.cell(row=row_num, column=1).value = text

# Kopfzeile anpassen
ws_prompt.cell(row=7, column=1).value = "- Währung: EUR"
ws_prompt.cell(row=8, column=1).value = "- Unternehmenstyp: B2B SaaS / AI-First (LoopforgeLab GmbH)"

# ══════════════════════════════════════════════════════════════════════════════
# Speichern
# ══════════════════════════════════════════════════════════════════════════════
wb_dst.save(OUTPUT)
print(f"\n✓ Datei gespeichert: {OUTPUT}")

# ── Validierung / Stress-Test (C13-Standard) ─────────────────────────────────
print("\n═══════════════════════════════════════════════════════")
print("VALIDIERUNG (C13 Stress-Test)")
print("═══════════════════════════════════════════════════════")

# 1. Cash-Check
neg_months = [(i+1, ending_cash[i]) for i in range(52) if ending_cash[i] < 0]
if neg_months:
    print(f"⚠  WARNUNG: {len(neg_months)} Monat(e) mit negativem Cash-Stand:")
    for m, c in neg_months[:5]:
        print(f"     M{m}: {c:,.0f} €")
    if len(neg_months) > 5:
        print(f"     … und {len(neg_months)-5} weitere")
    # Check M1–M9 (first 9 months)
    early_neg = [m for m, _ in neg_months if m <= 9]
    if early_neg:
        print(f"  ⚠  INSOLVENZRISIKO: Monate {early_neg} haben negatives Cash (M1–M9)")
else:
    print("✓  Cash-Check: Alle Monate positiv")

# 2. R&D-Amortisation (Hinweis)
print("ℹ  R&D Intangible: Nicht aktiviert (alle Kosten direkt im OpEx)")
print("   → Empfehlung: Einmalige Software-Entwicklungskosten ggf. bilanzieren")

# 3. Founder-Boni
first_profit_month = next((i+1 for i in range(52) if net_income[i] > 0), None)
if first_profit_month:
    print(f"✓  Founder Bonus: Erst ab M{first_profit_month} möglich (erster Gewinnmonat)")
else:
    print("ℹ  Founder Bonus: In 52 Monaten kein positives Net Income erreicht")

# Summary KPIs
print()
print("═══════════════════════════════════════════════════════")
print("KEY METRICS SUMMARY")
print("═══════════════════════════════════════════════════════")
print(f"  Gesamtumsatz M1–M52:    {total_rev_all:>14,.0f} €")
print(f"  Kumulierter Net Income: {cum_ni_all:>14,.0f} €")
print(f"  Gesamtes Eigenkapital:  {total_eq_all:>14,.0f} €")
print(f"  Höchster Cash:          {peak_cash:>14,.0f} €")
print(f"  Niedrigster Cash:       {lowest_cash:>14,.0f} €")
print(f"  Break-even Jahr:        {breakeven_year}")
print(f"  Ø Monatl. Burn:         {avg_burn:>14,.0f} €")
print("═══════════════════════════════════════════════════════")
print(f"\nOutput: {os.path.basename(OUTPUT)}")
