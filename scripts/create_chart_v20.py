"""
Chart-Erstellung: Revenue vs Net Profit vs Year-end Cash (Annual, 2026–2030)
Quelle: LFL_BM_C13_Normal_v19_20260315_20260315_1025.xlsx
Ausgabe: LFL_BM_C13_Normal_v20_20260315.xlsx  (neues Sheet "Revenue_Profit_Cash_Chart")

Chartdesign orientiert sich am Screenshot im Graphs-Sheet:
  - Blaue Linie  : Revenue (Jahresumsatz)
  - Rosa Linie   : Net Profit (Jahres-Nettoeinkommen)
  - Orange Linie : Year-end Cash
  - Grüne gestrichelte Vertikallinie: Break-even Jahr
"""

import openpyxl
from openpyxl.drawing.image import Image as XLImage
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.ticker as mticker
import io
import os
from datetime import datetime

# ── Pfade ─────────────────────────────────────────────────────────────────────
BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SOURCE   = os.path.join(BASE_DIR, 'LFL_BM_C13_Normal_v19_20260315_20260315_1025.xlsx')
OUTPUT   = os.path.join(BASE_DIR, 'LFL_BM_C13_Normal_v20_20260315.xlsx')

# ── Daten aus Annual-Sheet lesen ──────────────────────────────────────────────
print(f"Lade: {os.path.basename(SOURCE)}")
wb = openpyxl.load_workbook(SOURCE, data_only=True)
ws_a = wb['Annual']

# Zeile 1 = Header, Zeilen 2–6 = 2026–2030
years       = []
revenue     = []
net_profit  = []
year_end_cash = []

for row in range(2, 7):   # Rows 2–6 = 2026–2030
    yr  = ws_a.cell(row=row, column=1).value   # A: Year
    rev = ws_a.cell(row=row, column=2).value   # B: Revenue
    ni  = ws_a.cell(row=row, column=11).value  # K: Net Profit
    csh = ws_a.cell(row=row, column=15).value  # O: Cash (year-end)
    years.append(int(yr) if yr else 2026 + (row - 2))
    revenue.append(float(rev) if rev else 0.0)
    net_profit.append(float(ni) if ni else 0.0)
    year_end_cash.append(float(csh) if csh else 0.0)

print(f"\nJahreswerte:")
print(f"  {'Jahr':>6} | {'Revenue':>14} | {'Net Profit':>14} | {'Year-end Cash':>14}")
print(f"  {'-'*6}-+-{'-'*14}-+-{'-'*14}-+-{'-'*14}")
for i, yr in enumerate(years):
    print(f"  {yr:>6} | {revenue[i]:>14,.0f} | {net_profit[i]:>14,.0f} | {year_end_cash[i]:>14,.0f}")

# Break-even Jahr = erstes Jahr mit Net Profit > 0
breakeven_year = None
for i, np_ in enumerate(net_profit):
    if np_ > 0:
        breakeven_year = years[i]
        break
print(f"\nBreak-even Jahr: {breakeven_year}")

# ── Chart erstellen (matplotlib, Stil wie im Screenshot) ──────────────────────
fig, ax = plt.subplots(figsize=(13, 7))
fig.patch.set_facecolor('white')
ax.set_facecolor('white')

# Farben wie im Screenshot
COL_REV   = '#4A90D9'   # Blau – Revenue
COL_NP    = '#E86B7A'   # Rosa/Rot – Net Profit
COL_CASH  = '#F5A623'   # Orange – Year-end Cash

# Linien
ax.plot(years, revenue,     color=COL_REV,  linewidth=2.5, marker='o', markersize=6,
        label='Revenue', zorder=3)
ax.plot(years, net_profit,  color=COL_NP,   linewidth=2.0, marker='o', markersize=6,
        linestyle='--', label='Net Profit', zorder=3)
ax.plot(years, year_end_cash, color=COL_CASH, linewidth=2.0, marker='o', markersize=6,
        linestyle='--', label='Year-end Cash', zorder=3)

# Break-even Vertikallinie (grün gestrichelt)
if breakeven_year:
    ax.axvline(x=breakeven_year, color='#2ECC71', linewidth=1.5,
               linestyle='--', zorder=2, alpha=0.85)
    ax.text(breakeven_year + 0.05, ax.get_ylim()[1] * 0.97,
            'Break-even', color='#2ECC71', fontsize=9,
            va='top', ha='left', style='italic')

# Nulllinie
ax.axhline(y=0, color='#CCCCCC', linewidth=0.8, zorder=1)

# Y-Achse: Euro-Format
def eur_fmt(x, pos):
    if abs(x) >= 1_000_000:
        return f'€{x/1_000_000:.0f}M'
    elif abs(x) >= 1_000:
        return f'€{x/1_000:.0f}K'
    else:
        return f'€{x:.0f}'

ax.yaxis.set_major_formatter(mticker.FuncFormatter(eur_fmt))

# X-Achse: nur ganze Jahre
ax.set_xticks(years)
ax.set_xticklabels([str(y) for y in years], fontsize=10)

# Gitter
ax.yaxis.grid(True, linestyle='--', alpha=0.4, color='#DDDDDD')
ax.set_axisbelow(True)
for spine in ['top', 'right']:
    ax.spines[spine].set_visible(False)
ax.spines['left'].set_color('#CCCCCC')
ax.spines['bottom'].set_color('#CCCCCC')
ax.tick_params(colors='#555555', length=0)

# Titel & Untertitel
fig.text(0.06, 0.97, 'REVENUE VS NET PROFIT VS CASH',
         fontsize=14, fontweight='bold', color='#2C3E50',
         va='top', ha='left')
fig.text(0.06, 0.91,
         'Annual view. Cash is year-end balance; revenue and net profit are full-year totals.\n'
         'A green dashed line marks the first year where annual net profit is non-negative.',
         fontsize=8.5, color='#7F8C8D', va='top', ha='left')

# Legende unten mittig
ax.legend(loc='upper center', bbox_to_anchor=(0.5, -0.10),
          ncol=3, frameon=False, fontsize=10,
          handlelength=2.0,
          handler_map={})

# Datenpunkte beschriften (Revenue, Net Profit, Cash am letzten Jahr)
for i, yr in enumerate(years):
    for vals, col in [(revenue, COL_REV), (net_profit, COL_NP), (year_end_cash, COL_CASH)]:
        v = vals[i]
        if abs(v) >= 1_000_000:
            lbl = f'€{v/1_000_000:.1f}M'
        elif abs(v) >= 1_000:
            lbl = f'€{v/1_000:.0f}K'
        else:
            lbl = f'€{v:.0f}'
        # Nur letztes Jahr beschriften um Überfüllung zu vermeiden
        if i == len(years) - 1:
            ax.annotate(lbl, (yr, v),
                        textcoords='offset points', xytext=(8, 0),
                        fontsize=8, color=col, va='center')

plt.tight_layout(rect=[0, 0.05, 1, 0.88])

# ── Als PNG in BytesIO speichern ──────────────────────────────────────────────
buf = io.BytesIO()
plt.savefig(buf, format='png', dpi=150, bbox_inches='tight',
            facecolor='white', edgecolor='none')
buf.seek(0)
plt.close()
print(f"\nChart erstellt: {len(buf.getvalue()):,} Bytes")

# ── Neues Sheet "Revenue_Profit_Cash_Chart" in Workbook einfügen ──────────────
# Workbook in Formel-Modus laden um Sheets/Daten zu erhalten
wb2 = openpyxl.load_workbook(SOURCE, data_only=False)

# Neues Sheet erzeugen (oder überschreiben falls vorhanden)
CHART_SHEET = 'Revenue_Profit_Cash_Chart'
if CHART_SHEET in wb2.sheetnames:
    del wb2[CHART_SHEET]
ws_chart = wb2.create_sheet(CHART_SHEET, index=0)  # Als erstes Sheet

# Sheet-Tab-Farbe (blau passend zum Chart)
ws_chart.sheet_properties.tabColor = "4A90D9"

# Titel-Zelle
ws_chart['A1'] = 'Revenue vs Net Profit vs Year-end Cash — Normal-Szenario (2026–2030)'
ws_chart['A1'].font = openpyxl.styles.Font(bold=True, size=13, color='2C3E50')

ws_chart['A2'] = f'Quelle: {os.path.basename(SOURCE)} | Erstellt: {datetime.now().strftime("%d.%m.%Y %H:%M")}'
ws_chart['A2'].font = openpyxl.styles.Font(size=9, color='7F8C8D', italic=True)

# Datentabelle unterhalb (für Transparenz)
ws_chart['A4'] = 'Jahr';          ws_chart['A4'].font = openpyxl.styles.Font(bold=True)
ws_chart['B4'] = 'Revenue (€)';   ws_chart['B4'].font = openpyxl.styles.Font(bold=True)
ws_chart['C4'] = 'Net Profit (€)'; ws_chart['C4'].font = openpyxl.styles.Font(bold=True)
ws_chart['D4'] = 'Year-end Cash (€)'; ws_chart['D4'].font = openpyxl.styles.Font(bold=True)

num_fmt = '#,##0'
for i, yr in enumerate(years):
    r = 5 + i
    ws_chart.cell(row=r, column=1).value = yr
    ws_chart.cell(row=r, column=2).value = revenue[i]
    ws_chart.cell(row=r, column=3).value = net_profit[i]
    ws_chart.cell(row=r, column=4).value = year_end_cash[i]
    for c in [2, 3, 4]:
        ws_chart.cell(row=r, column=c).number_format = num_fmt

# Spaltenbreiten
ws_chart.column_dimensions['A'].width = 18
ws_chart.column_dimensions['B'].width = 18
ws_chart.column_dimensions['C'].width = 18
ws_chart.column_dimensions['D'].width = 20

# Chart-Bild einfügen (ab Zeile 11, Spalte A)
img = XLImage(buf)
img.anchor = 'A11'
img.width  = 940   # px  ≈ 13 Zoll × 72 dpi
img.height = 510   # px  ≈  7 Zoll × 72 dpi
ws_chart.add_image(img)

# Zeilenhöhe für Chart-Bereich setzen
for row_idx in range(11, 50):
    ws_chart.row_dimensions[row_idx].height = 14.5

# ── Speichern ─────────────────────────────────────────────────────────────────
wb2.save(OUTPUT)
print(f"✓ Datei gespeichert: {os.path.basename(OUTPUT)}")
print(f"  Sheet '{CHART_SHEET}' als erstes Sheet eingefügt")
print(f"  Datentabelle: Zeilen 4–{4 + len(years)}")
print(f"  Chart-Bild:   ab Zeile 11")
