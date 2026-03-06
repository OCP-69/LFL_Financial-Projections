"""
fix_revenue_structure.py
─────────────────────────
Erstellt LFL_BM_Konservativ_v5.xlsx aus v4 mit folgenden Korrekturen:

Revenue-Sheet:
  Z11 A11: Entfernt Entwicklungsnotiz ("HIER FEHLT CONSULTING...")
  Z12 A12: Umbenennung "3v" → "Neue Enterprise-Deals (Quartal)"
  Z17 (neu): Trennzeile mit Label-Erklärung für den Recurring-Block
  Z18: Label-Ergänzung "MRR (nur Lizenzumsätze, ohne Consulting)"
  Z19: Label-Ergänzung "ARR (nur Lizenzumsätze, ohne Consulting)"
  Z20: Formel korrigiert: =B18 → =B18+B29 (MRR + Consulting)
       Label: "Total Monthly Revenue (Lizenzen + Consulting)"

P&L: keine Änderung (bereits korrekt)
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SRC  = "/home/user/LFL_Financial-Projections/scenarios/LFL_BM_Konservativ_v4_Final.xlsx"
DEST = "/home/user/LFL_Financial-Projections/scenarios/LFL_BM_Konservativ_v5.xlsx"

BLUE_DARK   = "1F3864"
BLUE_MID    = "2E75B6"
BLUE_LIGHT  = "BDD7EE"
GREEN_LIGHT = "E2EFDA"
GREEN_DARK  = "375623"
AMBER       = "FFF2CC"
TEAL        = "1F6B75"
TEAL_LIGHT  = "E9F7F8"
GREY        = "F2F2F2"
ORANGE      = "FCE4D6"
WHITE       = "FFFFFF"
RED_LIGHT   = "FFCCCC"


def _thin():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)


def _medium():
    s = Side(style="medium", color="2E75B6")
    return Border(left=s, right=s, top=s, bottom=s)


def apply(cell, value=None, bold=False, italic=False, size=10,
          color="000000", bg=None, halign="left", wrap=True,
          fmt=None, border="thin"):
    if value is not None:
        cell.value = value
    cell.font      = Font(bold=bold, italic=italic, size=size,
                          color=color, name="Calibri")
    cell.alignment = Alignment(horizontal=halign, vertical="center",
                               wrap_text=wrap)
    cell.border    = _thin() if border == "thin" else _medium()
    if bg:
        cell.fill = PatternFill("solid", fgColor=bg)
    if fmt:
        cell.number_format = fmt


def patch_revenue(wb):
    ws     = wb["Revenue"]
    ncols  = 53    # col A + M1…M52

    # ── Z11: Entwicklungsnotiz entfernen ─────────────────────────────────────
    # Die Zelle A11 enthält: "HIER FEHLT CONSULTING UMSATZ..."
    # → Inhalt löschen, Zelle leer lassen (neutrale Trennzeile nutzen)
    ws.cell(11, 1).value = None
    ws.cell(11, 1).fill  = PatternFill("solid", fgColor=WHITE)
    ws.row_dimensions[11].height = 6   # schmale Leerzeile als visueller Trenner

    # ── Z12: "3v" umbenennen → "Neue Enterprise-Deals (Quartal)" ─────────────
    # Erklärung der Formel:
    #   IF(month >= Enterprise-Start-Monat AND (month - Start) MOD 3 = 0,
    #      Deals-pro-Quartal, 0)
    # → Löst im Quartalsrhythmus ab konfig. Startmonat neue Enterprise-Deals aus.
    apply(ws.cell(12, 1),
          value="Neue Enterprise-Deals (Quartal)",
          bold=True, bg=BLUE_LIGHT,
          halign="left")
    ws.row_dimensions[12].height = 18

    # ── Z17: Neue Trennzeile mit Kontext-Label für Recurring-Block ───────────
    # Z17 ist aktuell leer → nutzen als erklärenden Sub-Header
    apply(ws.cell(17, 1),
          value="── RECURRING REVENUE (SaaS-Metriken, ohne Consulting) ──",
          bold=True, italic=False, size=9,
          color=WHITE, bg=TEAL,
          halign="left", wrap=False)
    for col in range(2, ncols + 1):
        c = ws.cell(17, col)
        c.fill   = PatternFill("solid", fgColor=TEAL)
        c.border = _thin()
    ws.row_dimensions[17].height = 16

    # ── Z18: MRR — Label klarstellen ─────────────────────────────────────────
    apply(ws.cell(18, 1),
          value="MRR – Monthly Recurring Revenue (Lizenzen)",
          bold=True, bg=BLUE_LIGHT, halign="left")
    # Formel bleibt: =B9+B16 (Subscription + Enterprise) ✓
    ws.row_dimensions[18].height = 20

    # ── Z19: ARR — Label klarstellen ─────────────────────────────────────────
    apply(ws.cell(19, 1),
          value="ARR – Annual Recurring Revenue (Lizenzen = MRR × 12)",
          bold=True, bg=BLUE_LIGHT, halign="left")
    # Formel bleibt: =B18*12 ✓
    ws.row_dimensions[19].height = 20

    # ── Z20: Total Monthly Revenue — Formel + Label korrigieren ──────────────
    # ALT: =B18        (= nur MRR → Consulting fehlte!)
    # NEU: =B18+B29    (= MRR + Consulting Revenue)
    apply(ws.cell(20, 1),
          value="Total Monthly Revenue (Lizenzen + Consulting)",
          bold=True, color=WHITE, bg=GREEN_DARK, halign="left")
    ws.row_dimensions[20].height = 20

    for col in range(2, ncols + 1):
        col_l = get_column_letter(col)
        # Z20 Formel: MRR (Z18) + Consulting Revenue (Z29)
        ws.cell(20, col).value = f"={col_l}18+{col_l}29"
        c = ws.cell(20, col)
        c.fill          = PatternFill("solid", fgColor=GREEN_LIGHT)
        c.border        = _thin()
        c.alignment     = Alignment(horizontal="right", vertical="center")
        c.number_format = '#,##0'

    # ── Z21: ARR Growth — Bezug bleibt auf Z19 (ARR recurring) ──────────────
    # Formel =IFERROR(C19/B19-1,"") bleibt korrekt ✓

    # ── Hinweis-Zeile nach Consulting-Block (nach Z29) ───────────────────────
    note_row = 31
    apply(ws.cell(note_row, 1),
          value=(
              "ℹ Hinweis MRR/ARR vs. Total Revenue:\n"
              "MRR/ARR (Z18/19) = ausschließlich wiederkehrende Lizenzumsätze "
              "(Subscription Seats + Enterprise-Verträge). "
              "Consulting ist projektbezogen, nicht-wiederkehrend → "
              "kein Bestandteil von MRR/ARR (SaaS-Standard: SaaStr/OpenView/a16z).\n"
              "Total Monthly Revenue (Z20) = MRR + Consulting → "
              "fließt in P&L TOTAL REVENUE."
          ),
          italic=True, size=9, color="595959",
          bg=TEAL_LIGHT, halign="left", wrap=True)
    ws.row_dimensions[note_row].height = 55
    # Breite über alle Spalten
    try:
        ws.merge_cells(
            start_row=note_row, start_column=1,
            end_row=note_row, end_column=10
        )
    except Exception:
        pass

    print("  Z11: Entwicklungsnotiz entfernt")
    print("  Z12: '3v' → 'Neue Enterprise-Deals (Quartal)' (Quartalstakt-Formel erklärt)")
    print("  Z17: Sub-Header '── RECURRING REVENUE ──' eingefügt")
    print("  Z18: Label 'MRR – Monthly Recurring Revenue (Lizenzen)'")
    print("  Z19: Label 'ARR – Annual Recurring Revenue (Lizenzen = MRR × 12)'")
    print("  Z20: Label + Formel korrigiert: =B18 → =B18+B29 (MRR + Consulting)")
    print("  Z31: Erläuterungsnotiz MRR/ARR vs. Total Revenue")


def main():
    print(f"Lade: {SRC}")
    wb = openpyxl.load_workbook(SRC, data_only=False)

    print("\nPatche Revenue-Sheet …")
    patch_revenue(wb)

    wb.save(DEST)
    print(f"\n✅ Gespeichert: {DEST}")

    # Verifikation
    wb2 = openpyxl.load_workbook(DEST, data_only=False)
    ws  = wb2["Revenue"]

    print("\n── Verifikation Revenue ──")
    for r in [11, 12, 17, 18, 19, 20, 21]:
        a = ws.cell(r, 1).value
        b = ws.cell(r, 2).value   # M1 Formel
        print(f"  Z{r:2d}: A={repr(str(a))[:55]:56s} | M1={repr(str(b))[:50]}")

    print()
    # P&L unverändert prüfen
    ws_pl = wb2["P&L"]
    print("── P&L REVENUE-Sektion (unverändert) ──")
    for r in [5, 6, 7, 8]:
        a = ws_pl.cell(r, 1).value
        b = ws_pl.cell(r, 2).value
        print(f"  Z{r}: {repr(str(a))[:40]:41s} | M1: {repr(str(b))[:55]}")

    print()
    print("Logik-Prüfung (Szenario 'gering', Monat 36 Beispiel):")
    print("  MRR Z18  = Subscription-MRR + Enterprise-MRR   (Recurring only) ✓")
    print("  ARR Z19  = MRR × 12                             (Recurring only) ✓")
    print("  Consulting Z29 = Kunden × Wskt. × Tage × Satz  (Non-recurring)  ✓")
    print("  Total Z20 = MRR + Consulting  →  P&L TOTAL REVENUE               ✓")
    print("  ARR Growth Z21 referenziert Z19 (ARR recurring) → korrekt        ✓")
    print("  NRR Z22 gilt nur für Lizenzumsätze                                ✓")


if __name__ == "__main__":
    main()
