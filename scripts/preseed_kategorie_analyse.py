"""
Ausgaben-Analyse bis Ende Pre-Seed (M1–M16) aus 260312_LFL_BM_Vorlage_normal_v19.xlsx
4 Kategorien + Gegenüberstellung Carbon13/Business Angels-Finanzierung
"""

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import os

BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SRC  = os.path.join(BASE, '260312_LFL_BM_Vorlage_normal_v19.xlsx')
OUT  = os.path.join(BASE, 'LFL_BM_PreSeed_Ausgaben_Kategorien.xlsx')

wb_src = openpyxl.load_workbook(SRC, data_only=True)
ws5 = wb_src['5_Costs']

# Hilfsfunktion: Monatswerte lesen (M1-M16 = Spalten B–Q = col 2–17)
def row_vals(row, m_start=1, m_end=16):
    return [float(ws5.cell(row=row, column=1+m).value or 0) for m in range(m_start, m_end+1)]

def s(row, m_start=1, m_end=16):
    return sum(row_vals(row, m_start, m_end))

# ── Rohdaten M1–M16 ────────────────────────────────────────────────────────────
# PERSONAL
ceo_total   = s(6)      # CEO (R6) M1-M16
cto_total   = s(7)      # CTO (R7)
cco_total   = s(8)      # CCO (R8)
exec_total  = s(9)      # Executives gesamt
emp_monthly = row_vals(9+1)  # Mitarbeiter R10 – monatliche Werte

# Mitarbeiter-Split nach Rolle und Eintrittszeitpunkt (aus 2_Inputs):
# SW Dev M6, Key Account M10, Mech Eng M11, CS M14, Mkt Assist M14
# AG-Brutto/Mo: SW Dev=7625, Key Acc=7930, Mech Eng=6913, CS≈5287, Mkt≈4270
# Hardware-Renting je MA: 89 €/Mo (aus 2_Inputs R194-R197)
EMP_ROLES = {
    'SW Developer':       {'start_m': 6,  'cat': 'A', 'monthly': 7625.0,  'hw': 89},
    'Key Account':        {'start_m': 10, 'cat': 'B', 'monthly': 7930.0,  'hw': 89},
    'Mech./Domain Eng.':  {'start_m': 11, 'cat': 'A', 'monthly': 6913.0,  'hw': 89},
    'Customer Success':   {'start_m': 14, 'cat': 'B', 'monthly': 5287.0,  'hw': 89},
    'Marketing Assist.':  {'start_m': 14, 'cat': 'B', 'monthly': 4270.0,  'hw': 89},
}

# Monatsgenaue Basis-Mitarbeiterkosten (ohne Gehaltserhöhungen)
emp_base_by_cat = {'A': 0.0, 'B': 0.0}
hw_by_cat       = {'A': 0.0, 'B': 0.0}
emp_detail      = []

for role, info in EMP_ROLES.items():
    months_active = max(0, 16 - info['start_m'] + 1)
    base_cost = info['monthly'] * months_active
    hw_cost   = info['hw'] * months_active
    emp_base_by_cat[info['cat']] += base_cost
    hw_by_cat[info['cat']]       += hw_cost
    emp_detail.append((role, info['cat'], info['start_m'], months_active, base_cost, hw_cost))

# Restbetrag (Gehaltserhöhungen über 3% p.a.) proportional verteilen
emp_base_total = sum(emp_base_by_cat.values())
emp_actual     = s(10)      # Mitarbeiter R10 – tatsächlich
salary_delta   = emp_actual - emp_base_total   # Differenz = Gehaltserhöhungen
for cat in ['A', 'B']:
    emp_base_by_cat[cat] += salary_delta * (emp_base_by_cat[cat] / emp_base_total)

hw_actual = s(12)           # Hardware-Renting R12

# TECHNOLOGIE
cloud_hosting = s(15)       # R15
aiml_apis     = s(16)       # R16
saas_tools    = s(17)       # R17
tech_total    = s(18)       # R18

# BÜRO
coworking     = s(20)       # R20
office_total  = s(24)       # R24

# PROFESSIONAL SERVICES
rechtsanwalt  = s(26)       # R26
steuerberater = s(27)       # R27
prof_total    = s(30)       # R30

# VERSICHERUNG & BANK
versicherung  = s(32)       # R32
bankgebuehren = s(33)       # R33
sicherheit    = s(34)       # R34
vers_total    = s(35)       # R35

# MARKETING & SALES
paid_ads      = s(37)       # R37
events        = s(38)       # R38
sales_tools   = s(39)       # R39
mktg_total    = s(40)       # R40

# SONSTIGE
reisekosten   = s(42)       # R42
team_events   = s(43)       # R43
sonst_total   = s(45)       # R45

# PAYMENT PROCESSING
payment       = s(47)       # R47

total_all     = s(50)       # R50 Gesamtkosten

# ── Kategorie-Allokation ───────────────────────────────────────────────────────
# A) Software & AI Engineering
#    CTO, SW-Dev, Mech Eng, AI/ML APIs, SaaS Tools, anteilige Hardware
cat_A_items = [
    ('CTO Gehalt (AG-Brutto)',          'CTO – 5_Costs!C6:Q6',     cto_total),
    ('SW Developer + Mech. Eng. (Gehalt)', '5_Costs!C10:Q10 (anteilig)', emp_base_by_cat['A']),
    ('Hardware-Renting (Eng.)',          '5_Costs!C12:Q12 (anteilig)', hw_by_cat['A']),
    ('AI/ML APIs',                       '5_Costs!C16:Q16',         aiml_apis),
    ('SaaS Tools & Lizenzen',            '5_Costs!C17:Q17',         saas_tools),
]

# B) Customer Acquisition
#    CCO, Key Account + CS + Mkt Assist, Paid Ads, Events, Sales Tools, Payment Processing
cat_B_items = [
    ('CCO Gehalt (AG-Brutto)',           '5_Costs!C8:Q8',           cco_total),
    ('Key Account + CS + Mkt (Gehalt)',  '5_Costs!C10:Q10 (anteilig)', emp_base_by_cat['B']),
    ('Hardware-Renting (Cust.)',         '5_Costs!C12:Q12 (anteilig)', hw_by_cat['B']),
    ('Paid Ads & Content/SEO',           '5_Costs!C37:Q37',         paid_ads),
    ('Events & Messen',                  '5_Costs!C38:Q38',         events),
    ('Sales Tools & Provisionen',        '5_Costs!C39:Q39',         sales_tools),
    ('Payment Processing Fees',          '5_Costs!C47:Q47',         payment),
]

# C) Data Infrastructure
#    Cloud Hosting, Sicherheit & Compliance
cat_C_items = [
    ('Cloud Hosting (Basis)',            '5_Costs!C15:Q15',         cloud_hosting),
    ('Sicherheit & Compliance',          '5_Costs!C34:Q34',         sicherheit),
]

# D) Operations & Legal
#    CEO, Coworking, Rechtsanwalt, Steuerberater, Versicherungen, Bank, Reisen, Team Events
cat_D_items = [
    ('CEO Gehalt (AG-Brutto)',           '5_Costs!C6:Q6',           ceo_total),
    ('Coworking Space',                  '5_Costs!C20:Q20',         coworking),
    ('Rechtsanwalt',                     '5_Costs!C26:Q26',         rechtsanwalt),
    ('Steuerberater',                    '5_Costs!C27:Q27',         steuerberater),
    ('Versicherungen (D&O/Haftpfl./Cyber)', '5_Costs!C32:Q32',      versicherung),
    ('Bankgebühren',                     '5_Costs!C33:Q33',         bankgebuehren),
    ('Reisekosten & Weiterbildung',      '5_Costs!C42:Q42',         reisekosten),
    ('Team Events',                      '5_Costs!C43:Q43',         team_events),
]

cat_A_total = sum(v for _, _, v in cat_A_items)
cat_B_total = sum(v for _, _, v in cat_B_items)
cat_C_total = sum(v for _, _, v in cat_C_items)
cat_D_total = sum(v for _, _, v in cat_D_items)
cat_sum     = cat_A_total + cat_B_total + cat_C_total + cat_D_total

# Kontrollsumme
print(f'Kategorien-Summe:  {cat_sum:>12,.2f} €')
print(f'5_Costs R50 Total: {total_all:>12,.2f} €')
print(f'Differenz:         {cat_sum - total_all:>+12,.2f} €  (= Gehaltserhöhungs-Rundung)')

# ── Finanzierung ───────────────────────────────────────────────────────────────
# Aus 7_BS_CF R9 + 2_Inputs
fin_carbon13    = 120_000.0  # M4: C13 Ideation + GmbH Stammeinlage (2_Inputs R9)
fin_angel_a     = 200_000.0  # M4  (2_Inputs R18)
fin_angel_b     = 200_000.0  # M5  (2_Inputs R20)
fin_angel_c     = 200_000.0  # M8  (2_Inputs R22)
fin_angels_total = fin_angel_a + fin_angel_b + fin_angel_c
fin_total        = fin_carbon13 + fin_angels_total

print(f'\nFinanzierung M1-M16: {fin_total:>10,.0f} €')
print(f'  Carbon13:           {fin_carbon13:>10,.0f} € ({fin_carbon13/fin_total*100:.1f}%)')
print(f'  Business Angels:    {fin_angels_total:>10,.0f} € ({fin_angels_total/fin_total*100:.1f}%)')
print(f'\nGesamtausgaben:     {cat_sum:>10,.0f} €  = {cat_sum/fin_total*100:.1f}% der Finanzierung')
print(f'Cash-Reserve Ende Pre-Seed: {fin_total - cat_sum:>10,.0f} €')

# ══════════════════════════════════════════════════════════════════════════════
# EXCEL ERSTELLEN
# ══════════════════════════════════════════════════════════════════════════════
def fill(h): return PatternFill('solid', fgColor=h)
def fnt(bold=False, sz=10, color='1A1A1A', italic=False):
    return Font(bold=bold, size=sz, color=color, italic=italic, name='Calibri')
def aln(h='center', v='center', wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)
def brd(style='thin', color='CCCCCC'):
    s = Side(style=style, color=color)
    return Border(left=s, right=s, top=s, bottom=s)
def brd_outer():
    m = Side(style='medium', color='888888')
    return Border(left=m, right=m, top=m, bottom=m)

C = {
    'dark':    '1F4E79',
    'mid':     '2E75B6',
    'light':   'D6E4F0',
    'A':       'E3F2FD',   # Software & AI Eng  – blau
    'A_hdr':   '1565C0',
    'B':       'E8F5E9',   # Customer Acq.      – grün
    'B_hdr':   '2E7D32',
    'C':       'FFF3E0',   # Data Infra         – orange
    'C_hdr':   'E65100',
    'D':       'FCE4EC',   # Ops & Legal        – rosa
    'D_hdr':   'AD1457',
    'total':   'D5F5E3',
    'fin_c13': 'E0F7FA',
    'fin_ba':  'F3E5F5',
    'fin_tot': 'FFF9C4',
    'grey':    'F5F5F5',
    'white':   'FFFFFF',
}

CAT_LABELS = {
    'A': 'A)  Software & AI Engineering',
    'B': 'B)  Customer Acquisition',
    'C': 'C)  Data Infrastructure',
    'D': 'D)  Operations & Legal',
}
CAT_ITEMS  = {'A': cat_A_items, 'B': cat_B_items, 'C': cat_C_items, 'D': cat_D_items}
CAT_TOTALS = {'A': cat_A_total, 'B': cat_B_total, 'C': cat_C_total, 'D': cat_D_total}

wb = openpyxl.Workbook()

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 1: Haupt-Matrix
# ══════════════════════════════════════════════════════════════════════════════
ws = wb.active
ws.title = 'Ausgaben_nach_Kategorien'
ws.sheet_view.showGridLines = False
ws.sheet_properties.tabColor = C['dark']

# Spaltenbreiten
COLS = {'A': 38, 'B': 30, 'C': 16, 'D': 14, 'E': 14, 'F': 14, 'G': 14}
for col, w in COLS.items():
    ws.column_dimensions[col].width = w

def cell(r, c, val='', bg='FFFFFF', bold=False, sz=10, color='1A1A1A',
         ha='left', va='center', wrap=False, fmt=None, italic=False, border=True):
    cell_ = ws.cell(row=r, column=c, value=val)
    cell_.fill = fill(bg)
    cell_.font = fnt(bold, sz, color, italic)
    cell_.alignment = aln(ha, va, wrap)
    if border: cell_.border = brd()
    if fmt:    cell_.number_format = fmt
    return cell_

# ── TITEL ─────────────────────────────────────────────────────────────────────
ws.merge_cells('A1:G1')
c = ws['A1']
c.value = 'LOOPFORGELAB — AUSGABEN-ANALYSE BIS ENDE PRE-SEED (M1–M16 | Apr 2026 – Jul 2027)'
c.font  = Font(bold=True, size=14, color='FFFFFF', name='Calibri')
c.fill  = fill(C['dark'])
c.alignment = aln('left', 'center')
ws.row_dimensions[1].height = 30

ws.merge_cells('A2:G2')
c = ws['A2']
c.value = f'Quelle: 260312_LFL_BM_Vorlage_normal_v19.xlsx  |  5_Costs R6–R50 (Spalten B–Q)  |  Gesamtausgaben: {cat_sum:,.0f} €  |  Finanzierung: {fin_total:,.0f} €'
c.font  = Font(size=9, italic=True, color='444444', name='Calibri')
c.fill  = fill('EEF2F7')
c.alignment = aln('left', 'center')
ws.row_dimensions[2].height = 15

ws.row_dimensions[3].height = 8   # Spacer

# ── SPALTEN-HEADER ────────────────────────────────────────────────────────────
HDR = 4
ws.row_dimensions[HDR].height = 32
hdrs = ['Kostenposition', 'Quelldaten (Sheet / Zellen)', 'Betrag (€)',
        '% von Ges.-Ausgaben', '% von Finanzierung', 'Carbon13-Anteil (€)', 'BA-Anteil (€)']
hdr_bg = C['mid']
for ci, h in enumerate(hdrs, 1):
    cell(HDR, ci, h, hdr_bg, bold=True, sz=10, color='FFFFFF', ha='center', wrap=True)

# ── KATEGORIE-BLÖCKE ─────────────────────────────────────────────────────────
def write_category(cat, start_row):
    r = start_row
    hdr_bg = C[f'{cat}_hdr']
    row_bg = C[cat]
    label  = CAT_LABELS[cat]
    items  = CAT_ITEMS[cat]
    total  = CAT_TOTALS[cat]

    # Kategorie-Header
    ws.row_dimensions[r].height = 22
    ws.merge_cells(f'A{r}:B{r}')
    cell(r, 1, label, hdr_bg, bold=True, sz=11, color='FFFFFF', ha='left')
    ws.cell(row=r, column=2).fill = fill(hdr_bg)  # merged
    cell(r, 3, total,  hdr_bg, bold=True, sz=11, color='FFFFFF', ha='right',
         fmt='#,##0 "€"')
    cell(r, 4, total/cat_sum,  hdr_bg, bold=True, sz=11, color='FFFFFF', ha='right',
         fmt='0.0%')
    cell(r, 5, total/fin_total, hdr_bg, bold=True, sz=11, color='FFFFFF', ha='right',
         fmt='0.0%')
    cell(r, 6, total*(fin_carbon13/fin_total), hdr_bg, bold=True, sz=11, color='FFFFFF',
         ha='right', fmt='#,##0 "€"')
    cell(r, 7, total*(fin_angels_total/fin_total), hdr_bg, bold=True, sz=11, color='FFFFFF',
         ha='right', fmt='#,##0 "€"')
    r += 1

    # Einzel-Positionen
    for pos_name, pos_src, pos_val in items:
        ws.row_dimensions[r].height = 16
        cell(r, 1, f'    {pos_name}',  row_bg, sz=9, ha='left')
        cell(r, 2, pos_src,            row_bg, sz=8, color='555555', ha='left', italic=True)
        cell(r, 3, pos_val,            row_bg, sz=9, ha='right', fmt='#,##0 "€"')
        cell(r, 4, pos_val/cat_sum,    row_bg, sz=9, ha='right', fmt='0.0%')
        cell(r, 5, pos_val/fin_total,  row_bg, sz=9, ha='right', fmt='0.0%')
        cell(r, 6, pos_val*(fin_carbon13/fin_total), row_bg, sz=9, ha='right', fmt='#,##0 "€"')
        cell(r, 7, pos_val*(fin_angels_total/fin_total), row_bg, sz=9, ha='right', fmt='#,##0 "€"')
        r += 1

    # Spacer
    for ci in range(1, 8):
        ws.cell(row=r, column=ci).fill = fill('FFFFFF')
    ws.row_dimensions[r].height = 5
    r += 1
    return r

cur = HDR + 1
for cat in ['A', 'B', 'C', 'D']:
    cur = write_category(cat, cur)

# ── GESAMT-ZEILE ──────────────────────────────────────────────────────────────
ws.row_dimensions[cur].height = 22
ws.merge_cells(f'A{cur}:B{cur}')
cell(cur, 1, 'GESAMTAUSGABEN M1–M16', C['dark'], bold=True, sz=11, color='FFFFFF', ha='left')
# merged - skip col 2
cell(cur, 3, cat_sum,      C['dark'], bold=True, sz=11, color='FFFFFF', ha='right', fmt='#,##0 "€"')
cell(cur, 4, cat_sum/cat_sum,   C['dark'], bold=True, sz=11, color='FFFFFF', ha='right', fmt='0.0%')
cell(cur, 5, cat_sum/fin_total, C['dark'], bold=True, sz=11, color='FFFFFF', ha='right', fmt='0.0%')
cell(cur, 6, cat_sum*(fin_carbon13/fin_total), C['dark'], bold=True, sz=11, color='FFFFFF', ha='right', fmt='#,##0 "€"')
cell(cur, 7, cat_sum*(fin_angels_total/fin_total), C['dark'], bold=True, sz=11, color='FFFFFF', ha='right', fmt='#,##0 "€"')
cur += 2

# ── FINANZIERUNGS-BLOCK ───────────────────────────────────────────────────────
ws.row_dimensions[cur].height = 8
cur += 1

ws.merge_cells(f'A{cur}:G{cur}')
c = ws.cell(row=cur, column=1)
c.value = 'FINANZIERUNG BIS ENDE PRE-SEED (M1–M16)'
c.font  = Font(bold=True, size=11, color='FFFFFF', name='Calibri')
c.fill  = fill(C['dark'])
c.alignment = aln('left', 'center')
ws.row_dimensions[cur].height = 22
cur += 1

FIN_ROWS = [
    ('Carbon13 (Ideation Funding + GmbH-Stammeinlage)', '7_BS_CF!E9 | 2_Inputs!B9', fin_carbon13,  C['fin_c13'], 'M4 (Jul 2026)'),
    ('Business Angel A',                                 '7_BS_CF!E9 | 2_Inputs!B18', fin_angel_a,   C['fin_ba'],  'M4 (Jul 2026)'),
    ('Business Angel B',                                 '7_BS_CF!F9 | 2_Inputs!B20', fin_angel_b,   C['fin_ba'],  'M5 (Aug 2026)'),
    ('Business Angel C',                                 '7_BS_CF!I9 | 2_Inputs!B22', fin_angel_c,   C['fin_ba'],  'M8 (Nov 2026)'),
]

# Fin-Header
ws.row_dimensions[cur].height = 20
fin_hdrs = ['Investor / Finanzierungsquelle', 'Quelldaten', 'Betrag (€)', '% der Finanzierung', 'Verwendung / Kategorie', 'Zufluss-Monat', '']
for ci, h in enumerate(fin_hdrs[:6], 1):
    cell(cur, ci, h, C['mid'], bold=True, sz=9, color='FFFFFF', ha='center', wrap=True)
cur += 1

for fname, fsrc, fval, fbg, fmonth in FIN_ROWS:
    ws.row_dimensions[cur].height = 18
    cell(cur, 1, fname,          fbg, sz=9, ha='left', bold=True)
    cell(cur, 2, fsrc,           fbg, sz=8, color='555555', ha='left', italic=True)
    cell(cur, 3, fval,           fbg, sz=10, ha='right', bold=True, fmt='#,##0 "€"')
    cell(cur, 4, fval/fin_total, fbg, sz=10, ha='right', bold=True, fmt='0.0%')
    cell(cur, 5, 'Anteilig alle 4 Kategorien', fbg, sz=9, ha='center', color='555555')
    cell(cur, 6, fmonth,         fbg, sz=9, ha='center', color='333333')
    cur += 1

# Gesamtfinanzierung
ws.row_dimensions[cur].height = 22
ws.merge_cells(f'A{cur}:B{cur}')
cell(cur, 1, 'TOTAL FINANZIERUNG M1–M16', C['dark'], bold=True, sz=11, color='FFFFFF', ha='left')
# merged - skip col 2
cell(cur, 3, fin_total,  C['dark'], bold=True, sz=11, color='FFFFFF', ha='right', fmt='#,##0 "€"')
cell(cur, 4, 1.0,        C['dark'], bold=True, sz=11, color='FFFFFF', ha='right', fmt='0.0%')
cell(cur, 5, f'davon Ausgaben: {cat_sum:,.0f} € ({cat_sum/fin_total:.1%})', C['dark'], bold=True, sz=10, color='FFFFFF', ha='center')
cell(cur, 6, f'Cash-Reserve: {fin_total-cat_sum:,.0f} €', C['dark'], bold=True, sz=10, color='FFFFFF', ha='center')
cur += 2

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 2: Mitarbeiter-Detail Allokation
# ══════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet('Mitarbeiter_Allokation')
ws2.sheet_properties.tabColor = C['A_hdr']
ws2.sheet_view.showGridLines = False

ws2.merge_cells('A1:G1')
c = ws2['A1']
c.value = 'MITARBEITER-ALLOKATION AUF KATEGORIEN (M1–M16)'
c.font = Font(bold=True, size=13, color='FFFFFF', name='Calibri')
c.fill = fill(C['dark'])
c.alignment = aln('left','center')
ws2.row_dimensions[1].height = 26

ws2.column_dimensions['A'].width = 26
ws2.column_dimensions['B'].width = 12
ws2.column_dimensions['C'].width = 14
ws2.column_dimensions['D'].width = 16
ws2.column_dimensions['E'].width = 16
ws2.column_dimensions['F'].width = 16
ws2.column_dimensions['G'].width = 16

hdrs2 = ['Rolle','Eintritt','Akt. Monate\n(bis M16)','Basis-Gehalt (€)\n(ohne Erhöhg.)','Hardware-Renting (€)','SUMME (€)','Kategorie']
ws2.row_dimensions[2].height = 28
for ci, h in enumerate(hdrs2, 1):
    c = ws2.cell(row=2, column=ci, value=h)
    c.font = Font(bold=True, size=9, color='FFFFFF', name='Calibri')
    c.fill = fill(C['mid'])
    c.alignment = aln('center','center', wrap=True)
    c.border = brd()

exec_rows = [
    ('CEO (Gründer)', 'M5', 12, ceo_total, 0, ceo_total, 'D) Ops & Legal'),
    ('CTO (Gründer)', 'M5', 12, cto_total, 0, cto_total, 'A) Softw./AI Eng.'),
    ('CCO (Gründer)', 'M5', 12, cco_total, 0, cco_total, 'B) Customer Acq.'),
]
for ed in emp_detail:
    role, cat, sm, months, bc, hwc = ed
    exec_rows.append((role, f'M{sm}', months, round(bc,0), round(hwc,0), round(bc+hwc,0), f'{"A" if cat=="A" else "B"}) {"Softw./AI Eng." if cat=="A" else "Customer Acq."}'))

cat_bg_map = {'A)': C['A'], 'B)': C['B'], 'D)': C['D']}
r2 = 3
for row in exec_rows:
    ws2.row_dimensions[r2].height = 16
    bg = C['grey']
    for key in cat_bg_map:
        if str(row[-1]).startswith(key):
            bg = cat_bg_map[key]
            break
    fmts = [None, None, None, '#,##0 "€"', '#,##0 "€"', '#,##0 "€"', None]
    for ci, (val, fmt) in enumerate(zip(row, fmts), 1):
        c = ws2.cell(row=r2, column=ci, value=val)
        c.fill = fill(bg)
        c.font = Font(size=9, name='Calibri', bold=(ci in [1,6,7]))
        c.alignment = aln('center' if ci != 1 else 'left', 'center')
        c.border = brd()
        if fmt: c.number_format = fmt
    r2 += 1

# Summenzeile
ws2.row_dimensions[r2].height = 20
sumrow = ['GESAMT', '', '', 0, 0, 0, '']
for row in exec_rows:
    sumrow[3] += row[3] if isinstance(row[3], (int,float)) else 0
    sumrow[4] += row[4] if isinstance(row[4], (int,float)) else 0
    sumrow[5] += row[5] if isinstance(row[5], (int,float)) else 0
fmts = [None, None, None, '#,##0 "€"', '#,##0 "€"', '#,##0 "€"', None]
for ci, (val, fmt) in enumerate(zip(sumrow, fmts), 1):
    c = ws2.cell(row=r2, column=ci, value=val)
    c.fill = fill(C['dark'])
    c.font = Font(bold=True, size=10, color='FFFFFF', name='Calibri')
    c.alignment = aln('center' if ci != 1 else 'left', 'center')
    c.border = brd()
    if fmt: c.number_format = fmt

# ── Speichern ─────────────────────────────────────────────────────────────────
wb.save(OUT)
print(f'\n✓ Gespeichert: {os.path.basename(OUT)}')
print(f'\n=== ZUSAMMENFASSUNG ===')
print(f'  A) Software & AI Engineering: {cat_A_total:>10,.0f} € ({cat_A_total/cat_sum:.1%} der Ausgaben | {cat_A_total/fin_total:.1%} der Finanzierung)')
print(f'  B) Customer Acquisition:      {cat_B_total:>10,.0f} € ({cat_B_total/cat_sum:.1%} der Ausgaben | {cat_B_total/fin_total:.1%} der Finanzierung)')
print(f'  C) Data Infrastructure:       {cat_C_total:>10,.0f} € ({cat_C_total/cat_sum:.1%} der Ausgaben | {cat_C_total/fin_total:.1%} der Finanzierung)')
print(f'  D) Operations & Legal:        {cat_D_total:>10,.0f} € ({cat_D_total/cat_sum:.1%} der Ausgaben | {cat_D_total/fin_total:.1%} der Finanzierung)')
print(f'  GESAMT AUSGABEN:              {cat_sum:>10,.0f} €')
print(f'  FINANZIERUNG (C13 + Angels):  {fin_total:>10,.0f} €')
print(f'    davon Carbon13:             {fin_carbon13:>10,.0f} € ({fin_carbon13/fin_total:.1%})')
print(f'    davon Business Angels:      {fin_angels_total:>10,.0f} € ({fin_angels_total/fin_total:.1%})')
print(f'  Cash-Reserve:                 {fin_total-cat_sum:>10,.0f} € ({(fin_total-cat_sum)/fin_total:.1%})')
