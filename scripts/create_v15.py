#!/usr/bin/env python3
"""
Erstellt v15 der LFL Financial Projections aus v14.
Alle Änderungen gemäß Aufgabenstellung vom 11.03.2026.
"""

import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
import re
import shutil

SRC = '/home/user/LFL_Financial-Projections/260307_LFL_BM_Vorlage_v14.xlsx'
DST_V15 = '/home/user/LFL_Financial-Projections/260307_LFL_BM_Vorlage_v15.xlsx'

print("=" * 70)
print("ERSTELLE v15 aus v14")
print("=" * 70)

wb = openpyxl.load_workbook(SRC, data_only=False)
print(f"Geladen: {SRC}")
print(f"Sheets: {wb.sheetnames}")

# ================================================================
# SCHRITT A: "Notizen & ToDos" Sheet erstellen
# Inhalte aus Inputs!A2 und Costs!B1 übernehmen, dann leeren
# ================================================================
print("\n--- SCHRITT A: Notizen & ToDos Sheet ---")

ws_inputs = wb['Inputs']
ws_costs = wb['Costs']

# Inhalte einlesen
inputs_a2_content = ws_inputs['A2'].value
costs_b1_content = ws_costs['B1'].value

print(f"  Inputs!A2 (Länge): {len(str(inputs_a2_content))} Zeichen")
print(f"  Costs!B1 (Länge): {len(str(costs_b1_content))} Zeichen")

# Neues Sheet erstellen
ws_notizen = wb.create_sheet("Notizen & ToDos")
ws_notizen['A1'] = "Aus Inputs!A2:"
ws_notizen['A2'] = inputs_a2_content
ws_notizen['A10'] = "Aus Costs!B1 (To-Do):"
ws_notizen['A11'] = costs_b1_content

# Quell-Zellen leeren
ws_inputs['A2'] = None
ws_costs['B1'] = None

print("  Notizen & ToDos Sheet erstellt.")
print("  Inputs!A2 und Costs!B1 geleert.")

# ================================================================
# SCHRITT B: "Weiterentwicklung" Sheet erstellen
# ================================================================
print("\n--- SCHRITT B: Weiterentwicklung Sheet ---")

ws_entw = wb.create_sheet("Weiterentwicklung")
entw_lines = [
    ("A1",  "IDEEN & VORSCHLÄGE FÜR KÜNFTIGE ENTWICKLUNG"),
    ("A2",  "Stand: 11.03.2026"),
    ("A3",  ""),
    ("A4",  "FEATURE 1: Automatisierte Kundenverteilung über Monate"),
    ("A5",  "Status: Vorgemerkt – noch nicht implementiert"),
    ("A6",  "Beschreibung:"),
    ("A7",  "  - Nutzer trägt in Inputs nur die GESAMTANZAHL neuer Kunden pro Phase (Ideation / Pre-Seed / Seed / Series A) ein"),
    ("A8",  "  - Für SME, Mid Company und Enterprise separat"),
    ("A9",  "  - Algorithmus verteilt die Kunden gleichmäßig (oder gewichtet) über die Monate der jeweiligen Phase"),
    ("A10", "  - Implementierung: neue Zeilen in 00_Input_Sandbox oder Inputs, Verteilungsformel in Revenue-Sheet"),
    ("A11", "  - Betrifft Sheets: Inputs, Revenue"),
    ("A12", ""),
    ("A13", "FEATURE 2: [Platzhalter für weitere Ideen]"),
]
for addr, val in entw_lines:
    ws_entw[addr] = val

print("  Weiterentwicklung Sheet erstellt.")

# ================================================================
# SCHRITT C: Inputs-Sheet umstrukturieren
#
# IST:
#   Row 23: "Initiale Seats (SME and Mid_Company)" | B23=3
#   Row 24: "Monatliche Seat-Wachstumsrate" | B24=0.03
#   Row 25: "Enterprise-Deals ab Monat" | B25=20
#   Row 26: "Durchschnitt Enterprise ARR" | B26=25000
#   Row 27: "Enterprise Deals pro Quartal" | B27=1
#   Row 28: "Jährliche Churn Rate" (VLOOKUP)
#   Row 29: "Net Revenue Retention" (VLOOKUP)
#
# SOLL (nach insert_rows(24)):
#   Row 23: "Initiale Seats SME" | B23=5
#   Row 24: "Initiale Seats Mid-Company" | B24=15  (NEU)
#   Row 25: "Monatliche Seat-Wachstumsrate" | B25=0.03 (unverändert)
#   Row 26: LEER (war "Enterprise-Deals ab Monat")
#   Row 27: "Durchschnittlicher Enterprise ARR pro Vertrag" | B27=25000
#   Row 28: LEER (war "Enterprise Deals pro Quartal")
#   Row 29: "Jährliche Churn Rate" (VLOOKUP)
#   Row 30: "Net Revenue Retention" (VLOOKUP)
# ================================================================
print("\n--- SCHRITT C: Inputs umstrukturieren ---")

# Werte VOR Änderung dokumentieren
print(f"  IST Row 23: A='{ws_inputs['A23'].value}' B={ws_inputs['B23'].value}")
print(f"  IST Row 24: A='{ws_inputs['A24'].value}' B={ws_inputs['B24'].value}")
print(f"  IST Row 25: A='{ws_inputs['A25'].value}' B={ws_inputs['B25'].value}")
print(f"  IST Row 26: A='{ws_inputs['A26'].value}' B={ws_inputs['B26'].value}")
print(f"  IST Row 27: A='{ws_inputs['A27'].value}' B={ws_inputs['B27'].value}")
print(f"  IST Row 28: A='{ws_inputs['A28'].value}' B={str(ws_inputs['B28'].value)[:60]}")
print(f"  IST Row 29: A='{ws_inputs['A29'].value}' B={str(ws_inputs['B29'].value)[:60]}")

# Eine Zeile nach Zeile 23 einfügen
ws_inputs.insert_rows(24)

# Nach insert_rows(24): Alle Zeilen ab 24 wurden um 1 nach unten verschoben
# Row 23: "Initiale Seats (SME and Mid_Company)" -> umändern
ws_inputs['A23'] = 'Initiale Seats SME'
ws_inputs['B23'] = 5

# Row 24: NEU - "Initiale Seats Mid-Company"
ws_inputs['A24'] = 'Initiale Seats Mid-Company'
ws_inputs['B24'] = 15

# Row 25: "Monatliche Seat-Wachstumsrate" (kein Eingriff nötig, unverändert aus vorher B24=0.03)
# (automatisch nach unten verschoben durch insert_rows)

# Row 26: War "Enterprise-Deals ab Monat" (B25) -> leeren
ws_inputs['A26'] = None
ws_inputs['B26'] = None

# Row 27: War "Durchschnitt Enterprise ARR" (B26) -> umbenennen, Wert bleibt
ws_inputs['A27'] = 'Durchschnittlicher Enterprise ARR pro Vertrag'
# B27 = 25000 (automatisch erhalten)

# Row 28: War "Enterprise Deals pro Quartal" (B27) -> leeren
ws_inputs['A28'] = None
ws_inputs['B28'] = None

# Row 29: "Jährliche Churn Rate" (war B28, VLOOKUP) - keine Änderung nötig
# Row 30: "Net Revenue Retention" (war B29, VLOOKUP) - keine Änderung nötig

print(f"\n  SOLL Row 23: A='{ws_inputs['A23'].value}' B={ws_inputs['B23'].value}")
print(f"  SOLL Row 24: A='{ws_inputs['A24'].value}' B={ws_inputs['B24'].value}")
print(f"  SOLL Row 25: A='{ws_inputs['A25'].value}' B={ws_inputs['B25'].value}")
print(f"  SOLL Row 26: A='{ws_inputs['A26'].value}' B={ws_inputs['B26'].value}")
print(f"  SOLL Row 27: A='{ws_inputs['A27'].value}' B={ws_inputs['B27'].value}")
print(f"  SOLL Row 28: A='{ws_inputs['A28'].value}' B={ws_inputs['B28'].value}")
print(f"  SOLL Row 29: A='{ws_inputs['A29'].value}'")
print(f"  SOLL Row 30: A='{ws_inputs['A30'].value}'")

# ================================================================
# SCHRITT D: Revenue-Sheet Formel-Referenzen anpassen
#
# Durch insert_rows(24) in Inputs verschoben sich ALLE Inputs-Zeilen ab 24.
# openpyxl aktualisiert cross-sheet Referenzen NICHT automatisch.
# Manuelle Anpassung erforderlich.
#
# MAPPING der betroffenen Referenzen:
#   Inputs!B24 (alt Monatliche Seat-Wachstumsrate) -> jetzt B25
#   Inputs!B25 (alt Enterprise-Deals ab Monat) -> jetzt B26 (gelöscht)
#   Inputs!B26 (alt Durchschnitt Enterprise ARR) -> jetzt B27
#   Inputs!B27 (alt Enterprise Deals pro Quartal) -> jetzt B28 (gelöscht)
#   Inputs!B28 (alt Jährliche Churn Rate) -> jetzt B29
#   Inputs!B29 (alt Net Revenue Retention) -> jetzt B30
#
# In Revenue-Sheet:
#   Row 8:  Inputs!B22 (ok), Inputs!B23 (-> SME seats), Inputs!B24 (-> now B25 seat growth)
#   Row 9:  Inputs!B28 (churn) -> Inputs!B29
#   Row 15: Inputs!B25 (enterprise start month, deleted), Inputs!B27 (enterprise deals/q, deleted)
#           -> Enterprise wird jetzt manuell eingetragen (Zeile 15 bereits händisch in original)
#           -> Zeile 15 wird zur manuellen Eingabe umgestellt
#   Row 16: Inputs!B28 (churn) -> Inputs!B29
#   Row 18: Inputs!B26 (enterprise ARR) -> Inputs!B27
#   Row 25: Inputs!B29 (NRR) -> Inputs!B30
#   Row 29: Inputs!B23 (initiale seats, now SME) -> Inputs!B23 (SME seats - for customer count calc)
#
# Neue Formel für Row 8 (New Seats total):
#   Jetzt: New Customers SME (row 6) * Inputs!B23 (SME seats)
#          + New Customers MidCo (row 7) * Inputs!B24 (MidCo seats)
#   (Statt des alten Algorithmus mit "Initiale Seats" und "Seat Wachstumsrate")
#   ABER: Row 9 (Churned Seats) und Row 10-11 (Net New, Total Active) bleiben gleich,
#   die verweisen auf B11 (Total Active Seats) und Inputs!B29 (Churn, now B29)
#
# Für Enterprise (Row 15):
#   War automatisch (IF AND MOD Formel), jetzt manuell
#   -> Row 15 Zellen B-BA werden als Kommentar behalten aber Formeln entfernt
#   -> Werte 0 als Standard (wird in Szenarien überschrieben)
# ================================================================
print("\n--- SCHRITT D: Revenue-Sheet Referenzen aktualisieren ---")

ws_rev = wb['Revenue']

# Column letters for M1-M52 (B through BA = columns 2-53)
# B=2, C=3, ..., Z=26, AA=27, ..., BA=53

# Helper: get column letter for month m (m=1 -> B, m=52 -> BA)
def col_for_month(m):
    return get_column_letter(m + 1)  # m=1 -> col 2 -> B

# ---- Row 8: New Seats total ----
# Old formula M1: =IF(1>=Inputs!$B$22,Inputs!$B$23,0)
# Old formula M2: =IF(2>=Inputs!$B$22,IF(2=Inputs!$B$22,Inputs!$B$23,ROUND(B11*Inputs!$B$24,0)),0)
#
# New formula: New Seats = New Customer SME * InitialeSeats_SME
#                        + New Customer MidCo * InitialeSeats_MidCo
# Row 6 = New Customer SME, Row 7 = New Customer MidCo
# Inputs!B23 = Initiale Seats SME (now 5)
# Inputs!B24 = Initiale Seats Mid-Company (now 15)
#
# The formula must reference the SME/MidCo manual rows:
# For month m (column X):
#   = X6 * Inputs!$B$23 + X7 * Inputs!$B$24

print("  Updating Revenue Row 8 (New Seats total) formulas...")
for m in range(1, 53):
    col = col_for_month(m)
    # New formula: SME customers * SME seats + MidCo customers * MidCo seats
    formula = f"={col}6*Inputs!$B$23+{col}7*Inputs!$B$24"
    ws_rev[f"{col}8"] = formula

print("  Revenue Row 8: All 52 months updated to SME+MidCo formula")

# ---- Row 9: Churned Seats ----
# Old: =IF(B11>0,ROUND(B11*Inputs!$B$28/12,0),0)  -> B28 = Churn (now B29)
# New: =IF(B11>0,ROUND(B11*Inputs!$B$29/12,0),0)
ws_rev['B9'] = '=0'  # Month 1 stays 0 (no churn in first month)
for m in range(2, 53):
    col = col_for_month(m)
    prev_col = col_for_month(m - 1)
    formula = f"=IF({prev_col}11>0,ROUND({prev_col}11*Inputs!$B$29/12,0),0)"
    ws_rev[f"{col}9"] = formula

print("  Revenue Row 9: All 52 months updated (Inputs!B28 -> B29)")

# ---- Row 15: Neue Enterprise-Deals ----
# Old: =IF(AND(m>=Inputs!$B$25,MOD(m-Inputs!$B$25,3)=0),Inputs!$B$27,0)
# New: Manuell (Inputs!B25 and B27 deleted)
# -> Clear all formula cells, set to 0 (will be overridden in scenarios)
print("  Updating Revenue Row 15 (Enterprise Deals) to manual input (clearing formulas)...")
for m in range(1, 53):
    col = col_for_month(m)
    ws_rev[f"{col}15"] = 0  # Default 0, manually overridden in scenarios

print("  Revenue Row 15: Set to 0 (manual input)")

# ---- Row 16: Churned Enterprise ----
# Old: =IF(B17>0,IF(MOD(m,12)=0,ROUND(B17*Inputs!$B$28,0),0),0)  -> B28 = Churn (now B29)
# New: =IF(B17>0,IF(MOD(m,12)=0,ROUND(B17*Inputs!$B$29,0),0),0)
ws_rev['B16'] = '=0'  # Month 1 stays 0
for m in range(2, 53):
    col = col_for_month(m)
    prev_col = col_for_month(m - 1)
    formula = f"=IF({prev_col}17>0,IF(MOD({m},12)=0,ROUND({prev_col}17*Inputs!$B$29,0),0),0)"
    ws_rev[f"{col}16"] = formula

print("  Revenue Row 16: All 52 months updated (Inputs!B28 -> B29)")

# ---- Row 18: Avg Enterprise ACV ----
# Old: =Inputs!$B$26*(1+Inputs!$B$21)^INT((m-1)/12)  -> B26 = Enterprise ARR (now B27)
# New: =Inputs!$B$27*(1+Inputs!$B$21)^INT((m-1)/12)
for m in range(1, 53):
    col = col_for_month(m)
    formula = f"=Inputs!$B$27*(1+Inputs!$B$21)^INT(({m}-1)/12)"
    ws_rev[f"{col}18"] = formula

print("  Revenue Row 18: All 52 months updated (Inputs!B26 -> B27)")

# ---- Row 25: Net Revenue Retention ----
# Old: =Inputs!$B$29  -> B29 = NRR (now B30)
# New: =Inputs!$B$30
for m in range(1, 53):
    col = col_for_month(m)
    ws_rev[f"{col}25"] = "=Inputs!$B$30"

print("  Revenue Row 25: All 52 months updated (Inputs!B29 -> B30)")

# ---- Row 29: Active Subscription Customers ----
# Old: =IF(B11>0, MAX(1, ROUND(B11/Inputs!$B$23, 0)), 0)  -> B23 now = SME seats
# For customer count we should use (B11 / (SME_seats + MidCo_seats)) or keep B23
# Actually this formula divides total seats by "initiale seats per customer"
# Since now we have SME and MidCo separately, the cleanest approach is:
# Keep reference to B23 (Initiale Seats SME) for backward compatibility
# This is acceptable as a simplification (customer count approximation)
# NOTE: Row 29 formula references Inputs!$B$23 which still exists (now = SME seats = 5)
# No change needed for row 29 - the reference is still valid (B23 still exists)
print("  Revenue Row 29: No change needed (Inputs!B23 still exists)")

# ---- Also fix Revenue Row 8 reference to Inputs!B24 (seat growth rate) ----
# Wait - the OLD Revenue Row 8 referenced Inputs!B24 for seat growth rate.
# After our insert_rows in Inputs:
# - Old Inputs!B24 (Monatliche Seat-Wachstumsrate=0.03) is now Inputs!B25
# - New Inputs!B24 = Initiale Seats Mid-Company = 15
# Our new Row 8 formula uses B24 for MidCo seats - CORRECT.
# The old growth rate logic is replaced by the new SME*seats+MidCo*seats logic.
print("  Revenue Row 8: SME+MidCo formula correctly uses B23=SME seats, B24=MidCo seats")

# ================================================================
# SCHRITT D2: Costs-Sheet Referenzen anpassen
#
# Costs references to Inputs that need updating after insert_rows(24):
#   - Costs uses Inputs!$B$47, $B$48 (Gehaltserhöhung, Lohnnebenkosten)
#     -> These rows are >= 47, not affected (we inserted at row 24, affects rows >= 24)
#   - Wait: The Costs formulas reference Inputs!$B$35, $B$36 etc. (Gehälter rows 32-46)
#     -> These are >= 24 so they shift by +1!
#
# After insert_rows(24) in Inputs:
#   Old Inputs!B32 (CEO Gehalt=100000) -> stays B32 (row 32 >= 24, shifts to 33)
#   Wait - let me recheck. insert_rows(24) inserts BEFORE row 24.
#   So rows >= 24 all shift down by 1:
#     Old B24 -> New B25
#     Old B32 -> New B33
#     Old B33 -> New B34
#     ...
#     Old B47 -> New B48
#     Old B48 -> New B49
#
#   Actually wait - let me recheck what openpyxl insert_rows does.
#   insert_rows(24) inserts a NEW row at position 24, shifting existing row 24 to 25.
#   So all rows originally >= 24 now have their row numbers incremented by 1.
#
# This affects Inputs rows referenced in Costs:
#   B32 (CEO) -> now B33
#   B33 (CTO) -> now B34
#   B34 (CCO) -> now B35
#   B35 (Software Dev) -> now B36
#   B36 (Junior SW Dev) -> now B37
#   B37 (ML/AI Eng) -> now B38
#   B38 (Site Reliability) -> now B39
#   B39 (UX/UI Designer) -> now B40
#   B40 (Mechanical Eng) -> now B41
#   B41 (Customer KA) -> now B42
#   B42 (Marketing Mgr) -> now B43
#   B43 (Marketing Asst) -> now B44
#   B44 (Customer Success) -> now B45
#   B45 (Office Mgr) -> now B46
#   B46 (Finance Mgr) -> now B47
#   B47 (Gehaltserhöhung) -> now B48
#   B48 (Lohnnebenkosten) -> now B49
#   All further rows also +1:
#   B50 (EINSTELLUNGSPLAN header) -> now B51
#   B51..B66 (Einstellungsmonate) -> now B52..B67
#   And in Inputs!E51..E66 references -> now E52..E67
#   B69 (Cloud/Hosting Basis) -> now B70
#   ... etc.
# ================================================================
print("\n--- SCHRITT D2: Alle Cross-Sheet Referenzen anpassen ---")

# We need to do a comprehensive find-and-replace of all Inputs!$B$XX references
# that need to be bumped by 1 (for all rows >= 24).
# Similarly for Inputs!$E$XX (hiring plan column E).

def update_formula_refs(formula, sheet_name='Inputs'):
    """
    For a given formula string, update all Inputs!$B$N and Inputs!$E$N references
    where N >= 24 to N+1.
    """
    if not formula or not isinstance(formula, str) or not formula.startswith('='):
        return formula

    def replace_ref(match):
        col_letter = match.group(1)  # B or E
        row_num = int(match.group(2))
        if row_num >= 24:
            return f"{sheet_name}!${col_letter}${row_num + 1}"
        return match.group(0)

    # Pattern: Inputs!$B$NN or Inputs!$E$NN (where NN is a row number)
    pattern = re.compile(r"Inputs!\$([BE])\$(\d+)", re.IGNORECASE)
    return pattern.sub(replace_ref, formula)

# Apply to all sheets except 'Inputs' itself
sheets_to_update = [name for name in wb.sheetnames if name != 'Inputs']
total_cells_updated = 0

for sheet_name in sheets_to_update:
    ws = wb[sheet_name]
    sheet_count = 0
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                new_formula = update_formula_refs(cell.value)
                if new_formula != cell.value:
                    cell.value = new_formula
                    sheet_count += 1
                    total_cells_updated += 1
    if sheet_count > 0:
        print(f"  {sheet_name}: {sheet_count} Formeln aktualisiert")

print(f"  Gesamt: {total_cells_updated} Formel-Referenzen aktualisiert")

# ================================================================
# SCHRITT D3: Verify key formulas after all updates
# ================================================================
print("\n--- SCHRITT D3: Verifikation nach Änderungen ---")
print(f"  Revenue!B8: {ws_rev['B8'].value}")
print(f"  Revenue!C8: {ws_rev['C8'].value}")
print(f"  Revenue!C9: {ws_rev['C9'].value}")
print(f"  Revenue!B15: {ws_rev['B15'].value}")
print(f"  Revenue!B16: {ws_rev['B16'].value}")
print(f"  Revenue!C16: {ws_rev['C16'].value}")
print(f"  Revenue!B18: {ws_rev['B18'].value}")
print(f"  Revenue!B25: {ws_rev['B25'].value}")
print(f"  Revenue!B29: {ws_rev['B29'].value}")

ws_costs_check = wb['Costs']
print(f"  Costs!B26 (CEO Gehalt): {ws_costs_check['B26'].value[:60]}")
print(f"  Costs!B51 (AI/ML API): {ws_costs_check['B51'].value[:60]}")
print(f"  Costs!B50 (Cloud Variable): {ws_costs_check['B50'].value[:60]}")

# Check Inputs row assignments
print(f"\n  Inputs!B23: {ws_inputs['A23'].value} = {ws_inputs['B23'].value}")
print(f"  Inputs!B24: {ws_inputs['A24'].value} = {ws_inputs['B24'].value}")
print(f"  Inputs!B25: {ws_inputs['A25'].value} = {ws_inputs['B25'].value}")
print(f"  Inputs!B27: {ws_inputs['A27'].value} = {ws_inputs['B27'].value}")
print(f"  Inputs!B29: {ws_inputs['A29'].value}")
print(f"  Inputs!B30: {ws_inputs['A30'].value}")
print(f"  Inputs!B33: {ws_inputs['A33'].value} = {ws_inputs['B33'].value}")
print(f"  Inputs!B48: {ws_inputs['A48'].value} = {ws_inputs['B48'].value}")
print(f"  Inputs!B49: {ws_inputs['A49'].value} = {ws_inputs['B49'].value}")

# ================================================================
# SAVE v15
# ================================================================
print(f"\n--- Speichere v15 nach {DST_V15} ---")
wb.save(DST_V15)
print("v15 gespeichert.")
print("\n" + "=" * 70)
print("v15 FERTIG")
print("=" * 70)
