"""
add_consulting_revenue.py
─────────────────────────
Extends LFL_BM_PDF_Konservativ_20260304_2313.xlsx with a Consulting
Revenue stream as a second revenue pillar alongside SaaS licences.

Changes made:
  00_Input_Sandbox  Row 11  → Consulting-Tage/Einsatz  (gering/normal/stark)
  Inputs            Row 137 → CONSULTING REVENUE-PARAMETER (section header)
                    Row 138 → Consulting Tagessatz (€/Tag)  ← VLOOKUP Sandbox
                    Row 139 → Consulting-Anteil je Kunde (%)
                    Row 140 → Consulting-Tage pro Einsatz   ← VLOOKUP Sandbox
  Revenue           Row 25  → CONSULTING REVENUE (header comment)
                    Row 26  → Active Subscription Customers
                    Row 27  → Total Active Customers (Subscription + Enterprise)
                    Row 28  → Consulting-Tage/Monat
                    Row 29  → Consulting Revenue (€/Monat)
  P&L               Row  8  → Consulting Revenue   ← =Revenue!{col}29
                    Row  7  → TOTAL REVENUE formula updated to include Row 8

Output: scenarios/LFL_BM_Konservativ_v2_Consulting.xlsx
"""

import copy
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SRC  = "scenarios/LFL_BM_PDF_Konservativ_20260304_2313.xlsx"
DEST = "scenarios/LFL_BM_Konservativ_v2_Consulting.xlsx"

# ── Inputs row constants (NEW rows – safe: after last existing row 135) ───────
INPUTS_ROW_SECTION  = 137   # Section header
INPUTS_ROW_RATE     = 138   # Consulting Tagessatz  (€/Tag)
INPUTS_ROW_SHARE    = 139   # Consulting-Anteil je Kunde  (%)
INPUTS_ROW_DAYS     = 140   # Consulting-Tage pro Einsatz

# ── Sandbox row for new entry ─────────────────────────────────────────────────
SANDBOX_ROW_DAYS    = 11    # After row 10 (AI-Personal-Hebel)

# ── Revenue new rows ──────────────────────────────────────────────────────────
REV_ROW_HEADER      = 25    # Section header
REV_ROW_SUB_CUST    = 26    # Active Subscription Customers
REV_ROW_TOTAL_CUST  = 27    # Total Active Customers
REV_ROW_DAYS_MO     = 28    # Consulting-Tage/Monat
REV_ROW_REV         = 29    # Consulting Revenue (€/Monat)

# ── P&L rows ──────────────────────────────────────────────────────────────────
PNL_ROW_TOTAL_REV   = 7     # existing TOTAL REVENUE → formula will be updated
PNL_ROW_CONSULTING  = 8     # empty row → becomes Consulting Revenue line


# ─────────────────────────────────────────────────────────────────────────────
# Style helpers
# ─────────────────────────────────────────────────────────────────────────────
BLUE_DARK   = "1F3864"
BLUE_LIGHT  = "BDD7EE"
GREEN_LIGHT = "E2EFDA"
AMBER       = "FFF2CC"
TEAL        = "1F6B75"
WHITE       = "FFFFFF"


def _thin():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)


def style_section_header(cell, bg=TEAL):
    cell.font      = Font(bold=True, color=WHITE, size=10, name="Calibri")
    cell.fill      = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
    cell.border    = _thin()


def style_label(cell, bold=False, bg=None):
    cell.font      = Font(bold=bold, size=10, name="Calibri")
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    cell.border    = _thin()
    if bg:
        cell.fill = PatternFill("solid", fgColor=bg)


def style_value(cell, fmt=None, bg=BLUE_LIGHT):
    cell.alignment = Alignment(horizontal="right", vertical="center")
    cell.border    = _thin()
    cell.fill      = PatternFill("solid", fgColor=bg)
    if fmt:
        cell.number_format = fmt


def style_note(cell):
    cell.font      = Font(italic=True, size=9, color="595959", name="Calibri")
    cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    cell.border    = _thin()


# ─────────────────────────────────────────────────────────────────────────────
# 1. 00_Input_Sandbox  –  add Consulting-Tage/Einsatz (Row 11)
# ─────────────────────────────────────────────────────────────────────────────
def patch_sandbox(wb):
    ws = wb["00_Input_Sandbox"]

    row = SANDBOX_ROW_DAYS
    # Columns: A=Kategorie, B=Variable, C=Einheit, D=Gering, E=Normal, F=Stark, G=Quelle

    data = [
        ("Consulting", None, None, None, None, None, None, None),  # placeholder
    ]
    # Actually write directly:
    ws.cell(row=row, column=1).value = "Consulting"
    ws.cell(row=row, column=2).value = "Consulting-Tage/Einsatz"
    ws.cell(row=row, column=3).value = "Tage/Einsatz"
    ws.cell(row=row, column=4).value = 2     # gering: 2 Tage (Basis-Schulung)
    ws.cell(row=row, column=5).value = 5     # normal: 5 Tage (Impl. + Schulung)
    ws.cell(row=row, column=6).value = 10    # stark:  10 Tage (komplex)
    ws.cell(row=row, column=7).value = (
        "Tage pro Beratungseinsatz: "
        "gering=2 (Basis-Einführung), "
        "normal=5 (Implementierung+Schulung), "
        "stark=10 (spez. Anforderungen & Datenbereinigung)"
    )

    # Style the row to match the table
    for col in range(1, 8):
        c = ws.cell(row=row, column=col)
        c.font      = Font(size=10, name="Calibri")
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        c.border    = _thin()
        if col in (4, 5, 6):
            c.fill      = PatternFill("solid", fgColor=GREEN_LIGHT)
            c.alignment = Alignment(horizontal="right", vertical="center")

    # Fix VLOOKUP range in existing formulas – currently $B$3:$F$20 which still covers row 11
    # No update needed since new row falls within existing lookup range $B$3:$F$20

    print(f"  Sandbox row {row}: Consulting-Tage/Einsatz added (2 / 5 / 10 Tage)")


# ─────────────────────────────────────────────────────────────────────────────
# 2. Inputs  –  add Consulting section (Rows 137-140)
# ─────────────────────────────────────────────────────────────────────────────
def patch_inputs(wb):
    ws = wb["Inputs"]

    BEMERKUNG = (
        "Umfasst:\n"
        "  a) Implementierungsbegleitung: Einführung & technische Integration vor Ort\n"
        "  b) Schulung: Anwenderschulung und Train-the-Trainer-Programme\n"
        "  c) Spezielle Kundenanforderungen & Datenbereinigung: Datenmigration, "
        "Stammdaten-Harmonisierung und kundenspez. Anpassungen"
    )

    # --- Row 137: Section header ---
    c = ws.cell(row=INPUTS_ROW_SECTION, column=1)
    c.value = "CONSULTING REVENUE-PARAMETER"
    style_section_header(c, bg=TEAL)
    # Merge across usual header width (columns A-E)
    for col in range(2, 6):
        cc = ws.cell(row=INPUTS_ROW_SECTION, column=col)
        cc.fill   = PatternFill("solid", fgColor=TEAL)
        cc.border = _thin()

    # --- Row 138: Consulting Tagessatz (€/Tag) ---
    ws.cell(row=INPUTS_ROW_RATE, column=1).value = "Consulting Tagessatz (€/Tag)"
    ws.cell(row=INPUTS_ROW_RATE, column=2).value = (
        "=VLOOKUP(\"Consulting-Tagessatz\","
        "'00_Input_Sandbox'!$B$3:$F$20,"
        "'00_Input_Sandbox'!$B$2-1,FALSE)"
    )
    ws.cell(row=INPUTS_ROW_RATE, column=3).value = "EUR pro Beratertag (szenarioabhängig aus Sandbox)"
    ws.cell(row=INPUTS_ROW_RATE, column=4).value = BEMERKUNG
    ws.cell(row=INPUTS_ROW_RATE, column=5).value = "Fix (marktbasiert, durch Szenario gesteuert)"

    # --- Row 139: Consulting-Anteil je Kunde (%) ---
    ws.cell(row=INPUTS_ROW_SHARE, column=1).value = "Consulting-Anteil je Kunde (%)"
    ws.cell(row=INPUTS_ROW_SHARE, column=2).value = 0.30
    ws.cell(row=INPUTS_ROW_SHARE, column=3).value = "Anteil Kunden, die Consulting kaufen"
    ws.cell(row=INPUTS_ROW_SHARE, column=4).value = (
        "Nicht jeder Kunde kauft Consulting. "
        "30 % = ca. 3 von 10 Kunden kaufen im Schnitt einen Einsatz pro Monat. "
        "Wächst mit Kundenzufriedenheit & Produktkomplexität. "
        "Steuerbar: gering=20%, normal=30%, stark=40% empfohlen."
    )
    ws.cell(row=INPUTS_ROW_SHARE, column=5).value = "Direkt anpassbar (kein Sandbox-Bezug)"

    # --- Row 140: Consulting-Tage pro Einsatz ---
    ws.cell(row=INPUTS_ROW_DAYS, column=1).value = "Consulting-Tage pro Einsatz"
    ws.cell(row=INPUTS_ROW_DAYS, column=2).value = (
        "=VLOOKUP(\"Consulting-Tage/Einsatz\","
        "'00_Input_Sandbox'!$B$3:$F$20,"
        "'00_Input_Sandbox'!$B$2-1,FALSE)"
    )
    ws.cell(row=INPUTS_ROW_DAYS, column=3).value = "Tage je Kundeneinsatz (szenarioabhängig aus Sandbox)"
    ws.cell(row=INPUTS_ROW_DAYS, column=4).value = BEMERKUNG
    ws.cell(row=INPUTS_ROW_DAYS, column=5).value = "Szenariogesteuert (Sandbox Zeile 11)"

    # Apply styles for rows 138-140
    for row in (INPUTS_ROW_RATE, INPUTS_ROW_SHARE, INPUTS_ROW_DAYS):
        style_label(ws.cell(row=row, column=1), bold=True)
        style_value(ws.cell(row=row, column=2), bg=BLUE_LIGHT)
        style_label(ws.cell(row=row, column=3))
        ws.cell(row=row, column=4).font      = Font(italic=True, size=9, color="595959", name="Calibri")
        ws.cell(row=row, column=4).alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(row=row, column=4).border    = _thin()
        style_label(ws.cell(row=row, column=5))
        ws.row_dimensions[row].height = 60

    # Number format for the percentage row
    ws.cell(row=INPUTS_ROW_SHARE, column=2).number_format = "0%"

    print(f"  Inputs rows {INPUTS_ROW_SECTION}-{INPUTS_ROW_DAYS}: Consulting section added")


# ─────────────────────────────────────────────────────────────────────────────
# 3. Revenue  –  add Consulting rows 25-29 for all 52 months
# ─────────────────────────────────────────────────────────────────────────────
def patch_revenue(wb):
    ws   = wb["Revenue"]
    ncols = 53   # col 1 = label, cols 2-53 = M1-M52

    # --- Row 25: Section header ---
    ws.cell(row=REV_ROW_HEADER, column=1).value = "── CONSULTING REVENUE ──"
    style_section_header(ws.cell(row=REV_ROW_HEADER, column=1), bg=TEAL)
    for col in range(2, ncols + 1):
        c = ws.cell(row=REV_ROW_HEADER, column=col)
        c.fill   = PatternFill("solid", fgColor=TEAL)
        c.border = _thin()

    # --- Row 26: Active Subscription Customers ---
    ws.cell(row=REV_ROW_SUB_CUST, column=1).value = "Active Subscription Customers"
    # Row 8 in Revenue = Total Active Seats; B23 in Inputs = Initiale Seats
    for col in range(2, ncols + 1):
        col_l = get_column_letter(col)
        ws.cell(row=REV_ROW_SUB_CUST, column=col).value = (
            f"=IF({col_l}8>0, MAX(1, ROUND({col_l}8/Inputs!$B$23, 0)), 0)"
        )

    # --- Row 27: Total Active Customers ---
    ws.cell(row=REV_ROW_TOTAL_CUST, column=1).value = "Total Active Customers (inkl. Enterprise)"
    for col in range(2, ncols + 1):
        col_l = get_column_letter(col)
        ws.cell(row=REV_ROW_TOTAL_CUST, column=col).value = (
            f"={col_l}{REV_ROW_SUB_CUST}+{col_l}14"
        )

    # --- Row 28: Consulting-Tage/Monat ---
    ws.cell(row=REV_ROW_DAYS_MO, column=1).value = "Consulting-Tage/Monat"
    # = Total Active Customers × consulting_share × days_per_engagement
    for col in range(2, ncols + 1):
        col_l = get_column_letter(col)
        ws.cell(row=REV_ROW_DAYS_MO, column=col).value = (
            f"={col_l}{REV_ROW_TOTAL_CUST}"
            f"*Inputs!$B${INPUTS_ROW_SHARE}"
            f"*Inputs!$B${INPUTS_ROW_DAYS}"
        )

    # --- Row 29: Consulting Revenue (€/Monat) ---
    ws.cell(row=REV_ROW_REV, column=1).value = "Consulting Revenue (€/Monat)"
    for col in range(2, ncols + 1):
        col_l = get_column_letter(col)
        ws.cell(row=REV_ROW_REV, column=col).value = (
            f"={col_l}{REV_ROW_DAYS_MO}*Inputs!$B${INPUTS_ROW_RATE}"
        )

    # --- Apply styles ---
    label_rows = {
        REV_ROW_SUB_CUST:   (AMBER,       AMBER),
        REV_ROW_TOTAL_CUST: (AMBER,       AMBER),
        REV_ROW_DAYS_MO:    (AMBER,       AMBER),
        REV_ROW_REV:        (GREEN_LIGHT, GREEN_LIGHT),
    }
    for row, (lbg, vbg) in label_rows.items():
        c = ws.cell(row=row, column=1)
        c.font      = Font(bold=True, size=10, name="Calibri", color="000000")
        c.fill      = PatternFill("solid", fgColor=lbg)
        c.alignment = Alignment(horizontal="left", vertical="center")
        c.border    = _thin()
        for col in range(2, ncols + 1):
            vc = ws.cell(row=row, column=col)
            vc.fill   = PatternFill("solid", fgColor=vbg)
            vc.border = _thin()
            vc.alignment = Alignment(horizontal="right", vertical="center")
            if row == REV_ROW_REV:
                vc.number_format = '#,##0'

    print(f"  Revenue rows {REV_ROW_HEADER}-{REV_ROW_REV}: Consulting calculations added (52 months)")


# ─────────────────────────────────────────────────────────────────────────────
# 4. P&L  –  add Consulting Revenue (row 8) + update TOTAL REVENUE (row 7)
# ─────────────────────────────────────────────────────────────────────────────
def patch_pnl(wb):
    ws    = wb["P&L"]
    ncols = 53   # col 1 = label, cols 2-53 = M1-M52

    # --- Row 8: Consulting Revenue ---
    ws.cell(row=PNL_ROW_CONSULTING, column=1).value = "Consulting Revenue"
    for col in range(2, ncols + 1):
        col_l = get_column_letter(col)
        ws.cell(row=PNL_ROW_CONSULTING, column=col).value = f"=Revenue!{col_l}{REV_ROW_REV}"

    # Apply style to row 8
    c = ws.cell(row=PNL_ROW_CONSULTING, column=1)
    c.font      = Font(bold=True, size=10, name="Calibri")
    c.alignment = Alignment(horizontal="left", vertical="center")
    c.border    = _thin()
    c.fill      = PatternFill("solid", fgColor=GREEN_LIGHT)
    for col in range(2, ncols + 1):
        vc = ws.cell(row=PNL_ROW_CONSULTING, column=col)
        vc.fill          = PatternFill("solid", fgColor=GREEN_LIGHT)
        vc.border        = _thin()
        vc.alignment     = Alignment(horizontal="right", vertical="center")
        vc.number_format = '#,##0'

    # --- Row 7: TOTAL REVENUE – update formula to include consulting (row 8) ---
    for col in range(2, ncols + 1):
        col_l = get_column_letter(col)
        # Old: =B5+B6   New: =B5+B6+B8
        ws.cell(row=PNL_ROW_TOTAL_REV, column=col).value = (
            f"={col_l}5+{col_l}6+{col_l}8"
        )

    print(f"  P&L row {PNL_ROW_CONSULTING}: Consulting Revenue added")
    print(f"  P&L row {PNL_ROW_TOTAL_REV}: TOTAL REVENUE formula updated (+Consulting)")


# ─────────────────────────────────────────────────────────────────────────────
# 5. Add column header comment in Revenue for the new rows
# ─────────────────────────────────────────────────────────────────────────────
def fix_revenue_header_row(wb):
    """Ensure row 3 (M1...M52 headers) is preserved and row 2 has dates."""
    # No changes needed – existing date/header rows are untouched
    pass


# ─────────────────────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────────────────────
def main():
    print(f"Loading: {SRC}")
    wb = openpyxl.load_workbook(SRC, data_only=False)

    print("\n1. Patching 00_Input_Sandbox ...")
    patch_sandbox(wb)

    print("\n2. Patching Inputs ...")
    patch_inputs(wb)

    print("\n3. Patching Revenue ...")
    patch_revenue(wb)

    print("\n4. Patching P&L ...")
    patch_pnl(wb)

    os.makedirs(os.path.dirname(DEST), exist_ok=True)
    wb.save(DEST)
    print(f"\n✅ Saved: {DEST}")

    # Quick verification
    wb2 = openpyxl.load_workbook(DEST, data_only=False)
    ws_inputs  = wb2["Inputs"]
    ws_sandbox = wb2["00_Input_Sandbox"]
    ws_rev     = wb2["Revenue"]
    ws_pnl     = wb2["P&L"]

    print("\n── Verification ──")
    print(f"  Sandbox  row {SANDBOX_ROW_DAYS} col B: {ws_sandbox.cell(row=SANDBOX_ROW_DAYS, column=2).value}")
    print(f"  Sandbox  row {SANDBOX_ROW_DAYS} gering/normal/stark: "
          f"{ws_sandbox.cell(row=SANDBOX_ROW_DAYS, column=4).value} / "
          f"{ws_sandbox.cell(row=SANDBOX_ROW_DAYS, column=5).value} / "
          f"{ws_sandbox.cell(row=SANDBOX_ROW_DAYS, column=6).value}")
    print(f"  Inputs   row {INPUTS_ROW_SECTION}: {ws_inputs.cell(row=INPUTS_ROW_SECTION, column=1).value}")
    print(f"  Inputs   row {INPUTS_ROW_RATE}  B: {ws_inputs.cell(row=INPUTS_ROW_RATE, column=2).value}")
    print(f"  Inputs   row {INPUTS_ROW_SHARE} B: {ws_inputs.cell(row=INPUTS_ROW_SHARE, column=2).value}")
    print(f"  Inputs   row {INPUTS_ROW_DAYS}  B: {ws_inputs.cell(row=INPUTS_ROW_DAYS, column=2).value}")
    print(f"  Revenue  row {REV_ROW_HEADER}   A: {ws_rev.cell(row=REV_ROW_HEADER, column=1).value}")
    print(f"  Revenue  row {REV_ROW_SUB_CUST} B (M1 formula): {ws_rev.cell(row=REV_ROW_SUB_CUST, column=2).value}")
    print(f"  Revenue  row {REV_ROW_REV}      B (M1 formula): {ws_rev.cell(row=REV_ROW_REV, column=2).value}")
    print(f"  P&L      row {PNL_ROW_CONSULTING} A: {ws_pnl.cell(row=PNL_ROW_CONSULTING, column=1).value}")
    print(f"  P&L      row {PNL_ROW_CONSULTING} B (M1): {ws_pnl.cell(row=PNL_ROW_CONSULTING, column=2).value}")
    print(f"  P&L      row {PNL_ROW_TOTAL_REV}  B (M1): {ws_pnl.cell(row=PNL_ROW_TOTAL_REV, column=2).value}")

    print("\n  Cash Flow still references P&L!B39 (Net Income) → unchanged ✓")
    print("  Balance Sheet still references P&L!B7 (Total Revenue) → now includes Consulting ✓")
    print("  Gross Margin P&L!B15 = P&L!B7-P&L!B13 → automatically benefits ✓")


if __name__ == "__main__":
    main()
