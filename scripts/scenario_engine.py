"""
LFL Financial Projection Scenario Engine
Version: 1.0
Basis: template_v0.4.xlsx

Verwendung:
  python3 scripts/scenario_engine.py --szenario stark
  python3 scripts/scenario_engine.py --changes "B28=0.12,B21=0.03" --name HighChurn
  python3 scripts/scenario_engine.py --baseline
"""

import openpyxl
import argparse
import json
import os
from datetime import datetime
from copy import copy

TEMPLATE_PATH        = "templates/template_v0.4.xlsx"
CUSTOM_BASELINE_PATH = "templates/baseline_custom.xlsx"
SCENARIOS_DIR        = "scenarios"
REPORTS_DIR          = "reports"


def get_active_template():
    """Gibt den Pfad der aktiven Basis zurück (Custom oder Original)."""
    if os.path.exists(CUSTOM_BASELINE_PATH):
        return CUSTOM_BASELINE_PATH
    return TEMPLATE_PATH


def save_as_custom_baseline(changes_inputs):
    """
    Schreibt geänderte Inputs in eine Custom-Basis-Datei (kein Szenario-Präfix).
    Gibt den Pfad zurück.
    """
    import shutil
    # Vom Original-Template starten (nicht von einer ggf. alten Custom-Basis)
    wb = openpyxl.load_workbook(TEMPLATE_PATH, data_only=False)
    ws_inputs = wb['Inputs']

    for cell_ref, new_val in changes_inputs.items():
        row = int(cell_ref[1:])
        old_val = ws_inputs.cell(row=row, column=2).value
        if isinstance(old_val, str) and old_val.startswith('='):
            continue  # Formeln nicht überschreiben
        ws_inputs.cell(row=row, column=2).value = new_val

    os.makedirs(os.path.dirname(CUSTOM_BASELINE_PATH), exist_ok=True)
    wb.save(CUSTOM_BASELINE_PATH)
    return CUSTOM_BASELINE_PATH


def delete_custom_baseline():
    """Löscht die Custom-Basis — Tool fällt zurück auf Original-Template."""
    if os.path.exists(CUSTOM_BASELINE_PATH):
        os.remove(CUSTOM_BASELINE_PATH)


def is_custom_baseline_active():
    return os.path.exists(CUSTOM_BASELINE_PATH)

# Mapping: editierbare Input-Zellen (Inputs-Sheet, Spalte B)
INPUT_CELL_LABELS = {
    "B4":  "Firmenname",
    "B5":  "Startdatum",
    "B6":  "Währung",
    "B7":  "Steuersatz",
    "B10": "Ideation Phase Betrag",
    "B11": "Ideation Phase Monat",
    "B12": "Pre-Seed Betrag",
    "B13": "Pre-Seed Monat",
    "B14": "Seed Betrag",
    "B15": "Seed Monat",
    "B16": "Series A Betrag",
    "B17": "Series A Monat",
    # B20 + B22: VLOOKUP aus Sandbox – NICHT direkt editieren
    "B21": "Preiserhöhung/Jahr",
    "B23": "Initiale Seats",
    "B24": "Monatliche Seat-Wachstumsrate",
    "B25": "Enterprise-Deals ab Monat",
    "B26": "Durchschnitt Enterprise ARR",
    "B27": "Enterprise Deals pro Quartal",
    "B28": "Jährliche Churn Rate",
    "B29": "Net Revenue Retention",
    "B32": "CEO Gehalt",
    "B33": "CTO Gehalt",
    "B34": "CCO Gehalt",
    "B35": "Senior Engineer Gehalt",
    "B36": "Junior Engineer Gehalt",
    "B37": "ML/AI Engineer Gehalt",
    "B38": "Product Manager Gehalt",
    "B39": "Sales Representative Gehalt",
    "B40": "Marketing Manager Gehalt",
    "B41": "Customer Success Gehalt",
    "B42": "Office/Admin Gehalt",
    "B43": "Finance Manager Gehalt",
    "B44": "Jährliche Gehaltserhöhung",
    "B45": "Lohnnebenkosten",
    "B82": "Cloud/Hosting Basis",
    "B83": "Cloud Skalierung pro Seat",
    "B84": "AI/ML APIs Basis",
    "B85": "AI Kosten Wachstum/Monat",
    "B86": "SaaS Tools intern Basis",
    "B87": "SaaS Tools pro MA",
    "B88": "Software-Lizenzen Dev",
    "B89": "Sicherheit & Compliance",
    "B92": "Laptop pro MA",
    "B93": "Monitor & Peripherie",
    "B94": "Ersatz-Zyklus Monate",
    "B95": "Sonstige IT-Ausstattung/Jahr",
    "B98": "Büromiete",
    "B99": "Büro-Upgrade ab Monat",
    "B100": "Neue Büromiete",
    "B101": "Nebenkosten",
    "B102": "Internet & Telefon",
    "B103": "Büroausstattung Initial",
    "B104": "Bürobedarf pro MA",
    "B107": "Rechtsanwalt Basis/Jahr",
    "B108": "RA pro Finanzierung",
    "B109": "Steuerberater",
    "B110": "Wirtschaftsprüfer/Jahr",
    "B111": "WP ab Monat",
    "B112": "Unternehmensberater/Jahr",
    "B115": "D&O Versicherung/Jahr",
    "B116": "Betriebshaftpflicht/Jahr",
    "B117": "Cyber-Versicherung/Jahr",
    "B118": "Bankgebühren",
    "B119": "Payment Processing",
    "B122": "Paid Ads Budget Initial",
    "B123": "Ads Budget Wachstum/Monat",
    "B124": "Content & SEO",
    "B125": "Events & Messen/Jahr",
    "B126": "Sales Tools",
    "B127": "Sales Provision",
    "B128": "Reisekosten Sales/MA/Monat",
    "B131": "Reisekosten allgemein/MA/Jahr",
    "B132": "Weiterbildung/MA/Jahr",
    "B133": "Team Events/MA/Jahr",
    "B134": "Sonstiges/Puffer",
    "B135": "Abschreibungsdauer Jahre",
}

SANDBOX_CELL_LABELS = {
    "D4": "Anteil Packaging (Gering)",
    "E4": "Anteil Packaging (Normal)",
    "F4": "Anteil Packaging (Stark)",
    "D5": "Startmonat Kunden (Gering)",
    "E5": "Startmonat Kunden (Normal)",
    "F5": "Startmonat Kunden (Stark)",
    "D6": "Startpreis Seat/Jahr (Gering)",
    "E6": "Startpreis Seat/Jahr (Normal)",
    "F6": "Startpreis Seat/Jahr (Stark)",
    "D7": "Consulting-Tagessatz (Gering)",
    "E7": "Consulting-Tagessatz (Normal)",
    "F7": "Consulting-Tagessatz (Stark)",
    "D9": "Onboarding-Aufwand (Gering)",
    "E9": "Onboarding-Aufwand (Normal)",
    "F9": "Onboarding-Aufwand (Stark)",
    "D10": "AI-Personal-Hebel (Gering)",
    "E10": "AI-Personal-Hebel (Normal)",
    "F10": "AI-Personal-Hebel (Stark)",
}

# Schlüssel-KPI Zellreferenzen (data_only=True Baseline)
# Spalten für M12=M, M24=Y, M36=AK, M52=BA (0-indexed: B=col2, M12=col14)
# In openpyxl: Spalte B = Index 2, M1=col2, M12=col13, M24=col25, M36=col37, M52=col53
KPI_MONTHS = {
    "M12": 13,   # Spalte N in openpyxl (1-indexed)
    "M24": 25,   # Spalte Z
    "M36": 37,   # Spalte AL
    "M52": 53,   # Spalte BB
}

# Zeilen der KPIs in den jeweiligen Sheets (1-indexed)
KPI_CELLS = {
    "Total ARR":       ("Revenue",    10),  # Zeile 10 = Subscription ARR
    "Total Headcount": ("Costs",       4),  # Zeile 4 = Total Headcount (anpassen wenn nötig)
    "EBITDA":          ("P&L",        21),  # Zeile 21 = EBITDA
    "Ending Cash":     ("Cash Flow",  19),  # Zeile 19 = Ending Cash
    "Burn Rate":       ("Cash Flow",  22),  # Zeile 22 = Burn Rate
    "Runway":          ("Cash Flow",  23),  # Zeile 23 = Runway
}


def detect_kpi_rows(wb_calc):
    """Findet KPI-Zeilenindizes automatisch anhand der Zeilenbeschriftungen."""
    kpi_map = {}
    sheet_kpis = {
        "Revenue":    ["Subscription ARR", "Total ARR", "TOTAL MRR"],
        "Costs":      ["TOTAL HEADCOUNT", "Total Headcount"],
        "P&L":        ["EBITDA", "Ebitda"],
        "Cash Flow":  ["Ending Cash", "ENDING CASH", "Burn Rate", "BURN RATE", "Runway", "RUNWAY"],
    }

    for sheet_name, keywords in sheet_kpis.items():
        if sheet_name not in wb_calc.sheetnames:
            continue
        ws = wb_calc[sheet_name]
        for row in ws.iter_rows(min_col=1, max_col=1):
            cell = row[0]
            if cell.value and isinstance(cell.value, str):
                for kw in keywords:
                    if kw.lower() in cell.value.lower():
                        kpi_map[f"{sheet_name}:{cell.value.strip()}"] = cell.row
    return kpi_map


def read_baseline_kpis(wb_calc):
    """Liest KPI-Werte aus einer data_only=True geladenen Workbook."""
    kpi_map = detect_kpi_rows(wb_calc)

    results = {}
    month_cols = {"M12": 13, "M24": 25, "M36": 37, "M52": 53}

    kpi_lookups = [
        ("Total ARR",       "Revenue",   ["subscription arr", "total arr"]),
        ("Total MRR",       "Revenue",   ["total mrr", "total monthly"]),
        ("Total Headcount", "Costs",     ["total headcount", "gesamt headcount"]),
        ("EBITDA",          "P&L",       ["ebitda"]),
        ("Net Income",      "P&L",       ["net income", "jahresüberschuss"]),
        ("Ending Cash",     "Cash Flow", ["ending cash", "kassenbestand"]),
        ("Burn Rate",       "Cash Flow", ["burn rate"]),
        ("Runway",          "Cash Flow", ["runway"]),
    ]

    for kpi_name, sheet_name, search_terms in kpi_lookups:
        if sheet_name not in wb_calc.sheetnames:
            continue
        ws = wb_calc[sheet_name]
        row_idx = None
        for row in ws.iter_rows(min_col=1, max_col=1):
            cell = row[0]
            if cell.value and isinstance(cell.value, str):
                for term in search_terms:
                    if term in cell.value.lower():
                        row_idx = cell.row
                        break
            if row_idx:
                break

        if row_idx is None:
            results[kpi_name] = {m: "n/a (Zeile nicht gefunden)" for m in month_cols}
            continue

        results[kpi_name] = {}
        for month_label, col_idx in month_cols.items():
            try:
                val = ws.cell(row=row_idx, column=col_idx).value
                results[kpi_name][month_label] = val
            except Exception:
                results[kpi_name][month_label] = None

    return results


def read_current_inputs(wb):
    """Liest aktuelle Werte der editierbaren Inputs."""
    ws_inputs = wb['Inputs']
    ws_sandbox = wb['00_Input_Sandbox']

    current = {}
    for cell_ref, label in INPUT_CELL_LABELS.items():
        row = int(cell_ref[1:])
        cell = ws_inputs.cell(row=row, column=2)
        val = cell.value
        # Formeln als Formel-String kennzeichnen
        if isinstance(val, str) and val.startswith('='):
            val = f"[FORMEL: {val[:50]}...]"
        current[cell_ref] = {"label": label, "value": val}

    current["SANDBOX_B1"] = {"label": "Aktives Szenario", "value": ws_sandbox['B1'].value}
    return current


def apply_scenario(changes_dict, szenario=None):
    """
    Hauptfunktion: Lädt Template, ändert Inputs, speichert neue Datei.

    changes_dict: {"Inputs": {"B28": 0.12, "B21": 0.03}, "Sandbox": {"B1": "stark"}}
    szenario: "gering" | "normal" | "stark" (Shortcut für Sandbox!B1)
    """
    active_tpl = get_active_template()
    if not os.path.exists(active_tpl):
        raise FileNotFoundError(f"Template nicht gefunden: {active_tpl}")

    # Template laden (Formeln erhalten)
    wb = openpyxl.load_workbook(active_tpl, data_only=False)
    ws_inputs = wb['Inputs']
    ws_sandbox = wb['00_Input_Sandbox']

    applied_changes = []

    # Sandbox-Szenario setzen
    if szenario:
        old_val = ws_sandbox['B1'].value
        ws_sandbox['B1'] = szenario
        applied_changes.append({
            "sheet": "00_Input_Sandbox",
            "cell": "B1",
            "label": "Aktives Szenario",
            "old": old_val,
            "new": szenario
        })

    # Input-Änderungen anwenden
    if "Inputs" in changes_dict:
        for cell_ref, new_val in changes_dict["Inputs"].items():
            row = int(cell_ref[1:])
            old_val = ws_inputs.cell(row=row, column=2).value
            label = INPUT_CELL_LABELS.get(cell_ref, cell_ref)

            # Sicherheitscheck: Keine Formelzellen überschreiben
            if isinstance(old_val, str) and old_val.startswith('='):
                print(f"  WARNUNG: {cell_ref} ({label}) ist eine Formelzelle – übersprungen!")
                continue

            ws_inputs.cell(row=row, column=2).value = new_val
            applied_changes.append({
                "sheet": "Inputs",
                "cell": cell_ref,
                "label": label,
                "old": old_val,
                "new": new_val
            })

    # Sandbox-Werte direkt ändern (z.B. einzelne Szenario-Spalten)
    if "Sandbox" in changes_dict:
        for cell_ref, new_val in changes_dict["Sandbox"].items():
            if cell_ref == "B1":
                old_val = ws_sandbox['B1'].value
                ws_sandbox['B1'] = new_val
                applied_changes.append({
                    "sheet": "00_Input_Sandbox",
                    "cell": cell_ref,
                    "label": "Aktives Szenario",
                    "old": old_val,
                    "new": new_val
                })
            else:
                old_val = ws_sandbox[cell_ref].value
                ws_sandbox[cell_ref] = new_val
                label = SANDBOX_CELL_LABELS.get(cell_ref, cell_ref)
                applied_changes.append({
                    "sheet": "00_Input_Sandbox",
                    "cell": cell_ref,
                    "label": label,
                    "old": old_val,
                    "new": new_val
                })

    return wb, applied_changes


def save_scenario(wb, scenario_name):
    """Speichert Szenario-Datei mit Namenskonvention."""
    os.makedirs(SCENARIOS_DIR, exist_ok=True)
    timestamp = datetime.now().strftime('%Y%m%d_%H%M')
    filename = f"LFL_BM_{scenario_name}_{timestamp}.xlsx"
    filepath = os.path.join(SCENARIOS_DIR, filename)
    wb.save(filepath)
    return filepath


def generate_delta_report(scenario_name, applied_changes, baseline_kpis, filepath):
    """Erstellt Delta-Bericht als Markdown."""
    os.makedirs(REPORTS_DIR, exist_ok=True)
    timestamp = datetime.now().strftime('%Y%m%d_%H%M')
    report_path = os.path.join(REPORTS_DIR, f"Report_{scenario_name}_{timestamp}.md")

    lines = [
        "═══════════════════════════════════════════════════════",
        f"SZENARIO: {scenario_name}",
        f"Basis: v0.4 Status quo | Erstellt: {datetime.now().strftime('%d.%m.%Y %H:%M')}",
        "═══════════════════════════════════════════════════════",
        "",
        "GEÄNDERTE INPUTS:",
    ]

    for change in applied_changes:
        old = change['old']
        new = change['new']

        # Delta berechnen wenn numerisch
        delta_str = ""
        if isinstance(old, (int, float)) and isinstance(new, (int, float)) and old != 0:
            delta_pct = (new - old) / abs(old) * 100
            delta_str = f"  (Δ {delta_pct:+.1f}%)"

        lines.append(
            f"  [{change['sheet']}] {change['cell']} {change['label']}: "
            f"{old} → {new}{delta_str}"
        )

    lines += [
        "",
        "KEY METRICS (Baseline aus template_v0.4.xlsx, data_only=True):",
        "  HINWEIS: Werte sind nur verfügbar wenn Excel die Formeln berechnet hat.",
        "  Öffne die generierte Datei in Excel/Google Sheets für aktuelle Werte.",
        "",
        f"{'KPI':<25} {'M12':>12} {'M24':>12} {'M36':>12} {'M52':>12}",
        "-" * 75,
    ]

    for kpi_name, month_vals in baseline_kpis.items():
        def fmt(v):
            if v is None:
                return "n/a"
            if isinstance(v, float):
                if abs(v) >= 1_000_000:
                    return f"€{v/1_000_000:.1f}M"
                elif abs(v) >= 1_000:
                    return f"€{v/1_000:.0f}k"
                else:
                    return f"{v:.1f}"
            return str(v)

        m12 = fmt(month_vals.get("M12"))
        m24 = fmt(month_vals.get("M24"))
        m36 = fmt(month_vals.get("M36"))
        m52 = fmt(month_vals.get("M52"))
        lines.append(f"{kpi_name:<25} {m12:>12} {m24:>12} {m36:>12} {m52:>12}")

    lines += [
        "",
        f"GESPEICHERTE DATEI: {filepath}",
        "",
        "NÄCHSTE SCHRITTE:",
        "  1. Datei in Excel/Google Sheets öffnen → Strg+Shift+F9 (Neuberechnung)",
        "  2. Werte mit Baseline vergleichen",
        "  3. Mögliche Folge-Szenarien: HighChurn, NoAI, Conservative",
        "═══════════════════════════════════════════════════════",
    ]

    report_content = "\n".join(lines)

    with open(report_path, 'w', encoding='utf-8') as f:
        f.write(report_content)

    return report_path, report_content


def run_baseline():
    """Erstellt Baseline-Kopie ohne Änderungen und liest KPIs aus."""
    print("=== BASELINE-LAUF ===")
    print(f"Template: {TEMPLATE_PATH}")

    wb_formula = openpyxl.load_workbook(TEMPLATE_PATH, data_only=False)

    # Aktuelles Szenario lesen
    ws_sandbox = wb_formula['00_Input_Sandbox']
    current_szenario = ws_sandbox['B1'].value
    print(f"Aktives Szenario: {current_szenario}")

    # Aktuelle Inputs lesen
    current_inputs = read_current_inputs(wb_formula)
    print("\nKERN-INPUTS (Status quo):")
    key_inputs = ["B28", "B21", "B23", "B24", "B25", "B26", "B27", "B29", "B12", "B14"]
    for cell_ref in key_inputs:
        info = current_inputs.get(cell_ref, {})
        print(f"  {cell_ref} {info.get('label', '')}: {info.get('value', 'n/a')}")

    # Baseline speichern
    filepath = save_scenario(wb_formula, "Baseline")
    print(f"\nBaseline gespeichert: {filepath}")

    # KPIs aus data_only lesen
    print("\nLade berechnete Werte (data_only=True)...")
    wb_calc = openpyxl.load_workbook(TEMPLATE_PATH, data_only=True)
    baseline_kpis = read_baseline_kpis(wb_calc)

    print(f"\n{'KPI':<25} {'M12':>12} {'M24':>12} {'M36':>12} {'M52':>12}")
    print("-" * 75)

    def fmt(v):
        if v is None:
            return "None*"
        if isinstance(v, float):
            if abs(v) >= 1_000_000:
                return f"€{v/1_000_000:.1f}M"
            elif abs(v) >= 1_000:
                return f"€{v/1_000:.0f}k"
            return f"{v:.1f}"
        return str(v)

    for kpi_name, month_vals in baseline_kpis.items():
        m12 = fmt(month_vals.get("M12"))
        m24 = fmt(month_vals.get("M24"))
        m36 = fmt(month_vals.get("M36"))
        m52 = fmt(month_vals.get("M52"))
        print(f"{kpi_name:<25} {m12:>12} {m24:>12} {m36:>12} {m52:>12}")

    print("\n* None = Datei wurde noch nicht in Excel geöffnet/berechnet.")
    print("  → Öffne template_v0.4.xlsx in Excel, speichere, dann erneut ausführen.")

    # KPIs als JSON speichern
    baseline_file = os.path.join(REPORTS_DIR, "baseline_kpis.json")
    os.makedirs(REPORTS_DIR, exist_ok=True)
    with open(baseline_file, 'w', encoding='utf-8') as f:
        json.dump({
            "created": datetime.now().isoformat(),
            "szenario": current_szenario,
            "kpis": {k: {m: str(v) for m, v in vals.items()}
                     for k, vals in baseline_kpis.items()}
        }, f, indent=2, ensure_ascii=False)

    print(f"\nBaseline-KPIs gespeichert: {baseline_file}")
    return baseline_kpis


def parse_changes_string(changes_str):
    """Parst 'B28=0.12,B21=0.03' in ein Changes-Dict."""
    changes = {"Inputs": {}}
    for part in changes_str.split(','):
        part = part.strip()
        if '=' in part:
            cell, val = part.split('=', 1)
            cell = cell.strip().upper()
            val = val.strip()
            # Typ-Erkennung
            try:
                val = float(val)
                if val == int(val):
                    val = int(val)
            except ValueError:
                pass  # String bleibt String
            changes["Inputs"][cell] = val
    return changes


def main():
    parser = argparse.ArgumentParser(description='LFL Scenario Engine')
    parser.add_argument('--szenario', choices=['gering', 'normal', 'stark'],
                       help='Sandbox-Szenario wechseln')
    parser.add_argument('--changes', type=str,
                       help='Direkte Input-Änderungen, z.B. "B28=0.12,B21=0.03"')
    parser.add_argument('--name', type=str, default=None,
                       help='Name für Szenario-Datei')
    parser.add_argument('--baseline', action='store_true',
                       help='Baseline-Lauf ohne Änderungen')

    args = parser.parse_args()

    if args.baseline:
        run_baseline()
        return

    # Name bestimmen
    name_parts = []
    if args.szenario:
        name_parts.append(f"Szenario_{args.szenario.capitalize()}")
    if args.name:
        name_parts.append(args.name)
    if not name_parts:
        name_parts.append("Custom")
    scenario_name = "_".join(name_parts)

    # Änderungen sammeln
    changes = {}
    if args.changes:
        changes = parse_changes_string(args.changes)

    # Vorschau anzeigen
    print(f"=== SZENARIO: {scenario_name} ===")
    print("Geplante Änderungen:")

    if args.szenario:
        print(f"  [00_Input_Sandbox] B1 Aktives Szenario: → '{args.szenario}'")
        print(f"  → B20 (Startpreis) und B22 (Startmonat) ändern sich per VLOOKUP")
        print(f"  → KI-Strategie Spalte F wird neu berechnet")

    if changes.get("Inputs"):
        for cell_ref, new_val in changes["Inputs"].items():
            label = INPUT_CELL_LABELS.get(cell_ref, cell_ref)
            print(f"  [Inputs] {cell_ref} {label}: → {new_val}")

    print("\nFortfahren? [j/N] ", end='')
    confirm = input().strip().lower()
    if confirm not in ('j', 'ja', 'y', 'yes'):
        print("Abgebrochen.")
        return

    # Szenario anwenden
    wb, applied_changes = apply_scenario(changes, szenario=args.szenario)

    # Speichern
    filepath = save_scenario(wb, scenario_name)
    print(f"\n✓ Datei gespeichert: {filepath}")

    # Baseline-KPIs laden
    wb_calc = openpyxl.load_workbook(TEMPLATE_PATH, data_only=True)
    baseline_kpis = read_baseline_kpis(wb_calc)

    # Delta-Bericht
    report_path, report_content = generate_delta_report(
        scenario_name, applied_changes, baseline_kpis, filepath
    )
    print(f"✓ Bericht gespeichert: {report_path}")
    print("\n" + report_content)


if __name__ == '__main__':
    main()
