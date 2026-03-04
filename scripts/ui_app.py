"""
LFL Financial Projection — Streamlit UI
Startet mit: streamlit run scripts/ui_app.py
"""

import streamlit as st
import openpyxl
import json
import os
import sys
from datetime import datetime
from pathlib import Path

# Pfad-Setup: Projekt-Root ermitteln
ROOT = Path(__file__).parent.parent
sys.path.insert(0, str(ROOT / "scripts"))

from scenario_engine import (
    apply_scenario,
    save_scenario,
    generate_delta_report,
    read_current_inputs,
    read_baseline_kpis,
    save_as_custom_baseline,
    delete_custom_baseline,
    is_custom_baseline_active,
    get_active_template,
    INPUT_CELL_LABELS,
    SANDBOX_CELL_LABELS,
    TEMPLATE_PATH        as _TEMPLATE_PATH,
    CUSTOM_BASELINE_PATH as _CUSTOM_BASELINE_PATH,
)

TEMPLATE_PATH        = str(ROOT / _TEMPLATE_PATH)
CUSTOM_BASELINE_PATH = str(ROOT / _CUSTOM_BASELINE_PATH)
REPORTS_DIR          = str(ROOT / "reports")
SCENARIOS_DIR        = str(ROOT / "scenarios")

# ── Hilfsfunktionen ──────────────────────────────────────────────────────────

def _active_tpl_path():
    """Gibt den absoluten Pfad der aktiv genutzten Basis zurück."""
    if Path(CUSTOM_BASELINE_PATH).exists():
        return CUSTOM_BASELINE_PATH
    return TEMPLATE_PATH

@st.cache_data(show_spinner=False)
def load_current_inputs(_cache_key=0):
    wb = openpyxl.load_workbook(_active_tpl_path(), data_only=False)
    return read_current_inputs(wb)

@st.cache_data(show_spinner=False)
def load_baseline_kpis(_cache_key=0):
    wb = openpyxl.load_workbook(_active_tpl_path(), data_only=True)
    return read_baseline_kpis(wb)

def fmt_eur(v):
    if v is None:
        return "n/a"
    if isinstance(v, (int, float)):
        if abs(v) >= 1_000_000:
            return f"€ {v/1_000_000:.1f} M"
        if abs(v) >= 1_000:
            return f"€ {v/1_000:.0f} k"
        return f"€ {v:.0f}"
    return str(v)

def fmt_val(v, cell_ref=""):
    """Formatiert Werte je nach Typ/Zelle."""
    if v is None:
        return "—"
    if isinstance(v, float):
        # Prozentzellen
        pct_cells = {"B7","B21","B24","B28","B29","B44","B45","B85","B119","B123","B127","B134"}
        if cell_ref in pct_cells:
            return f"{v*100:.1f} %"
        if v > 100:
            return f"€ {v:,.0f}"
        return f"{v:.4f}".rstrip('0').rstrip('.')
    if isinstance(v, datetime):
        return v.strftime("%d.%m.%Y")
    return str(v)

def run_scenario(scenario_name, szenario, changes_inputs, changes_sandbox):
    changes = {}
    if changes_inputs:
        changes["Inputs"] = changes_inputs
    if changes_sandbox:
        changes["Sandbox"] = changes_sandbox

    wb, applied = apply_scenario(changes, szenario=szenario if szenario != "(unverändert)" else None)

    # Pfade relativ zu ROOT setzen
    orig_scenarios = os.getcwd()
    os.chdir(ROOT)
    filepath = save_scenario(wb, scenario_name)
    wb_calc = openpyxl.load_workbook(TEMPLATE_PATH, data_only=True)
    kpis = read_baseline_kpis(wb_calc)
    report_path, report_md = generate_delta_report(scenario_name, applied, kpis, filepath)
    os.chdir(orig_scenarios)

    return filepath, report_path, report_md, applied, kpis

# ── Seiten-Konfiguration ─────────────────────────────────────────────────────

st.set_page_config(
    page_title="LFL Financial Projections",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded",
)

st.markdown("""
<style>
.metric-card {
    background: #1e1e2e;
    border-radius: 10px;
    padding: 16px 20px;
    margin: 6px 0;
    border-left: 4px solid #7c6af7;
}
.metric-card h4 { color: #cdd6f4; margin: 0 0 4px 0; font-size: 0.82rem; }
.metric-card .val { color: #cba6f7; font-size: 1.25rem; font-weight: 700; }
.section-header {
    color: #89b4fa;
    font-size: 1.05rem;
    font-weight: 600;
    border-bottom: 1px solid #313244;
    padding-bottom: 4px;
    margin: 18px 0 10px 0;
}
</style>
""", unsafe_allow_html=True)

# ── Sidebar ───────────────────────────────────────────────────────────────────

with st.sidebar:
    st.image("https://via.placeholder.com/240x60/7c6af7/ffffff?text=LoopforgeLab", use_container_width=True)
    st.markdown("## LFL Financial Model")
    st.caption("v0.4 · April 2026 – M52")

    st.divider()
    st.markdown("**Navigation**")
    page = st.radio(
        "",
        ["Dashboard", "Szenario erstellen", "Variablen editor", "Assistent"],
        label_visibility="collapsed",
    )

    st.divider()
    st.markdown("**Anthropic API-Key**")
    api_key_input = st.text_input(
        "API-Key",
        value=st.session_state.get("anthropic_api_key", ""),
        type="password",
        placeholder="sk-ant-api03-...",
        label_visibility="collapsed",
        help="Nur für den Assistent-Tab nötig. Key wird nur im Browser gespeichert, nie übertragen.",
    )
    if api_key_input:
        st.session_state["anthropic_api_key"] = api_key_input
        os.environ["ANTHROPIC_API_KEY"] = api_key_input
        st.caption("Key gesetzt.")
    elif os.environ.get("ANTHROPIC_API_KEY"):
        st.caption("Key aus Umgebung geladen.")
    else:
        st.caption("Kein Key — Assistent inaktiv.")

    st.divider()
    st.markdown("**Aktive Basis**")
    if Path(CUSTOM_BASELINE_PATH).exists():
        st.success("Custom-Basis aktiv")
        st.caption("Inputs wurden überschrieben.")
        if st.button("Zurück zu Original-Basis", use_container_width=True):
            os.chdir(ROOT)
            delete_custom_baseline()
            st.session_state["basis_cache_key"] = st.session_state.get("basis_cache_key", 0) + 1
            st.cache_data.clear()
            st.rerun()
    else:
        st.info("Original-Basis (template_v0.4)")

    st.divider()
    if st.button("Cache leeren / Neu laden", use_container_width=True):
        st.cache_data.clear()
        st.rerun()

# ─────────────────────────────────────────────────────────────────────────────
# PAGE 1: DASHBOARD
# ─────────────────────────────────────────────────────────────────────────────

_ck = st.session_state.get("basis_cache_key", 0)

if page == "Dashboard":
    is_custom = Path(CUSTOM_BASELINE_PATH).exists()
    basis_label = "Custom-Basis (geänderte Inputs)" if is_custom else "template_v0.4.xlsx (Original)"
    st.title("Dashboard — LFL Financial Projections")
    st.caption(f"Basis: {basis_label} · Aktualisiert: {datetime.now().strftime('%d.%m.%Y %H:%M')}")
    if is_custom:
        st.info("Custom-Basis aktiv — Inputs wurden manuell geändert. Szenarien basieren auf diesen Werten."
                " Zurücksetzen über die Sidebar.")

    with st.spinner("Lade Modell-Daten..."):
        inputs   = load_current_inputs(_ck)
        kpis     = load_baseline_kpis(_ck)

    # ── Aktives Szenario ──────────────────────────────────────────────────────
    szenario_val = inputs.get("SANDBOX_B1", {}).get("value", "?")
    churn        = inputs.get("B28", {}).get("value", "?")
    price_inc    = inputs.get("B21", {}).get("value", "?")
    init_seats   = inputs.get("B23", {}).get("value", "?")
    ent_start    = inputs.get("B25", {}).get("value", "?")

    col1, col2, col3, col4, col5 = st.columns(5)
    col1.metric("Aktives Szenario", str(szenario_val).capitalize())
    col2.metric("Churn Rate/Jahr", fmt_val(churn, "B28"))
    col3.metric("Preiserhöhung/Jahr", fmt_val(price_inc, "B21"))
    col4.metric("Initiale Seats", str(init_seats))
    col5.metric("Enterprise ab Monat", str(ent_start))

    st.divider()

    # ── KPI-Tabelle ───────────────────────────────────────────────────────────
    st.markdown('<p class="section-header">Key Performance Indicators (Baseline)</p>', unsafe_allow_html=True)
    st.caption("Werte aus gecachten Excel-Formeln. Öffne template_v0.4.xlsx in Excel → Strg+Shift+F9 für aktuelle Werte.")

    months = ["M12", "M24", "M36", "M52"]
    kpi_rows = []
    for kpi_name, month_vals in kpis.items():
        row = {"KPI": kpi_name}
        for m in months:
            v = month_vals.get(m)
            if kpi_name in ("Total ARR", "Total MRR", "EBITDA", "Net Income", "Ending Cash", "Burn Rate"):
                row[m] = fmt_eur(v)
            elif kpi_name == "Total Headcount":
                row[m] = str(int(v)) if isinstance(v, (int, float)) else "—"
            elif kpi_name == "Runway":
                row[m] = f"{v:.1f} Mo" if isinstance(v, (int, float)) else "—"
            else:
                row[m] = fmt_eur(v)
        kpi_rows.append(row)

    if kpi_rows:
        import pandas as pd
        df = pd.DataFrame(kpi_rows).set_index("KPI")
        st.dataframe(df, use_container_width=True)
    else:
        st.info("Keine KPI-Daten gefunden. Template in Excel öffnen und speichern.")

    # ── Generierte Szenarien ──────────────────────────────────────────────────
    st.divider()
    st.markdown('<p class="section-header">Generierte Szenario-Dateien</p>', unsafe_allow_html=True)
    scen_dir = ROOT / "scenarios"
    files = sorted(scen_dir.glob("*.xlsx"), key=os.path.getmtime, reverse=True) if scen_dir.exists() else []
    if files:
        for f in files[:10]:
            mtime = datetime.fromtimestamp(f.stat().st_mtime).strftime("%d.%m.%Y %H:%M")
            st.markdown(f"`{f.name}` · {mtime}")
    else:
        st.info("Noch keine Szenarien generiert.")

# ─────────────────────────────────────────────────────────────────────────────
# PAGE 2: SZENARIO ERSTELLEN
# ─────────────────────────────────────────────────────────────────────────────

elif page == "Szenario erstellen":
    st.title("Szenario erstellen")

    with st.spinner("Lade aktuelle Werte..."):
        inputs = load_current_inputs()

    col_left, col_right = st.columns([1, 1], gap="large")

    with col_left:
        st.markdown('<p class="section-header">Basis-Szenario</p>', unsafe_allow_html=True)

        szenario_choice = st.selectbox(
            "Sandbox-Szenario (steuert Startpreis, Startmonat, KI-Hebel)",
            ["(unverändert)", "gering", "normal", "stark"],
            help="Wechsel zwischen den 3 vordefinierten Szenarien. Ändert B20, B22 und Spalte F im Einstellungsplan per VLOOKUP."
        )

        szenario_desc = {
            "(unverändert)": "",
            "gering":  "Automotive-Fokus · Startpreis €800 · Startmonat 14 · AI-Hebel: 0",
            "normal":  "Hybrid-Fokus · Startpreis €1.000 · Startmonat 8 · AI-Hebel: +6 Mo",
            "stark":   "Packaging-Fokus · Startpreis €1.200 · Startmonat 4 · AI-Hebel: 99 (KI ersetzt Stellen)",
        }
        if szenario_choice != "(unverändert)":
            st.caption(szenario_desc[szenario_choice])

        st.markdown('<p class="section-header">Schnell-Inputs</p>', unsafe_allow_html=True)
        st.caption("Die häufigsten Stellschrauben — alle Werte optional.")

        c1, c2 = st.columns(2)
        with c1:
            churn_new = st.number_input(
                "Churn Rate/Jahr (%)",
                min_value=0.0, max_value=100.0,
                value=float(inputs.get("B28", {}).get("value", 0.08) or 0.08) * 100,
                step=0.5, format="%.1f",
            )
            price_inc_new = st.number_input(
                "Preiserhöhung/Jahr (%)",
                min_value=0.0, max_value=50.0,
                value=float(inputs.get("B21", {}).get("value", 0.08) or 0.08) * 100,
                step=0.5, format="%.1f",
            )
            ent_start_new = st.number_input(
                "Enterprise-Deals ab Monat",
                min_value=1, max_value=52,
                value=int(inputs.get("B25", {}).get("value", 24) or 24),
                step=1,
            )
        with c2:
            nrr_new = st.number_input(
                "Net Revenue Retention (%)",
                min_value=50.0, max_value=200.0,
                value=float(inputs.get("B29", {}).get("value", 1.18) or 1.18) * 100,
                step=1.0, format="%.1f",
            )
            ent_arr_new = st.number_input(
                "Ø Enterprise ARR (€)",
                min_value=0, max_value=1_000_000,
                value=int(inputs.get("B26", {}).get("value", 150000) or 150000),
                step=10000,
            )
            seed_new = st.number_input(
                "Seed Betrag (€)",
                min_value=0, max_value=20_000_000,
                value=int(inputs.get("B14", {}).get("value", 6000000) or 6000000),
                step=500000,
            )

        st.markdown('<p class="section-header">Szenario-Name</p>', unsafe_allow_html=True)
        scenario_name = st.text_input(
            "Name der Szenario-Datei",
            value=f"Szenario_{szenario_choice.capitalize()}" if szenario_choice != "(unverändert)" else "Custom",
            placeholder="z.B. HighChurn_NoAI",
        )

    with col_right:
        st.markdown('<p class="section-header">Vorschau der Änderungen</p>', unsafe_allow_html=True)

        # Delta berechnen
        orig = {
            "B28": float(inputs.get("B28", {}).get("value", 0.08) or 0.08),
            "B21": float(inputs.get("B21", {}).get("value", 0.08) or 0.08),
            "B25": int(inputs.get("B25", {}).get("value", 24) or 24),
            "B29": float(inputs.get("B29", {}).get("value", 1.18) or 1.18),
            "B26": int(inputs.get("B26", {}).get("value", 150000) or 150000),
            "B14": int(inputs.get("B14", {}).get("value", 6000000) or 6000000),
        }
        new_vals = {
            "B28": churn_new / 100,
            "B21": price_inc_new / 100,
            "B25": ent_start_new,
            "B29": nrr_new / 100,
            "B26": ent_arr_new,
            "B14": seed_new,
        }

        changes_inputs = {}
        any_change = False

        if szenario_choice != "(unverändert)":
            st.markdown(f"**Sandbox B1:** `{inputs.get('SANDBOX_B1',{}).get('value','?')}` → `{szenario_choice}`")
            any_change = True

        for cell_ref, new_val in new_vals.items():
            old_val = orig[cell_ref]
            label   = INPUT_CELL_LABELS.get(cell_ref, cell_ref)
            if abs(float(new_val) - float(old_val)) > 1e-9:
                delta_pct = (new_val - old_val) / abs(old_val) * 100 if old_val else 0
                arrow = "▲" if new_val > old_val else "▼"
                st.markdown(f"**{cell_ref}** {label}  \n`{old_val}` → `{new_val}` ({arrow} {abs(delta_pct):.1f}%)")
                changes_inputs[cell_ref] = new_val
                any_change = True

        if not any_change:
            st.info("Keine Änderungen gegenüber dem Template.")

        st.divider()

        generate_btn = st.button(
            "Szenario generieren & speichern",
            type="primary",
            use_container_width=True,
            disabled=(not any_change),
        )

        if generate_btn:
            with st.spinner(f"Generiere '{scenario_name}'..."):
                try:
                    os.chdir(ROOT)
                    wb, applied = apply_scenario(
                        {"Inputs": changes_inputs} if changes_inputs else {},
                        szenario=szenario_choice if szenario_choice != "(unverändert)" else None,
                    )
                    filepath = save_scenario(wb, scenario_name)
                    wb_calc  = openpyxl.load_workbook(TEMPLATE_PATH, data_only=True)
                    kpis     = read_baseline_kpis(wb_calc)
                    _, report_path, report_md = generate_delta_report(
                        scenario_name, applied, kpis, filepath
                    ) if False else (None, None, None)
                    rp, rc   = None, None
                    rp, rc = generate_delta_report(scenario_name, applied, kpis, filepath)
                except Exception as e:
                    st.error(f"Fehler: {e}")
                    st.stop()

            st.success(f"Gespeichert: `{filepath}`")
            st.text_area("Delta-Bericht", rc, height=400)
            st.cache_data.clear()

# ─────────────────────────────────────────────────────────────────────────────
# PAGE 3: VARIABLEN-EDITOR
# ─────────────────────────────────────────────────────────────────────────────

elif page == "Variablen editor":
    st.title("Variablen-Editor")
    st.caption("Alle editierbaren Inputs des Modells. Änderungen werden erst beim Generieren eines Szenarios angewendet.")

    with st.spinner("Lade Inputs..."):
        inputs = load_current_inputs()

    tab_inputs, tab_sandbox, tab_new = st.tabs(["Inputs Sheet", "Sandbox / Szenariowerte", "Neue Variable hinzufügen"])

    # ── Tab: Inputs ───────────────────────────────────────────────────────────
    with tab_inputs:
        import pandas as pd

        # Gruppen
        groups = {
            "Finanzierung": ["B10","B11","B12","B13","B14","B15","B16","B17"],
            "Revenue": ["B21","B23","B24","B25","B26","B27","B28","B29"],
            "Gehälter": ["B32","B33","B34","B35","B36","B37","B38","B39","B40","B41","B42","B43","B44","B45"],
            "Technologie & Cloud": ["B82","B83","B84","B85","B86","B87","B88","B89"],
            "Hardware": ["B92","B93","B94","B95"],
            "Büro": ["B98","B99","B100","B101","B102","B103","B104"],
            "Professional Services": ["B107","B108","B109","B110","B111","B112"],
            "Versicherungen & Bank": ["B115","B116","B117","B118","B119"],
            "Marketing & Sales": ["B122","B123","B124","B125","B126","B127","B128"],
            "Sonstiges": ["B131","B132","B133","B134","B135"],
        }

        selected_group = st.selectbox("Gruppe filtern", ["Alle"] + list(groups.keys()))

        rows = []
        for cell_ref, label in INPUT_CELL_LABELS.items():
            if selected_group != "Alle":
                if cell_ref not in groups.get(selected_group, []):
                    continue
            info = inputs.get(cell_ref, {})
            val  = info.get("value")
            pending_val = st.session_state.get("pending_changes", {}).get(cell_ref)
            if isinstance(val, str) and val.startswith("[FORMEL"):
                rows.append({"Zelle": cell_ref, "Bezeichnung": label,
                              "Aktueller Wert": "FORMEL (VLOOKUP)", "Geändert auf": "—", "Editierbar": False})
            else:
                rows.append({"Zelle": cell_ref, "Bezeichnung": label,
                              "Aktueller Wert": val,
                              "Geändert auf": str(pending_val) if pending_val is not None else "—",
                              "Editierbar": True})

        df = pd.DataFrame(rows).drop(columns=["Editierbar"])
        st.dataframe(df, use_container_width=True, height=380)

        st.divider()
        st.markdown("**Wert ändern**")
        c1, c2, c3 = st.columns([1, 2, 1])
        editable_cells = [r["Zelle"] for r in rows if r["Editierbar"]]
        with c1:
            edit_cell = st.selectbox("Zelle", editable_cells, key="edit_cell_select")
        with c2:
            current_label = next((r["Bezeichnung"] for r in rows if r["Zelle"] == edit_cell), "")
            current_val   = inputs.get(edit_cell, {}).get("value", "")
            new_edit_val  = st.text_input(current_label,
                                          value=str(current_val) if current_val is not None else "",
                                          key="edit_val_input")
        with c3:
            st.markdown("<br>", unsafe_allow_html=True)
            if st.button("Hinzufügen", use_container_width=True, key="btn_add_change"):
                if "pending_changes" not in st.session_state:
                    st.session_state["pending_changes"] = {}
                try:
                    v = float(new_edit_val)
                    if v == int(v): v = int(v)
                except ValueError:
                    v = new_edit_val
                st.session_state["pending_changes"][edit_cell] = v
                st.rerun()

        # ── Ausstehende Änderungen ─────────────────────────────────────────
        pending = st.session_state.get("pending_changes", {})
        if pending:
            st.divider()
            st.markdown(f"**Ausstehende Änderungen ({len(pending)})**")
            for k, v in list(pending.items()):
                lbl = INPUT_CELL_LABELS.get(k, k)
                col_l, col_r = st.columns([7, 1])
                col_l.markdown(f"- `{k}` **{lbl}**: {inputs.get(k,{}).get('value','?')} → **{v}**")
                if col_r.button("✕", key=f"del_{k}"):
                    del st.session_state["pending_changes"][k]
                    st.rerun()

            st.divider()

            # ── Option A: Als neue Basis setzen ───────────────────────────
            st.markdown("### Option A — Als neue Basiswerte setzen")
            st.caption(
                "Die geänderten Inputs werden zur neuen Basis. "
                "Dashboard und alle Szenarien nutzen dann diese Werte."
            )
            if st.button("✅ Als neue Basis übernehmen", type="primary", key="btn_set_basis"):
                with st.spinner("Speichere Custom-Basis..."):
                    try:
                        os.chdir(ROOT)
                        save_as_custom_baseline(st.session_state["pending_changes"])
                        st.session_state["pending_changes"] = {}
                        st.session_state["basis_cache_key"] = st.session_state.get("basis_cache_key", 0) + 1
                        st.cache_data.clear()
                        st.success("Custom-Basis gespeichert. Dashboard und Szenarien nutzen jetzt diese Werte.")
                        st.rerun()
                    except Exception as e:
                        st.error(f"Fehler: {e}")

            st.divider()

            # ── Option B: Als benanntes Szenario exportieren ───────────────
            st.markdown("### Option B — Als benanntes Szenario exportieren (Excel)")
            st.caption("Erstellt eine Excel-Datei mit diesen Inputs, ohne die Basis dauerhaft zu ändern.")
            sc1, sc2 = st.columns([3, 1])
            with sc1:
                export_name = st.text_input("Szenario-Name", value="Mein_Szenario",
                                            key="export_name", label_visibility="collapsed")
            with sc2:
                export_btn = st.button("💾 Exportieren", use_container_width=True, key="btn_export_named")
            if export_btn:
                with st.spinner("Erstelle Excel..."):
                    try:
                        os.chdir(ROOT)
                        wb_e, applied_e = apply_scenario({"Inputs": pending}, szenario=None)
                        fp_e = save_scenario(wb_e, export_name.strip() or "Custom")
                        wb_k = openpyxl.load_workbook(_active_tpl_path(), data_only=True)
                        kpis_e = read_baseline_kpis(wb_k)
                        _, rpt_e = generate_delta_report(export_name.strip(), applied_e, kpis_e, fp_e)
                        st.success(f"Gespeichert: `{fp_e}`")
                        with open(fp_e, "rb") as fh:
                            st.download_button("⬇ Herunterladen", fh, Path(fp_e).name,
                                               mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                        with st.expander("Delta-Bericht"):
                            st.text(rpt_e)
                    except Exception as e:
                        st.error(f"Fehler: {e}")

            st.divider()
            if st.button("Alle Änderungen verwerfen", key="btn_discard"):
                st.session_state["pending_changes"] = {}
                st.rerun()

        # ── Szenarien aus aktiver Basis generieren ─────────────────────────
        st.divider()
        st.markdown("### Szenarien aus aktiver Basis exportieren")
        is_custom = Path(CUSTOM_BASELINE_PATH).exists()
        basis_info = "Custom-Basis" if is_custom else "Original-Basis (template_v0.4)"
        st.caption(f"Aktive Basis: **{basis_info}**. Wähle Szenarien zum Exportieren:")

        col_g, col_n, col_s = st.columns(3)
        do_gering = col_g.checkbox("Gering", value=True, key="chk_gering")
        do_normal = col_n.checkbox("Normal", value=True, key="chk_normal")
        do_stark  = col_s.checkbox("Stark",  value=True, key="chk_stark")

        if st.button("📊 Gewählte Szenarien erstellen", type="primary", key="btn_gen_szenarien"):
            selected = [s for s, do in [("gering", do_gering), ("normal", do_normal), ("stark", do_stark)] if do]
            if not selected:
                st.warning("Mindestens ein Szenario auswählen.")
            else:
                os.chdir(ROOT)
                results = {}
                for sz in selected:
                    with st.spinner(f"Erstelle Szenario '{sz}'..."):
                        try:
                            wb_sz, applied_sz = apply_scenario({}, szenario=sz)
                            fp_sz = save_scenario(wb_sz, f"Szenario_{sz.capitalize()}")
                            wb_kk = openpyxl.load_workbook(_active_tpl_path(), data_only=True)
                            kpis_sz = read_baseline_kpis(wb_kk)
                            _, rpt_sz = generate_delta_report(f"Szenario_{sz.capitalize()}", applied_sz, kpis_sz, fp_sz)
                            results[sz] = (fp_sz, rpt_sz)
                        except Exception as e:
                            st.error(f"Fehler bei '{sz}': {e}")

                if results:
                    st.success(f"{len(results)} Szenario(s) erstellt.")
                    for sz, (fp_sz, rpt_sz) in results.items():
                        with st.expander(f"📥 {sz.capitalize()} — {Path(fp_sz).name}"):
                            with open(fp_sz, "rb") as fh:
                                st.download_button(
                                    f"⬇ {sz.capitalize()} herunterladen",
                                    fh, Path(fp_sz).name,
                                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                    key=f"dl_{sz}",
                                )
                            st.text(rpt_sz)

    # ── Tab: Sandbox ──────────────────────────────────────────────────────────
    with tab_sandbox:
        wb_s = openpyxl.load_workbook(TEMPLATE_PATH, data_only=False)
        ws_s = wb_s['00_Input_Sandbox']

        rows_sb = []
        for row in ws_s.iter_rows(min_row=3, max_row=15, min_col=1, max_col=6, values_only=False):
            row_data = [c.value for c in row]
            if any(v is not None for v in row_data):
                rows_sb.append({
                    "Zeile": row[0].row,
                    "Kategorie": row_data[0],
                    "Variable": row_data[1],
                    "Einheit": row_data[2],
                    "Gering (D)": row_data[3],
                    "Normal (E)": row_data[4],
                    "Stark (F)": row_data[5],
                })

        if rows_sb:
            import pandas as pd
            df_sb = pd.DataFrame(rows_sb).set_index("Zeile")
            st.dataframe(df_sb, use_container_width=True)

        st.divider()
        st.markdown("**Sandbox-Wert ändern**")
        sandbox_cell = st.selectbox("Zelle (D=Gering, E=Normal, F=Stark)", list(SANDBOX_CELL_LABELS.keys()))
        sb_label  = SANDBOX_CELL_LABELS.get(sandbox_cell, sandbox_cell)
        sb_cur    = ws_s[sandbox_cell].value
        sb_new    = st.number_input(f"{sb_label} (aktuell: {sb_cur})", value=float(sb_cur) if isinstance(sb_cur, (int, float)) else 0.0)

        if st.button("Sandbox-Änderung zur Warteschlange hinzufügen"):
            if "pending_sandbox" not in st.session_state:
                st.session_state["pending_sandbox"] = {}
            st.session_state["pending_sandbox"][sandbox_cell] = sb_new
            st.success(f"Sandbox {sandbox_cell} = {sb_new} hinzugefügt.")

        if st.session_state.get("pending_sandbox"):
            st.markdown("**Ausstehende Sandbox-Änderungen:**")
            for k, v in st.session_state["pending_sandbox"].items():
                st.markdown(f"- `{k}` {SANDBOX_CELL_LABELS.get(k,k)}: **{v}**")

    # ── Tab: Neue Variable ────────────────────────────────────────────────────
    with tab_new:
        st.markdown("**Neue Zeile in der Sandbox hinzufügen**")
        st.caption("Fügt einen neuen Parameter in das Sandbox-Sheet ein (Zeilen 11–15 sind noch frei).")

        col_a, col_b = st.columns(2)
        with col_a:
            new_var_kategorie = st.text_input("Kategorie", placeholder="z.B. Kosten")
            new_var_name      = st.text_input("Variable", placeholder="z.B. Cloud-Kosten-Faktor")
            new_var_einheit   = st.text_input("Einheit", placeholder="z.B. Faktor")
        with col_b:
            new_var_gering = st.number_input("Wert: Gering", value=0.8, step=0.1, format="%.2f")
            new_var_normal = st.number_input("Wert: Normal", value=1.0, step=0.1, format="%.2f")
            new_var_stark  = st.number_input("Wert: Stark",  value=1.5, step=0.1, format="%.2f")

        if st.button("Variable in Sandbox schreiben", type="primary"):
            wb_new = openpyxl.load_workbook(TEMPLATE_PATH, data_only=False)
            ws_new = wb_new['00_Input_Sandbox']

            # Freie Zeile finden (ab Zeile 11)
            free_row = None
            for r in range(11, 20):
                if ws_new.cell(r, 1).value is None and ws_new.cell(r, 2).value is None:
                    free_row = r
                    break

            if free_row is None:
                st.error("Keine freie Zeile in der Sandbox gefunden (Zeilen 11-19 sind alle belegt).")
            else:
                ws_new.cell(free_row, 1).value = new_var_kategorie
                ws_new.cell(free_row, 2).value = new_var_name
                ws_new.cell(free_row, 3).value = new_var_einheit
                ws_new.cell(free_row, 4).value = new_var_gering
                ws_new.cell(free_row, 5).value = new_var_normal
                ws_new.cell(free_row, 6).value = new_var_stark
                wb_new.save(TEMPLATE_PATH)
                st.cache_data.clear()
                st.success(f"Variable '{new_var_name}' in Zeile {free_row} der Sandbox geschrieben.")
                st.warning("Template wurde direkt geändert. Bitte in Excel öffnen und prüfen.")
                st.rerun()

# ─────────────────────────────────────────────────────────────────────────────
# PAGE 4: ASSISTENT
# ─────────────────────────────────────────────────────────────────────────────

elif page == "Assistent":
    st.title("Finanzmodell-Assistent")
    st.caption("Frag den KI-Assistenten zu den Projektionen, Annahmen und Szenario-Auswirkungen.")

    try:
        from model_assistant import get_assistant_response, MODEL_CONTEXT
        assistant_available = True
    except ImportError:
        assistant_available = False

    if not assistant_available:
        st.warning("Assistent-Modul nicht geladen. Stelle sicher, dass `scripts/model_assistant.py` existiert.")
        st.stop()

    # API-Key prüfen
    if not os.environ.get("ANTHROPIC_API_KEY"):
        st.warning(
            "**Kein API-Key gesetzt.**\n\n"
            "Gib deinen Anthropic API-Key in der **Sidebar links** ein (Feld 'API-Key').\n\n"
            "Key besorgen: https://console.anthropic.com/settings/keys"
        )
        st.stop()

    # Chat-History
    if "messages" not in st.session_state:
        st.session_state["messages"] = []

    # Aktuelle Inputs als Kontext
    with st.spinner("Lade Modell-Kontext..."):
        inputs = load_current_inputs()
        kpis   = load_baseline_kpis()

    # Kontext-Zusammenfassung für Assistent
    context_summary = f"""
Aktives Szenario: {inputs.get('SANDBOX_B1',{}).get('value','?')}
Churn Rate: {inputs.get('B28',{}).get('value','?')}
Preiserhöhung/Jahr: {inputs.get('B21',{}).get('value','?')}
Enterprise-Start: Monat {inputs.get('B25',{}).get('value','?')}
Pre-Seed: €{inputs.get('B12',{}).get('value','?')}
Seed: €{inputs.get('B14',{}).get('value','?')}
KPI M24 Ending Cash: {kpis.get('Ending Cash',{}).get('M24','?')}
KPI M24 Burn Rate: {kpis.get('Burn Rate',{}).get('M24','?')}
KPI M24 Runway: {kpis.get('Runway',{}).get('M24','?')}
"""

    # Chat-Verlauf anzeigen
    for msg in st.session_state["messages"]:
        with st.chat_message(msg["role"]):
            st.markdown(msg["content"])

    # Schnell-Fragen
    with st.expander("Schnell-Fragen (klicken zum Einfügen)"):
        quick_qs = [
            "Was passiert mit dem Runway wenn wir die Churn Rate auf 15% erhöhen?",
            "Welche 3 Inputs haben den größten Hebel auf den Ending Cash in M36?",
            "Erkläre mir den Unterschied zwischen Szenario 'gering' und 'stark'.",
            "Ab wann wird das Modell EBITDA-positiv?",
            "Was ändert sich wenn wir Enterprise-Deals 6 Monate früher starten?",
            "Wie berechnet sich der AI-Personal-Hebel im Einstellungsplan?",
        ]
        cols = st.columns(2)
        for i, q in enumerate(quick_qs):
            if cols[i % 2].button(q, key=f"qq_{i}", use_container_width=True):
                st.session_state["quick_input"] = q
                st.rerun()

    # Chat-Input
    user_input = st.chat_input("Frag zum Finanzmodell...")

    # Schnell-Frage übernehmen
    if "quick_input" in st.session_state and st.session_state["quick_input"]:
        user_input = st.session_state.pop("quick_input")

    if user_input:
        st.session_state["messages"].append({"role": "user", "content": user_input})
        with st.chat_message("user"):
            st.markdown(user_input)

        with st.chat_message("assistant"):
            with st.spinner("Analysiere..."):
                response = get_assistant_response(
                    user_input,
                    st.session_state["messages"][:-1],
                    context_summary,
                )
            st.markdown(response)

        st.session_state["messages"].append({"role": "assistant", "content": response})

    if st.session_state.get("messages"):
        if st.button("Chat leeren"):
            st.session_state["messages"] = []
            st.rerun()
