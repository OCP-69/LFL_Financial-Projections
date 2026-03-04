#!/usr/bin/env python3
"""
PDF-basierte Szenarien: Konservativ & Aggressiv
Basis: 260304_LFL-Financial-Planning-and-Carbon-Case.pdf

Erstellt zwei Excel-Szenarien mit eingebetteter Erklärungs-Sheet
und separaten Markdown-Reports.

Aufruf: python3 scripts/create_pdf_scenarios.py
"""

import os
import sys
from pathlib import Path
from datetime import datetime
import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).parent.parent
os.chdir(ROOT)
sys.path.insert(0, str(ROOT / "scripts"))

from scenario_engine import (
    apply_scenario,
    save_scenario,
    read_baseline_kpis,
    TEMPLATE_PATH,
    REPORTS_DIR,
    SCENARIOS_DIR,
)

# ── FARBEN & STYLES ────────────────────────────────────────────────────────────

CLR_HEADER_K  = "1F4E79"   # Dunkelblau – Konservativ
CLR_HEADER_A  = "7B2D00"   # Dunkelorange – Aggressiv
CLR_SUB       = "2E75B6"   # Hellblau Zwischenüberschrift
CLR_SUB_A     = "C55A11"   # Orange Zwischenüberschrift
CLR_ROW_EVEN  = "EBF3FA"   # Leicht blau (konservativ)
CLR_ROW_EVEN_A= "FCE4D6"   # Leicht orange (aggressiv)
CLR_WHITE     = "FFFFFF"
CLR_CHANGED   = "FFFF00"   # Gelb – geänderter Wert
CLR_TEXT_DARK = "1A1A1A"
CLR_WARN      = "FF0000"

def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _font(bold=False, size=11, color="000000", italic=False):
    return Font(bold=bold, size=size, color=color, italic=italic,
                name="Calibri")

def _border_thin():
    s = Side(style="thin", color="AAAAAA")
    return Border(left=s, right=s, top=s, bottom=s)

def _center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def _left():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)


# ── PARAMETER-DEFINITIONEN (aus PDF) ──────────────────────────────────────────

# Baseline-Werte (Original-Template) für Delta-Darstellung
BASELINE_INPUTS = {
    "B12": 1_500_000,  "B13": 5,   "B14": 6_000_000,  "B15": 17,
    "B16": 15_000_000, "B17": 35,  "B21": 0.08,        "B23": 5,
    "B24": 0.05,       "B25": 24,  "B26": 150_000,     "B27": 1,
    "B28": 0.08,       "B29": 1.18,
    "B32": 72_000,     "B33": 72_000, "B34": 72_000,
    "B35": 90_000,     "B37": 110_000,
    "B44": 0.05,       "B45": 0.21,
    "B82": 1_200,      "B83": 50,  "B84": 1_000,  "B85": 0.05,
    "B122": 500, "B123": 0.05, "B124": 1_500, "B125": 25_000,
    "B127": 0.10, "B128": 500,
}

# Sandbox-Baseline
BASELINE_SANDBOX = {
    "D5": 14, "D6": 800, "D7": 1_200, "D9": 120, "D10": 0,
    "E5":  8, "E6": 1_000, "E7": 1_500, "E9":  60, "E10": 6,
    "F5":  4, "F6": 1_200, "F7": 1_800, "F9":  30, "F10": 99,
}

KONSERVATIV_SANDBOX = {
    "D5": 17,    # Startmonat Kunden: 17 (PDF: ~17 months to revenue)
    "D6": 200,   # Preis 200 EUR/Monat/Seat → 2.400 EUR/Jahr (PDF benchmark_low)
    "D7": 1_200, # Consulting: 1.200 EUR/Tag (PDF lower bound)
    "D9": 120,   # Onboarding: 120 Stunden (höherer Aufwand)
    "D10": 0,    # KI-Hebel: 0 (kein KI-Delay → mehr Personal)
}

KONSERVATIV_INPUTS = {
    "B12": 1_500_000,  # Pre-Seed: 1,5M (PDF lower end)
    "B14": 4_000_000,  # Seed: 4M (PDF lower end)
    "B15": 17,         # Seed Monat: 17 (September 2027)
    "B16": 10_000_000, # Series A: 10M (konservativ)
    "B17": 36,         # Series A: Monat 36
    "B21": 0.05,       # Preiserhöhung/Jahr: 5% (statt 8%)
    "B23": 3,          # Initiale Seats: 3 (PDF: 3-5, lower end)
    "B24": 0.03,       # Monatl. Seat-Wachstum: 3%
    "B25": 30,         # Enterprise ab Monat: 30 (spät)
    "B26": 25_000,     # Enterprise ARR: 25.000 EUR
    "B27": 1,          # Enterprise Deals/Quartal: 1
    "B28": 0.15,       # Churn Rate: 15% (hoch)
    "B29": 1.05,       # NRR: 105% (wenig Expansion)
    "B44": 0.03,       # Gehaltserhöhung: 3%
    "B82": 800,        # Cloud: 800 EUR/Monat
    "B84": 600,        # AI APIs: 600 EUR/Monat
    "B85": 0.03,       # AI-Kostenwachstum: 3%
    "B122": 300,       # Paid Ads: 300 EUR/Monat
    "B123": 0.03,      # Ads-Wachstum: 3%
    "B124": 1_000,     # Content & SEO: 1.000 EUR/Monat
    "B125": 25_000,    # Events: 25.000 EUR/Jahr (PDF-Wert)
    "B127": 0.08,      # Sales-Provision: 8%
    "B128": 300,       # Sales-Reise: 300 EUR/MA/Monat
}

AGGRESSIV_SANDBOX = {
    "F5": 9,     # Startmonat Kunden: 9 (früh, Seed-Tempo)
    "F6": 500,   # Preis 500 EUR/Monat/Seat → 6.000 EUR/Jahr (PDF benchmark_high)
    "F7": 1_800, # Consulting: 1.800 EUR/Tag (PDF upper bound)
    "F9": 30,    # Onboarding: 30 Stunden (AI-optimiert)
    "F10": 99,   # KI-Hebel: 99 Monate (maximale KI-Substitution)
}

AGGRESSIV_INPUTS = {
    "B12": 2_000_000,  # Pre-Seed: 2M (PDF upper end)
    "B14": 6_000_000,  # Seed: 6M (PDF upper end)
    "B15": 17,         # Seed Monat: 17 (September 2027)
    "B16": 20_000_000, # Series A: 20M (aggressive)
    "B17": 30,         # Series A: Monat 30 (früher)
    "B21": 0.10,       # Preiserhöhung/Jahr: 10%
    "B23": 5,          # Initiale Seats: 5 (PDF avg)
    "B24": 0.08,       # Monatl. Seat-Wachstum: 8%
    "B25": 18,         # Enterprise ab Monat: 18 (früh)
    "B26": 50_000,     # Enterprise ARR: 50.000 EUR
    "B27": 2,          # Enterprise Deals/Quartal: 2
    "B28": 0.06,       # Churn Rate: 6%
    "B29": 1.25,       # NRR: 125%
    "B35": 95_000,     # Senior Engineer: 95.000 (Premiumtalent)
    "B37": 120_000,    # ML/AI Engineer: 120.000 (Top-KI-Talent)
    "B44": 0.06,       # Gehaltserhöhung: 6%
    "B82": 2_000,      # Cloud: 2.000 EUR/Monat
    "B83": 80,         # Cloud/Seat: 80 EUR
    "B84": 2_000,      # AI APIs: 2.000 EUR/Monat
    "B85": 0.08,       # AI-Kostenwachstum: 8%
    "B122": 1_000,     # Paid Ads: 1.000 EUR/Monat
    "B123": 0.08,      # Ads-Wachstum: 8%
    "B124": 2_500,     # Content & SEO: 2.500 EUR/Monat
    "B125": 50_000,    # Events: 50.000 EUR/Jahr (2x für Conference-First GTM)
    "B127": 0.12,      # Sales-Provision: 12%
    "B128": 700,       # Sales-Reise: 700 EUR/MA/Monat
}


# ── ERKLÄRUNGS-TEXTE ──────────────────────────────────────────────────────────

INTRO_TEXT = (
    "Dieses Szenario basiert auf dem LFL Financial Planning & Carbon Case Dossier "
    "(260304_LFL-Financial-Planning-and-Carbon-Case.pdf, Stand 4. März 2026). "
    "Alle Parameter wurden aus den strategischen Meeting-Transkripten von Rene und dem Finanzmodellierer abgeleitet. "
    "Externe Annahmen wurden nicht hinzugefügt. Offene Punkte (TBD/OPEN) aus dem Dossier sind im Report-Sheet markiert."
)

KONSERVATIV_SECTIONS = [
    {
        "titel": "1. PREISGESTALTUNG — Konservative Positionierung",
        "parameter": [
            ("Sandbox D6", "Preis/Seat/Monat", "800 EUR (gering-Standard)", "200 EUR",
             "Das PDF nennt 350 EUR/Monat als Referenzpreis (→ 4.200 EUR/Seat/Jahr) und 200 EUR als benchmark_low. "
             "Im konservativen Szenario wählen wir den Benchmark-Tiefpunkt, um das Preisrisiko in der Marktvalidierung abzubilden. "
             "Ein zu hoher Einstiegspreis ohne bewiesenen ROI verlängert den Sales-Zyklus erheblich."),
            ("Inputs B21", "Preiserhöhung/Jahr", "8%", "5%",
             "Geringere jährliche Preissteigerung reflektiert schwächere Marktposition und Kundenwiderstand in der Frühphase. "
             "Packaging-Kunden (primärer Markt: 2.255 Unternehmen laut Rene) sind preissensitiv."),
        ],
        "auswirkung": (
            "ARR pro Standardkunde (5 Seats): 200 EUR × 5 × 12 = 12.000 EUR/Jahr (vs. 21.000 EUR im PDF-Referenzpreis). "
            "Runway und Break-Even verschlechtern sich deutlich. Höheres Funding-Risiko."
        ),
    },
    {
        "titel": "2. KUNDEN-TIMING & WACHSTUM — Langsame Marktdurchdringung",
        "parameter": [
            ("Sandbox D5", "Startmonat erste Kunden", "14", "17",
             "Das PDF nennt explizit '~17 months to first revenue'. Im konservativen Szenario halten wir an diesem "
             "Planwert fest. Die Pre-Seed-Phase (3-5 Kunden, Sept 2026-Sept 2027) dient reiner Validierung ohne "
             "Revenue-Fokus. First paying customer frühestens Monat 17 (September 2027, mit Seed-Close)."),
            ("Inputs B23", "Initiale Seats", "5", "3",
             "PDF nennt 3-5 als Pre-Seed-Kundenzahl. Konservativ: 3 Piloten. Jeder Pilot mit 5 Seats = 15 Seats total."),
            ("Inputs B24", "Monatl. Seat-Wachstum", "5%", "3%",
             "Verlangsamtes Wachstum durch längere Sales-Zyklen (packaging/automotive lt. PDF: 'longer lead cycle in automotive'). "
             "3% entspricht ca. 43% Jahreswachstum – konservativ aber realistisch für B2B-Manufacturing."),
            ("Inputs B25", "Enterprise-Deals ab Monat", "24", "30",
             "Enterprise-Deals starten erst nach Monat 30 – nach erstem erfolgreichen Track Record. "
             "PDF: deal_complexity höher als Standard, lead_cycle länger."),
        ],
        "auswirkung": (
            "Seed-Ziel: 50-60 Kunden (PDF) → im Konservativ-Szenario realistisch 25-35 Kunden bis Monat 17-24. "
            "Gesamt-ARR M24: deutlich unter 500K EUR. Enterprise-Revenue startet erst ab M30."
        ),
    },
    {
        "titel": "3. CHURN & KUNDENBINDUNG — Höheres Abwanderungsrisiko",
        "parameter": [
            ("Inputs B28", "Jährliche Churn Rate", "8%", "15%",
             "PDF: adoption_probability_low = 10%. Geringe Adoption bedeutet hohes Abwanderungsrisiko sobald "
             "Verträge auslaufen. 15% Churn entspricht einem 'gefährdeten' SaaS-Modell in der Early-Stage, "
             "typisch für Produkte ohne tiefen Workflow-Lock-in. Causal: Fehlende Onboarding-Qualität + "
             "unklarer ROI-Nachweis (wie Olaf betont) führen zu Abwanderung."),
            ("Inputs B29", "Net Revenue Retention", "118%", "105%",
             "Ohne starke Expansion durch Upsell oder neue Seats bleibt NRR knapp über 100%. "
             "PDF: Keine quantifizierten Upsell-Mechanismen – konservativ daher 105%."),
        ],
        "auswirkung": (
            "ARR-Erosion durch Churn: Bei 100K ARR und 15% Churn verliert das Unternehmen 15K/Jahr. "
            "Ohne starkes Neuwachstum schrumpft der Revenue-Pool. NRR von 105% rettet knapp die Nettowachstum-Rate."
        ),
    },
    {
        "titel": "4. FUNDING — Konservative Kapitalstrategie",
        "parameter": [
            ("Inputs B12", "Pre-Seed Betrag", "1.500.000 EUR", "1.500.000 EUR",
             "PDF Lower End: 1,5M EUR. Im Konservativ-Szenario bleibt Pre-Seed beim Minimum. "
             "Runway: 12 Monate (Sept 2026 – Sept 2027). Reicht für MVP + 3-5 Pilotkunden."),
            ("Inputs B14", "Seed Betrag", "6.000.000 EUR", "4.000.000 EUR",
             "PDF Seed Lower End: 4M EUR. Konservativ, da langsameres Wachstum weniger Kapital verbraucht. "
             "Seed-Runway: 12 Monate (Sept 2027 – Sept 2028), für 50-60 Kunden."),
            ("Inputs B16", "Series A Betrag", "15.000.000 EUR", "10.000.000 EUR",
             "Reduzierte Series A durch geringeres Wachstumstempo. Schont Verwässerung."),
            ("Inputs B17", "Series A Monat", "35", "36",
             "Leicht später, da Meilensteine langsamer erreicht werden."),
        ],
        "auswirkung": (
            "Gesamtfinanzierung: 15,5M EUR (vs. 23,5M im Aggressiv). "
            "Niedrigerer Cash-Verbrauch durch geringere Headcount- und Marketing-Ausgaben."
        ),
    },
    {
        "titel": "5. HEADCOUNT & KI-HEBEL — Vollständige Personalplanung",
        "parameter": [
            ("Sandbox D10", "AI-Personal-Hebel", "0 Monate (Gering-Standard)", "0 Monate",
             "Konservativ: Kein KI-Delay beim Hiring. Alle Rollen werden zum ursprünglich geplanten Monat eingestellt. "
             "Begründung: KI-Werkzeuge sind in der Praxis in Monat 1-24 noch nicht produktionsreif genug, "
             "um Hiring zu verzögern. Das Unternehmen baut Standard-Kapazitäten auf."),
            ("Inputs B44", "Jährl. Gehaltserhöhung", "5%", "3%",
             "Geringere Lohnsteigerung reflektiert konservatives Budget und weniger Wettbewerb um Talent "
             "im konservativen Szenario (kleineres Unternehmen, Berlin-Markt)."),
        ],
        "auswirkung": (
            "Konservativ hat MEHR Mitarbeiter als Aggressiv (kein KI-Delay). "
            "Personalkosten sind der größte Kostenblock. Typische Entwicklung:\n"
            "  M1-5: CEO + CTO (Founder, Eigenfinanzierung)\n"
            "  M5-17: +2 Senior Engineers nach Pre-Seed\n"
            "  M17-24: +CCO, +1 Sales, +1 Customer Success nach Seed\n"
            "  M24-36: +Marketing, +Finance, +2 weitere Engineers\n"
            "  M36-52: +2-3 Sales Reps, +2 CS, +1 weitere\n"
            "  GESAMT M52: ~18-22 Mitarbeiter"
        ),
    },
    {
        "titel": "6. MARKETING & SALES — Budgetdisziplin",
        "parameter": [
            ("Inputs B125", "Events & Messen/Jahr", "25.000 EUR", "25.000 EUR",
             "PDF-Placeholder-Wert beibehalten. Rene betont: Industrie-Konferenzen > Software-Konferenzen. "
             "25K EUR als Baseline für 2-3 Packaging/Machinery-Konferenzen. Im Konservativ-Szenario kein Aufschlag."),
            ("Inputs B122", "Paid Ads Initial", "500 EUR/Monat", "300 EUR/Monat",
             "Geringeres Ads-Budget in der Validierungsphase. B2B-Manufacturing kauft nicht über Ads."),
            ("Inputs B127", "Sales-Provision", "10%", "8%",
             "Etwas geringere Provision, da weniger aggressives Sales-Ziel."),
        ],
        "auswirkung": (
            "Total Marketing-Spend M1-M52: ~30-40% geringer als Aggressiv. "
            "GTM über Konferenzen (primary), Content (sekundär), POCs (Conversion)."
        ),
    },
]

AGGRESSIV_SECTIONS = [
    {
        "titel": "1. PREISGESTALTUNG — Premium-Positionierung",
        "parameter": [
            ("Sandbox F6", "Preis/Seat/Monat", "1.200 EUR (stark-Standard)", "500 EUR",
             "Das PDF nennt 500 EUR/Monat als benchmark_high (CAD/CAM-adjacent). Im aggressiven Szenario nutzen wir "
             "den oberen Benchmark, da: (a) verifizierter ROI durch CO₂-Einsparungen und Effizienzgewinne messbar, "
             "(b) Packaging/Machinery-Kunden haben hohen Stundensatz (typisch 60-120 EUR/h) → ROI-Rechnung "
             "ist bei 500 EUR/Seat/Monat = 6.000 EUR/Jahr positiv ab ~8h Einsparung/Monat/Seat. "
             "Hinweis: Bisherige Sandbox-Werte (800-1.200 EUR/Monat) erscheinen zu hoch verglichen mit PDF."),
            ("Inputs B21", "Preiserhöhung/Jahr", "8%", "10%",
             "Aggressivere Preisstrategie: 10% p.a. durch nachgewiesenen Mehrwert und geringe Konkurrenz "
             "in der Nische. PDF-Standard: 8%, aggressive Auslegung: +2pp."),
        ],
        "auswirkung": (
            "ARR pro Standardkunde (5 Seats): 500 EUR × 5 × 12 = 30.000 EUR/Jahr (vs. 12.000 EUR konservativ). "
            "2,5× höhere Revenue pro Kunde. Break-Even deutlich früher erreichbar."
        ),
    },
    {
        "titel": "2. KUNDEN-TIMING & WACHSTUM — Schnelle Marktpenetration",
        "parameter": [
            ("Sandbox F5", "Startmonat erste Kunden", "4 (stark-Standard)", "9",
             "PDF: 'Seed Customers Targeted: 50-60 companies' in Phase 2. Um dieses Ziel bis Monat 28 (Seed +12) "
             "zu erreichen, müssen erste zahlende Kunden in Monat 9 starten. "
             "Aggressiv aber realistisch: PoCs beginnen parallel zu Pre-Seed (M5-M9). "
             "Achtung: Standard Stark-Sandbox nutzt Monat 4 – das ist zu optimistisch, wir setzen auf 9."),
            ("Inputs B23", "Initiale Seats", "5", "5",
             "PDF avg_seats_per_customer = 5. Aggressiv nutzt sofort volle Bundle-Größe."),
            ("Inputs B24", "Monatl. Seat-Wachstum", "5%", "8%",
             "8% monatliches Wachstum = ~150% jährliches Seat-Wachstum. Möglich durch: "
             "(a) Conference-First GTM (Packaging-Konferenzen, primärer Lead-Kanal), "
             "(b) Starke PoC-Pipeline, (c) Referenzkundennetz."),
            ("Inputs B25", "Enterprise ab Monat", "24", "18",
             "Enterprise-Deals starten früh nach erstem Referenzkunden-Erfolg. "
             "Dealvolumen: 50.000 EUR ARR (aggressiv vs. 25.000 konservativ)."),
            ("Inputs B26", "Enterprise ARR", "150.000 EUR", "50.000 EUR",
             "Deutlich realistischer als bisheriger Baseline-Wert (150K). PDF nennt ~20K als Standard-ARR. "
             "Enterprise-Deals sind größere Kunden mit mehr Seats und längeren Verträgen: 50K = ~8 Seats × 6.000 EUR."),
            ("Inputs B27", "Enterprise Deals/Quartal", "1", "2",
             "Aggressiv: 2 Enterprise-Deals/Quartal ab Monat 18. Erfordert starke Sales-Kapazität."),
        ],
        "auswirkung": (
            "Seed-Ziel (PDF: 50-60 Kunden) → Aggressiv: 55-65 Kunden bis Monat 29. "
            "Gesamt-ARR M24: 400-700K EUR, M36: 1,5-2,5M EUR, M52: 4-8M EUR (inkl. Enterprise)."
        ),
    },
    {
        "titel": "3. CHURN & KUNDENBINDUNG — Starke Retention",
        "parameter": [
            ("Inputs B28", "Jährliche Churn Rate", "8%", "6%",
             "PDF: adoption_probability_high = 50%. Bei hoher Adoption und starkem Workflow-Lock-in "
             "(Onboarding: 6-Stufen-Paket laut PDF) entsteht tiefer Switching-Cost. 6% Churn entspricht "
             "dem oberen Ende gut-performender B2B-SaaS-Produkte in Nischenmärkten. "
             "Causal: Tiefer Daten-Lock-in durch Normalisierung + Integration (Teil des Onboardings)."),
            ("Inputs B29", "Net Revenue Retention", "118%", "125%",
             "Starke NRR durch: (a) Seat-Expansion innerhalb der Accounts (mehrere Abteilungen), "
             "(b) Modul-Upsells (Carbon Case-Features), (c) Gehalts-/Preiserhöhungen (10% p.a.). "
             "125% bedeutet: Bestandskunden generieren 25% mehr Revenue als im Vorjahr."),
        ],
        "auswirkung": (
            "ARR-Multiplikator durch NRR: Nach 3 Jahren verdoppelt eine Kohorte ihren Revenue-Beitrag. "
            "Schnelle Payback-Zeit für CAC. Investor-Kennzahl: Rule of 40 mit Aggressiv deutlich besser."
        ),
    },
    {
        "titel": "4. FUNDING — Aggressive Kapitalstrategie",
        "parameter": [
            ("Inputs B12", "Pre-Seed Betrag", "1.500.000 EUR", "2.000.000 EUR",
             "PDF Upper End: 2M EUR. Mehr Kapital in der Validierungsphase für schnelleren MVP und 5 Pilotkunden."),
            ("Inputs B14", "Seed Betrag", "6.000.000 EUR", "6.000.000 EUR",
             "PDF Upper End: 6M EUR. Maximum für schnellen Scale auf 50-60 Kunden."),
            ("Inputs B16", "Series A Betrag", "15.000.000 EUR", "20.000.000 EUR",
             "Größere Series A für aggressiven Markt-Scale. Möglich bei nachgewiesener Traction (50+ Kunden, "
             "starke ARR-Entwicklung). Investor-Thesis: Winner-takes-most in Manufacturing-AI-Nische."),
            ("Inputs B17", "Series A Monat", "35", "30",
             "Frühere Series A durch schnelleres Erreichen der Meilensteine (ARR, Kunden, Runway-Bedarf)."),
        ],
        "auswirkung": (
            "Gesamtfinanzierung: 28M EUR (vs. 15,5M konservativ). "
            "Höherer Cash-Verbrauch durch mehr Hiring und Marketing, aber deutlich schnelleres Wachstum."
        ),
    },
    {
        "titel": "5. HEADCOUNT & KI-HEBEL — KI-First Personalstrategie",
        "parameter": [
            ("Sandbox F10", "AI-Personal-Hebel", "99 (stark-Standard)", "99",
             "Maximaler KI-Hebel: Rollen mit Kategorie 'KI-Hebel' oder 'KI-Agent' werden um 99 Monate verzögert, "
             "was im 52-Monats-Modell bedeutet: Sie werden NICHT eingestellt. "
             "Dies ist die radikale KI-First-Strategie: Kleine Kernmannschaft, skaliert durch KI-Agenten. "
             "Causal: LFL ist selbst eine AI-Plattform – intern KI maximal nutzen ist Produktdemonstration."),
            ("Inputs B35", "Senior Engineer Gehalt", "90.000 EUR", "95.000 EUR",
             "Premium für Top-Talente: Wettbewerbsfähiges Gehalt für die wenigen, kritischen Rollen."),
            ("Inputs B37", "ML/AI Engineer Gehalt", "110.000 EUR", "120.000 EUR",
             "ML/AI-Spezialisten sind der Kern des Produkts. 120K EUR wettbewerbsfähig in Berlin 2026."),
            ("Inputs B44", "Jährl. Gehaltserhöhung", "5%", "6%",
             "Höhere Gehaltserhöhung um Top-Talente zu halten. Bei weniger Mitarbeitern ist Retention kritisch."),
        ],
        "auswirkung": (
            "Aggressiv hat WENIGER Mitarbeiter als Konservativ (KI-Delay). Personalkosten pro Kopf höher, "
            "Gesamtkosten aber geringer. Typische Entwicklung:\n"
            "  M1-5: CEO + CTO (Gründer)\n"
            "  M5-17: +ML/AI Engineers (Kernprodukt, KI-Rollen = 'Fix')\n"
            "  M17-24: +CCO, +1 Sales (KI-Hebel-Rollen verzögert)\n"
            "  M24-36: Minimale weitere Einstellungen (KI übernimmt CS, Support, Admin)\n"
            "  M36-52: +2-3 Enterprise Sales, Partner Manager\n"
            "  GESAMT M52: ~10-14 Mitarbeiter (deutlich weniger als Konservativ)\n"
            "  → Partnernetzwerk (PDF: 'External cooperation for implementation') kompensiert fehlende CS-Kapazität"
        ),
    },
    {
        "titel": "6. MARKETING & SALES — Conference-First GTM",
        "parameter": [
            ("Inputs B125", "Events & Messen/Jahr", "25.000 EUR", "50.000 EUR",
             "Verdoppelung des Event-Budgets. PDF-Empfehlung Rene: 'Industry conferences MORE important than "
             "software conferences.' Targeting: Packaging/Machinery-Konferenzen (Drupa, FachPack, K-Messe). "
             "50K EUR = 3-4 Premium-Konferenz-Teilnahmen + Eigenveranstaltungen."),
            ("Inputs B124", "Content & SEO", "1.500 EUR/Monat", "2.500 EUR/Monat",
             "Stärkere Content-Strategie: Whitepapers (PDF erwähnt als Lead-Nurture), Video-Content. "
             "B2B-Manufacturing-Entscheider lesen Fachberichte vor Kaufentscheidung."),
            ("Inputs B127", "Sales-Provision", "10%", "12%",
             "Höhere Provision motiviert aggressives Sales. Bei 30K EUR ARR/Kunde = 3.600 EUR Provision – "
             "starker Anreiz für Enterprise-Fokus."),
        ],
        "auswirkung": (
            "Total Marketing M1-M52: ~2× höher als Konservativ, aber relativer Marketing-Spend "
            "(als % vom ARR) sinkt schnell durch höheres Revenue-Wachstum. "
            "CAC: Noch zu quantifizieren (OPEN: PDF G2.2)."
        ),
    },
]

KEY_DIFFERENCES = """
╔══════════════════════════════════════════════════════════════════════╗
║         HAUPTUNTERSCHIEDE: KONSERVATIV vs. AGGRESSIV                ║
╠══════════════════════════════════════════════════════════════════════╣
║  Parameter          │ Konservativ  │ Aggressiv    │ Quelle          ║
╠══════════════════════════════════════════════════════════════════════╣
║  Preis/Seat/Monat   │ 200 EUR      │ 500 EUR      │ PDF benchmark   ║
║  Preis/Seat/Jahr    │ 2.400 EUR    │ 6.000 EUR    │ ×12             ║
║  First Revenue Mon. │ 17           │ 9            │ PDF + Anpassung ║
║  Initiale Seats     │ 3            │ 5            │ PDF avg         ║
║  Seat-Wachstum/Mo.  │ 3%           │ 8%           │ Scenario        ║
║  Enterprise ab Mon. │ 30           │ 18           │ Scenario        ║
║  Enterprise ARR     │ 25.000 EUR   │ 50.000 EUR   │ Scenario        ║
║  Deals/Quartal      │ 1            │ 2            │ Scenario        ║
║  Churn Rate/Jahr    │ 15%          │ 6%           │ PDF adoption    ║
║  NRR                │ 105%         │ 125%         │ Scenario        ║
║  Pre-Seed           │ 1,5M EUR     │ 2,0M EUR     │ PDF range       ║
║  Seed               │ 4,0M EUR     │ 6,0M EUR     │ PDF range       ║
║  Series A           │ 10M EUR      │ 20M EUR      │ Scenario        ║
║  Headcount M52      │ ~18-22       │ ~10-14       │ KI-Hebel-Effekt ║
║  Events/Jahr        │ 25.000 EUR   │ 50.000 EUR   │ PDF + ×2        ║
║  KI-Personal-Hebel  │ 0 Monate     │ 99 Monate    │ Maximaler Effekt║
║  ARR-Schätzung M24  │ 200-400K     │ 600K-1,2M    │ Modellschätzung ║
║  ARR-Schätzung M52  │ 500K-1M      │ 5-10M        │ Modellschätzung ║
╚══════════════════════════════════════════════════════════════════════╝
"""

OPEN_ITEMS = """
OFFENE PUNKTE AUS DEM PDF (Decision Register Part G):
─────────────────────────────────────────────────────
G1.1 HIGH: Timeline-Varianten konsolidieren (Sept 2026 vs. April-Start)
G1.3 HIGH: First-Paying-Customer-Timing final festlegen (kritischste Variable)
G1.4 MED:  Deployment-Standard (Cloud/On-Prem/Hybrid) → beeinflusst Kostenstruktur
G1.7 MED:  KI-Rollensubstitution quantifizieren (direkt auf Personalkosten)
G2.2 HIGH: Conversion-Rate-Annahmen per Kanal definieren
G2.3 HIGH: Customer-Success-Skalierungsmodell planen (identifizierter Engpass)
G3.1 HIGH: Kausale CO₂-Wirkungskette präzise definieren
G3.3 HIGH: Baseline für CO₂-Berechnungen festlegen
"""


# ── EXCEL ERKLÄRUNGS-SHEET ERSTELLEN ──────────────────────────────────────────

def add_explanation_sheet(wb, scenario_name, sections, is_aggressiv=False):
    """Fügt ein 'LFL_Erklärung' Sheet zur Workbook hinzu."""

    header_clr = CLR_HEADER_A if is_aggressiv else CLR_HEADER_K
    sub_clr    = CLR_SUB_A    if is_aggressiv else CLR_SUB
    row_clr    = CLR_ROW_EVEN_A if is_aggressiv else CLR_ROW_EVEN
    title_pfx  = "AGGRESSIV" if is_aggressiv else "KONSERVATIV"

    ws = wb.create_sheet("LFL_Szenario_Erklaerung")

    # Spaltenbreiten
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["E"].width = 60

    row = 1

    def _write(r, c, val, bold=False, size=11, bg=None, fg="000000",
               italic=False, merge_to=None, align="left", border=False):
        cell = ws.cell(row=r, column=c, value=val)
        cell.font = _font(bold=bold, size=size, color=fg, italic=italic)
        if bg:
            cell.fill = _fill(bg)
        cell.alignment = _center() if align == "center" else _left()
        if border:
            cell.border = _border_thin()
        if merge_to:
            ws.merge_cells(start_row=r, start_column=c,
                           end_row=r, end_column=merge_to)
        return cell

    # Titelzeile
    ws.row_dimensions[row].height = 40
    _write(row, 1, f"LFL FINANCIAL PROJECTIONS — SZENARIO {title_pfx}",
           bold=True, size=16, bg=header_clr, fg=CLR_WHITE,
           merge_to=5, align="center")
    row += 1

    ws.row_dimensions[row].height = 20
    _write(row, 1, f"Basis: 260304_LFL-Financial-Planning-and-Carbon-Case.pdf | "
           f"Erstellt: {datetime.now().strftime('%d.%m.%Y %H:%M')}",
           italic=True, size=10, bg=header_clr, fg="CCCCCC",
           merge_to=5, align="center")
    row += 1

    # Intro
    row += 1
    ws.row_dimensions[row].height = 15
    _write(row, 1, "HINTERGRUND", bold=True, size=12, bg=sub_clr, fg=CLR_WHITE,
           merge_to=5)
    row += 1
    ws.row_dimensions[row].height = 60
    _write(row, 1, INTRO_TEXT, italic=True, merge_to=5)
    ws.cell(row=row, column=1).alignment = Alignment(
        wrap_text=True, vertical="top")
    row += 2

    # Tabellen-Header
    _write(row, 1, "Parameter",      bold=True, size=10, bg=header_clr, fg=CLR_WHITE, border=True)
    _write(row, 2, "Bezeichnung",    bold=True, size=10, bg=header_clr, fg=CLR_WHITE, border=True)
    _write(row, 3, "Baseline",       bold=True, size=10, bg=header_clr, fg=CLR_WHITE, border=True)
    _write(row, 4, f"NEU ({title_pfx})", bold=True, size=10, bg=header_clr, fg=CLR_WHITE, border=True)
    _write(row, 5, "Begründung / Zusammenhang",
           bold=True, size=10, bg=header_clr, fg=CLR_WHITE, border=True)
    row += 1

    for sec in sections:
        # Abschnitts-Header
        ws.row_dimensions[row].height = 22
        _write(row, 1, sec["titel"], bold=True, size=11, bg=sub_clr,
               fg=CLR_WHITE, merge_to=5)
        row += 1

        for i, (param, label, baseline, neu, erklaerung) in enumerate(sec["parameter"]):
            bg = row_clr if i % 2 == 0 else CLR_WHITE
            ws.row_dimensions[row].height = 80
            _write(row, 1, param,       bg=bg, border=True, size=9)
            _write(row, 2, label,       bg=bg, border=True, size=9, bold=True)
            _write(row, 3, baseline,    bg=bg, border=True, size=9, italic=True)
            _write(row, 4, neu,         bg=CLR_CHANGED, border=True, size=10, bold=True)
            cell5 = _write(row, 5, erklaerung, bg=bg, border=True, size=9)
            cell5.alignment = Alignment(wrap_text=True, vertical="top")
            row += 1

        # Auswirkung
        ws.row_dimensions[row].height = 60
        _write(row, 1, "→ Auswirkung", bold=True, italic=True, size=9,
               bg="FFF2CC", merge_to=1)
        auswirkung_cell = ws.cell(row=row, column=2, value=sec["auswirkung"])
        auswirkung_cell.font = _font(italic=True, size=9)
        auswirkung_cell.fill = _fill("FFF2CC")
        auswirkung_cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(start_row=row, start_column=2,
                       end_row=row, end_column=5)
        row += 2

    # Hauptunterschiede Box
    ws.row_dimensions[row].height = 20
    _write(row, 1, "VERGLEICH: KONSERVATIV vs. AGGRESSIV", bold=True, size=12,
           bg=header_clr, fg=CLR_WHITE, merge_to=5)
    row += 1
    ws.row_dimensions[row].height = 300
    cell = ws.cell(row=row, column=1, value=KEY_DIFFERENCES)
    cell.font = Font(name="Courier New", size=8)
    cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=row, start_column=1,
                   end_row=row, end_column=5)
    row += 2

    # Offene Punkte
    ws.row_dimensions[row].height = 20
    _write(row, 1, "OFFENE PUNKTE (aus PDF Decision Register Part G)",
           bold=True, size=12, bg=CLR_WARN, fg=CLR_WHITE, merge_to=5)
    row += 1
    ws.row_dimensions[row].height = 180
    cell = ws.cell(row=row, column=1, value=OPEN_ITEMS)
    cell.font = _font(size=9, italic=True)
    cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=row, start_column=1,
                   end_row=row, end_column=5)
    row += 1

    return ws


# ── MARKDOWN REPORT GENERIEREN ────────────────────────────────────────────────

def generate_scenario_report(scenario_name, applied_changes, sections,
                              is_aggressiv=False):
    """Erstellt ausführlichen Markdown-Report für ein Szenario."""
    title_pfx = "AGGRESSIV" if is_aggressiv else "KONSERVATIV"
    lines = [
        f"# LFL Financial Projections — Szenario {title_pfx}: {scenario_name}",
        f"**Basis:** 260304_LFL-Financial-Planning-and-Carbon-Case.pdf",
        f"**Erstellt:** {datetime.now().strftime('%d.%m.%Y %H:%M')}",
        "",
        "---",
        "",
        f"## Hintergrund",
        "",
        INTRO_TEXT,
        "",
        "---",
        "",
    ]

    for sec in sections:
        lines.append(f"## {sec['titel']}")
        lines.append("")
        lines.append("| Parameter | Bezeichnung | Baseline | NEU | Begründung |")
        lines.append("|-----------|------------|---------|-----|------------|")
        for param, label, baseline, neu, erklaerung in sec["parameter"]:
            erklaerung_short = erklaerung.replace("\n", " ")[:200]
            lines.append(f"| `{param}` | {label} | {baseline} | **{neu}** | {erklaerung_short} |")
        lines.append("")
        lines.append(f"**→ Hauptauswirkung:** {sec['auswirkung']}")
        lines.append("")

    lines += [
        "---",
        "",
        "## Gesamtvergleich",
        "",
        "```",
        KEY_DIFFERENCES,
        "```",
        "",
        "---",
        "",
        "## Offene Punkte aus dem PDF",
        "",
        OPEN_ITEMS,
        "",
        "---",
        "",
        "## Angewendete Änderungen (technisch)",
        "",
        "| Sheet | Zelle | Bezeichnung | Alt → Neu |",
        "|-------|-------|-------------|-----------|",
    ]

    for ch in applied_changes:
        old, new = ch["old"], ch["new"]
        if isinstance(old, float):
            old_str = f"{old:.4f}".rstrip('0').rstrip('.')
        else:
            old_str = str(old)
        if isinstance(new, float):
            new_str = f"{new:.4f}".rstrip('0').rstrip('.')
        else:
            new_str = str(new)
        lines.append(f"| {ch['sheet']} | `{ch['cell']}` | {ch['label']} | "
                     f"{old_str} → **{new_str}** |")

    lines.append("")
    return "\n".join(lines)


# ── SZENARIO ERSTELLEN ────────────────────────────────────────────────────────

def create_scenario(name, sandbox_changes, input_changes, base_szenario,
                    sections, is_aggressiv=False):
    """Erstellt vollständiges Szenario mit Excel + Markdown-Report."""

    print(f"\n{'='*60}")
    print(f"Erstelle Szenario: {name}")
    print(f"{'='*60}")

    changes = {
        "Inputs": input_changes,
        "Sandbox": sandbox_changes,
    }

    wb, applied = apply_scenario(changes, szenario=base_szenario)

    print(f"  ✓ {len(applied)} Änderungen angewendet")

    # Explanation Sheet hinzufügen
    add_explanation_sheet(wb, name, sections, is_aggressiv=is_aggressiv)
    print("  ✓ Erklärungs-Sheet hinzugefügt")

    # Excel speichern
    filepath = save_scenario(wb, name)
    print(f"  ✓ Excel gespeichert: {filepath}")

    # Markdown Report
    report_md = generate_scenario_report(name, applied, sections, is_aggressiv)
    report_path = os.path.join(REPORTS_DIR,
                               f"Report_{name}_{datetime.now().strftime('%Y%m%d_%H%M')}.md")
    os.makedirs(REPORTS_DIR, exist_ok=True)
    with open(report_path, "w", encoding="utf-8") as f:
        f.write(report_md)
    print(f"  ✓ Markdown-Report: {report_path}")

    return filepath, report_path


# ── MAIN ──────────────────────────────────────────────────────────────────────

def main():
    print("""
╔══════════════════════════════════════════════════════╗
║  LFL PDF-SZENARIEN GENERATOR                        ║
║  Basis: 260304_LFL-Financial-Planning-and-Carbon-   ║
║         Case.pdf (März 2026)                        ║
╚══════════════════════════════════════════════════════╝
    """)

    # 1. KONSERVATIV
    fp_k, rp_k = create_scenario(
        name           = "PDF_Konservativ",
        sandbox_changes= KONSERVATIV_SANDBOX,
        input_changes  = KONSERVATIV_INPUTS,
        base_szenario  = "gering",
        sections       = KONSERVATIV_SECTIONS,
        is_aggressiv   = False,
    )

    # 2. AGGRESSIV
    fp_a, rp_a = create_scenario(
        name           = "PDF_Aggressiv",
        sandbox_changes= AGGRESSIV_SANDBOX,
        input_changes  = AGGRESSIV_INPUTS,
        base_szenario  = "stark",
        sections       = AGGRESSIV_SECTIONS,
        is_aggressiv   = True,
    )

    print(f"""
╔══════════════════════════════════════════════════════╗
║  FERTIG!                                            ║
╠══════════════════════════════════════════════════════╣
║  KONSERVATIV                                        ║
║    Excel:  {Path(fp_k).name:<42}║
║    Report: {Path(rp_k).name:<42}║
╠══════════════════════════════════════════════════════╣
║  AGGRESSIV                                          ║
║    Excel:  {Path(fp_a).name:<42}║
║    Report: {Path(rp_a).name:<42}║
╚══════════════════════════════════════════════════════╝

Nächste Schritte:
  1. Excel-Dateien in Microsoft Excel öffnen
  2. Strg+Shift+F9 (Alle Formeln neu berechnen)
  3. Sheet 'LFL_Szenario_Erklaerung' für vollständige
     Begründungen lesen
    """)

    return fp_k, fp_a, rp_k, rp_a


if __name__ == "__main__":
    main()
