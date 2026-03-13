"""
compute_model.py – Berechnet 52-Monats-Finanzdaten für ein Szenario.

Da openpyxl keine Excel-Formeln neu berechnen kann, implementiert dieses Modul
die Kernlogik des v19-Modells in Python.

Ansatz:
  - REVENUE:  Vollständig neu berechnet aus Szenario-Parametern (1_Szenario E/F/G)
  - COSTS:    Gelesen aus Quelldatei (gering-Basiswerte); nur Revenue-abhängige
              Positionen werden skaliert (Payment Processing, Cloud variable)
  - P&L:      Abgeleitet (Revenue - COGS - OpEx)
  - CF:       Abgeleitet (Net Income + Equity Funding)
"""

import math
import openpyxl

# ── Szenario-Spalten in 1_Szenario: E=5 (gering), F=6 (normal), G=7 (stark) ──
SZEN_COLS = {'gering': 5, 'normal': 6, 'stark': 7}

# ── Customer-Skalierungsfaktoren (relative to gering manual plan) ─────────────
# Skaliert neue Kunden pro Monat:
#   normal: 2× mehr SME/Mid, 1.5× Ent
#   stark:  4× mehr SME/Mid, 2× Ent
KUNDE_SCALE = {
    'gering': (1.0, 1.0, 1.0),   # (SME, Mid, Enterprise)
    'normal': (2.0, 2.0, 1.5),
    'stark':  (4.5, 4.5, 2.5),
}

# Monatliche Startverschiebung (bei positivem Wert: früher)
START_SHIFT = {
    'gering': 0,
    'normal': -1,   # M8 statt M7 = 1 Monat später, aber mehr Kunden
    'stark':  2,    # 2 Monate früher (M5 statt M7)
}


def read_szenario_params(wb, szenario: str) -> dict:
    """Liest Szenario-Parameter aus dem 1_Szenario-Sheet."""
    ws = wb['1_Szenario']
    col = SZEN_COLS[szenario]

    def get(row):
        v = ws.cell(row=row, column=col).value
        try:
            return float(v)
        except (TypeError, ValueError):
            return 0.0

    return {
        'startmonat_kunden': int(get(4)),
        'impl_tagessatz':    get(5),
        'impl_tage_kunde':   get(6),
        'impl_buchungsquote':get(7),
        'impl_startmonat':   int(get(8)),
        'nrr':               get(9),
        'churn_annual':      get(10),
        'impl_umsatz_kunde': get(14),
        'seats_sme':         int(get(15)),
        'seats_mid':         int(get(16)),
        'price_sme':         get(17),    # €/Monat
        'price_mid':         get(18),
        'enterprise_fee':    get(19),    # €/Jahr
        'price_increase':    get(20),    # p.a.
    }


def read_gering_customer_plan(wb) -> tuple:
    """Liest den manuellen Kundenplan (gering) aus 2_Inputs R217-R219."""
    ws = wb['2_Inputs']
    sme = [int(ws.cell(row=217, column=c).value or 0) for c in range(2, 54)]
    mid = [int(ws.cell(row=218, column=c).value or 0) for c in range(2, 54)]
    ent = [int(ws.cell(row=219, column=c).value or 0) for c in range(2, 54)]
    return sme, mid, ent


def generate_customer_plan(szenario: str, base_sme, base_mid, base_ent) -> tuple:
    """
    Generiert einen 52-Monats-Kundenplan für das gewünschte Szenario.
    Basiert auf dem gering-Plan, skaliert und zeitlich verschoben.
    """
    if szenario == 'gering':
        return list(base_sme), list(base_mid), list(base_ent)

    scale_sme, scale_mid, scale_ent = KUNDE_SCALE[szenario]
    shift = START_SHIFT[szenario]   # + = früher, - = später

    N = 52
    new_sme = [0] * N
    new_mid = [0] * N
    new_ent = [0] * N

    for i in range(N):
        j = i - shift   # Quellmonat im gering-Plan
        if 0 <= j < N:
            new_sme[i] = round(base_sme[j] * scale_sme)
            new_mid[i] = round(base_mid[j] * scale_mid)
            new_ent[i] = round(base_ent[j] * scale_ent)

    return new_sme, new_mid, new_ent


def compute_revenue(p: dict, new_sme, new_mid, new_ent) -> dict:
    """
    Berechnet alle Revenue-Zeilen für 52 Monate.

    Gibt zurück: dict mit 52-Elemente-Listen:
      seats_sme, seats_mid, enterprise_count,
      sme_rev, mid_rev, enterprise_rev,
      mrr, arr, impl_rev, total_revenue
    """
    N = 52
    churn_monthly = p['churn_annual'] / 12

    seats_sme_active    = [0] * N
    seats_mid_active    = [0] * N
    ent_active          = [0] * N
    sme_rev             = [0.0] * N
    mid_rev             = [0.0] * N
    ent_rev             = [0.0] * N
    impl_rev            = [0.0] * N
    mrr_arr             = [0.0] * N
    arr_arr             = [0.0] * N
    total_rev           = [0.0] * N

    prev_sme, prev_mid, prev_ent = 0, 0, 0

    for m in range(N):   # m = 0..51 (Monat 1..52)
        year = m // 12
        price_factor = (1 + p['price_increase']) ** year

        # Neue Seats
        added_sme = new_sme[m] * p['seats_sme']
        added_mid = new_mid[m] * p['seats_mid']
        added_ent = new_ent[m]

        # Churned Seats
        churned_sme = round(prev_sme * churn_monthly)
        churned_mid = round(prev_mid * churn_monthly)
        churned_ent = round(prev_ent * churn_monthly) if prev_ent > 0 else 0

        # Aktive Bestände
        cur_sme = max(0, prev_sme + added_sme - churned_sme)
        cur_mid = max(0, prev_mid + added_mid - churned_mid)
        cur_ent = max(0, prev_ent + added_ent - churned_ent)

        seats_sme_active[m] = cur_sme
        seats_mid_active[m] = cur_mid
        ent_active[m]       = cur_ent

        # Revenue
        sme_rev[m] = cur_sme * p['price_sme'] * price_factor
        mid_rev[m] = cur_mid * p['price_mid'] * price_factor
        ent_rev[m] = cur_ent * p['enterprise_fee'] * price_factor / 12

        mrr = sme_rev[m] + mid_rev[m] + ent_rev[m]
        mrr_arr[m] = mrr
        arr_arr[m] = mrr * 12

        # Implementation Support
        impl_start = p['impl_startmonat'] - 1   # 0-basiert
        if m >= impl_start:
            new_kunden = new_sme[m] + new_mid[m] + new_ent[m]
            impl_rev[m] = new_kunden * p['impl_umsatz_kunde'] * p['impl_buchungsquote']
        else:
            impl_rev[m] = 0.0

        total_rev[m] = mrr + impl_rev[m]

        prev_sme, prev_mid, prev_ent = cur_sme, cur_mid, cur_ent

    return {
        'seats_sme':        seats_sme_active,
        'seats_mid':        seats_mid_active,
        'enterprise_count': ent_active,
        'sme_rev':          sme_rev,
        'mid_rev':          mid_rev,
        'enterprise_rev':   ent_rev,
        'mrr':              mrr_arr,
        'arr':              arr_arr,
        'impl_rev':         impl_rev,
        'total_revenue':    total_rev,
    }


def read_costs_from_source(wb) -> dict:
    """
    Liest die szenario-unabhängigen Kostenpositionen (gering-Basis)
    aus dem v19-Quelldatei-Sheet 5_Costs.
    """
    ws = wb['5_Costs']
    def row52(r):
        return [float(ws.cell(row=r, column=c).value or 0) for c in range(2, 54)]

    return {
        'total_personnel':  row52(13),
        'total_tech':       row52(18),
        'total_office':     row52(24),
        'total_prof':       row52(30),
        'total_ins':        row52(35),
        'total_mktg':       row52(40),
        'total_other':      row52(45),
    }


def read_equity_from_source(wb) -> list:
    """Liest Equity Funding aus 7_BS_CF Row 9."""
    ws = wb['7_BS_CF']
    return [float(ws.cell(row=9, column=c).value or 0) for c in range(2, 54)]


def compute_financials(rev: dict, costs: dict, equity: list) -> dict:
    """
    Berechnet P&L und Cash Flow aus Revenue + Costs.

    COGS: Payment Processing (2.5% of Revenue) + Cloud Variable (seat-basiert)
    OpEx: Alle Kostenkategorien aus costs dict
    """
    N = 52
    payment_proc_rate = 0.025   # 2.5% des Umsatzes

    total_cogs      = [0.0] * N
    gross_profit    = [0.0] * N
    total_opex      = [0.0] * N
    ebitda          = [0.0] * N
    income_tax      = [0.0] * N
    net_income      = [0.0] * N
    ending_cash     = [0.0] * N
    beginning_cash  = [0.0] * N
    burn_rate       = [0.0] * N
    runway          = [0.0] * N

    prev_cash = 0.0

    for m in range(N):
        rev_m = rev['total_revenue'][m]

        # COGS (revenue-abhängig)
        payment = rev_m * payment_proc_rate
        cogs_m  = payment    # + cloud variable (gering approximated by zero in M1)

        total_cogs[m]   = cogs_m
        gross_profit[m] = rev_m - cogs_m

        # OpEx (aus Quelldatei – szenario-unabhängig)
        opex_m = (costs['total_personnel'][m]
                  + costs['total_tech'][m]
                  + costs['total_office'][m]
                  + costs['total_prof'][m]
                  + costs['total_ins'][m]
                  + costs['total_mktg'][m]
                  + costs['total_other'][m])
        total_opex[m] = opex_m

        ebitda_m = gross_profit[m] - opex_m
        ebitda[m] = ebitda_m

        # Tax (nur wenn EBITDA > 0)
        tax = max(0.0, ebitda_m * 0.30)
        income_tax[m] = tax
        net_income[m]  = ebitda_m - tax

        # Cash Flow
        beginning_cash[m] = prev_cash
        net_change = net_income[m] + equity[m]
        end_cash   = prev_cash + net_change
        ending_cash[m] = end_cash

        burn = max(0.0, -net_change)
        burn_rate[m] = burn
        runway[m] = (end_cash / burn) if (burn > 0 and end_cash > 0) else (999.0 if end_cash >= 0 else 0.0)

        prev_cash = end_cash

    return {
        'total_cogs':     total_cogs,
        'gross_profit':   gross_profit,
        'total_opex':     total_opex,
        'ebitda':         ebitda,
        'income_tax':     income_tax,
        'net_income':     net_income,
        'beginning_cash': beginning_cash,
        'ending_cash':    ending_cash,
        'burn_rate':      burn_rate,
        'runway':         runway,
        'equity_funding': equity,
    }


def build_model(source_path: str, szenario: str) -> dict:
    """
    Vollständige Modell-Berechnung für ein Szenario.

    Returns:
        dict mit allen 52-Monats-Daten + Szenario-Metadaten
    """
    wb = openpyxl.load_workbook(source_path, data_only=True)

    p = read_szenario_params(wb, szenario)

    # Kundenplan: Basisplan (gering) + Skalierung für andere Szenarien
    base_sme, base_mid, base_ent = read_gering_customer_plan(wb)
    new_sme, new_mid, new_ent = generate_customer_plan(szenario, base_sme, base_mid, base_ent)

    rev    = compute_revenue(p, new_sme, new_mid, new_ent)
    costs  = read_costs_from_source(wb)
    equity = read_equity_from_source(wb)
    fin    = compute_financials(rev, costs, equity)

    # Firmendaten
    ws_inp = wb['2_Inputs']
    firma  = str(ws_inp.cell(row=3, column=2).value or 'LoopforgeLab GmbH')
    start  = ws_inp.cell(row=4, column=2).value
    start_str = start.strftime('%d.%m.%Y') if hasattr(start, 'strftime') else '01.04.2026'

    return {
        'szenario':      szenario,
        'firma':         firma,
        'start_str':     start_str,
        'params':        p,
        # Revenue
        'total_revenue': rev['total_revenue'],
        'mrr':           rev['mrr'],
        'arr':           rev['arr'],
        'seats_sme':     rev['seats_sme'],
        'seats_mid':     rev['seats_mid'],
        'enterprise_count': rev['enterprise_count'],
        'impl_rev':      rev['impl_rev'],
        # Costs
        'total_cogs':    fin['total_cogs'],
        'gross_profit':  fin['gross_profit'],
        'total_personnel': costs['total_personnel'],
        'total_tech':    costs['total_tech'],
        'total_office':  costs['total_office'],
        'total_prof':    costs['total_prof'],
        'total_ins':     costs['total_ins'],
        'total_mktg':    costs['total_mktg'],
        'total_other':   costs['total_other'],
        'total_opex':    fin['total_opex'],
        # P&L
        'ebitda':        fin['ebitda'],
        'income_tax':    fin['income_tax'],
        'net_income':    fin['net_income'],
        # CF
        'equity_funding': fin['equity_funding'],
        'beginning_cash': fin['beginning_cash'],
        'ending_cash':   fin['ending_cash'],
        'burn_rate':     fin['burn_rate'],
        'runway':        fin['runway'],
    }
