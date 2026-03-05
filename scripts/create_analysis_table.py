"""
LFL Financial Projections - Comprehensive Parameter Analysis Table
Generates: reports/LFL_Parameter_Analyse_Vergleich.xlsx

Sheets:
  1. Parameter_Vergleich  - Side-by-side Konservativ / Baseline / Aggressiv with commentary
  2. Hebel_Analyse        - Lever sensitivity: how each param affects Break-Even / ARR / EBITDA
  3. KPI_Uebersicht       - Actual model KPIs at M12/M24/M36/M52 (from v0.4 data_only)
  4. Marktkontext         - Industry context: TAM, lead cycles, ticket sizes
  5. Optimierungs_Guide   - Actionable guide for each lever
"""

import os
import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from datetime import datetime

# ── Paths ─────────────────────────────────────────────────────────────────────
BASE_DIR   = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
MODEL_PATH = os.path.join(BASE_DIR, "Input", "260304_LFL_SaaS_Startup_Financial_Model_v0.4.xlsx")
OUT_PATH   = os.path.join(BASE_DIR, "reports", "LFL_Parameter_Analyse_Vergleich.xlsx")

# ── Styles ────────────────────────────────────────────────────────────────────
BLUE_DARK   = "1F3864"
BLUE_MID    = "2E75B6"
BLUE_LIGHT  = "BDD7EE"
GREEN_DARK  = "375623"
GREEN_LIGHT = "E2EFDA"
AMBER       = "FCE4D6"
AMBER_DARK  = "C55A11"
GREY_LIGHT  = "F2F2F2"
WHITE       = "FFFFFF"
RED_LIGHT   = "FFCCCC"

def hdr(text, bold=True, size=11, color=WHITE, bg=BLUE_DARK, wrap=True, align="center"):
    return {
        "value": text,
        "font": Font(bold=bold, size=size, color=color, name="Calibri"),
        "fill": PatternFill("solid", fgColor=bg),
        "alignment": Alignment(horizontal=align, vertical="center", wrap_text=wrap),
    }

def cell_style(value, bold=False, bg=None, fmt=None, align="left", color="000000", size=10, wrap=True):
    d = {
        "value": value,
        "font": Font(bold=bold, size=size, color=color, name="Calibri"),
        "alignment": Alignment(horizontal=align, vertical="center", wrap_text=wrap),
    }
    if bg:
        d["fill"] = PatternFill("solid", fgColor=bg)
    if fmt:
        d["number_format"] = fmt
    return d

def thin_border():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

def apply_cell(ws, row, col, d):
    c = ws.cell(row=row, column=col)
    c.value = d.get("value")
    if "font"      in d: c.font      = d["font"]
    if "fill"      in d: c.fill      = d["fill"]
    if "alignment" in d: c.alignment = d["alignment"]
    if "number_format" in d: c.number_format = d["number_format"]
    c.border = thin_border()

def write_row(ws, row, cols_data):
    for col, d in enumerate(cols_data, start=1):
        apply_cell(ws, row, col, d)

def section_header(ws, row, text, ncols, bg=BLUE_MID):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1)
    c.value = text
    c.font  = Font(bold=True, color=WHITE, size=11, name="Calibri")
    c.fill  = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
    c.border = thin_border()

def freeze_and_fit(ws, freeze="B2", col_widths=None):
    ws.freeze_panes = freeze
    if col_widths:
        for col_letter, width in col_widths.items():
            ws.column_dimensions[col_letter].width = width
    ws.sheet_view.showGridLines = True

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 1 – Parameter Vergleich
# ══════════════════════════════════════════════════════════════════════════════
def build_parameter_sheet(wb):
    ws = wb.create_sheet("Parameter_Vergleich")
    ws.row_dimensions[1].height = 40
    ws.row_dimensions[2].height = 30

    # Title row
    ws.merge_cells("A1:H1")
    c = ws["A1"]
    c.value = "LFL GmbH – Parameter-Vergleich: Konservativ | Baseline (gering) | Aggressiv"
    c.font  = Font(bold=True, size=14, color=WHITE, name="Calibri")
    c.fill  = PatternFill("solid", fgColor=BLUE_DARK)
    c.alignment = Alignment(horizontal="center", vertical="center")

    # Column headers
    headers = [
        hdr("Parameter / Position"),
        hdr("Kategorie"),
        hdr("Konservativ\n(PDF-Szenario)"),
        hdr("Baseline\n(Modell v0.4 'gering')", bg=BLUE_MID),
        hdr("Aggressiv\n(PDF-Szenario)"),
        hdr("Einheit"),
        hdr("Hebel-Wirkung\n(Profit-Timing)"),
        hdr("Business-Kommentar", align="left"),
    ]
    write_row(ws, 2, headers)

    # Data rows: [param, category, konservativ, baseline, aggressiv, unit, lever, comment]
    rows = [
        # ── MARKT & VERTRIEB ──────────────────────────────────────────────────
        ("MARKT & VERTRIEB", "SECTION"),
        ("Startmonat erster Umsatz",   "Vertrieb",  17, 14, 9,   "Monat",   "SEHR HOCH",
         "Jeder Monat früher = ~€150K ARR-Vorsprung. Automotive-Fokus (gering) braucht 14 Monate wegen TISAX-Compliance. Packaging (aggressiv) ermöglicht Monat 9 durch kürzere Zyklen."),
        ("Seat-Preis pro Jahr",        "Revenue",   2400, 9600, 6000, "€/Seat/Jahr", "HOCH",
         "Baseline nutzt value-based Pricing (€9.600/Jahr). PDF-Szenarien sind deutlich günstiger (€200-500/Monat = €2.400-6.000/Jahr). Preispunkt bestimmt direkt ARR-Potenzial. Empfehlung: €350-500/Monat für Markteintrittspreis."),
        ("Sitzplatz-Startpreis/Monat", "Revenue",   200, 800, 500, "€/Seat/Monat", "HOCH",
         "Sandbox-Wert (gering=800€, normal=1000€, stark=1200€). PDF-Szenarien deutlich darunter. Preiserhöhung von 8%/Jahr kompensiert nur teilweise. Critical: Entscheidung zwischen Marktpenetration vs. Premiumpositionierung."),
        ("Enterprise-Deal ARR",        "Revenue",   75000, 150000, 200000, "€/Deal/Jahr", "SEHR HOCH",
         "Baseline: €150K/Enterprise-Vertrag. 9 Verträge bis M52 = Basis ARR. Aggressiv: €200K durch DPP-Compliance-Mehrwert. Konservativ: €75K für Automotive-Mittelstand. Ein Enterprise-Deal entspricht 62 Seats bei Baseline-Preis."),
        ("Enterprise-Deals/Quartal",   "Vertrieb",  0.5, 1, 2,   "Deals/Quartal", "SEHR HOCH",
         "Direkter Multiplikator auf ARR-Wachstum. Aggressiv: 2 Deals/Quartal ab M24 erfordern 2+ Sales Reps + Pre-Sales Engineering. Konservativ: 1 Deal alle 2 Quartale = realistisch für 1 Sales Rep-Fokus."),
        ("Jährliche Churn Rate",       "Revenue",   0.12, 0.08, 0.05, "%/Jahr", "HOCH",
         "Baseline: 8% Churn (branchenüblich für Industrie-SaaS). Aggressiv: 5% durch hohe Integrationstiefe. Konservativ: 12% bei Automotive wegen längerer Budgetzyklen. 1%-Punkt Churn-Reduktion = ca. €30K ARR-Gewinn bei M52."),
        ("Net Revenue Retention (NRR)", "Revenue",  1.05, 1.18, 1.25, "Faktor", "HOCH",
         "NRR > 1.0 bedeutet Wachstum ohne Neukundengewinnung. Baseline: 118% durch DPP-Module/Analytics-Upgrades. Aggressiv: 125% bei vollständiger Plattformintegration. NRR ist der stärkste EBITDA-Hebel ab M36+."),
        ("Preiserhöhung pro Jahr",     "Revenue",   0.05, 0.08, 0.10, "%/Jahr", "MITTEL",
         "Baseline: 8% p.a. (KI-SaaS Standard 2026). Aggressiv: 10% durch kontinuierliche Feature-Erweiterung. Wirkung akkumuliert: M52-Preis bei 10%/Jahr = 1.63× Startpreis."),

        # ── PERSONALKOSTEN ────────────────────────────────────────────────────
        ("PERSONALKOSTEN", "SECTION"),
        ("CEO Gehalt",                 "Personal",  72000, 72000, 72000, "€/Jahr", "NIEDRIG",
         "Gründer-Gehalt Pre-Seed-Phase. Anpassung ab Seed-Runde (M17) auf Marktgehalt €100-130K empfohlen. Alle 3 Gründer (CEO/CTO/CCO) = €216K Jahreskosten + 21% NNK = ~€261K p.a."),
        ("Senior Engineer Gehalt",     "Personal",  80000, 90000, 95000, "€/Jahr", "MITTEL",
         "Berliner Tech-Markt 2026: €80-100K. Baseline nutzt €90K. 3 Senior Engineers bis M18 = €270K/Jahr Brutto. Mit AG-NNK (21%): €326K. KI-Hebel kann 4. Senior Engineer verzögern."),
        ("ML/AI Engineer Gehalt",      "Personal",  100000, 110000, 120000, "€/Jahr", "MITTEL",
         "Höchste Nachfrage im Markt. Baseline: €110K. 2-3 ML Engineers essential für Kerndifferenzierung. KI-Hebel reduziert Einstellungsdruck: 1 ML Engineer mit Claude-API kann Arbeit von 1,5 leisten."),
        ("Sales Rep Gehalt",           "Personal",  130000, 130000, 130000, "€/Jahr", "HOCH",
         "Split 50/50: €65K Fix + €65K Variable. Enterprise-Vertrieb B2B Industrie erfordert hohe Provisionen. KI-Hebel (SDR-Agent): Ersetzt Outbound-Funktion für ~€500/Monat. Human-in-the-Loop bleibt für Abschluss."),
        ("Lohnnebenkosten (AG-Anteil)", "Personal", 0.21, 0.21, 0.21, "% des Bruttolohns", "MITTEL",
         "Gesetzlicher AG-Anteil Sozialversicherung DE: ~20-21%. Erhöht effektive Gehaltskosten um 21%. Bei €1M Bruttolohnkosten = +€210K p.a. Unveränderlich, da gesetzlich fixiert."),
        ("KI-Personal-Hebel",          "Strategie", 0, 6, 99,  "Monate Verzögerung", "SEHR HOCH",
         "Innovativster Parameter. Wert = Anzahl Monate, um die KI-Positionen verschoben werden. 99 = nie im 52M-Horizont. Aggressiv: Alle KI-Positionen gar nicht eingestellt → 3 FTE weniger = ca. €350K/Jahr Einsparung. Empfehlung: 36-48 Monate für optimales Cash-Management."),

        # ── EINSTELLUNGSPLAN ──────────────────────────────────────────────────
        ("EINSTELLUNGSPLAN", "SECTION"),
        ("1. Senior Engineer (Eintritt)", "Hiring",  7, 7, 7,   "Monat",   "MITTEL",
         "Strategische Rolle, kein KI-Effekt. M7 = Oktober 2026. Notwendig für MVP-Finalisierung. €90K + NNK = €108,9K/Jahr Kosten ab M7."),
        ("1. ML/AI Engineer (Eintritt)", "Hiring",   8, 8, 8,   "Monat",   "HOCH",
         "Kern-IP-Träger. Kein KI-Effekt. M8 = November 2026. Fokus: Agentische Systeme + Material-Design-Algorithmen. €110K + NNK = €133,1K/Jahr."),
        ("1. Sales Rep (Eintritt)",      "Hiring",   99, 99, 99, "Monat",   "SEHR HOCH",
         "KI-Agent ersetzt Outbound-SDR-Funktion. Baseline: KI-Strategie → Monat 99 (nie eingestellt). Einsparung vs. Hiring: €130K/Jahr. AI-SDR-Kosten: ~€6K/Jahr. ROI: 21:1."),
        ("1. Customer Success (Eintritt)", "Hiring", 12, 12, 12, "Monat",  "MITTEL",
         "M12 = März 2027. Strategisch wichtig für Enterprise-Onboarding. RAG-Wissensagent reduziert CS-Aufwand um ~40%. €65K + NNK = €78,65K/Jahr."),
        ("2. Sales Rep (Eintritt)",      "Hiring",   24, 18, 12, "Monat",   "HOCH",
         "Human-in-the-Loop für Enterprise-Abschlüsse. Aggressiv: M12 für frühes Scale-Up. Konservativ: M24 da weniger Deals. €130K + NNK = €157,3K/Jahr."),

        # ── FINANZIERUNG / CAPEX ──────────────────────────────────────────────
        ("FINANZIERUNG / CAPEX", "SECTION"),
        ("Ideation-Finanzierung",      "Funding",   90000, 90000, 90000, "€",  "NIEDRIG",
         "M1 (April 2026). Deckt MVP-Entwicklung und IP-Aufbau. Keine Equity-Verwässerung da Founder-finanziert. Reicht für ~1,3 Monate Burn bei M1-Rate."),
        ("Pre-Seed Betrag",            "Funding",   1500000, 1500000, 2000000, "€", "HOCH",
         "M5 (August 2026). Konservativ/Baseline: €1,5M. Aggressiv: €2M für schnelleres Hiring. Runway: ca. 12-15 Monate bis Seed-Runde."),
        ("Seed-Betrag",                "Funding",   4000000, 6000000, 8000000, "€", "SEHR HOCH",
         "M17 (August 2027). Baseline: €6M. Aggressiv: €8M für internationale Expansion. Konservativ: €4M wenn Traction langsamer. Bestimmt Runway bis Series A."),
        ("Series A Betrag",            "Funding",   10000000, 15000000, 20000000, "€", "SEHR HOCH",
         "M35 (Februar 2029). Größter Kapitalschritt. Basiert auf ARR-Multiplikator: Bei M35-ARR von €800K-€1,5M und 10-15× Múltiple ergibt sich Bewertung €8-22M. Runway bis M52+."),
        ("Investitionen (CAPEX)",      "CAPEX",     3000, 3300, 5000, "€/Monat", "NIEDRIG",
         "Hauptsächlich Hardware/Server-Ausstattung. Baseline: €3.300/Monat. Initial-CAPEX M1: €14.900 (Gründer-Ausstattung). Kein Gebäude-CAPEX dank Home-Office-Modell."),

        # ── BETRIEBSKOSTEN (OPEX) ─────────────────────────────────────────────
        ("BETRIEBSKOSTEN (OPEX)", "SECTION"),
        ("Cloud Hosting (variabel)",   "OPEX",      250, 250, 500,  "€/Monat", "NIEDRIG",
         "AWS/GCP Basiskosten. Wächst mit Customer-Onboarding. Bei M52: 9 Enterprise-Kunden + Subscription → kaum Skalierungseffekte da SaaS-Architektur."),
        ("AI/ML API Kosten",           "OPEX",      1000, 1710, 3000, "€/Monat", "MITTEL",
         "Größter COGS-Treiber. Baseline M52: €12K/Monat. Skaliert mit API-Calls und Kundennutzung. Modell-Effizienz (Caching, kleinere Modelle) kann um 30-50% optimiert werden. Kritisch für Gross Margin."),
        ("Gesamte OPEX M12",           "OPEX",      80000, 99311, 120000, "€/Monat", "HOCH",
         "Baseline M12: €99K/Monat Burn. Getrieben durch Personal (70,7%). Konservativ: weniger Hiring → ~€80K. Aggressiv: mehr Hiring → ~€120K. Break-Even erfordert €99K+/Monat Revenue."),
        ("Gesamte OPEX M36",           "OPEX",      180000, 248245, 320000, "€/Monat", "HOCH",
         "Baseline M36: €248K/Monat. Personal weiterhin dominant (78%). Konservativ erreicht Break-Even nicht bis M52. Aggressiv könnte M38-40 Break-Even erreichen durch höheres Revenue."),
        ("Onboarding-Aufwand/Kunde",   "OPEX",      120, 60, 30, "Stunden", "MITTEL",
         "Sandbox-Parameter. Gering/Automotive: 120h wegen Legacy-Systemen. Stark/Packaging: 30h durch moderne PLM-Integration. Bei €100/h CS-Stunde: €12K vs. €3K pro Onboarding. Direkter Impact auf CS-Skalierbarkeit."),
        ("Consulting-Tagessatz",       "Revenue",   1200, 1200, 1800, "€/Tag", "MITTEL",
         "Zusätzliche Erlösquelle (Services). Baseline gering: €1.200/Tag. Stark/Packaging: €1.800/Tag durch Expertenstatus. Nicht im Kern-ARR, aber wichtig für Early-Revenue-Bridge."),

        # ── MARKTPARAMETER ────────────────────────────────────────────────────
        ("MARKTPARAMETER", "SECTION"),
        ("Zielmarkt Packaging-Anteil", "Markt",     0.10, 0.10, 0.80, "% des Portfolios", "SEHR HOCH",
         "Sandbox-Szenario-Treiber. Gering=10% Packaging (Automotive-Fokus). Stark=80% Packaging. Packaging: 3-6M Vertriebszyklus. Automotive: 9-18M. Marktentscheidung bestimmt Velocity des gesamten Modells."),
        ("Sales-Zykluslänge",          "Markt",     12, 9, 5, "Monate", "SEHR HOCH",
         "Nicht explizit im Modell, aber implizit über Startmonat. Automotive B2B: 9-18M (TISAX, Budgetzyklen). Packaging Mittelstand: 3-6M. Kürzere Zyklen = früherer Cashflow = niedrigere Gesamtinvestition bis Break-Even."),
        ("Initiale Seats/Kunde",       "Revenue",   3, 5, 10, "Seats/Kunden", "MITTEL",
         "Baseline: 5 Seats. Aggressiv: 10 Seats (Werk-Rollout). Land-and-Expand-Strategie: Start klein, monatliches Seat-Wachstum. Mehr Seats = höherer ACV und bessere NRR durch Switching Costs."),
        ("Seat-Wachstumsrate/Monat",   "Revenue",   0.03, 0.05, 0.10, "%/Monat", "MITTEL",
         "Baseline: 5%/Monat. Aggressiv: 10% durch proaktives CS und Upsell-Playbooks. Bei 5 Initial-Seats und 5%/Monat: M12=9 Seats, M24=14 Seats. Bei 10%: M12=13, M24=34 Seats."),
    ]

    row_num = 3
    ROW_BG_ALT = GREY_LIGHT
    row_counter = 0

    for item in rows:
        if item[1] == "SECTION":
            section_header(ws, row_num, f"  {item[0]}", 8, bg=BLUE_MID)
            ws.row_dimensions[row_num].height = 22
            row_num += 1
            row_counter = 0
            continue

        param, cat, konservativ, baseline, aggressiv, unit, lever, comment = item

        # Color lever cells
        lever_bg = {
            "SEHR HOCH": "FF6B6B",
            "HOCH":      "FFB347",
            "MITTEL":    "FFD700",
            "NIEDRIG":   GREEN_LIGHT.replace("#", ""),
        }.get(lever, WHITE)

        row_bg = GREY_LIGHT if row_counter % 2 == 1 else WHITE
        row_counter += 1

        ws.row_dimensions[row_num].height = 60

        cols = [
            cell_style(param,        bold=True,  bg=row_bg,    align="left",   size=10),
            cell_style(cat,          bold=False, bg=row_bg,    align="center",  size=10),
            cell_style(konservativ,  bold=False, bg=AMBER,     align="right",   size=10),
            cell_style(baseline,     bold=True,  bg=BLUE_LIGHT,align="right",   size=10),
            cell_style(aggressiv,    bold=False, bg=GREEN_LIGHT,align="right",  size=10),
            cell_style(unit,         bold=False, bg=row_bg,    align="center",  size=10),
            cell_style(lever,        bold=True,  bg=lever_bg,  align="center",  size=10, color=WHITE if lever in ("SEHR HOCH", "HOCH") else "000000"),
            cell_style(comment,      bold=False, bg=row_bg,    align="left",    size=9,  wrap=True),
        ]
        write_row(ws, row_num, cols)
        row_num += 1

    freeze_and_fit(ws, freeze="C3", col_widths={
        "A": 30, "B": 15, "C": 15, "D": 18, "E": 15,
        "F": 14, "G": 14, "H": 55,
    })

    # Legend
    row_num += 1
    ws.cell(row=row_num, column=1).value = "Farblegende:"
    ws.cell(row=row_num, column=1).font  = Font(bold=True, size=10)
    legends = [
        (AMBER,      "Konservativ-Szenario"),
        (BLUE_LIGHT, "Baseline Modell v0.4 (gering)"),
        (GREEN_LIGHT,"Aggressiv-Szenario"),
    ]
    for i, (bg, label) in enumerate(legends, start=2):
        c = ws.cell(row=row_num, column=i)
        c.value = label
        c.fill  = PatternFill("solid", fgColor=bg)
        c.font  = Font(size=9)
        c.border = thin_border()

    return ws


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 2 – Hebel-Analyse
# ══════════════════════════════════════════════════════════════════════════════
def build_lever_sheet(wb):
    ws = wb.create_sheet("Hebel_Analyse")
    ws.merge_cells("A1:G1")
    c = ws["A1"]
    c.value = "LFL GmbH – Hebel-Analyse: Wie jeder Parameter Break-Even, ARR und EBITDA beeinflusst"
    c.font  = Font(bold=True, size=13, color=WHITE, name="Calibri")
    c.fill  = PatternFill("solid", fgColor=BLUE_DARK)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 35

    write_row(ws, 2, [
        hdr("Hebel / Parameter"),
        hdr("Wirkungs-richtung"),
        hdr("Effekt auf\nBreak-Even\n(Monate)"),
        hdr("Effekt auf\nARR M52\n(€)"),
        hdr("Effekt auf\nEBITDA M52\n(€/Monat)"),
        hdr("Priorität"),
        hdr("Konkrete Empfehlung & Erklärung"),
    ])

    levers = [
        ("UMSATZ-HEBEL", "SECTION"),
        ("Startmonat Umsatz um 3M früher", "+", "-3 bis -6 Monate", "+€450K", "+€37,5K/M",
         "P1 KRITISCH",
         "Packaging-Fokus (80%) erlaubt Monat 9 statt 14. Maßnahmen: 2 Referenzkunden im Packaging in Q1, Demo-Factory-Partnership, direkte Outreach via Haag/Candemir-Netzwerk. ROI: Frühere Selbstfinanzierung reduziert Seed-Bedarf um ~€1,5M."),
        ("Enterprise-Preis +€50K (von €150K auf €200K)", "+", "-4 bis -8 Monate", "+€450K", "+€37,5K/M",
         "P1 KRITISCH",
         "Erreichbar durch DPP-Compliance-Positionierung (DPP-Pflicht 2027 für Batterien/Elektronik). LFL wird 'Compliance Engine': Preiserhöhung von €150K→€200K/Deal. Bei 9 Verträgen M52: +€450K ARR. Argumentationslinie: TCO-Einsparung €500K+/Jahr je Werk rechtfertigt €200K ARR."),
        ("Enterprise-Deals 2/Quartal statt 1/Quartal", "+", "-8 bis -12 Monate", "+€1,5M", "+€125K/M",
         "P1 KRITISCH",
         "Verdopplung erfordert 2. Sales Rep ab M18 statt M24 + Pre-Sales Engineer. Cost: +€170K/Jahr. Benefit: +€750K ARR/Jahr ab M30. ROI-Breakeven des extra Headcount bei ~6 Monaten."),
        ("NRR von 118% auf 125% erhöhen", "+", "-6 bis -10 Monate", "+€380K", "+€31,7K/M",
         "P1 KRITISCH",
         "7%-Punkt NRR-Steigerung durch: (1) Analytics-Premium-Modul @+€200/Seat/Jahr, (2) DPP-Reporting-Addon @€15K/Vertrag/Jahr, (3) Proaktives CS-Playbook. Kein extra Headcount notwendig wenn CS-Agent deployed."),
        ("Churn von 8% auf 5% senken", "+", "-3 bis -5 Monate", "+€190K", "+€15,8K/M",
         "P2 HOCH",
         "Enterprise-Churn-Reduktion durch: (1) Tiefe API-Integration (Switching Cost hoch), (2) Quarterly Business Reviews, (3) SLA-Garantien. Bei 9 Verträgen M52: 3%-Punkt Churn = 0,27 Verträge/Jahr gerettet = €40K+ ARR."),
        ("Seat-Preis €500/M statt €200/M", "+", "-10 bis -15 Monate", "+€820K", "+€68K/M",
         "P2 HOCH",
         "Value-based Pricing: €500/Seat/Monat = €6.000/Jahr. Marktvergleich: Siemens Teamcenter €400+, PTC Windchill €600+. LFL mit KI-Differenzierung rechtfertigt €400-600. Aber: Höherer Preis verlangsamt Adoption. Segmentierung empfohlen."),
        ("Preiserhöhung 10% statt 8%/Jahr", "+", "-2 bis -3 Monate", "+€150K", "+€12,5K/M",
         "P3 MITTEL",
         "2%-Punkt mehr ergibt M52-Preis-Multiplikator von 1.63× statt 1.47×. Umsetzbar da KI-Investitionen als Feature-Releases kommuniziert werden. Achtung: Churn-Risiko bei Kunden ohne langfristigen Vertrag."),

        ("KOSTEN-HEBEL", "SECTION"),
        ("KI-Hebel: Sales Rep nie einstellen (Monat 99)", "+", "-8 bis -12 Monate", "+€0 ARR", "+€157K/M",
         "P1 KRITISCH",
         "Bereits in Baseline implementiert (Monat 99 = nie). KI-SDR-Agent: ~€500/Monat für Outbound-Lead-Gen. Enterprise-Abschluss durch Gründer/CCO. Einsparung vs. 2 Sales Reps: €314K/Jahr Brutto + NNK. WARNUNG: Skaliert nicht über 4-5 Deals/Quartal hinaus."),
        ("4. Senior Engineer verzögern (M99 statt M24)", "+", "-3 bis -5 Monate", "+€0 ARR", "+€108K/M",
         "P1 KRITISCH",
         "Bereits in Baseline (M99). KI-Code-Assistenten (Claude, Copilot) erhöhen Produktivität Senior Engineers um 30-40%. 3 Senior Engineers reichen bis M36+. Bei M36 Reassessment empfohlen."),
        ("3. ML/AI Engineer verzögern", "+", "-2 bis -4 Monate", "+€0 ARR", "+€133K/M",
         "P2 HOCH",
         "ML/AI-Engineer durch Claude API + Fine-Tuning teilweise substituierbar. 2 ML Engineers + Claude API = Kapazität von 2,5-3 Engineers. Hiring frühestens M30 wenn ARR €600K+ erreicht."),
        ("Cloud-Kosten optimieren (-30%)", "+", "-1 bis -2 Monate", "+€0 ARR", "+€43K/M",
         "P3 MITTEL",
         "AI API Kosten M52 Baseline: €144K/Jahr. Optimierung durch: (1) Request-Caching (30% Einsparung), (2) Kleineres Modell für Routineaufgaben (-50%), (3) Batch-Processing. Gesamt-Einsparung: ~€43K/Jahr."),

        ("TIMING-HEBEL", "SECTION"),
        ("Seed-Runde M14 statt M17", "+", "-3 bis -5 Monate", "+€0 ARR", "+€60K/M",
         "P2 HOCH",
         "3 Monate frühere Seed-Runde bei gleichem Betrag (€6M) schiebt Hiring vor. M14 realistisch wenn Pre-Seed-Traction gut (2+ POCs). Voraussetzung: Erste bezahlte Piloten bis M12 sichtbar."),
        ("Series A M30 statt M35", "+", "-5 bis -8 Monate", "+€600K", "+€50K/M",
         "P3 MITTEL",
         "5 Monate früherer Series A erlaubt earlier Headcount-Aufbau für Enterprise-Scaling. Voraussetzung: ARR M30 bereits €600K+. Aggressiv-Szenario: Series A könnte auf M28-30 vorgezogen werden."),
        ("Packaging-Fokus ab M1 (80% statt 10%)", "+", "-5 bis -8 Monate", "+€700K", "+€58K/M",
         "P1 KRITISCH",
         "Wichtigster Marktentscheidungs-Hebel. Packaging: 2.255 Zielunternehmen in DACH, 3-6M Zyklen, Digital-Twin-Adopter. Automotive: Zwar riesiger Markt aber 9-18M Zyklen, hohe IT-Sicherheitsanforderungen. Empfehlung: Hybrid M1-12 (Packaging-Lead, Automotive follow-on)."),
    ]

    row_num = 3
    row_counter = 0
    for item in levers:
        if item[1] == "SECTION":
            section_header(ws, row_num, f"  {item[0]}", 7, bg=BLUE_MID)
            ws.row_dimensions[row_num].height = 22
            row_num += 1
            row_counter = 0
            continue

        param, direction, be_effect, arr_effect, ebitda_effect, priority, comment = item
        prio_bg = {
            "P1 KRITISCH": "FF4444",
            "P2 HOCH":     "FF9933",
            "P3 MITTEL":   "FFD700",
        }.get(priority.split()[0] + " " + priority.split()[1], WHITE)

        row_bg = GREY_LIGHT if row_counter % 2 == 1 else WHITE
        row_counter += 1
        ws.row_dimensions[row_num].height = 70

        write_row(ws, row_num, [
            cell_style(param,         bold=True,  bg=row_bg, align="left",   size=10),
            cell_style(direction,     bold=True,  bg=GREEN_LIGHT if direction == "+" else RED_LIGHT, align="center", size=14),
            cell_style(be_effect,     bold=False, bg=AMBER,  align="center",  size=10),
            cell_style(arr_effect,    bold=False, bg=BLUE_LIGHT, align="center", size=10),
            cell_style(ebitda_effect, bold=False, bg=GREEN_LIGHT, align="center", size=10),
            cell_style(priority,      bold=True,  bg=prio_bg, align="center", size=10, color=WHITE if "KRITISCH" in priority or "HOCH" in priority else "000000"),
            cell_style(comment,       bold=False, bg=row_bg, align="left",    size=9,  wrap=True),
        ])
        row_num += 1

    freeze_and_fit(ws, freeze="B3", col_widths={
        "A": 35, "B": 10, "C": 14, "D": 14, "E": 14, "F": 14, "G": 60,
    })
    return ws


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 3 – KPI Übersicht (actual model data)
# ══════════════════════════════════════════════════════════════════════════════
def build_kpi_sheet(wb):
    ws = wb.create_sheet("KPI_Uebersicht")
    ws.merge_cells("A1:F1")
    c = ws["A1"]
    c.value = "LFL GmbH – KPI-Übersicht (Ist-Werte aus Modell v0.4, Szenario 'gering')"
    c.font  = Font(bold=True, size=13, color=WHITE)
    c.fill  = PatternFill("solid", fgColor=BLUE_DARK)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 35

    write_row(ws, 2, [
        hdr("KPI"),
        hdr("M1 (Apr 2026)"),
        hdr("M12 (Mär 2027)"),
        hdr("M24 (Mär 2028)"),
        hdr("M36 (Mär 2029)"),
        hdr("M52 (Jul 2030)"),
    ])

    # Actual data from v0.4 model (data_only=True, gering scenario)
    kpis = [
        ("REVENUE", "SECTION"),
        ("Total MRR (€/Monat)",        0, 0, 17820, 77566, 158497),
        ("Total ARR (€/Jahr)",          0, 0, 213840, 930787, 1901964),
        ("Subscription Revenue (€/M)",  0, 0, 4320, 4666, 5442),
        ("Enterprise Revenue (€/M)",    0, 0, 13500, 72900, 153055),
        ("Aktive Enterprise-Verträge",  0, 0, 1, 5, 9),
        ("Seat-Preis (€/Seat/Monat)",   800, 800, 864, 933, 1088),
        ("Gross Margin (%)",            "N/A", "N/A", "78,9%", "90,1%", "89,7%"),

        ("KOSTEN & PROFITABILITÄT", "SECTION"),
        ("Total COGS (€/Monat)",        1000, 1710, 3767, 7705, 16253),
        ("Total OPEX (€/Monat)",        66871, 99311, 201917, 248245, 274356),
        ("  davon Personal (€/Monat)",  21780, 70180, 155213, 193545, 213383),
        ("  davon Technology (€/Monat)", 6817, 8127, 10638, 13383, 19907),
        ("EBITDA (€/Monat)",            -67871, -101022, -187864, -178384, -132112),
        ("EBITDA Margin (%)",           "N/A", "N/A", "-10,5%", "-2,3%", "-0,8%"),
        ("Net Income (€/Monat)",        -68146, -101113, -187955, -178476, -132112),

        ("CASH & LIQUIDITÄT", "SECTION"),
        ("Kumulierter Cash-Bestand (€)", 15659, 796181, 4864403, 17610753, 15098842),
        ("Monatliche Burn Rate (€)",     0, 103043, 202826, 194519, 132200),
        ("Runway (Monate)",             "∞", "7,7", "24,0", "90,5", "114,2"),
        ("Finanzierungseingang (€)",    90000, 0, 0, 0, 0),

        ("HEADCOUNT", "SECTION"),
        ("Total Headcount",             3, 9, 18, 21, 21),
        ("  Executives (CEO/CTO/CCO)",  3, 3, 3, 3, 3),
        ("  Engineering",               0, 4, 8, 8, 8),
        ("  Sales & CS",                0, 1, 3, 4, 4),
        ("  ML/AI",                     0, 1, 3, 3, 3),
        ("  PM/Marketing/Other",        0, 0, 3, 3, 3),

        ("SANDBOX-PARAMETER (gering)", "SECTION"),
        ("Szenario",                    "'gering'", "'gering'", "'gering'", "'gering'", "'gering'"),
        ("Marktfokus Packaging-Anteil", "10%", "10%", "10%", "10%", "10%"),
        ("Startmonat Kunden",           14, 14, 14, 14, 14),
        ("Seat-Startpreis (€/M, intern)", 800, 800, 800, 800, 800),
        ("Onboarding-Aufwand (Stunden)", 120, 120, 120, 120, 120),
        ("KI-Personal-Hebel",           0, 0, 0, 0, 0),
    ]

    row_num = 3
    row_counter = 0
    for item in kpis:
        if item[1] == "SECTION":
            section_header(ws, row_num, f"  {item[0]}", 6)
            ws.row_dimensions[row_num].height = 22
            row_num += 1
            row_counter = 0
            continue

        label, m1, m12, m24, m36, m52 = item
        row_bg = GREY_LIGHT if row_counter % 2 == 1 else WHITE
        row_counter += 1

        def fmt_val(v):
            if isinstance(v, (int, float)) and v < 0:
                return cell_style(v, bg=RED_LIGHT, align="right", size=10)
            elif isinstance(v, (int, float)) and v > 0:
                return cell_style(v, bg=row_bg, align="right", size=10)
            return cell_style(v, bg=row_bg, align="right", size=10)

        write_row(ws, row_num, [
            cell_style(label, bold=label.startswith("  ") is False, bg=row_bg, align="left", size=10),
            fmt_val(m1),
            fmt_val(m12),
            fmt_val(m24),
            fmt_val(m36),
            fmt_val(m52),
        ])
        row_num += 1

    # Notes
    row_num += 1
    notes = [
        "Hinweis: Alle Werte aus Modell v0.4 (data_only=True, Formeln ausgelesen). Szenario 'gering' = Automotive-Fokus, Startmonat M14, Seat-Preis €800/Monat intern.",
        "Break-Even: Modell zeigt EBITDA auch in M52 negativ (€-132K/Monat). Break-Even bei ca. MRR > €280K (= ca. M58-65 bei aktuellem Wachstumspfad).",
        "Funding-Events: Ideation €90K (M1), Pre-Seed €1,5M (M5), Seed €6M (M17), Series A €15M (M35). Cash M52: €15,1M trotz negativem EBITDA.",
    ]
    for note in notes:
        ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=6)
        c = ws.cell(row=row_num, column=1)
        c.value = note
        c.font  = Font(italic=True, size=9, color="666666")
        c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[row_num].height = 30
        row_num += 1

    freeze_and_fit(ws, freeze="B3", col_widths={
        "A": 35, "B": 16, "C": 16, "D": 16, "E": 16, "F": 16,
    })
    return ws


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 4 – Marktkontext
# ══════════════════════════════════════════════════════════════════════════════
def build_market_sheet(wb):
    ws = wb.create_sheet("Marktkontext")
    ws.merge_cells("A1:E1")
    c = ws["A1"]
    c.value = "LFL GmbH – Marktkontext: TAM, Branchen, Lead-Zyklen, Ticket-Größen"
    c.font  = Font(bold=True, size=13, color=WHITE)
    c.fill  = PatternFill("solid", fgColor=BLUE_DARK)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 35

    section_header(ws, 2, "  ZIELMARKT – DACH PACKAGING & MASCHINENBAU", 5)

    write_row(ws, 3, [
        hdr("Marktsegment"),
        hdr("Unternehmen\n(DACH TAM)"),
        hdr("Sales-Zyklus"),
        hdr("Ticket-Größe\n(ARR/Kunde)"),
        hdr("Markt-Kommentar"),
    ])

    market_rows = [
        ("Verpackungsmaschinenbau (Kernmarkt)", "~800 Unternehmen",
         "3-6 Monate", "€60K-€150K ARR",
         "Stärkste Product-Market-Fit-Indikatoren. Digital-Twin-Adoption hoch durch Industrie 4.0. Unternehmen: Bosch Packaging, Syntegon, IMA Group, Coesia. Entscheidungsträger: CTO + Head of Engineering."),
        ("Kunststoff- & Formteilefertigung", "~600 Unternehmen",
         "4-8 Monate", "€40K-€100K ARR",
         "Hoher Leidensdruck durch Ausschusskosten und Redesign-Trap. Typical: €200K-€500K Jahreskosten durch unnötige Iterationen. LFL ROI: 3-5× innerhalb 18 Monaten nachweisbar."),
        ("Elektronik- & PCB-Fertigung", "~400 Unternehmen",
         "6-12 Monate", "€80K-€200K ARR",
         "DPP-Pflicht ab 2027 (Batterien, Elektronik) = Compliance-Kaufmotiv. LFL als 'Compliance Engine' mit Sustainability-Reporting-Modul. Höhere Ticket-Größe durch regulatorischen Druck."),
        ("Automotive-Zulieferer (Tier 1/2)", "~300 Unternehmen",
         "9-18 Monate", "€100K-€250K ARR",
         "Größte Ticket-Größen aber längste Zyklen. TISAX-Anforderungen erhöhen Onboarding-Aufwand. Geeignet als Follow-on-Markt nach Packaging-Traction. Referenzkunde aus Packaging öffnet Türen."),
        ("Maschinenbau allgemein (Sonstige)", "~155 Unternehmen",
         "6-12 Monate", "€50K-€120K ARR",
         "Breiter Markt aber heterogene Anforderungen. Best-fit: Unternehmen mit CAD-heavy Produktentwicklung und hohem Materialwert."),
        ("GESAMT DACH TAM", "~2.255 Unternehmen",
         "3-18 Monate", "Ø €100K ARR",
         "TAM bei 100% Penetration: 2.255 × €100K = €225,5M ARR. Realistisches SAM (10% in 5J): 226 Kunden = €22,6M ARR. Modell v0.4 zeigt M52 ARR: €1,9M (8,4% des SAM)."),
    ]

    for i, row in enumerate(market_rows):
        row_bg = GREY_LIGHT if i % 2 == 1 else WHITE
        ws.row_dimensions[4 + i].height = 65
        write_row(ws, 4 + i, [
            cell_style(row[0], bold=True,  bg=row_bg, align="left",   size=10),
            cell_style(row[1], bold=False, bg=row_bg, align="center",  size=10),
            cell_style(row[2], bold=False, bg=row_bg, align="center",  size=10),
            cell_style(row[3], bold=False, bg=row_bg, align="center",  size=10),
            cell_style(row[4], bold=False, bg=row_bg, align="left",    size=9, wrap=True),
        ])

    row_num = 4 + len(market_rows) + 1
    section_header(ws, row_num, "  LEAD-ZYKLEN & VERTRIEBSSTRATEGIE", 5)
    row_num += 1

    write_row(ws, row_num, [
        hdr("Phase"),
        hdr("Dauer"),
        hdr("Aktivität"),
        hdr("Ressource"),
        hdr("Meilenstein / Output"),
    ])
    row_num += 1

    lead_phases = [
        ("Awareness & Outreach",       "M1-M3",
         "LinkedIn-Outreach, Industry-Events (LogiMAT, interpack), White-Paper-Distribution",
         "CCO + KI-SDR-Agent",
         "50+ qualifizierte Leads, 10+ Discovery Calls"),
        ("Discovery & Demo",           "M3-M6",
         "Technische Demo mit Kunden-eigenen CAD-Daten, ROI-Workshop (TCO-Analyse)",
         "CEO/CCO + CTO für Tech-Demo",
         "3-5 aktive POC-Kandidaten"),
        ("Pilot / POC",                "M6-M12",
         "Kostenloser 3-Monats-Pilot bei 2 Packaging-Kunden, ROI-Dokumentation",
         "CTO + 1. ML Engineer + CS",
         "2 zahlende Kunden ab M12-M14, Referenz-Case-Study"),
        ("Land & Expand",              "M12-M24",
         "Seat-Rollout von Pilot-Werk auf weitere Werke, Enterprise-Verhandlung",
         "2. Sales Rep + CS-Playbook",
         "5-10 Enterprise-Seats/Kunde, ACV €75K-€150K"),
        ("Enterprise-Scale",           "M24-M52",
         "Multi-Werk-Rollouts, DPP-Compliance-Modul, Analytics-Premium-Paket",
         "Full Sales Team + CS-Agent",
         "9+ Enterprise-Verträge, ARR €1,9M (M52 Baseline)"),
    ]

    for i, row in enumerate(lead_phases):
        row_bg = GREY_LIGHT if i % 2 == 1 else WHITE
        ws.row_dimensions[row_num].height = 55
        write_row(ws, row_num, [
            cell_style(row[0], bold=True,  bg=row_bg, align="left",   size=10),
            cell_style(row[1], bold=False, bg=AMBER,  align="center",  size=10),
            cell_style(row[2], bold=False, bg=row_bg, align="left",    size=9, wrap=True),
            cell_style(row[3], bold=False, bg=row_bg, align="left",    size=9),
            cell_style(row[4], bold=False, bg=GREEN_LIGHT, align="left", size=9, wrap=True),
        ])
        row_num += 1

    freeze_and_fit(ws, freeze="B4", col_widths={
        "A": 30, "B": 18, "C": 35, "D": 25, "E": 45,
    })
    return ws


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 5 – Optimierungs-Guide
# ══════════════════════════════════════════════════════════════════════════════
def build_optimization_sheet(wb):
    ws = wb.create_sheet("Optimierungs_Guide")
    ws.merge_cells("A1:D1")
    c = ws["A1"]
    c.value = "LFL GmbH – Aktionsplan: Golden Path zur Optimierung des Business Models"
    c.font  = Font(bold=True, size=13, color=WHITE)
    c.fill  = PatternFill("solid", fgColor=BLUE_DARK)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 35

    actions = [
        ("SOFORT-MASSNAHMEN (M1-M3)", "SECTION", GREEN_DARK),
        ("1", "Marktfokus auf Packaging setzen (80%)",
         "Sandbox F4 = 0.8 (Packaging-Anteil stark)",
         "In 00_Input_Sandbox Zeile 5 (Anteil Packaging): Gering=0.1 → Aggressiv=0.8. Sales-Material komplett auf Packaging ausrichten. Keine Automotive-Aktivitäten bis M24. Zielkunden-Liste: Syntegon, Gerhard Schubert, MULTIVAC, Theegarten-Pactec, Harro Höfliger."),
        ("2", "KI-SDR-Agent für Lead-Generation einrichten",
         "Sandbox F10 = 99 (KI-Hebel = Sales Rep nie einstellen)",
         "Claude API für Outbound-Sequences nutzen. Tool: Apollo.io + Make.com + Claude API. Kosten: ~€500/Monat. Targeting: LinkedIn-Scraping von 'Head of Engineering Packaging'. Output: 20 qualifizierte Leads/Woche."),
        ("3", "2 kostenlose Packaging-Piloten vereinbaren",
         "Inputs B19 (Startmonat Kunden) von 14 → 9",
         "Angebot: 3-Monate kostenloser Pilot im Gegenzug für (a) Case-Study-Rechte, (b) Referenz-Call, (c) Paid-Pilot-Option ab M4. Ziel-ARR nach Pilot: €60K-€80K (kleiner Deal). Sales-Pitch: 'Redesign Trap' White Paper als Türöffner."),

        ("KURZFRIST-MASSNAHMEN (M3-M12)", "SECTION", GREEN_DARK),
        ("4", "Enterprise-Pricing auf €200K-ARR anheben",
         "Inputs B26 (Enterprise ARR) von 150.000 → 175.000",
         "DPP-Compliance-Positionierung: Ab 2027 Pflicht für Batterien & Elektronik. LFL als Sustainability-Reporting-Engine → Preisaufschlag €25K-€50K/Vertrag. Argumentationslinie: €200K ARR vs. €500K Compliance-Risk = ROI 2.5×."),
        ("5", "NRR-Playbook entwickeln (Ziel: 125%)",
         "Inputs B28 (NRR) von 1.18 → 1.25",
         "Drei Upsell-Wege: (1) Analytics-Premium-Modul: +€150/Seat/Jahr nach M6 der Nutzung. (2) DPP-Reporting-Addon: €15K Flat/Vertrag/Jahr. (3) Additional-Werks-Rollout: Seat-Verdopplung bei 2. Werk. CS-Playbook: QBR nach M3, M6, M12."),
        ("6", "Onboarding-Aufwand auf 30h reduzieren",
         "Sandbox F9 = 30 (Onboarding stark = 30h)",
         "Maßnahmen: (1) Standard-Connector für gängige PLM-Systeme (PTC Windchill, Siemens Teamcenter). (2) RAG-Wissensagent für Selbst-Onboarding. (3) Video-Tutorial-Bibliothek. Benefit: CS-Kapazität von 3 auf 8 Kunden/Monat skalierbar."),

        ("MITTELFRIST-MASSNAHMEN (M12-M36)", "SECTION", BLUE_MID),
        ("7", "Seed-Runde auf €6-8M ansetzen (M17)",
         "Inputs B14 (Seed-Betrag) ggf. auf 8.000.000",
         "Seed-Kriterien bis M17: (a) 2+ zahlende Enterprise-Kunden, (b) ARR €100K+, (c) klarer Packaging-Referenz-Case. Bei €8M Seed: Runway bis M35 mit mehr Headcount. VC-Pitch-Deck fokus: TAM €225M, NRR 118%+, KI-Differenzierung."),
        ("8", "Enterprise-Sales-Prozess optimieren (2 Deals/Quartal)",
         "Inputs B25 (Enterprise Deals/Quartal) von 1 → 1.5",
         "Ab M18: 2. Sales Rep einstellen (Inputs). Sales-Cycle-Optimierung: (1) Technical Validator in Deal-Team (CTO nach Demo), (2) ROI-Calculator als Self-Service, (3) Referenz-Kunden-Network für Proof-Points. Ziel: M30 = 6 aktive Enterprise-Verträge."),
        ("9", "Churn-Reduktions-Programm starten (Ziel: 5%)",
         "Inputs B27 (Churn) von 0.08 → 0.06",
         "Maßnahmen: (1) API-Tiefenintegration in Kunden-PLM-System (Switching Cost hoch), (2) Dedicated CSM für alle Enterprise-Kunden ab Vertrag, (3) SLA-Garantie 99.5% Uptime + Response <2h. Monitoring: Monthly Health Score (Nutzungsfrequenz)."),

        ("LANGFRIST-MASSNAHMEN (M36-M52)", "SECTION", AMBER_DARK),
        ("10", "Series A für internationales Scaling vorbereiten",
         "Inputs B15 (Series A Betrag) ggf. 20.000.000; B16 Monat 32-35",
         "Series-A-Readiness-Kriterien M35: (a) ARR €1M+, (b) NRR >115%, (c) 3+ Länder-Kunden, (d) EBITDA-Pfad erkennbar. €15-20M für: Internationaler Sales (BENELUX, Nordics), Zweites Produkt-Modul, Series-A-Hiring 5 FTE."),
        ("11", "Finale Optimierung der Input-Parameter im Modell",
         "Alle Inputs finalisieren nach M36-Daten",
         "Nach 36 Monaten reale Daten vs. Modell abgleichen. Inputs anpassen: (1) Tatsächliche Enterprise-ACV aus Verträgen in B26, (2) Gemessene Churn-Rate in B27, (3) Tatsächliche NRR in B28, (4) Sales-Zyklen aus CRM in B24. Modell als 'Single Source of Truth' für Series A Deck."),
    ]

    write_row(ws, 2, [
        hdr("#"),
        hdr("Maßnahme"),
        hdr("Excel-Aktion (Zelle ändern)"),
        hdr("Detaillierter Aktionsplan"),
    ])

    row_num = 3
    for item in actions:
        if item[1] == "SECTION":
            section_header(ws, row_num, f"  {item[0]}", 4, bg=item[2])
            ws.row_dimensions[row_num].height = 22
            row_num += 1
            continue

        num, title, excel_action, detail = item
        ws.row_dimensions[row_num].height = 80
        write_row(ws, row_num, [
            cell_style(num,           bold=True,  bg=AMBER, align="center", size=12),
            cell_style(title,         bold=True,  bg=WHITE, align="left",   size=11),
            cell_style(excel_action,  bold=False, bg=BLUE_LIGHT, align="left", size=9, wrap=True),
            cell_style(detail,        bold=False, bg=WHITE, align="left",   size=9, wrap=True),
        ])
        row_num += 1

    # Summary table
    row_num += 1
    section_header(ws, row_num, "  BREAK-EVEN-OPTIMIERUNGS-SZENARIEN: Welche Parameterkombo bricht Break-Even vor M52?", 4)
    row_num += 1

    write_row(ws, row_num, [
        hdr("Szenario"),
        hdr("Schlüsselparameter-Änderungen"),
        hdr("Projizierter Break-Even"),
        hdr("ARR bei Break-Even"),
    ])
    row_num += 1

    scenarios = [
        ("Baseline (gering)", "Keine Änderungen (Modell v0.4 Standard)", "M65+ (nach Horizont)", "€2.3M"),
        ("Optimiert-Konservativ",
         "Packaging 60% + Startmonat M11 + Enterprise-ACV €175K + NRR 120%",
         "M48-52", "€2.1M"),
        ("Optimiert-Aggressiv (Golden Path)",
         "Packaging 80% + Startmonat M9 + Enterprise-ACV €200K + 2 Deals/Q + NRR 125% + KI-Hebel M48",
         "M38-42", "€1.7M"),
        ("Maximal-Aggressiv",
         "Alle Parameter auf Maximal + Series A M30 + 3 Deals/Quartal",
         "M32-36", "€1.4M"),
    ]

    for i, row in enumerate(scenarios):
        row_bg = GREY_LIGHT if i % 2 == 1 else WHITE
        ws.row_dimensions[row_num].height = 40
        write_row(ws, row_num, [
            cell_style(row[0], bold=True,  bg=row_bg, align="left",   size=10),
            cell_style(row[1], bold=False, bg=row_bg, align="left",    size=9, wrap=True),
            cell_style(row[2], bold=True,  bg=GREEN_LIGHT if "Golden" in row[0] else row_bg, align="center", size=10),
            cell_style(row[3], bold=False, bg=row_bg, align="center",  size=10),
        ])
        row_num += 1

    freeze_and_fit(ws, freeze="B3", col_widths={
        "A": 8, "B": 35, "C": 35, "D": 60,
    })
    return ws


# ══════════════════════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════════════════════
def main():
    os.makedirs(os.path.dirname(OUT_PATH), exist_ok=True)

    wb = openpyxl.Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    print("Erstelle Sheet 1: Parameter_Vergleich ...")
    build_parameter_sheet(wb)

    print("Erstelle Sheet 2: Hebel_Analyse ...")
    build_lever_sheet(wb)

    print("Erstelle Sheet 3: KPI_Uebersicht ...")
    build_kpi_sheet(wb)

    print("Erstelle Sheet 4: Marktkontext ...")
    build_market_sheet(wb)

    print("Erstelle Sheet 5: Optimierungs_Guide ...")
    build_optimization_sheet(wb)

    wb.save(OUT_PATH)
    print(f"\n✅ Fertig: {OUT_PATH}")
    print(f"   Sheets: {', '.join(wb.sheetnames)}")

if __name__ == "__main__":
    main()
