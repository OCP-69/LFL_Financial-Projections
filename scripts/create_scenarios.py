#!/usr/bin/env python3
"""
Erstellt zwei Szenario-Dateien aus v15.
Szenario 1: Gering
Szenario 2: Normal
"""

import openpyxl
from openpyxl.utils import get_column_letter
import shutil

SRC_V15 = '/home/user/LFL_Financial-Projections/260307_LFL_BM_Vorlage_v15.xlsx'

# ================================================================
# Helper
# ================================================================
def col_for_month(m):
    """Month m (1-52) -> Column letter (B=M1, BA=M52)"""
    return get_column_letter(m + 1)


def fill_manual_revenue_rows(ws_rev, sme_values, midco_values, enterprise_values):
    """
    Fill Revenue rows 6 (SME), 7 (MidCo), 15 (Enterprise) with manual values.
    Values dicts: {month_number: value}
    Missing months = 0
    """
    for m in range(1, 53):
        col = col_for_month(m)
        ws_rev[f"{col}6"] = sme_values.get(m, 0)
        ws_rev[f"{col}7"] = midco_values.get(m, 0)
        ws_rev[f"{col}15"] = enterprise_values.get(m, 0)


# ================================================================
# Szenario 1: Gering
# ================================================================
print("=" * 60)
print("SZENARIO 1: Gering")
print("=" * 60)

dst_gering = '/home/user/LFL_Financial-Projections/LFL_BM_Szenario_Gering_20260311_1000.xlsx'
shutil.copy2(SRC_V15, dst_gering)

wb_g = openpyxl.load_workbook(dst_gering, data_only=False)
ws_sandbox_g = wb_g['00_Input_Sandbox']
ws_rev_g = wb_g['Revenue']

# Setze Szenario
ws_sandbox_g['B1'] = 'gering'
print(f"  Sandbox!B1 = 'gering'")

# SME: Monate 1-6=0, 7-12=1, 13-24=2, 25-36=3, 37-52=4
sme_gering = {}
for m in range(1, 53):
    if m <= 6:
        sme_gering[m] = 0
    elif m <= 12:
        sme_gering[m] = 1
    elif m <= 24:
        sme_gering[m] = 2
    elif m <= 36:
        sme_gering[m] = 3
    else:
        sme_gering[m] = 4

# MidCo: Monate 1-12=0, 13-24=1, 25-36=1, 37-52=2
midco_gering = {}
for m in range(1, 53):
    if m <= 12:
        midco_gering[m] = 0
    elif m <= 24:
        midco_gering[m] = 1
    elif m <= 36:
        midco_gering[m] = 1
    else:
        midco_gering[m] = 2

# Enterprise: Monate 1-23=0, dann alle 6 Monate 1 Deal
enterprise_gering = {}
for m in range(1, 53):
    if m >= 24 and (m - 24) % 6 == 0:
        enterprise_gering[m] = 1
    else:
        enterprise_gering[m] = 0

fill_manual_revenue_rows(ws_rev_g, sme_gering, midco_gering, enterprise_gering)

# Print summary
total_sme = sum(sme_gering.values())
total_midco = sum(midco_gering.values())
total_enterprise = sum(enterprise_gering.values())
print(f"  SME Kunden gesamt (52 Monate): {total_sme}")
print(f"  MidCo Kunden gesamt (52 Monate): {total_midco}")
print(f"  Enterprise Deals gesamt (52 Monate): {total_enterprise}")
print(f"  Enterprise Monate: {sorted([m for m, v in enterprise_gering.items() if v > 0])}")

wb_g.save(dst_gering)
print(f"  Gespeichert: {dst_gering}")

# ================================================================
# Szenario 2: Normal
# ================================================================
print()
print("=" * 60)
print("SZENARIO 2: Normal")
print("=" * 60)

dst_normal = '/home/user/LFL_Financial-Projections/LFL_BM_Szenario_Normal_20260311_1000.xlsx'
shutil.copy2(SRC_V15, dst_normal)

wb_n = openpyxl.load_workbook(dst_normal, data_only=False)
ws_sandbox_n = wb_n['00_Input_Sandbox']
ws_rev_n = wb_n['Revenue']

# Setze Szenario
ws_sandbox_n['B1'] = 'normal'
print(f"  Sandbox!B1 = 'normal'")

# SME: Monate 1-6=0, 7-12=2, 13-24=4, 25-36=6, 37-52=8
sme_normal = {}
for m in range(1, 53):
    if m <= 6:
        sme_normal[m] = 0
    elif m <= 12:
        sme_normal[m] = 2
    elif m <= 24:
        sme_normal[m] = 4
    elif m <= 36:
        sme_normal[m] = 6
    else:
        sme_normal[m] = 8

# MidCo: Monate 1-9=0, 10-24=1, 25-36=2, 37-52=3
midco_normal = {}
for m in range(1, 53):
    if m <= 9:
        midco_normal[m] = 0
    elif m <= 24:
        midco_normal[m] = 1
    elif m <= 36:
        midco_normal[m] = 2
    else:
        midco_normal[m] = 3

# Enterprise: Monate 1-17=0, dann alle 4 Monate 1 Deal (ab Series A = Monat 36)
# Spez: ab Monat 18, alle 4 Monate
enterprise_normal = {}
for m in range(1, 53):
    if m >= 18 and (m - 18) % 4 == 0:
        enterprise_normal[m] = 1
    else:
        enterprise_normal[m] = 0

fill_manual_revenue_rows(ws_rev_n, sme_normal, midco_normal, enterprise_normal)

# Print summary
total_sme_n = sum(sme_normal.values())
total_midco_n = sum(midco_normal.values())
total_enterprise_n = sum(enterprise_normal.values())
print(f"  SME Kunden gesamt (52 Monate): {total_sme_n}")
print(f"  MidCo Kunden gesamt (52 Monate): {total_midco_n}")
print(f"  Enterprise Deals gesamt (52 Monate): {total_enterprise_n}")
print(f"  Enterprise Monate: {sorted([m for m, v in enterprise_normal.items() if v > 0])}")

wb_n.save(dst_normal)
print(f"  Gespeichert: {dst_normal}")

print()
print("=" * 60)
print("BEIDE SZENARIEN ERSTELLT")
print("=" * 60)
