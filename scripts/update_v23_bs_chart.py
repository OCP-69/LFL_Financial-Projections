"""
Update LFL_BM_C13_Normal_redacted_v23.xlsx:
  1. Fill BalanceSheet with year-end values (2026–2030)
  2. Add Revenue_Profit_Cash_Chart sheet (phase-based, starting at Pre-Seed)

Output: LFL_BM_C13_Normal_redacted_v24_20260315.xlsx
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.ticker as mticker
from matplotlib.patches import FancyBboxPatch
import numpy as np
from io import BytesIO
from datetime import datetime
import shutil, os

SRC_V23 = 'LFL_BM_C13_Normal_redacted_v23.xlsx'
OUT     = 'LFL_BM_C13_Normal_redacted_v24_20260315.xlsx'

# ── Styles ──────────────────────────────────────────────────────────────────
DARK_BLUE  = '1F3864'
MID_BLUE   = '2E75B6'
LIGHT_BLUE = 'D6E4F0'
WHITE      = 'FFFFFF'
LIGHT_GREY = 'F2F2F2'

def fill(hex_color):
    return PatternFill('solid', fgColor=hex_color)

def border_thin():
    s = Side(style='thin', color='BFBFBF')
    return Border(left=s, right=s, top=s, bottom=s)

FMT_EUR = '#,##0.00'
FMT_INT = '#,##0'

# ── Load file ────────────────────────────────────────────────────────────────
print(f'Copying {SRC_V23} → {OUT}')
shutil.copy2(SRC_V23, OUT)
wb = openpyxl.load_workbook(OUT)

# ── Read Monthly data (source for all calculations) ──────────────────────────
ws_mo = wb['Monthly']

# Collect all 48 data rows (rows 2–49)
months = []
for r in range(2, 50):
    yr   = ws_mo.cell(r, 1).value
    m    = ws_mo.cell(r, 2).value
    if yr is None or m is None:
        continue
    rev  = float(ws_mo.cell(r, 3).value  or 0)
    cor  = float(ws_mo.cell(r, 6).value  or 0)
    opex = float(ws_mo.cell(r, 7).value  or 0)
    tax  = float(ws_mo.cell(r, 9).value  or 0)
    eq   = float(ws_mo.cell(r, 5).value  or 0)   # Other Cash In (equity)
    beg  = float(ws_mo.cell(r, 11).value or 0)
    end  = float(ws_mo.cell(r, 12).value or 0)
    deb  = float(ws_mo.cell(r, 14).value or 0)
    cred = float(ws_mo.cell(r, 15).value or 0)
    net_profit = rev - cor - opex - tax
    months.append({
        'row': r, 'year': yr, 'month': m,
        'rev': rev, 'cor': cor, 'opex': opex, 'tax': tax,
        'equity': eq, 'beg': beg, 'end': end,
        'debtors': deb, 'creditors': cred,
        'net_profit': net_profit,
    })

# ── 1. Fill BalanceSheet ─────────────────────────────────────────────────────
print('Filling BalanceSheet...')
ws_bs = wb['BalanceSheet']

# Headers with styling
bs_headers = {
    1: 'Year',
    2: 'Cash (year-end)',
    3: 'Debtors',
    4: 'R&D Intangible',
    5: 'Creditors',
    6: 'Debt',
    7: 'Net Assets',
    8: 'Total Equity Investment',
    9: 'Cumulative Net Profit',
    10: 'Quick Ratio',
}
for c, label in bs_headers.items():
    cell = ws_bs.cell(1, c)
    cell.value = label
    cell.font  = Font(name='Calibri', bold=True, color=WHITE, size=9)
    cell.fill  = fill(DARK_BLUE)
    cell.alignment = Alignment(horizontal='center', wrap_text=True)
    cell.border = border_thin()

ws_bs.row_dimensions[1].height = 30
ws_bs.column_dimensions['A'].width = 10
for c in range(2, 11):
    ws_bs.column_dimensions[get_column_letter(c)].width = 18

# Build year-groups: last row per calendar year
year_end_rows = {}
for d in months:
    y = d['year']
    year_end_rows[y] = d   # last one wins → year-end

# Compute cumulative equity & cumulative net profit per year
years_sorted = sorted(year_end_rows.keys())
cum_equity = 0.0
cum_profit = 0.0

for yr_idx, year in enumerate(years_sorted):
    # All months in this year
    yr_months = [d for d in months if d['year'] == year]
    cum_equity += sum(d['equity']     for d in yr_months)
    cum_profit += sum(d['net_profit'] for d in yr_months)
    d_end = year_end_rows[year]

    cash      = d_end['end']
    debtors   = d_end['debtors']
    creditors = d_end['creditors']
    net_assets = cash + debtors - creditors
    quick_ratio = (cash / creditors) if creditors > 0 else 0

    row = yr_idx + 2
    # Add row if needed (template only has 2026–2029)
    bg = LIGHT_BLUE if yr_idx % 2 == 0 else WHITE

    row_data = [
        year, cash, debtors, 0.0, creditors, 0.0,
        net_assets, cum_equity, cum_profit, quick_ratio,
    ]
    for c_idx, val in enumerate(row_data, 1):
        cell = ws_bs.cell(row, c_idx)
        cell.value = val
        cell.fill  = fill(bg)
        cell.border = border_thin()
        if c_idx == 1:
            cell.font = Font(name='Calibri', bold=True, size=9)
            cell.alignment = Alignment(horizontal='center')
            cell.number_format = '0'
        elif c_idx == 10:   # Quick Ratio
            cell.font = Font(name='Calibri', size=9)
            cell.alignment = Alignment(horizontal='right')
            cell.number_format = '0.0'
        else:
            cell.font = Font(name='Calibri', size=9)
            cell.alignment = Alignment(horizontal='right')
            cell.number_format = FMT_EUR

    # Color net assets negative in red
    na_cell = ws_bs.cell(row, 7)
    if net_assets < 0:
        na_cell.font = Font(name='Calibri', size=9, color='C00000', bold=True)
    # Color cumulative profit
    cp_cell = ws_bs.cell(row, 9)
    if cum_profit < 0:
        cp_cell.font = Font(name='Calibri', size=9, color='C00000', bold=True)
    else:
        cp_cell.font = Font(name='Calibri', size=9, color='375623', bold=True)

# Note row
note_row = len(years_sorted) + 3
ws_bs.cell(note_row, 1).value = (
    'Note: All values are year-end. Counting starts at Pre-Seed Phase '
    '(Source M5 = Target Month 1, Aug 2026). 2026 = 5 months only (Aug–Dec).'
)
ws_bs.cell(note_row, 1).font = Font(name='Calibri', size=8, italic=True, color='595959')
ws_bs.merge_cells(f'A{note_row}:J{note_row}')

print(f'  BalanceSheet: {len(years_sorted)} years filled (2026–{max(years_sorted)})')

# ── 2. Build Revenue_Profit_Cash_Chart ───────────────────────────────────────
print('Building Revenue_Profit_Cash_Chart...')

# Phase definitions (target month ranges, 1-based)
PHASES = [
    {'name': 'Pre-Seed', 'label': 'Pre-Seed\n(M1–M12)', 'src': 'M5–M16',  'm_s': 1,  'm_e': 12},
    {'name': 'Seed',     'label': 'Seed\n(M13–M24)',    'src': 'M17–M28', 'm_s': 13, 'm_e': 24},
    {'name': 'Series A', 'label': 'Series A\n(M25–M36)','src': 'M29–M40', 'm_s': 25, 'm_e': 36},
    {'name': 'Series B', 'label': 'Series B\n(M37–M48)','src': 'M41–M52', 'm_s': 37, 'm_e': 48},
]

phase_data = []
for ph in PHASES:
    ph_months = [d for d in months if ph['m_s'] <= d['month'] <= ph['m_e']]
    rev_total  = sum(d['rev']        for d in ph_months)
    net_total  = sum(d['net_profit'] for d in ph_months)
    end_cash   = ph_months[-1]['end'] if ph_months else 0
    phase_data.append({
        'name':     ph['name'],
        'label':    ph['label'],
        'src':      ph['src'],
        'revenue':  rev_total,
        'net':      net_total,
        'cash':     end_cash,
        'm_range':  f"M{ph['m_s']}–M{ph['m_e']}",
    })
    print(f"  {ph['name']}: Rev={rev_total:,.0f}  Net={net_total:,.0f}  EndCash={end_cash:,.0f}")

# ── Chart sheet ──────────────────────────────────────────────────────────────
if 'Revenue_Profit_Cash_Chart' in wb.sheetnames:
    del wb['Revenue_Profit_Cash_Chart']
ws_chart = wb.create_sheet('Revenue_Profit_Cash_Chart', 0)

# Header
ws_chart.merge_cells('A1:F1')
ws_chart['A1'] = 'Revenue vs Net Profit vs Year-end Cash — Normal-Szenario (Pre-Seed-Start)'
ws_chart['A1'].font  = Font(name='Calibri', bold=True, size=13, color=WHITE)
ws_chart['A1'].fill  = fill(DARK_BLUE)
ws_chart['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws_chart.row_dimensions[1].height = 28

ws_chart.merge_cells('A2:F2')
ws_chart['A2'] = (
    f'Quelle: 260315_LFL_BM_Vorlage_normal_redacted_final.xlsx  |  '
    f'Erstellt: {datetime.now().strftime("%d.%m.%Y %H:%M")}  |  '
    f'Monat 1 = Pre-Seed-Start (Quelle M5, Aug 2026)'
)
ws_chart['A2'].font      = Font(name='Calibri', size=9, italic=True, color='595959')
ws_chart['A2'].alignment = Alignment(horizontal='center')
ws_chart.row_dimensions[2].height = 16

# Table header R4
table_headers = ['Phase', 'Ziel-Monate', 'Quell-Monate', 'Revenue (€)', 'Net Profit (€)', 'Phase-End Cash (€)']
for c_idx, hdr in enumerate(table_headers, 1):
    cell = ws_chart.cell(4, c_idx)
    cell.value     = hdr
    cell.font      = Font(name='Calibri', bold=True, color=WHITE, size=9)
    cell.fill      = fill(MID_BLUE)
    cell.alignment = Alignment(horizontal='center')
    cell.border    = border_thin()

PHASE_COLORS = ['9DC3E6', 'A9D18E', 'FFD966', 'F4B183']

# Table data rows R5–R8
for i, pd_ in enumerate(phase_data):
    row = i + 5
    bg = PHASE_COLORS[i]
    row_vals = [pd_['name'], pd_['m_range'], pd_['src'],
                pd_['revenue'], pd_['net'], pd_['cash']]
    for c_idx, val in enumerate(row_vals, 1):
        cell = ws_chart.cell(row, c_idx)
        cell.value  = val
        cell.fill   = fill(bg)
        cell.border = border_thin()
        if c_idx <= 3:
            cell.font      = Font(name='Calibri', bold=(c_idx == 1), size=9)
            cell.alignment = Alignment(horizontal='center')
        else:
            cell.font          = Font(name='Calibri', size=9,
                                      color=('C00000' if val < 0 else '000000'))
            cell.alignment     = Alignment(horizontal='right')
            cell.number_format = '#,##0'

# Column widths
widths = [14, 14, 14, 20, 20, 20]
for c_idx, w in enumerate(widths, 1):
    ws_chart.column_dimensions[get_column_letter(c_idx)].width = w

# ── Matplotlib chart ──────────────────────────────────────────────────────────
fig, ax = plt.subplots(figsize=(13, 7))
fig.patch.set_facecolor('#FAFAFA')
ax.set_facecolor('#F8F9FA')

x_labels = [pd_['label'] for pd_ in phase_data]
revenues  = [pd_['revenue']  for pd_ in phase_data]
nets      = [pd_['net']      for pd_ in phase_data]
cashes    = [pd_['cash']     for pd_ in phase_data]
x         = np.arange(len(x_labels))

# Phase background bands
phase_bg = ['#EBF3FB', '#EBF7E4', '#FFF9E6', '#FDF3EA']
for xi, bg in enumerate(phase_bg):
    ax.axvspan(xi - 0.4, xi + 0.4, alpha=0.25, color=bg, zorder=0)

# Lines
lw = 2.5
ax.plot(x, revenues, 'o-',  color='#2E75B6', linewidth=lw, markersize=8,
        label='Revenue', zorder=5)
ax.plot(x, nets,     's-',  color='#C00000', linewidth=lw, markersize=8,
        label='Net Profit', zorder=5)
ax.plot(x, cashes,   '^--', color='#ED7D31', linewidth=lw, markersize=8,
        label='Phase-End Cash', zorder=5)

# Break-even line at 0
ax.axhline(0, color='#70AD47', linestyle='--', linewidth=1.5, alpha=0.8, label='Break-even (0)')

# Value labels on data points
def fmt_val(v):
    if abs(v) >= 1e6:
        return f'€{v/1e6:.1f}M'
    elif abs(v) >= 1e3:
        return f'€{v/1e3:.0f}K'
    else:
        return f'€{v:.0f}'

for xi, (r, n, c) in enumerate(zip(revenues, nets, cashes)):
    offset_rev = 0.06 * max(cashes)
    ax.annotate(fmt_val(r), (xi, r), textcoords='offset points',
                xytext=(0, 10), ha='center', fontsize=8.5,
                color='#2E75B6', fontweight='bold')
    ax.annotate(fmt_val(n), (xi, n), textcoords='offset points',
                xytext=(0, -16), ha='center', fontsize=8.5,
                color='#C00000', fontweight='bold')
    ax.annotate(fmt_val(c), (xi, c), textcoords='offset points',
                xytext=(0, 10), ha='center', fontsize=8.5,
                color='#ED7D31', fontweight='bold')

ax.set_xticks(x)
ax.set_xticklabels(x_labels, fontsize=10, fontweight='bold')
ax.yaxis.set_major_formatter(mticker.FuncFormatter(
    lambda v, _: f'€{v/1e6:.1f}M' if abs(v) >= 1e6 else f'€{v/1e3:.0f}K'
))
ax.set_ylabel('EUR', fontsize=10)
ax.set_title(
    'Revenue vs Net Profit vs Phase-End Cash\n'
    'Normal-Szenario | Zählung ab Pre-Seed (Monat 1 = Quelle M5, Aug 2026)',
    fontsize=12, fontweight='bold', color='#1F3864', pad=14
)
ax.legend(loc='upper left', fontsize=9, framealpha=0.85)
ax.grid(axis='y', alpha=0.4, linestyle='--')
ax.spines['top'].set_visible(False)
ax.spines['right'].set_visible(False)
ax.spines['left'].set_color('#CCCCCC')
ax.spines['bottom'].set_color('#CCCCCC')

# Phase labels inside chart
for xi, ph in enumerate(phase_data):
    ax.text(xi, ax.get_ylim()[0] * 0.85 if ax.get_ylim()[0] < 0 else -0.05 * max(cashes),
            ph['name'], ha='center', fontsize=8, color='#595959', style='italic')

plt.tight_layout(pad=2.0)

# Save to BytesIO
buf = BytesIO()
plt.savefig(buf, format='png', dpi=150, bbox_inches='tight',
            facecolor=fig.get_facecolor())
plt.close(fig)
buf.seek(0)

# Embed in chart sheet
img = XLImage(buf)
img.anchor = 'A10'
img.width  = 900
img.height = 480
ws_chart.add_image(img)

# ── Save ─────────────────────────────────────────────────────────────────────
print(f'Saving: {OUT}')
wb.save(OUT)
size_kb = os.path.getsize(OUT) / 1024
print(f'✓ Fertig: {OUT}  ({size_kb:.0f} KB)')
print()
print('=== BalanceSheet Summary ===')
for yr in years_sorted:
    yr_months = [d for d in months if d['year'] == yr]
    d_end     = year_end_rows[yr]
    print(f'  {yr}: Cash={d_end["end"]:>14,.0f}  Debtors={d_end["debtors"]:>10,.0f}  '
          f'Creditors={d_end["creditors"]:>8,.0f}  Net Assets={d_end["end"]+d_end["debtors"]-d_end["creditors"]:>14,.0f}')
print()
print('=== Chart Phase Aggregation ===')
for pd_ in phase_data:
    print(f'  {pd_["name"]:10s}: Rev={pd_["revenue"]:>12,.0f}  Net={pd_["net"]:>12,.0f}  '
          f'EndCash={pd_["cash"]:>14,.0f}')
