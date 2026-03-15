"""
Data Merge Script v2 – Phasen-basierte Aggregation
Quelle: 260315_LFL_BM_Vorlage_normal_v19.xlsx
Ziel:   LFL_BM_C13_Normal_v21_20260315.xlsx

Korrekte Phasen-Abgrenzung (gemäß 2_Inputs / Startup-Phasen-Zeile in P&L):
  Ideation : M1–M4   (4 Monate)
  Pre-Seed : M5–M16  (12 Monate)
  Seed     : M17–M28 (12 Monate)
  Series A : M29–M40 (12 Monate)
  Series B : M41–M52 (12 Monate)

Monat→Spalte: M1=B(2), M4=E(5), M5=F(6), M16=Q(17),
              M17=R(18), M28=AC(29), M29=AD(30), M40=AO(41),
              M41=AP(42), M52=BA(53)
"""

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.ticker as mticker
import io, os
from datetime import datetime
from copy import deepcopy

BASE_DIR   = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SOURCE     = os.path.join(BASE_DIR, '260315_LFL_BM_Vorlage_normal_v19.xlsx')
TEMPLATE   = os.path.join(BASE_DIR, 'C13_Template_financial_projections_neu.xlsx')
OUTPUT     = os.path.join(BASE_DIR, 'LFL_BM_C13_Normal_v21_20260315.xlsx')

# ── Phasen-Definitionen ────────────────────────────────────────────────────────
# idx = 0-basierter Monatsindex (M1=0, M52=51)
# col = Spaltenindex in Quelldatei (M1=col2, M52=col53)
PHASES = [
    {'name': 'Ideation',  'label': 'Ideation\n(M1–M4)',   'idx_s':  0, 'idx_e':  4},
    {'name': 'Pre-Seed',  'label': 'Pre-Seed\n(M5–M16)',  'idx_s':  4, 'idx_e': 16},
    {'name': 'Seed',      'label': 'Seed\n(M17–M28)',     'idx_s': 16, 'idx_e': 28},
    {'name': 'Series A',  'label': 'Series A\n(M29–M40)', 'idx_s': 28, 'idx_e': 40},
    {'name': 'Series B',  'label': 'Series B\n(M41–M52)', 'idx_s': 40, 'idx_e': 52},
]

def idx_to_col(idx):
    """0-basierter Monatsindex → Spaltenindex in Quelldatei (M1=idx0=col2)"""
    return idx + 2

def read_row(ws, row_num, num_months=52):
    return [float(ws.cell(row=row_num, column=2+i).value or 0) for i in range(num_months)]

# ── Source laden ───────────────────────────────────────────────────────────────
print(f"Lade Quelldatei: {os.path.basename(SOURCE)}")
wb_src = openpyxl.load_workbook(SOURCE, data_only=True)
ws_rev  = wb_src['4_Revenue']
ws_pl   = wb_src['6_P&L']
ws_cf   = wb_src['7_BS_CF']

total_revenue  = read_row(ws_rev, 32)
total_cogs     = read_row(ws_pl,  14)
total_opex     = read_row(ws_pl,  27)
net_income     = read_row(ws_pl,  37)
income_tax     = read_row(ws_pl,  35)
ebitda_vals    = read_row(ws_pl,  29)
equity_funding = read_row(ws_cf,   9)
beginning_cash = read_row(ws_cf,  14)
ending_cash    = read_row(ws_cf,  15)

debtors   = list(total_revenue)
creditors = [(total_cogs[i] + total_opex[i]) * 0.10 for i in range(52)]

# Kumulative Werte
cum_equity = []
cum_profit = []
s = 0.0
for v in equity_funding:
    s += v; cum_equity.append(s)
s = 0.0
for v in net_income:
    s += v; cum_profit.append(s)

# ── Phasen-Aggregate berechnen ─────────────────────────────────────────────────
print("\nPhasen-Aggregate:")
print(f"  {'Phase':12s} | {'Monate':8s} | {'Spalten':10s} | {'Revenue':>14} | {'Net Profit':>14} | {'Year-end Cash':>14}")
print(f"  {'-'*12}-+-{'-'*8}-+-{'-'*10}-+-{'-'*14}-+-{'-'*14}-+-{'-'*14}")

phase_data = []
for ph in PHASES:
    s_idx, e_idx = ph['idx_s'], ph['idx_e']
    n = e_idx - s_idx
    col_s = idx_to_col(s_idx)
    col_e = idx_to_col(e_idx - 1)   # letzter Monat der Phase

    rev  = sum(total_revenue[s_idx:e_idx])
    cogs = sum(total_cogs[s_idx:e_idx])
    opex = sum(total_opex[s_idx:e_idx])
    gp   = rev - cogs
    ebit = gp - opex
    tax  = sum(income_tax[s_idx:e_idx])
    ni   = sum(net_income[s_idx:e_idx])
    eq   = sum(equity_funding[s_idx:e_idx])

    cash_end  = ending_cash[e_idx - 1]
    deb_end   = debtors[e_idx - 1]
    cred_end  = creditors[e_idx - 1]
    net_ass   = cash_end + deb_end - cred_end
    qr        = (cash_end / cred_end) if cred_end > 0 else 0.0
    min_cash  = min(ending_cash[s_idx:e_idx])

    gross_burns = [total_cogs[i] + total_opex[i] for i in range(s_idx, e_idx)]
    avg_gb = sum(gross_burns) / n

    ebit_loss = [ebitda_vals[i] for i in range(s_idx, e_idx) if ebitda_vals[i] < 0]
    avg_eb = sum(ebit_loss) / len(ebit_loss) if ebit_loss else 0.0

    cum_eq_end = cum_equity[e_idx - 1]
    cum_pr_end = cum_profit[e_idx - 1]

    phase_data.append({
        'phase': ph, 'n': n,
        'col_s': col_s, 'col_e': col_e,
        'rev': rev, 'cogs': cogs, 'opex': opex, 'gp': gp,
        'ebit': ebit, 'tax': tax, 'ni': ni, 'eq': eq,
        'cash_end': cash_end, 'deb_end': deb_end, 'cred_end': cred_end,
        'net_ass': net_ass, 'qr': qr, 'min_cash': min_cash,
        'avg_gb': avg_gb, 'avg_eb': avg_eb,
        'cum_eq': cum_eq_end, 'cum_pr': cum_pr_end,
    })

    col_range = f"{get_column_letter(col_s)}–{get_column_letter(col_e)}"
    print(f"  {ph['name']:12s} | M{s_idx+1:2d}–M{e_idx:2d}  | {col_range:10s} | "
          f"{rev:>14,.0f} | {ni:>14,.0f} | {cash_end:>14,.0f}")

# ── Template laden ─────────────────────────────────────────────────────────────
print(f"\nLade Template: {os.path.basename(TEMPLATE)}")
wb_dst = openpyxl.load_workbook(TEMPLATE, data_only=False)

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 1: Monthly – 52 Monate + Phase-Label in Spalte A
# ══════════════════════════════════════════════════════════════════════════════
print("Befülle Monthly-Sheet …")
ws_m = wb_dst['Monthly']

# Phase-Label und Monatsnummer je Monat
phase_for_month = {}
month_num_in_phase = {}
for ph in PHASES:
    for idx in range(ph['idx_s'], ph['idx_e']):
        phase_for_month[idx] = ph['name']
        month_num_in_phase[idx] = idx - ph['idx_s'] + 1

# Zeile 1 = Header (schon vorhanden)
# Leere Monate Jan–Mär 2026: Rows 2–4 → Nullen
for row in range(2, 5):
    ws_m.cell(row=row, column=1).value = '2026 (Jan–Mär)'
    ws_m.cell(row=row, column=2).value = row - 1   # Jan=1, Feb=2, Mär=3
    for col in range(3, 16):
        ws_m.cell(row=row, column=col).value = 0.0

# M1–M52: Rows 5–56
for m_idx in range(52):
    row = 5 + m_idx
    ws_m.cell(row=row, column=1).value  = phase_for_month[m_idx]        # A: Phase
    ws_m.cell(row=row, column=2).value  = month_num_in_phase[m_idx]      # B: Monat in Phase
    ws_m.cell(row=row, column=3).value  = round(total_revenue[m_idx], 2)
    ws_m.cell(row=row, column=4).value  = round(total_revenue[m_idx], 2)
    ws_m.cell(row=row, column=5).value  = round(equity_funding[m_idx], 2)
    ws_m.cell(row=row, column=6).value  = round(total_cogs[m_idx], 2)
    ws_m.cell(row=row, column=7).value  = round(total_opex[m_idx], 2)
    ws_m.cell(row=row, column=8).value  = 0.0
    ws_m.cell(row=row, column=9).value  = round(income_tax[m_idx], 2)
    ws_m.cell(row=row, column=10).value = 0.0
    ws_m.cell(row=row, column=11).value = round(beginning_cash[m_idx], 2)
    ws_m.cell(row=row, column=12).value = round(ending_cash[m_idx], 2)
    ws_m.cell(row=row, column=13).value = 0.0
    ws_m.cell(row=row, column=14).value = round(debtors[m_idx], 2)
    ws_m.cell(row=row, column=15).value = round(creditors[m_idx], 2)

print(f"  → 52 Monate eingetragen (Rows 5–56)")

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 2: Annual → Phasen-Aggregation
# ══════════════════════════════════════════════════════════════════════════════
print("Befülle Annual-Sheet (Phasen-Aggregation) …")
ws_a = wb_dst['Annual']

# Header Spalte A auf "Phase" umbenennen
ws_a.cell(row=1, column=1).value = 'Phase'

for i, pd_ in enumerate(phase_data):
    row = 2 + i
    ws_a.cell(row=row, column=1).value  = pd_['phase']['name']
    ws_a.cell(row=row, column=2).value  = round(pd_['rev'],   2)
    ws_a.cell(row=row, column=3).value  = round(pd_['cogs'],  2)
    ws_a.cell(row=row, column=4).value  = round(pd_['gp'],    2)
    ws_a.cell(row=row, column=5).value  = round(pd_['opex'],  2)
    ws_a.cell(row=row, column=6).value  = round(pd_['ebit'],  2)
    ws_a.cell(row=row, column=7).value  = 0.0
    ws_a.cell(row=row, column=8).value  = round(pd_['tax'],   2)
    ws_a.cell(row=row, column=9).value  = 0.0
    ws_a.cell(row=row, column=10).value = 0.0
    ws_a.cell(row=row, column=11).value = round(pd_['ni'],    2)
    ws_a.cell(row=row, column=12).value = round(pd_['min_cash'], 2)
    ws_a.cell(row=row, column=13).value = round(pd_['avg_gb'],   2)
    ws_a.cell(row=row, column=14).value = round(pd_['avg_eb'],   2)
    ws_a.cell(row=row, column=15).value = round(pd_['cash_end'], 2)
    ws_a.cell(row=row, column=16).value = round(pd_['deb_end'],  2)
    ws_a.cell(row=row, column=17).value = 0.0
    ws_a.cell(row=row, column=18).value = round(pd_['cred_end'], 2)
    ws_a.cell(row=row, column=19).value = 0.0
    ws_a.cell(row=row, column=20).value = round(pd_['net_ass'],  2)
    ws_a.cell(row=row, column=21).value = round(pd_['cum_eq'],   2)
    ws_a.cell(row=row, column=22).value = round(pd_['cum_pr'],   2)
    ws_a.cell(row=row, column=23).value = round(pd_['qr'],       4)

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 3: BalanceSheet
# ══════════════════════════════════════════════════════════════════════════════
print("Befülle BalanceSheet …")
ws_bs = wb_dst['BalanceSheet']
ws_bs.cell(row=1, column=1).value = 'Phase'

for i, pd_ in enumerate(phase_data):
    row = 2 + i
    ws_bs.cell(row=row, column=1).value  = pd_['phase']['name']
    ws_bs.cell(row=row, column=2).value  = round(pd_['cash_end'], 2)
    ws_bs.cell(row=row, column=3).value  = round(pd_['deb_end'],  2)
    ws_bs.cell(row=row, column=4).value  = 0.0
    ws_bs.cell(row=row, column=5).value  = round(pd_['cred_end'], 2)
    ws_bs.cell(row=row, column=6).value  = 0.0
    ws_bs.cell(row=row, column=7).value  = round(pd_['net_ass'],  2)
    ws_bs.cell(row=row, column=8).value  = round(pd_['cum_eq'],   2)
    ws_bs.cell(row=row, column=9).value  = round(pd_['cum_pr'],   2)
    ws_bs.cell(row=row, column=10).value = round(pd_['qr'],       4)

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 4: Prompt Business Model_Slides
# ══════════════════════════════════════════════════════════════════════════════
print("Aktualisiere Prompt-Sheet …")
ws_prompt = wb_dst['Prompt Business Model_Slides']
total_rev_all = sum(total_revenue)
avg_burn      = sum(total_cogs[i] + total_opex[i] for i in range(52)) / 52
neg_e         = [ebitda_vals[i] for i in range(52) if ebitda_vals[i] < 0]
avg_ebit_burn = sum(neg_e) / len(neg_e) if neg_e else 0.0
first_neg     = next((i for i, c in enumerate(ending_cash) if c < 0), None)
first_neg_lbl = f"M{first_neg+1} (Ideation)" if first_neg is not None else "nie"
peak_cash     = max(ending_cash)
lowest_cash   = min(ending_cash)
cum_ni_all    = sum(net_income)

# Break-even Phase = erste Phase mit positivem Net Income
breakeven_phase = "nicht erreicht"
for pd_ in phase_data:
    if pd_['ni'] > 0:
        breakeven_phase = pd_['phase']['name']
        break

# Series-B Werte für "letztes Jahr"
sb = phase_data[4]

updates = {
    7:  "- Währung: EUR",
    8:  "- Unternehmenstyp: B2B SaaS / AI-First (LoopforgeLab GmbH)",
    11: f"- Gesamtumsatz über den Zeitraum M1–M52: {total_rev_all:,.0f} €",
    12: f"- Umsatz Series B (M41–M52): {sb['rev']:,.0f} €",
    13: f"- Net Income Series B (M41–M52): {sb['ni']:,.0f} €",
    14: f"- Kumuliertes Net Income M1–M52: {cum_ni_all:,.0f} €",
    15: f"- Gesamtes Eigenkapital M1–M52: {sum(equity_funding):,.0f} €",
    16: "- Nicht-verwässernde Förderung: 0 €",
    19: f"- Höchster Cash-Stand: {peak_cash:,.0f} €",
    20: f"- Niedrigster Cash-Stand: {lowest_cash:,.0f} €",
    21: f"- Erster Monat mit negativem Cash: {first_neg_lbl}",
    22: f"- Durchschnittliche monatliche Gesamtausgaben: {avg_burn:,.0f} €",
    23: f"- Durchschnittlicher EBIT-Burn (Verlustmonate): {avg_ebit_burn:,.0f} €",
    27: f"- Break-even Phase (Net Income > 0): {breakeven_phase}",
}
for r, t in updates.items():
    ws_prompt.cell(row=r, column=1).value = t

# ══════════════════════════════════════════════════════════════════════════════
# CHART: Revenue vs Net Profit vs Year-end Cash (Phasen-X-Achse)
# ══════════════════════════════════════════════════════════════════════════════
print("Erstelle Chart …")

phase_labels   = [pd_['phase']['label'] for pd_ in phase_data]
phase_names    = [pd_['phase']['name']  for pd_ in phase_data]
rev_vals_ph    = [pd_['rev']       for pd_ in phase_data]
ni_vals_ph     = [pd_['ni']        for pd_ in phase_data]
cash_vals_ph   = [pd_['cash_end']  for pd_ in phase_data]
x              = list(range(len(phase_labels)))

# Break-even Phase Index
be_idx = next((i for i, pd_ in enumerate(phase_data) if pd_['ni'] > 0), None)

COL_REV  = '#4A90D9'
COL_NP   = '#E86B7A'
COL_CASH = '#F5A623'

fig, ax = plt.subplots(figsize=(13, 7))
fig.patch.set_facecolor('white')
ax.set_facecolor('white')

ax.plot(x, rev_vals_ph,  color=COL_REV,  linewidth=2.5, marker='o', markersize=7,
        label='Revenue', zorder=3)
ax.plot(x, ni_vals_ph,   color=COL_NP,   linewidth=2.0, marker='o', markersize=7,
        linestyle='--', label='Net Profit', zorder=3)
ax.plot(x, cash_vals_ph, color=COL_CASH, linewidth=2.0, marker='o', markersize=7,
        linestyle='--', label='Year-end Cash', zorder=3)

# Break-even Vertikallinie
if be_idx is not None:
    ax.axvline(x=be_idx, color='#2ECC71', linewidth=1.5, linestyle='--', zorder=2, alpha=0.85)
    y_top = max(max(rev_vals_ph), max(cash_vals_ph)) * 0.97
    ax.text(be_idx + 0.07, y_top, 'Break-even', color='#2ECC71',
            fontsize=9, va='top', ha='left', style='italic')

ax.axhline(y=0, color='#CCCCCC', linewidth=0.8, zorder=1)

def eur_fmt(v, pos):
    if abs(v) >= 1_000_000: return f'€{v/1_000_000:.0f}M'
    if abs(v) >= 1_000:     return f'€{v/1_000:.0f}K'
    return f'€{v:.0f}'

ax.yaxis.set_major_formatter(mticker.FuncFormatter(eur_fmt))
ax.set_xticks(x)
ax.set_xticklabels(phase_labels, fontsize=10)
ax.yaxis.grid(True, linestyle='--', alpha=0.4, color='#DDDDDD')
ax.set_axisbelow(True)
for sp in ['top', 'right']:
    ax.spines[sp].set_visible(False)
ax.spines['left'].set_color('#CCCCCC')
ax.spines['bottom'].set_color('#CCCCCC')
ax.tick_params(colors='#555555', length=0)

# Datenpunkt-Labels (letzter Wert je Linie)
for vals, col in [(rev_vals_ph, COL_REV), (ni_vals_ph, COL_NP), (cash_vals_ph, COL_CASH)]:
    for i, v in enumerate(vals):
        if i == len(vals) - 1:
            lbl = f'€{v/1_000_000:.1f}M' if abs(v) >= 1_000_000 else f'€{v/1_000:.0f}K'
            ax.annotate(lbl, (i, v), textcoords='offset points',
                        xytext=(8, 0), fontsize=8.5, color=col, va='center')

fig.text(0.06, 0.97, 'REVENUE VS NET PROFIT VS CASH',
         fontsize=14, fontweight='bold', color='#2C3E50', va='top', ha='left')
fig.text(0.06, 0.91,
         'Phasenbasierte Aggregation (Normal-Szenario). Cash = Phasenend-Saldo;\n'
         'Revenue und Net Profit = Summe je Phase. Grün gestrichelt = Break-even Phase.',
         fontsize=8.5, color='#7F8C8D', va='top', ha='left')

ax.legend(loc='upper center', bbox_to_anchor=(0.5, -0.12),
          ncol=3, frameon=False, fontsize=10)
plt.tight_layout(rect=[0, 0.06, 1, 0.88])

buf = io.BytesIO()
plt.savefig(buf, format='png', dpi=150, bbox_inches='tight',
            facecolor='white', edgecolor='none')
buf.seek(0)
plt.close()
print(f"  Chart erstellt: {len(buf.getvalue()):,} Bytes")

# ── Neues Sheet einfügen ──────────────────────────────────────────────────────
CHART_SHEET = 'Revenue_Profit_Cash_Chart'
if CHART_SHEET in wb_dst.sheetnames:
    del wb_dst[CHART_SHEET]
ws_chart = wb_dst.create_sheet(CHART_SHEET, index=0)
ws_chart.sheet_properties.tabColor = '4A90D9'

ws_chart['A1'] = 'Revenue vs Net Profit vs Year-end Cash — Normal-Szenario (Phasen-Aggregation)'
ws_chart['A1'].font = openpyxl.styles.Font(bold=True, size=13, color='2C3E50')
ws_chart['A2'] = f'Quelle: {os.path.basename(SOURCE)} | Erstellt: {datetime.now().strftime("%d.%m.%Y %H:%M")}'
ws_chart['A2'].font = openpyxl.styles.Font(size=9, color='7F8C8D', italic=True)

headers = ['Phase', 'Monate', 'Spalten (Quelle)', 'Revenue (€)', 'Net Profit (€)', 'Year-end Cash (€)']
for c, h in enumerate(headers, 1):
    cell = ws_chart.cell(row=4, column=c, value=h)
    cell.font = openpyxl.styles.Font(bold=True)

num_fmt = '#,##0'
for i, pd_ in enumerate(phase_data):
    r = 5 + i
    col_s_ltr = get_column_letter(idx_to_col(pd_['phase']['idx_s']))
    col_e_ltr = get_column_letter(idx_to_col(pd_['phase']['idx_e'] - 1))
    ws_chart.cell(row=r, column=1).value = pd_['phase']['name']
    ws_chart.cell(row=r, column=2).value = f"M{pd_['phase']['idx_s']+1}–M{pd_['phase']['idx_e']}"
    ws_chart.cell(row=r, column=3).value = f"{col_s_ltr}–{col_e_ltr}"
    ws_chart.cell(row=r, column=4).value = round(pd_['rev'], 0)
    ws_chart.cell(row=r, column=5).value = round(pd_['ni'],  0)
    ws_chart.cell(row=r, column=6).value = round(pd_['cash_end'], 0)
    for c in [4, 5, 6]:
        ws_chart.cell(row=r, column=c).number_format = num_fmt

for col, width in [(1,16),(2,12),(3,16),(4,18),(5,18),(6,20)]:
    ws_chart.column_dimensions[get_column_letter(col)].width = width

img = XLImage(buf)
img.anchor = 'A11'
img.width, img.height = 940, 510
ws_chart.add_image(img)
for ri in range(11, 50):
    ws_chart.row_dimensions[ri].height = 14.5

# ── Speichern ─────────────────────────────────────────────────────────────────
wb_dst.save(OUTPUT)
print(f"\n✓ Gespeichert: {os.path.basename(OUTPUT)}")
print(f"  Sheets: {wb_dst.sheetnames}")
