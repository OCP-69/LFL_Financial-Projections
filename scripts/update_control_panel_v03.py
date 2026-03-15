"""
Update Control Panel + add LFL_Parameters sheet in v02 → save as v03.

Changes:
  1. Create 'LFL_Parameters' sheet with all key LFL source parameters + lineage
  2. Update Control Panel (B1-B16) with correct LFL values and formulas
     referencing LFL_Parameters where possible
  3. Override R3 (Avg Sale) and R4 (Customers) in all Year sheets per month
     with actual values from source data → fixes Performance Dashboard Customers

Source:   260315_LFL_BM_Szen_normal_final.xlsx
Base:     C13_Template_financial_projections_ext_2026_v02.xlsx
Output:   C13_Template_financial_projections_ext_2026_v03.xlsx
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import shutil

SRC = '260315_LFL_BM_Szen_normal_final.xlsx'
IN  = 'C13_Template_financial_projections_ext_2026_v02.xlsx'
OUT = 'C13_Template_financial_projections_ext_2026_v03.xlsx'

# ── Load source ────────────────────────────────────────────────────────────────
print('Loading source data...')
wb_src  = openpyxl.load_workbook(SRC, data_only=True)
ws_rev  = wb_src['4_Revenue']
ws_inp  = wb_src['2_Inputs']
ws_pl   = wb_src['6_P&L']
ws_cf   = wb_src['7_BS_CF']

def inp(row): return ws_inp.cell(row, 2).value or 0

# Revenue model parameters (from 2_Inputs)
sme_seat_price   = inp(35)   # 460 €/Mo/Seat
sme_seats_kunde  = inp(33)   # 6 Seats/Kunde
mid_seat_price   = inp(36)   # 530 €/Mo/Seat
mid_seats_kunde  = inp(34)   # 12 Seats/Kunde
ent_fee_pa       = inp(37)   # 80,000 €/Jahr
price_increase   = inp(38)   # 4% p.a.
first_cust_month = inp(39)   # M8 (src) = M4 (tgt)
churn_pa         = inp(40)   # 5% p.a.
tagessatz        = inp(43)   # 1,500 €/Tag
tax_rate         = inp(6)    # 30%
salary_raise     = inp(84)   # 3% p.a.
ag_aufschlag     = inp(85)   # 22%

# Financing (from 2_Inputs: Pre-Seed + Business Angels = M5, Seed = M17, Series A = M29)
preseed  = (inp(11) + inp(18) + inp(20) + inp(22))  # Pre-Seed + 3 Business Angels
seed     = inp(13)
series_a = inp(15)

# Totals for CoR%
total_rev  = sum(ws_pl.cell(8,  c).value or 0 for c in range(6, 54))
total_cogs = sum(ws_pl.cell(14, c).value or 0 for c in range(6, 54))
cor_pct    = total_cogs / total_rev if total_rev > 0 else 0

# Derived: avg MRR per customer (initial SME)
sme_monthly_rev_per_cust = sme_seat_price * sme_seats_kunde  # 2,760

print(f'  Pre-Seed={preseed:,.0f}  Seed={seed:,.0f}  Series A={series_a:,.0f}')
print(f'  SME avg rev/cust: {sme_monthly_rev_per_cust:.0f} €/Mo')
print(f'  Actual CoR%: {cor_pct:.2%}')

# ── Build per-month customer + avg-sale data (tgt M1-M48) ────────────────────
customer_data = []
for m in range(5, 53):  # source M5–M52 → target M1–M48
    col = m + 1
    sme_seats  = int(ws_rev.cell(8,  col).value or 0)
    mid_seats  = int(ws_rev.cell(15, col).value or 0)
    ent        = int(ws_rev.cell(21, col).value or 0)
    mrr        = float(ws_rev.cell(25, col).value or 0)
    total_cust = sme_seats // int(sme_seats_kunde) + mid_seats // int(mid_seats_kunde) + ent
    avg_mrr    = round(mrr / max(1, total_cust), 2) if total_cust > 0 else sme_monthly_rev_per_cust
    customer_data.append({
        'tgt_m':     m - 4,
        'total_cust': total_cust,
        'avg_mrr':   avg_mrr,
    })

# ── Copy v02 → v03 ────────────────────────────────────────────────────────────
print(f'Copying {IN} → {OUT}')
shutil.copy2(IN, OUT)
wb = openpyxl.load_workbook(OUT)  # data_only=False to preserve formulas

# ── Helper styles ──────────────────────────────────────────────────────────────
BOLD       = Font(bold=True)
HEADER_BG  = PatternFill("solid", fgColor="1F4E79")  # dark blue
HEADER_FG  = Font(bold=True, color="FFFFFF")
SECT_BG    = PatternFill("solid", fgColor="D6E4F0")   # light blue
SECT_FG    = Font(bold=True, color="1F4E79")
EURO       = '#,##0'
PCT        = '0.0%'
THIN       = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin'),
)

def hdr_cell(ws, row, col, val):
    c = ws.cell(row, col, val)
    c.font = HEADER_FG; c.fill = HEADER_BG
    c.alignment = Alignment(horizontal='center', wrap_text=True)

def sect_cell(ws, row, col, val):
    c = ws.cell(row, col, val)
    c.font = SECT_FG; c.fill = SECT_BG
    c.alignment = Alignment(horizontal='left')

def val_cell(ws, row, col, val, fmt=None):
    c = ws.cell(row, col, val)
    if fmt: c.number_format = fmt
    return c

# ══════════════════════════════════════════════════════════════════════════════
# STEP 1 – Create 'LFL_Parameters' sheet
# ══════════════════════════════════════════════════════════════════════════════
print('Creating LFL_Parameters sheet...')
if 'LFL_Parameters' in wb.sheetnames:
    del wb['LFL_Parameters']

# Insert as second sheet (after Performance Dashboard)
ws_p = wb.create_sheet('LFL_Parameters', 1)
ws_p.column_dimensions['A'].width = 42
ws_p.column_dimensions['B'].width = 18
ws_p.column_dimensions['C'].width = 18
ws_p.column_dimensions['D'].width = 38

# Title
ws_p.merge_cells('A1:D1')
c = ws_p['A1']
c.value = 'LFL Financial Model – Schlüsselparameter & Datenherkunft'
c.font = Font(bold=True, size=13, color='FFFFFF')
c.fill = HEADER_BG
c.alignment = Alignment(horizontal='center')
ws_p.row_dimensions[1].height = 22

ws_p['A2'] = 'Quelldatei:'
ws_p['B2'] = SRC
ws_p['A2'].font = BOLD
ws_p['A3'] = 'Zieldatei (v03):'
ws_p['B3'] = OUT
ws_p['A3'].font = BOLD

# ── Header row ──
row = 5
for col, txt in enumerate(['Parameter', 'Wert', 'Einheit', 'Quelle / Erläuterung'], 1):
    hdr_cell(ws_p, row, col, txt)

row = 6
# ── Section: Revenue Model ──
sect_cell(ws_p, row, 1, '▶ KUNDENMODELL – REVENUE-ANNAHMEN'); row += 1

params = [
    ('SME Seat-Preis',                     sme_seat_price,         '€/Monat/Seat',  '2_Inputs!B35'),
    ('SME Seats je Kunde',                 sme_seats_kunde,        'Seats/Kunde',   '2_Inputs!B33'),
    ('SME Monatsumsatz je Kunde',          None,                   '€/Monat',       '= B8 × B9  (Seat-Preis × Seats)'),
    ('Mid-Company Seat-Preis',             mid_seat_price,         '€/Monat/Seat',  '2_Inputs!B36'),
    ('Mid-Company Seats je Kunde',         mid_seats_kunde,        'Seats/Kunde',   '2_Inputs!B34'),
    ('Mid-Company Monatsumsatz je Kunde',  None,                   '€/Monat',       '= B11 × B12  (Seat-Preis × Seats)'),
    ('Enterprise Jahres-Fee',              ent_fee_pa,             '€/Jahr',        '2_Inputs!B37'),
    ('Enterprise Monatsumsatz je Kunde',   None,                   '€/Monat',       '= B14 / 12'),
    ('Jährliche Preiserhöhung',            price_increase,         '% p.a.',        '2_Inputs!B38'),
    ('Erster zahlender Kunde (Ziel-Mo.)',  int(first_cust_month) - 4, 'Ziel-Monat','2_Inputs!B39 (Quell-M8 → Ziel-M4)'),
    ('Jährliche Churn Rate',               churn_pa,               '% p.a.',        '2_Inputs!B40'),
]

sme_price_row = sme_seats_row = mid_price_row = mid_seats_row = ent_fee_row = None
for label, val, unit, source in params:
    val_cell(ws_p, row, 1, label)
    if label == 'SME Seat-Preis':
        sme_price_row = row
        val_cell(ws_p, row, 2, val, EURO)
    elif label == 'SME Seats je Kunde':
        sme_seats_row = row
        val_cell(ws_p, row, 2, val)
    elif label == 'SME Monatsumsatz je Kunde':
        # Formula: = sme_price_row × sme_seats_row
        ws_p.cell(row, 2).value = f'=B{sme_price_row}*B{sme_seats_row}'
        ws_p.cell(row, 2).number_format = EURO
        ws_p.cell(row, 2).font = Font(bold=True)
    elif label == 'Mid-Company Seat-Preis':
        mid_price_row = row
        val_cell(ws_p, row, 2, val, EURO)
    elif label == 'Mid-Company Seats je Kunde':
        mid_seats_row = row
        val_cell(ws_p, row, 2, val)
    elif label == 'Mid-Company Monatsumsatz je Kunde':
        ws_p.cell(row, 2).value = f'=B{mid_price_row}*B{mid_seats_row}'
        ws_p.cell(row, 2).number_format = EURO
        ws_p.cell(row, 2).font = Font(bold=True)
    elif label == 'Enterprise Jahres-Fee':
        ent_fee_row = row
        val_cell(ws_p, row, 2, val, EURO)
    elif label == 'Enterprise Monatsumsatz je Kunde':
        ws_p.cell(row, 2).value = f'=B{ent_fee_row}/12'
        ws_p.cell(row, 2).number_format = EURO
        ws_p.cell(row, 2).font = Font(bold=True)
    elif 'Preiserhöhung' in label or 'Churn' in label:
        val_cell(ws_p, row, 2, val, PCT)
    else:
        val_cell(ws_p, row, 2, val)
    val_cell(ws_p, row, 3, unit)
    val_cell(ws_p, row, 4, source)
    row += 1

sme_mrr_row = 9   # "SME Monatsumsatz je Kunde" (7th param, row 6+2=8... let me track)
# Actually let me just hardcode: row 6 is first param, so:
# Row 6: SME Seat-Preis → sme_price_row = 7 (0-indexed start at 7 since row=7 after sect_cell)
# I'll set sme_mrr_row properly in the LFL_Parameters B-cell for Control Panel reference

# Find the "SME Monatsumsatz je Kunde" row (it's the 3rd param = row offset 2 from start=7)
# Let me just compute it: section header at row 6, params start at 7
# param 1 (SME Seat-Preis) = row 7, param 3 (SME Monatsumsatz) = row 9
SME_MRR_ROW = 9   # LFL_Parameters row for "SME Monatsumsatz je Kunde"

row += 1  # blank
# ── Section: Payment Terms ──
sect_cell(ws_p, row, 1, '▶ ZAHLUNGSBEDINGUNGEN  (LFL: 100 % Barverkauf)'); row += 1

pay_params = [
    ('Zahlungseingang sofort (Bar-Anteil)',  1.0, 'Anteil 0–1',  'LFL: sofortige Zahlung, kein Zahlungsziel'),
    ('1-Wochen-Kredit (Kreditkarte)',        0,   'Anteil 0–1',  'LFL: keine Kreditkartenzahlungsziele'),
    ('1-Monats-Kredit (Rechnung/Invoice)',   0,   'Anteil 0–1',  'LFL: keine Rechnungszahlungsziele'),
    ('Lieferantenkredit (COGS, 1 Monat)',    0,   'Anteil 0–1',  'LFL: COGS werden sofort bezahlt'),
]
CASH_ROW = row
WK1_ROW  = row + 1
MO_ROW   = row + 2
SUP_ROW  = row + 3
for label, val, unit, src in pay_params:
    val_cell(ws_p, row, 1, label)
    val_cell(ws_p, row, 2, val, PCT)
    val_cell(ws_p, row, 3, unit)
    val_cell(ws_p, row, 4, src)
    row += 1

row += 1  # blank
# ── Section: COGS Ratio ──
sect_cell(ws_p, row, 1, '▶ UMSATZKOSTENQUOTE (COGS) – Gesamtperiode M5–M52'); row += 1

total_rev_row  = row
val_cell(ws_p, row, 1, 'Gesamtumsatz Pre-Seed bis Series B')
val_cell(ws_p, row, 2, round(total_rev, 2), EURO)
val_cell(ws_p, row, 3, '€')
val_cell(ws_p, row, 4, '6_P&L!R8, Quell-Spalten M5–M52 (= Ziel M1–M48)')
row += 1

total_cogs_row = row
val_cell(ws_p, row, 1, 'Gesamte COGS Pre-Seed bis Series B')
val_cell(ws_p, row, 2, round(total_cogs, 2), EURO)
val_cell(ws_p, row, 3, '€')
val_cell(ws_p, row, 4, '6_P&L!R14, Quell-Spalten M5–M52 (= Ziel M1–M48)')
row += 1

cor_formula_row = row
ws_p.cell(row, 1).value = 'Effektive COGS-Quote (gewichteter Durchschnitt)'
ws_p.cell(row, 2).value = f'=B{total_cogs_row}/B{total_rev_row}'
ws_p.cell(row, 2).number_format = PCT
ws_p.cell(row, 2).font = Font(bold=True)
val_cell(ws_p, row, 3, '% des Umsatzes')
val_cell(ws_p, row, 4, '= Gesamt-COGS / Gesamt-Umsatz  →  Quelle: berechnete Formel')
row += 1

row += 1  # blank
# ── Section: Financing ──
sect_cell(ws_p, row, 1, '▶ FINANZIERUNGSRUNDEN (Quelle: 2_Inputs + 7_BS_CF)'); row += 1

fin_params = [
    ('Pre-Seed + Business Angels (Ziel M1)', preseed,    '€', '2_Inputs: R11 + R18 + R20 + R22'),
    ('Seed (Ziel M13)',                       seed,       '€', '2_Inputs!B13'),
    ('Series A (Ziel M25)',                   series_a,   '€', '2_Inputs!B15'),
]
for label, val, unit, src in fin_params:
    val_cell(ws_p, row, 1, label)
    val_cell(ws_p, row, 2, val, EURO)
    val_cell(ws_p, row, 3, unit)
    val_cell(ws_p, row, 4, src)
    row += 1

row += 1  # blank
# ── Section: Personal / Tax ──
sect_cell(ws_p, row, 1, '▶ PERSONAL & STEUERN (Quelle: 2_Inputs)'); row += 1

misc_params = [
    ('Gehaltserhöhung p.a.',        salary_raise,   '% p.a.', '2_Inputs!B84'),
    ('AG-Aufschlag (SV + Umlagen)', ag_aufschlag,   '%',      '2_Inputs!B85'),
    ('Ertragssteuersatz',           tax_rate,        '%',      '2_Inputs!B6'),
]
for label, val, unit, src in misc_params:
    val_cell(ws_p, row, 1, label)
    val_cell(ws_p, row, 2, val, PCT)
    val_cell(ws_p, row, 3, unit)
    val_cell(ws_p, row, 4, src)
    row += 1

print(f'  LFL_Parameters built  (rows 1–{row})')

# ══════════════════════════════════════════════════════════════════════════════
# STEP 2 – Update Control Panel
# ══════════════════════════════════════════════════════════════════════════════
print('Updating Control Panel...')
ws_cp = wb['Control Panel']

# Widen col A for readability; add col C for source notes
ws_cp.column_dimensions['A'].width = 36
ws_cp.column_dimensions['B'].width = 16
ws_cp.column_dimensions['C'].width = 42

def cp_set(row, val, fmt=None, note=None):
    c = ws_cp.cell(row, 2, val)
    if fmt: c.number_format = fmt
    if note:
        nc = ws_cp.cell(row, 3, note)
        nc.font = Font(italic=True, color='595959')
        nc.alignment = Alignment(wrap_text=True)

# B1: Average Sale per Customer → SME Monatsumsatz/Kunde (initial)
ws_cp.cell(1, 2).value = f"='LFL_Parameters'!B{SME_MRR_ROW}"
ws_cp.cell(1, 2).number_format = '#,##0'
ws_cp.cell(1, 3).value = f"LFL_Parameters!B{SME_MRR_ROW}: SME Seat-Preis × Seats/Kunde = {sme_seat_price} × {int(sme_seats_kunde)}"
ws_cp.cell(1, 3).font = Font(italic=True, color='595959')
ws_cp.cell(1, 1).value = 'Average Sale per Customer (SME)'

# B2: Base Number of Customers per Month → 0 (LFL starts with no customers)
ws_cp.cell(2, 2).value = 0
ws_cp.cell(2, 3).value = (f"LFL: kein Kundenstamm zu Start. Erster Kunde Ziel-M{int(first_cust_month)-4} "
                           f"(Quell-M{int(first_cust_month)}). "
                           f"Kundenzahlen je Monat direkt in R4 der Jahressheets hinterlegt.")
ws_cp.cell(2, 3).font = Font(italic=True, color='595959')

# B3: Potential Customers per Month → keep 10,000 (external market assumption)
ws_cp.cell(3, 2).value = 10000
ws_cp.cell(3, 3).value = ("Externe Marktannahme (TAM). Nicht in Quelldatei definiert. "
                           "Einflussgröße für Marktanteil-Berechnung im Performance Dashboard.")
ws_cp.cell(3, 3).font = Font(italic=True, color='595959')

# B4: Customer Growth per month → 5% (matches LFL's monthly seat growth)
ws_cp.cell(4, 2).value = 0.05
ws_cp.cell(4, 2).number_format = '0%'
ws_cp.cell(4, 3).value = ("Entspricht dem monatlichen Seat-Wachstum im LFL-Modell. "
                           "Achtung: R4-Werte (Kundenzahl/Monat) sind Ist-Werte aus der Quelldatei, "
                           "nicht aus dieser Wachstumsformel abgeleitet.")
ws_cp.cell(4, 3).font = Font(italic=True, color='595959')

# Row 7: 'Cash' label, B7 = Cash fraction = 1.0
ws_cp.cell(7, 2).value = f"='LFL_Parameters'!B{CASH_ROW}"
ws_cp.cell(7, 2).number_format = '0%'
ws_cp.cell(7, 3).value = (f"LFL_Parameters!B{CASH_ROW}: 100 % Bareingang im Buchungsmonat. "
                           f"Keine Zahlungsziele.")
ws_cp.cell(7, 3).font = Font(italic=True, color='595959')

# B8: 1 Week Credit → 0
ws_cp.cell(8, 2).value = f"='LFL_Parameters'!B{WK1_ROW}"
ws_cp.cell(8, 2).number_format = '0%'
ws_cp.cell(8, 3).value = f"LFL_Parameters!B{WK1_ROW}: Kein 1-Wochen-Kredit. LFL = 100 % sofort."
ws_cp.cell(8, 3).font = Font(italic=True, color='595959')

# B9: 1 Month Credit → 0
ws_cp.cell(9, 2).value = f"='LFL_Parameters'!B{MO_ROW}"
ws_cp.cell(9, 2).number_format = '0%'
ws_cp.cell(9, 3).value = f"LFL_Parameters!B{MO_ROW}: Kein Rechnungszahlungsziel. LFL = B2B SaaS, sofortige Zahlung."
ws_cp.cell(9, 3).font = Font(italic=True, color='595959')

# B11: Cost of Revenue → formula from Year Cash Budget sheets (actual weighted CoR%)
cor_formula = (
    "=IFERROR("
    "(SUM('Year 1 Cash Budget'!C18:N18)"
    "+SUM('Year 2 Cash Budget'!B18:M18)"
    "+SUM('Year 3 Cash Budget'!B18:M18)"
    "+SUM('Year 4 Cash Budget'!B18:M18))"
    "/"
    "(SUM('Year 1 Cash Budget'!C7:N7)"
    "+SUM('Year 2 Cash Budget'!B7:M7)"
    "+SUM('Year 3 Cash Budget'!B7:M7)"
    "+SUM('Year 4 Cash Budget'!B7:M7))"
    ",0)"
)
ws_cp.cell(11, 2).value = cor_formula
ws_cp.cell(11, 2).number_format = '0.0%'
ws_cp.cell(11, 2).font = Font(bold=True)
ws_cp.cell(11, 3).value = (f"Formel: Σ COGS (R18, Jahressheets Y1–Y4) / Σ Umsatz (R7, Jahressheets Y1–Y4). "
                            f"Ist-Wert aus Quelldatei: {cor_pct:.1%}")
ws_cp.cell(11, 3).font = Font(italic=True, color='595959')
ws_cp.cell(11, 1).value = 'Cost of Revenue (COGS) — gewichtet'

# B13: 1 Month Supplier Credit → 0
ws_cp.cell(13, 2).value = f"='LFL_Parameters'!B{SUP_ROW}"
ws_cp.cell(13, 2).number_format = '0%'
ws_cp.cell(13, 3).value = (f"LFL_Parameters!B{SUP_ROW}: LFL zahlt COGS sofort (kein Lieferantenkredit). "
                            f"COGS = Cloud-Hosting (variabel) + AI/ML APIs + Payment Processing.")
ws_cp.cell(13, 3).font = Font(italic=True, color='595959')

# B15: Investor Return → keep 12% (add note)
ws_cp.cell(15, 3).value = "Externe Annahme: erwartete Rendite der Investoren. Nicht in Quelldatei definiert."
ws_cp.cell(15, 3).font = Font(italic=True, color='595959')

# B16: Founder Profit % → keep 20% (add note)
ws_cp.cell(16, 3).value = ("Externe Annahme: Gründer-Gewinnbeteiligung. "
                            "Nicht direkt aus Quelldatei; kann im LFL-Shareholder-Agreement definiert werden.")
ws_cp.cell(16, 3).font = Font(italic=True, color='595959')

# Add header for source column if not already present
if ws_cp.cell(1, 3).value is None or 'LFL_Parameters' in str(ws_cp.cell(1, 3).value):
    pass  # already set above
ws_cp.cell(5, 3).value = '◀ Quelle / Erläuterung'
ws_cp.cell(5, 3).font = Font(bold=True, color='1F4E79')

# ══════════════════════════════════════════════════════════════════════════════
# STEP 3 – Override R3 (Avg Sale) + R4 (Customers) in all Year sheets
# ══════════════════════════════════════════════════════════════════════════════
print('Overriding R3 (Avg Sale) and R4 (Customers) per month in Year sheets...')

SHEET_MAP = {
    'Year 1 Cash Budget': {'start_tgt': 1,  'end_tgt': 12, 'col_offset': 2},
    'Year 2 Cash Budget': {'start_tgt': 13, 'end_tgt': 24, 'col_offset': 1},
    'Year 3 Cash Budget': {'start_tgt': 25, 'end_tgt': 36, 'col_offset': 1},
    'Year 4 Cash Budget': {'start_tgt': 37, 'end_tgt': 48, 'col_offset': 1},
    'Year 5 Cash Budget': {'start_tgt': 49, 'end_tgt': 60, 'col_offset': 1},
}

def tgt_to_col(tgt_month, col_offset):
    m_in_yr = ((tgt_month - 1) % 12) + 1
    return m_in_yr + col_offset

for sheet_name, cfg in SHEET_MAP.items():
    ws = wb[sheet_name]
    start, end, offset = cfg['start_tgt'], cfg['end_tgt'], cfg['col_offset']

    for tgt_m in range(start, end + 1):
        col = tgt_to_col(tgt_m, offset)
        src_idx = tgt_m - 1  # 0-based

        if src_idx < len(customer_data):
            d = customer_data[src_idx]
            total_cust = d['total_cust']
            avg_mrr    = d['avg_mrr']
        else:
            total_cust = 0
            avg_mrr    = sme_monthly_rev_per_cust

        # R3: Average Sale per Customer (actual avg MRR per active customer)
        ws.cell(3, col).value = round(avg_mrr, 0)
        ws.cell(3, col).number_format = '#,##0'
        # R4: Customers (actual total active customers this month)
        ws.cell(4, col).value = total_cust
        ws.cell(4, col).number_format = '#,##0'

    print(f'  {sheet_name}: M{start}–{end}  done')

# Verify Year 1 Month 3 (first customer month): cust=1, avg=2760
ws1 = wb['Year 1 Cash Budget']
print(f'  Year 1 M3 (col 5): R4={ws1.cell(4, 5).value} (expect 1), R3={ws1.cell(3, 5).value} (expect 2760)')

# ── Save ───────────────────────────────────────────────────────────────────────
print(f'Saving {OUT}...')
wb.save(OUT)
import os
kb = os.path.getsize(OUT) / 1024
print(f'✓ Fertig: {OUT}  ({kb:.0f} KB)')
print()
print('=== Control Panel Korrekturen ===')
print(f'  B1  (Avg Sale/Customer):  {sme_monthly_rev_per_cust:,.0f} €/Mo  [war: 50]')
print(f'  B2  (Base Customers/Mo):  0                            [war: 300]')
print(f'  B7  (Cash fraction):      100%                         [war: leer/0]')
print(f'  B8  (1-Week Credit):      0%                           [war: 100%]')
print(f'  B9  (1-Month Credit):     0%                           [war: leer]')
print(f'  B11 (Cost of Revenue):    {cor_pct:.1%}  (Formel)     [war: 35%]')
print(f'  B13 (Supplier Credit):    0%                           [war: 100%]')
