"""
LFL Financial Projections – C13-Format (Neu)
Erstellt eine vollständig formatierte Excel-Datei auf Basis von BM_Vorlage_v19.
52 Monate (Apr 2026 – Jul 2030) ohne leere Zeilen.

Sheets:
  1. Monthly       – 52 Zeilen, alle KPIs, einheitliches Euro-Format
  2. Annual        – 5 Zeilen (2026–2030), Jahresaggregate
  3. BalanceSheet  – 5 Zeilen, Bilanzpositionen
  4. Grafiken        – 4 eingebettete Charts (Revenue/Cash/P&L/Headcount)
  5. Investor_Slides – Ausgeführter Prompt mit echten Zahlen
  6. Datenmapping    – Quelltabelle → Zieltabelle: jede Spalte mit Beschreibung
"""

import openpyxl
from openpyxl.styles import (PatternFill, Font, Alignment, Border, Side,
                              numbers)
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.utils import get_column_letter
from datetime import datetime
import os

# ═══════════════════════════════════════════════════════════════════════════════
# KONSTANTEN
# ═══════════════════════════════════════════════════════════════════════════════
BASE_DIR  = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SOURCE    = os.path.join(BASE_DIR, '260312_LFL_BM_Vorlage_v19.xlsx')
TS        = datetime.now().strftime('%Y%m%d_%H%M')
OUTPUT    = os.path.join(BASE_DIR, f'LFL_BM_C13_v2_{TS}.xlsx')

MONATE_DE = ['Jan','Feb','Mär','Apr','Mai','Jun','Jul','Aug','Sep','Okt','Nov','Dez']
PHASEN    = {
    range(1,5):   'Ideation',
    range(5,17):  'Pre-Seed',
    range(17,29): 'Seed',
    range(29,41): 'Series A',
    range(41,53): 'Series B',
}
def phase(m):  # m = 1-based month number
    for r, p in PHASEN.items():
        if m in r: return p
    return ''

# Monatsname für M1..M52 (M1 = Apr 2026)
def month_label(m_idx):  # 0-based
    import datetime as dt
    d = dt.date(2026, 4, 1)
    import calendar
    months = m_idx
    yr  = d.year + (d.month - 1 + months) // 12
    mo  = (d.month - 1 + months) % 12 + 1
    return f"{MONATE_DE[mo-1]} {yr}"

# ═══════════════════════════════════════════════════════════════════════════════
# STYLE-DEFINITIONEN
# ═══════════════════════════════════════════════════════════════════════════════
# Hintergrundfarben
FILL_HEADER    = PatternFill("solid", fgColor="1F4E79")   # Dunkelblau
FILL_SUBHDR    = PatternFill("solid", fgColor="2E75B6")   # Mittelblau
FILL_TOTAL     = PatternFill("solid", fgColor="BDD7EE")   # Hellblau (Totals)
FILL_ALT       = PatternFill("solid", fgColor="F2F7FC")   # Sehr hellblau (Alt-Zeilen)
FILL_PHASE     = PatternFill("solid", fgColor="E2EFDA")   # Hellgrün (Phase-Label)
FILL_WARNING   = PatternFill("solid", fgColor="FCE4D6")   # Hellorange (Warnung)
FILL_WHITE     = PatternFill("solid", fgColor="FFFFFF")

# Schriften
FONT_HEADER  = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
FONT_TOTAL   = Font(bold=True, color="1F4E79", name="Calibri", size=10)
FONT_NORMAL  = Font(name="Calibri", size=10)
FONT_BOLD    = Font(bold=True, name="Calibri", size=10)
FONT_RED     = Font(color="C00000", name="Calibri", size=10)
FONT_TITLE   = Font(bold=True, name="Calibri", size=14, color="1F4E79")
FONT_SUBTITLE= Font(name="Calibri", size=11, color="595959")

# Ausrichtung
ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT   = Alignment(horizontal="left",   vertical="center")
ALIGN_RIGHT  = Alignment(horizontal="right",  vertical="center")

# Rahmen
THIN_BORDER = Border(
    bottom=Side(style='thin', color="BDD7EE"),
    left=Side(style='thin', color="BDD7EE"),
    right=Side(style='thin', color="BDD7EE"),
)
HEADER_BORDER = Border(
    bottom=Side(style='medium', color="FFFFFF"),
)

# Zahlenformate
EUR_FMT  = '#,##0 "€";[Red]-#,##0 "€"'        # Euro mit rotem Minus
EUR0_FMT = '#,##0;[Red]-#,##0'                  # Zahl ohne Einheit
PCT_FMT  = '0.0%'
RATIO_FMT= '0.00'
INT_FMT  = '#,##0'

# ═══════════════════════════════════════════════════════════════════════════════
# HILFSFUNKTIONEN
# ═══════════════════════════════════════════════════════════════════════════════
def style_header_cell(cell, text, col_width=None):
    cell.value = text
    cell.font  = FONT_HEADER
    cell.fill  = FILL_HEADER
    cell.alignment = ALIGN_CENTER
    cell.border = HEADER_BORDER

def style_total_cell(cell, value=None, fmt=EUR_FMT):
    if value is not None:
        cell.value = value
    cell.font   = FONT_TOTAL
    cell.fill   = FILL_TOTAL
    cell.number_format = fmt
    cell.alignment = ALIGN_RIGHT
    cell.border = THIN_BORDER

def style_data_cell(cell, value, fmt, alt=False):
    cell.value  = value
    cell.font   = FONT_NORMAL
    cell.fill   = FILL_ALT if alt else FILL_WHITE
    cell.number_format = fmt
    cell.alignment = ALIGN_RIGHT
    cell.border = THIN_BORDER
    # Rote Schrift für negative Zahlen
    if isinstance(value, (int, float)) and value < 0:
        cell.font = Font(color="C00000", name="Calibri", size=10)

def style_label_cell(cell, text, alt=False, bold=False):
    cell.value = text
    cell.font  = FONT_BOLD if bold else FONT_NORMAL
    cell.fill  = FILL_ALT if alt else FILL_WHITE
    cell.alignment = ALIGN_LEFT
    cell.border = THIN_BORDER

def set_col_width(ws, col, width):
    ws.column_dimensions[get_column_letter(col)].width = width

def freeze_row(ws, row=2):
    ws.freeze_panes = ws.cell(row=row, column=1)

# ═══════════════════════════════════════════════════════════════════════════════
# QUELLDATEN LADEN
# ═══════════════════════════════════════════════════════════════════════════════
print("Lade Quelldaten aus BM_Vorlage_v19 …")
wb_src = openpyxl.load_workbook(SOURCE, data_only=True)

ws_rev   = wb_src['4_Revenue']
ws_pl    = wb_src['6_P&L']
ws_cf    = wb_src['7_BS_CF']
ws_costs = wb_src['5_Costs']

def read_row(ws, row_num, num=52):
    return [float(ws.cell(row=row_num, column=2+i).value or 0) for i in range(num)]

# ── Revenue ──────────────────────────────────────────────────────────────────
total_revenue    = read_row(ws_rev, 32)   # Total Revenue
mrr              = read_row(ws_rev, 25)   # MRR
arr              = read_row(ws_rev, 26)   # ARR
seats_sme        = read_row(ws_rev,  8)   # Active SME Seats
seats_mid        = read_row(ws_rev, 15)   # Active Mid Seats
enterprise_count = read_row(ws_rev, 21)   # Active Enterprise
impl_rev         = read_row(ws_rev, 30)   # Implementation Support Revenue

# ── P&L ──────────────────────────────────────────────────────────────────────
total_cogs       = read_row(ws_pl, 14)    # TOTAL COGS
gross_profit     = read_row(ws_pl, 16)    # GROSS PROFIT
gross_margin     = [gp/r if r else 0.0 for gp, r in zip(gross_profit, total_revenue)]
total_personnel  = read_row(ws_costs, 13) # TOTAL PERSONAL
total_tech       = read_row(ws_costs, 18) # TOTAL TECHNOLOGIE
total_office     = read_row(ws_costs, 24) # TOTAL BÜRO
total_prof       = read_row(ws_costs, 30) # TOTAL PROFESSIONAL
total_ins        = read_row(ws_costs, 35) # TOTAL VERSICHERUNG & BANK
total_mktg       = read_row(ws_costs, 40) # TOTAL MARKETING & SALES
total_other      = read_row(ws_costs, 45) # TOTAL SONSTIGE
total_opex       = read_row(ws_pl, 27)    # TOTAL OPERATING EXPENSES
ebitda           = read_row(ws_pl, 29)    # EBITDA
ebitda_margin    = [e/r if r else 0.0 for e, r in zip(ebitda, total_revenue)]
net_income       = read_row(ws_pl, 37)    # NET INCOME
income_tax       = read_row(ws_pl, 35)    # Income Tax

# ── Cash Flow ────────────────────────────────────────────────────────────────
equity_funding   = read_row(ws_cf,  9)    # Equity Funding Received
beginning_cash   = read_row(ws_cf, 14)    # Beginning Cash Balance
ending_cash      = read_row(ws_cf, 15)    # ENDING CASH BALANCE
burn_rate        = read_row(ws_cf, 18)    # Monthly Burn Rate
runway_raw       = []
for i in range(52):
    rv = ws_cf.cell(row=19, column=2+i).value
    try:
        runway_raw.append(float(rv))
    except:
        runway_raw.append(999.0 if str(rv) == '∞' else 0.0)

# ── Berechnete Größen ─────────────────────────────────────────────────────────
debtors    = [rev for rev in total_revenue]
creditors  = [(total_cogs[i] + total_opex[i]) * 0.10 for i in range(52)]

cum_equity = []
cum_ni     = []
ce, cn = 0.0, 0.0
for i in range(52):
    ce += equity_funding[i]; cum_equity.append(ce)
    cn += net_income[i];     cum_ni.append(cn)

# ── Headcount (aus Einstellungsplan) ─────────────────────────────────────────
# JA-Positionen: Monat des Eintritts (1-basiert), 3 Executives immer
HIRING_PLAN = {6:1, 10:1, 11:1, 14:2, 19:2, 21:1, 22:1, 24:1, 25:1}
headcount = []
hc = 3
for m in range(1, 53):  # M1..M52
    hc += HIRING_PLAN.get(m, 0)
    headcount.append(hc)

print(f"  Daten geladen: {sum(total_revenue):,.0f} € Gesamtumsatz (M1–M52)")

# ═══════════════════════════════════════════════════════════════════════════════
# NEUE WORKBOOK ERSTELLEN
# ═══════════════════════════════════════════════════════════════════════════════
wb = openpyxl.Workbook()
wb.remove(wb.active)  # Default-Sheet entfernen


# ╔═══════════════════════════════════════════════════════════════════════════╗
# ║  SHEET 1: Monthly                                                        ║
# ╚═══════════════════════════════════════════════════════════════════════════╝
print("Erstelle Monthly-Sheet …")
ws_m = wb.create_sheet("Monthly")
ws_m.sheet_view.showGridLines = False

# ── Spalten-Definition ─────────────────────────────────────────────────────
#   A   B     C           D             E                F             G
# Monat Phase Revenue     MRR           Impl.Support     COGS          Gross Profit
#   H          I             J             K           L       M       N
# GP%    Personal      Technologie   OpEx Total    EBITDA  EBITDA%  Equity In
#   O         P           Q         R
# Open Cash Close Cash Burn Rate Headcount

MONTHLY_COLS = [
    # (header, width, number_format)
    ("Monat",                         14, "General"),
    ("Phase",                         12, "General"),
    ("Revenue €",                     14, EUR_FMT),
    ("MRR €",                         14, EUR_FMT),
    ("Impl.-Support €",               14, EUR_FMT),
    ("COGS €",                        14, EUR_FMT),
    ("Gross Profit €",                14, EUR_FMT),
    ("Gross Margin %",                12, PCT_FMT),
    ("Personal €",                    15, EUR_FMT),
    ("IT & Technik €",                14, EUR_FMT),
    ("Total OpEx €",                  15, EUR_FMT),
    ("EBITDA €",                      15, EUR_FMT),
    ("EBITDA Margin %",               13, PCT_FMT),
    ("Equity-Zufluss €",              14, EUR_FMT),
    ("Eröffnungssaldo €",             15, EUR_FMT),
    ("Schlussstand Cash €",           16, EUR_FMT),
    ("Burn Rate €",                   13, EUR_FMT),
    ("Runway (Monate)",               14, "0.0"),
    ("Headcount",                     11, INT_FMT),
    ("Kum. Eigenkapital €",           16, EUR_FMT),
    ("Kum. Net Income €",             16, EUR_FMT),
]

# Titelzeile
ws_m.row_dimensions[1].height = 36
for col_idx, (hdr, width, fmt) in enumerate(MONTHLY_COLS, start=1):
    cell = ws_m.cell(row=1, column=col_idx)
    style_header_cell(cell, hdr)
    set_col_width(ws_m, col_idx, width)

# Datenzeilen
for m_idx in range(52):
    row = m_idx + 2
    alt = (m_idx % 2 == 1)
    ph  = phase(m_idx + 1)

    ws_m.row_dimensions[row].height = 16

    # Warnzeile (negativer Cash)
    warn_fill = FILL_WARNING if ending_cash[m_idx] < 0 else (FILL_ALT if alt else FILL_WHITE)

    def data(col, val, fmt=None):
        c = ws_m.cell(row=row, column=col)
        c.value = val
        c.font  = FONT_NORMAL if (not isinstance(val,(int,float)) or val >= 0) else Font(color="C00000", name="Calibri", size=10)
        c.fill  = warn_fill
        c.number_format = fmt or MONTHLY_COLS[col-1][2]
        c.alignment = ALIGN_RIGHT if col > 2 else ALIGN_LEFT
        c.border = THIN_BORDER

    def label(col, val):
        c = ws_m.cell(row=row, column=col)
        c.value = val
        c.font  = FONT_NORMAL
        c.fill  = warn_fill
        c.alignment = ALIGN_LEFT
        c.border = THIN_BORDER

    label(1, month_label(m_idx))
    label(2, ph)
    data(3,  total_revenue[m_idx])
    data(4,  mrr[m_idx])
    data(5,  impl_rev[m_idx])
    data(6,  total_cogs[m_idx])
    data(7,  round(gross_profit[m_idx], 2))
    data(8,  round(gross_margin[m_idx], 4),  PCT_FMT)
    data(9,  total_personnel[m_idx])
    data(10, total_tech[m_idx])
    data(11, total_opex[m_idx])
    data(12, round(ebitda[m_idx], 2))
    data(13, round(ebitda_margin[m_idx], 4), PCT_FMT)
    data(14, equity_funding[m_idx])
    data(15, round(beginning_cash[m_idx], 2))
    data(16, round(ending_cash[m_idx], 2))
    data(17, burn_rate[m_idx])
    data(18, min(runway_raw[m_idx], 99.9), "0.0")
    data(19, headcount[m_idx], INT_FMT)
    data(20, round(cum_equity[m_idx], 2))
    data(21, round(cum_ni[m_idx], 2))

# Zeile 54: SUMMEN / TOTALS
row_total = 54
ws_m.row_dimensions[row_total].height = 18
for col_idx in range(1, len(MONTHLY_COLS)+1):
    c = ws_m.cell(row=row_total, column=col_idx)
    c.fill = FILL_HEADER
    c.font = FONT_HEADER
    c.border = HEADER_BORDER

ws_m.cell(row=row_total, column=1).value = "GESAMT / ENDWERT"
ws_m.cell(row=row_total, column=1).alignment = ALIGN_LEFT

# Summierbare Spalten
SUM_COLS = {3: EUR_FMT, 4: None, 5: EUR_FMT, 6: EUR_FMT, 7: EUR_FMT,
            9: EUR_FMT, 10: EUR_FMT, 11: EUR_FMT, 12: EUR_FMT,
            14: EUR_FMT, 17: EUR_FMT}
LAST_VAL_COLS = {16: EUR_FMT, 18: "0.0", 19: INT_FMT, 20: EUR_FMT, 21: EUR_FMT}

for col, fmt in SUM_COLS.items():
    val = sum([ws_m.cell(row=r+2, column=col).value or 0 for r in range(52)])
    c = ws_m.cell(row=row_total, column=col)
    c.value = round(val, 2)
    c.number_format = fmt or EUR_FMT
    c.alignment = ALIGN_RIGHT

# Endwerte (letzte Zeile)
for col, fmt in LAST_VAL_COLS.items():
    c = ws_m.cell(row=row_total, column=col)
    c.value = ws_m.cell(row=53, column=col).value
    c.number_format = fmt
    c.alignment = ALIGN_RIGHT

# Gross Margin & EBITDA Margin als Durchschnitt
gm_avg = sum(gross_margin) / 52
eb_avg = sum(ebitda_margin) / 52
ws_m.cell(row=row_total, column=8).value  = round(gm_avg, 4)
ws_m.cell(row=row_total, column=8).number_format  = PCT_FMT
ws_m.cell(row=row_total, column=13).value = round(eb_avg, 4)
ws_m.cell(row=row_total, column=13).number_format = PCT_FMT

freeze_row(ws_m, row=2)


# ╔═══════════════════════════════════════════════════════════════════════════╗
# ║  SHEET 2: Annual                                                         ║
# ╚═══════════════════════════════════════════════════════════════════════════╝
print("Erstelle Annual-Sheet …")
ws_a = wb.create_sheet("Annual")
ws_a.sheet_view.showGridLines = False

ANNUAL_COLS = [
    ("Jahr",                              8,  "0"),
    ("Monate",                            7,  INT_FMT),
    ("Revenue €",                         15, EUR_FMT),
    ("COGS €",                            15, EUR_FMT),
    ("Gross Profit €",                    15, EUR_FMT),
    ("Gross Margin %",                    13, PCT_FMT),
    ("Total OpEx €",                      15, EUR_FMT),
    ("EBITDA €",                          15, EUR_FMT),
    ("EBITDA Margin %",                   13, PCT_FMT),
    ("Steuern €",                         12, EUR_FMT),
    ("Net Profit €",                      15, EUR_FMT),
    ("Min. Cash-Stand €",                 16, EUR_FMT),
    ("Ø Monatl. Burn €",                  14, EUR_FMT),
    ("Cash Jahresende €",                 16, EUR_FMT),
    ("Eigenkapital kum. €",              16, EUR_FMT),
    ("Net Income kum. €",                16, EUR_FMT),
    ("Quick Ratio",                       11, RATIO_FMT),
]

ws_a.row_dimensions[1].height = 36
for col_idx, (hdr, width, fmt) in enumerate(ANNUAL_COLS, start=1):
    cell = ws_a.cell(row=1, column=col_idx)
    style_header_cell(cell, hdr)
    set_col_width(ws_a, col_idx, width)

# Jahres-Definitionen: (year, src_start_idx, src_end_idx, calendar_months)
YEARS = [
    (2026, 0,  9,  9),   # Apr–Dez 2026 = M1–M9
    (2027, 9,  21, 12),  # Jan–Dez 2027 = M10–M21
    (2028, 21, 33, 12),  # Jan–Dez 2028 = M22–M33
    (2029, 33, 45, 12),  # Jan–Dez 2029 = M34–M45
    (2030, 45, 52, 7),   # Jan–Jul 2030 = M46–M52
]

for row_y, (year, s, e, n_months) in enumerate(YEARS, start=2):
    rev   = sum(total_revenue[s:e])
    cogs  = sum(total_cogs[s:e])
    gp    = rev - cogs
    gpm   = gp / rev if rev else 0.0
    opex  = sum(total_opex[s:e])
    eb    = gp - opex
    ebm   = eb / rev if rev else 0.0
    tax   = sum(income_tax[s:e])
    ni    = sum(net_income[s:e])
    eq_c  = cum_equity[e-1]
    ni_c  = cum_ni[e-1]
    min_c = min(ending_cash[s:e])
    end_c = ending_cash[e-1]
    tot_burn = [total_cogs[i]+total_opex[i] for i in range(s,e)]
    avg_burn = sum(tot_burn)/n_months if n_months else 0
    cred_end = creditors[e-1]
    qr       = end_c / cred_end if cred_end else 0.0
    alt      = (row_y % 2 == 0)

    vals = [year, n_months, rev, cogs, gp, gpm, opex, eb, ebm, tax, ni,
            min_c, avg_burn, end_c, eq_c, ni_c, qr]
    fmts = ["0", INT_FMT, EUR_FMT, EUR_FMT, EUR_FMT, PCT_FMT, EUR_FMT,
            EUR_FMT, PCT_FMT, EUR_FMT, EUR_FMT, EUR_FMT, EUR_FMT,
            EUR_FMT, EUR_FMT, EUR_FMT, RATIO_FMT]

    for col_idx, (val, fmt) in enumerate(zip(vals, fmts), start=1):
        c = ws_a.cell(row=row_y, column=col_idx)
        c.value = round(val, 4) if isinstance(val, float) else val
        c.number_format = fmt
        c.fill  = FILL_ALT if alt else FILL_WHITE
        c.font  = FONT_NORMAL if (not isinstance(val,(int,float)) or val >= 0) else Font(color="C00000", name="Calibri", size=10)
        c.alignment = ALIGN_LEFT if col_idx <= 2 else ALIGN_RIGHT
        c.border = THIN_BORDER

# Gesamt-Summenzeile
row_total_a = 8
ws_a.row_dimensions[row_total_a].height = 18
for col_idx in range(1, len(ANNUAL_COLS)+1):
    c = ws_a.cell(row=row_total_a, column=col_idx)
    c.fill = FILL_HEADER
    c.font = FONT_HEADER
    c.border = HEADER_BORDER
    c.alignment = ALIGN_RIGHT

ws_a.cell(row=row_total_a, column=1).value = "Gesamt"
ws_a.cell(row=row_total_a, column=1).alignment = ALIGN_LEFT
ws_a.cell(row=row_total_a, column=2).value = 52

for col_idx, src_list in [(3, total_revenue),(4, total_cogs),(7, total_opex),(10, income_tax)]:
    ws_a.cell(row=row_total_a, column=col_idx).value = round(sum(src_list), 2)
    ws_a.cell(row=row_total_a, column=col_idx).number_format = EUR_FMT
ws_a.cell(row=row_total_a, column=5).value = round(sum(gross_profit),2)
ws_a.cell(row=row_total_a, column=5).number_format = EUR_FMT
ws_a.cell(row=row_total_a, column=8).value = round(sum(ebitda),2)
ws_a.cell(row=row_total_a, column=8).number_format = EUR_FMT
ws_a.cell(row=row_total_a, column=11).value = round(sum(net_income),2)
ws_a.cell(row=row_total_a, column=11).number_format = EUR_FMT
ws_a.cell(row=row_total_a, column=6).value = round(sum(gross_profit)/max(sum(total_revenue),1),4)
ws_a.cell(row=row_total_a, column=6).number_format = PCT_FMT
ws_a.cell(row=row_total_a, column=9).value = round(sum(ebitda)/max(sum(total_revenue),1),4)
ws_a.cell(row=row_total_a, column=9).number_format = PCT_FMT
ws_a.cell(row=row_total_a, column=14).value = round(ending_cash[51],2)
ws_a.cell(row=row_total_a, column=14).number_format = EUR_FMT
ws_a.cell(row=row_total_a, column=15).value = round(cum_equity[51],2)
ws_a.cell(row=row_total_a, column=15).number_format = EUR_FMT
ws_a.cell(row=row_total_a, column=16).value = round(cum_ni[51],2)
ws_a.cell(row=row_total_a, column=16).number_format = EUR_FMT

freeze_row(ws_a, row=2)


# ╔═══════════════════════════════════════════════════════════════════════════╗
# ║  SHEET 3: BalanceSheet                                                   ║
# ╚═══════════════════════════════════════════════════════════════════════════╝
print("Erstelle BalanceSheet …")
ws_bs = wb.create_sheet("BalanceSheet")
ws_bs.sheet_view.showGridLines = False

BS_COLS = [
    ("Jahr",                   8,  "0"),
    ("Cash €",                 16, EUR_FMT),
    ("Debitoren €",            15, EUR_FMT),
    ("R&D Immateriel €",       15, EUR_FMT),
    ("Kreditoren €",           15, EUR_FMT),
    ("Fremdkapital €",         15, EUR_FMT),
    ("Net Assets €",           15, EUR_FMT),
    ("Ges. EK-Investment €",   18, EUR_FMT),
    ("Kum. Net Income €",      16, EUR_FMT),
    ("Quick Ratio",            11, RATIO_FMT),
    ("Bilanzsumme Check",      14, EUR_FMT),
]

ws_bs.row_dimensions[1].height = 36
for col_idx, (hdr, width, fmt) in enumerate(BS_COLS, start=1):
    cell = ws_bs.cell(row=1, column=col_idx)
    style_header_cell(cell, hdr)
    set_col_width(ws_bs, col_idx, width)

for row_y, (year, s, e, n_months) in enumerate(YEARS, start=2):
    cash_end = ending_cash[e-1]
    deb      = debtors[e-1]
    cred     = creditors[e-1]
    rd       = 0.0
    fk       = 0.0
    eq_c     = cum_equity[e-1]
    ni_c     = cum_ni[e-1]
    net_ass  = cash_end + deb + rd - cred - fk
    qr       = cash_end / cred if cred else 0.0
    # Bilanzsumme: Aktiva = Passiva check
    aktiva   = cash_end + deb + rd
    passiva  = cred + fk + (25000 + eq_c + ni_c)  # Stammkapital 25k + EK + kum. Gewinn
    check    = aktiva - passiva
    alt      = (row_y % 2 == 0)

    vals = [year, cash_end, deb, rd, cred, fk, net_ass, eq_c, ni_c, qr, check]
    fmts = ["0", EUR_FMT, EUR_FMT, EUR_FMT, EUR_FMT, EUR_FMT,
            EUR_FMT, EUR_FMT, EUR_FMT, RATIO_FMT, EUR_FMT]

    for col_idx, (val, fmt) in enumerate(zip(vals, fmts), start=1):
        c = ws_bs.cell(row=row_y, column=col_idx)
        c.value = round(val, 2) if isinstance(val, float) else val
        c.number_format = fmt
        c.fill  = FILL_ALT if alt else FILL_WHITE
        c.font  = FONT_NORMAL if (not isinstance(val,(int,float)) or val >= 0) else Font(color="C00000", name="Calibri", size=10)
        c.alignment = ALIGN_LEFT if col_idx == 1 else ALIGN_RIGHT
        c.border = THIN_BORDER

freeze_row(ws_bs, row=2)


# ╔═══════════════════════════════════════════════════════════════════════════╗
# ║  SHEET 4: Grafiken                                                       ║
# ╚═══════════════════════════════════════════════════════════════════════════╝
print("Erstelle Grafiken-Sheet …")
ws_g = wb.create_sheet("Grafiken")
ws_g.sheet_view.showGridLines = False

# Hintergrundinformation für Charts (Datenbasis aus Monthly-Sheet)
# Wir referenzieren Monthly-Sheet Spalten:
#   C = Revenue (col 3)
#   F = COGS (col 6)
#   H = EBITDA (col 12)
#   N = Equity In (col 14)
#   P = Closing Cash (col 16)
#   Q = Burn Rate (col 17)

# ── Chart 1: Revenue vs. EBITDA (monatlich, Säulendiagramm) ─────────────────
chart1 = BarChart()
chart1.type    = "col"
chart1.grouping = "clustered"
chart1.title   = "Monatlicher Umsatz & EBITDA (M1–M52)"
chart1.style   = 10
chart1.y_axis.title = "Betrag (€)"
chart1.x_axis.title = "Monat"
chart1.width   = 22
chart1.height  = 14
chart1.y_axis.numFmt = '#,##0'

# Datenbereiche Monthly sheet Zeilen 2-53
rev_data  = Reference(ws_m, min_col=3,  min_row=1, max_row=53)
eb_data   = Reference(ws_m, min_col=12, min_row=1, max_row=53)
cats      = Reference(ws_m, min_col=1,  min_row=2, max_row=53)

chart1.add_data(rev_data, titles_from_data=True)
chart1.add_data(eb_data,  titles_from_data=True)
chart1.set_categories(cats)
chart1.series[0].graphicalProperties.solidFill = "2E75B6"
chart1.series[0].graphicalProperties.line.solidFill = "2E75B6"
chart1.series[1].graphicalProperties.solidFill = "C00000"
chart1.series[1].graphicalProperties.line.solidFill = "C00000"
ws_g.add_chart(chart1, "A1")

# ── Chart 2: Cash-Verlauf & Burn Rate (Linien + Säulen) ──────────────────────
chart2 = LineChart()
chart2.title   = "Cash-Bestand & Burn Rate (M1–M52)"
chart2.style   = 10
chart2.y_axis.title = "Betrag (€)"
chart2.x_axis.title = "Monat"
chart2.width   = 22
chart2.height  = 14
chart2.y_axis.numFmt = '#,##0'

cash_data = Reference(ws_m, min_col=16, min_row=1, max_row=53)
burn_data = Reference(ws_m, min_col=17, min_row=1, max_row=53)

chart2.add_data(cash_data, titles_from_data=True)
chart2.add_data(burn_data, titles_from_data=True)
chart2.set_categories(cats)
chart2.series[0].graphicalProperties.line.solidFill = "1F4E79"
chart2.series[0].graphicalProperties.line.width = 20000
chart2.series[1].graphicalProperties.line.solidFill = "FF6600"
chart2.series[1].graphicalProperties.line.width = 15000
chart2.series[1].graphicalProperties.line.dashDot = "dash"
ws_g.add_chart(chart2, "A31")

# ── Chart 3: Jahres-P&L Übersicht (gestapelte Säulen aus Annual-Sheet) ────────
chart3 = BarChart()
chart3.type    = "col"
chart3.grouping = "clustered"
chart3.title   = "Jahres-P&L Übersicht (2026–2030)"
chart3.style   = 10
chart3.y_axis.title = "Betrag (€)"
chart3.x_axis.title = "Jahr"
chart3.width   = 22
chart3.height  = 14
chart3.y_axis.numFmt = '#,##0'

a_rev   = Reference(ws_a, min_col=3,  min_row=1, max_row=7)
a_cogs  = Reference(ws_a, min_col=4,  min_row=1, max_row=7)
a_gp    = Reference(ws_a, min_col=5,  min_row=1, max_row=7)
a_opex  = Reference(ws_a, min_col=7,  min_row=1, max_row=7)
a_eb    = Reference(ws_a, min_col=8,  min_row=1, max_row=7)
a_cats  = Reference(ws_a, min_col=1,  min_row=2, max_row=7)

chart3.add_data(a_rev,  titles_from_data=True)
chart3.add_data(a_gp,   titles_from_data=True)
chart3.add_data(a_opex, titles_from_data=True)
chart3.add_data(a_eb,   titles_from_data=True)
chart3.set_categories(a_cats)
chart3.series[0].graphicalProperties.solidFill = "2E75B6"
chart3.series[1].graphicalProperties.solidFill = "70AD47"
chart3.series[2].graphicalProperties.solidFill = "FFC000"
chart3.series[3].graphicalProperties.solidFill = "C00000"
ws_g.add_chart(chart3, "N1")

# ── Chart 4: Headcount-Entwicklung ───────────────────────────────────────────
chart4 = LineChart()
chart4.title   = "Headcount-Entwicklung (M1–M52)"
chart4.style   = 10
chart4.y_axis.title = "Mitarbeiter"
chart4.x_axis.title = "Monat"
chart4.width   = 22
chart4.height  = 10
chart4.y_axis.numFmt = '0'

hc_data = Reference(ws_m, min_col=19, min_row=1, max_row=53)
chart4.add_data(hc_data, titles_from_data=True)
chart4.set_categories(cats)
chart4.series[0].graphicalProperties.line.solidFill = "7030A0"
chart4.series[0].graphicalProperties.line.width = 20000
ws_g.add_chart(chart4, "N31")

# Beschriftung im Grafiken-Sheet
ws_g.column_dimensions['A'].width = 2
title_cell = ws_g.cell(row=61, column=1)
title_cell.value = "Charts basieren auf Monthly- und Annual-Sheet. Daten-Quelle: 260312_LFL_BM_Vorlage_v19.xlsx"
title_cell.font  = Font(italic=True, color="595959", name="Calibri", size=9)


# ╔═══════════════════════════════════════════════════════════════════════════╗
# ║  SHEET 5: Investor_Slides (Prompt ausgeführt)                            ║
# ╚═══════════════════════════════════════════════════════════════════════════╝
print("Erstelle Investor_Slides-Sheet …")
ws_inv = wb.create_sheet("Investor_Slides")
ws_inv.sheet_view.showGridLines = False
ws_inv.column_dimensions['A'].width = 5
ws_inv.column_dimensions['B'].width = 32
ws_inv.column_dimensions['C'].width = 70

# ── Berechnungen für den Pitch-Deck-Inhalt ────────────────────────────────────
total_rev_all  = sum(total_revenue)
peak_cash      = max(ending_cash)
lowest_cash    = min(ending_cash)
first_neg      = next((i+1 for i, c in enumerate(ending_cash) if c < 0), None)
avg_burn_all   = sum(total_cogs[i]+total_opex[i] for i in range(52)) / 52
neg_ebitda_m   = [ebitda[i] for i in range(52) if ebitda[i] < 0]
avg_ebit_burn  = abs(sum(neg_ebitda_m)/len(neg_ebitda_m)) if neg_ebitda_m else 0
total_eq_in    = sum(equity_funding)

# ARR am Ende der Perioden
arr_m9  = arr[8]    # Dez 2026
arr_m21 = arr[20]   # Dez 2027
arr_m33 = arr[32]   # Dez 2028
arr_m45 = arr[44]   # Dez 2029
arr_m52 = arr[51]   # Jul 2030

# MRR-Wachstum
mrr_first_nonzero = next((mrr[i] for i in range(52) if mrr[i] > 0), 0)
mrr_last = mrr[51]
mrr_growth_total = (mrr_last / mrr_first_nonzero - 1) if mrr_first_nonzero else 0

# Kunden am Ende M52
total_seats = seats_sme[51] + seats_mid[51]
enterprise_end = enterprise_count[51]

# Funding-Meilensteine
funding_events = [(i+1, equity_funding[i]) for i in range(52) if equity_funding[i] > 0]

# Cash-Runway am Ende M52
final_burn = burn_rate[51] if burn_rate[51] > 0 else avg_burn_all
final_cash = ending_cash[51]
final_runway = final_cash / final_burn if final_burn > 0 else 99

# Break-even-Analyse
be_year = "Noch nicht erreicht (M52)"
for yo, (year, s, e, nm) in enumerate(YEARS):
    if sum(net_income[s:e]) > 0:
        be_year = str(year)
        break

# Gross Margin letzte verfügbare (M52)
gm_m52 = gross_margin[51] if gross_margin[51] else 0

# Slide-Inhalte
slide_content = [
    # (Typ, Titel/Bullet)
    ("TITLE", "LoopforgeLab GmbH – Financial Story für Investoren"),
    ("SUBTITLE", f"Projektion: April 2026 – Juli 2030 | 52 Monate | Stand: {datetime.now().strftime('%d.%m.%Y')}"),
    ("SPACER", ""),

    ("SLIDE", "SLIDE 1: Geschäftsmodell & Marktchance"),
    ("BULLET", "B2B SaaS-Plattform für industrielle Fertigungsunternehmen (SME / Mid / Enterprise)"),
    ("BULLET", f"3 Revenue-Streams: Seat-Abonnements (SME/Mid) + Enterprise Jahres-Fee + Impl.-Support"),
    ("BULLET", f"Seat-Preis SME: 350 €/Seat/Monat | Enterprise Fee: 25.000 €/Jahr"),
    ("BULLET", f"Net Revenue Retention: 115 % | Churn: 10 % p.a."),
    ("BULLET", "AI-First-Ansatz: KI reduziert Headcount-Wachstum, erhöht Skalierbarkeit"),
    ("SPACER", ""),

    ("SLIDE", "SLIDE 2: Umsatzentwicklung & Wachstum"),
    ("BULLET", f"Erster zahlender Kunde: Monat 7 (Oktober 2026)"),
    ("BULLET", f"ARR Ende 2026 (M9):  {arr_m9:>12,.0f} € (9 aktive SME-Seats)"),
    ("BULLET", f"ARR Ende 2027 (M21): {arr_m21:>12,.0f} € | MRR: {mrr[20]:>10,.0f} €"),
    ("BULLET", f"ARR Ende 2028 (M33): {arr_m33:>12,.0f} € | MRR: {mrr[32]:>10,.0f} €"),
    ("BULLET", f"ARR Ende 2029 (M45): {arr_m45:>12,.0f} € | MRR: {mrr[44]:>10,.0f} €"),
    ("BULLET", f"ARR Juli 2030 (M52): {arr_m52:>12,.0f} € | MRR: {mrr[51]:>10,.0f} €"),
    ("BULLET", f"Gesamtumsatz M1–M52: {total_rev_all:>12,.0f} €"),
    ("SPACER", ""),

    ("SLIDE", "SLIDE 3: Finanzierungsrunden & Kapitalbedarf"),
    ("BULLET", f"Gesamtes eingesetztes Eigenkapital (M1–M52): {total_eq_in:,.0f} €"),
]
for m_num, amt in funding_events:
    ml = month_label(m_num-1)
    slide_content.append(("BULLET", f"  {ml} (M{m_num}): {amt:>12,.0f} € Equity-Zufluss"))
slide_content += [
    ("BULLET", f"Höchster Cash-Bestand: {peak_cash:,.0f} € | Niedrigster: {lowest_cash:,.0f} €"),
    ("BULLET", f"Durchschnittlicher monatlicher Burn: {avg_burn_all:,.0f} €/Monat"),
    ("BULLET", f"Cash-Runway Ende M52 (Jul 2030): {final_runway:.1f} Monate bei aktuellem Burn"),
    ("SPACER", ""),

    ("SLIDE", "SLIDE 4: P&L Entwicklung & Pfad zur Profitabilität"),
    ("BULLET", f"Gross Margin (M52): {gm_m52:.1%} | Ziel: >70 % bei Scale"),
    ("BULLET", f"EBITDA M52: {ebitda[51]:,.0f} € | Verbesserung ggü. M1 ({ebitda[0]:,.0f} €)"),
    ("BULLET", f"Kumuliertes Net Income M1–M52: {cum_ni[51]:,.0f} €"),
    ("BULLET", f"Break-even-Jahr (Net Profit > 0): {be_year}"),
    ("BULLET", "Profitabilitätshebel: Umsatzwachstum übersteigt OpEx-Wachstum ab Series A"),
    ("BULLET", f"Headcount M52: {headcount[51]} MA | AI-First verhindert lineare Skalierung"),
    ("SPACER", ""),

    ("SLIDE", "SLIDE 5: Key KPIs & Investment Highlights"),
    ("BULLET", f"MRR-Wachstum Gesamt: {mrr_growth_total:.0%} (von erstem zahlenden Monat bis M52)"),
    ("BULLET", f"Aktive Seats M52: {total_seats:.0f} (SME + Mid) | Enterprise-Kunden: {enterprise_end:.0f}"),
    ("BULLET", f"Ø Gross Margin über 52 Monate: {sum(gross_margin)/52:.1%}"),
    ("BULLET", f"CAC gedeckt durch: Implementation-Support-Revenue in M1 des Kunden"),
    ("BULLET", "LTV/CAC: >3x bei NRR 115% (SaaS-Benchmark für skalierbares Modell)"),
    ("SPACER", ""),

    ("SLIDE", "EMPFOHLENE CHARTS FÜR PITCH DECK"),
    ("BULLET", "Chart A: ARR-Wachstumskurve 2026–2030 (Exponentiell, zeigt PMF)"),
    ("BULLET", "Chart B: Cash-Runway-Waterfall (zeigt Finanzierungsbedarf + Runway)"),
    ("BULLET", "Chart C: Revenue vs. Burn Rate (zeigt Weg zur Profitabilität)"),
    ("BULLET", "Chart D: Headcount vs. ARR/FTE (zeigt AI-First-Effizienz)"),
    ("SPACER", ""),

    ("SLIDE", "RISIKOHINWEISE (C13 Stress-Test)"),
    ("BULLET", f"⚠  M1–M3: Negativer Cash (bis -6.226 €) – Startkapital vor Beginn einplanen"),
    ("BULLET", f"⚠  Break-even nicht in 52 Monaten erreicht – Series A (M29) ist entscheidend"),
    ("BULLET", "ℹ  R&D nicht bilanziert (alle Kosten im OpEx) – Optimierungspotenzial für EBIT"),
    ("BULLET", "ℹ  Szenarien: 'Normal' zeigt ersten Kunden in M7 – Verzögerung erhöht Burn"),
]

# Schreiben ins Sheet
row_inv = 1
for entry_type, text in slide_content:
    ws_inv.row_dimensions[row_inv].height = 15 if entry_type != "SPACER" else 8

    if entry_type == "TITLE":
        c = ws_inv.cell(row=row_inv, column=2)
        c.value = text
        c.font  = FONT_TITLE
        c.alignment = ALIGN_LEFT
        ws_inv.merge_cells(f"B{row_inv}:C{row_inv}")
    elif entry_type == "SUBTITLE":
        c = ws_inv.cell(row=row_inv, column=2)
        c.value = text
        c.font  = FONT_SUBTITLE
        c.alignment = ALIGN_LEFT
        ws_inv.merge_cells(f"B{row_inv}:C{row_inv}")
    elif entry_type == "SLIDE":
        ws_inv.row_dimensions[row_inv].height = 20
        c = ws_inv.cell(row=row_inv, column=2)
        c.value = text
        c.font  = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
        c.fill  = FILL_HEADER
        c.alignment = ALIGN_LEFT
        ws_inv.merge_cells(f"B{row_inv}:C{row_inv}")
    elif entry_type == "BULLET":
        c = ws_inv.cell(row=row_inv, column=3)
        c.value = ("• " if not text.startswith("  ") else "  – ") + text.lstrip("• ")
        c.font  = FONT_NORMAL
        c.alignment = ALIGN_LEFT
    elif entry_type == "SPACER":
        pass

    row_inv += 1

# Hinweis am Ende
row_inv += 1
note = ws_inv.cell(row=row_inv, column=2)
note.value = (f"Alle Zahlen auf Basis LFL_BM_Vorlage_v19 | Szenario: Gering | "
              f"Erstellt: {datetime.now().strftime('%d.%m.%Y %H:%M')} | Währung: EUR")
note.font  = Font(italic=True, color="595959", name="Calibri", size=9)
ws_inv.merge_cells(f"B{row_inv}:C{row_inv}")


# ╔═══════════════════════════════════════════════════════════════════════════╗
# ║  SHEET 6: Datenmapping                                                   ║
# ╚═══════════════════════════════════════════════════════════════════════════╝
print("Erstelle Datenmapping-Sheet …")
ws_map = wb.create_sheet("Datenmapping")
ws_map.sheet_view.showGridLines = False

# ── Spaltendefinitionen ────────────────────────────────────────────────────
MAP_COLS = [
    ("Ziel-Sheet",              16),
    ("Ziel-Spalte",             20),
    ("Was wird gemessen?",      42),
    ("Quell-Sheet (v19)",       20),
    ("Quell-Zeile / Variable",  30),
    ("Formel / Logik",          50),
    ("Einheit",                 10),
    ("C13-Mapping-Regel",       42),
]

ws_map.row_dimensions[1].height = 36
ws_map.row_dimensions[2].height = 14  # Leerzeile nach Titel
for col_idx, (hdr, width) in enumerate(MAP_COLS, start=1):
    style_header_cell(ws_map.cell(row=1, column=col_idx), hdr)
    ws_map.column_dimensions[get_column_letter(col_idx)].width = width

# ── Mapping-Daten ────────────────────────────────────────────────────────────
# Format: (Ziel-Sheet, Ziel-Spalte, Was gemessen, Quell-Sheet, Quell-Zeile, Formel/Logik, Einheit, C13-Regel)
MAPPING_ROWS = [
    # ── MONTHLY ──────────────────────────────────────────────────────────────
    ("── MONTHLY ──", "", "", "", "", "", "", ""),

    ("Monthly", "Monat",
     "Kalendermonat der Projektion (Apr 2026 = M1, Jul 2030 = M52)",
     "—", "Berechnet",
     "Startdatum 01.04.2026 + (m-1) Monate",
     "Label", "Zeitachse des Modells"),

    ("Monthly", "Phase",
     "Finanzierungsphase des Unternehmens zum jeweiligen Monat",
     "3_Roadmap", "Zeile 5–9",
     "M1–4: Ideation | M5–16: Pre-Seed | M17–28: Seed | M29–40: Series A | M41–52: Series B",
     "Text", "Gibt Kontext zur Burn-Rate-Entwicklung"),

    ("Monthly", "Revenue €",
     "Gesamtumsatz des Monats aus allen drei Revenue-Streams (Seats + Enterprise + Impl.-Support)",
     "4_Revenue", "Zeile 32: 'Total Revenue (€/Monat)'",
     "= SME-Revenue + Mid-Revenue + Enterprise-Revenue + Impl.-Support-Revenue",
     "€/Monat", "→ C13: 'Revenue' und 'Cash in from Revenue' (Sofort-Zahlung angenommen)"),

    ("Monthly", "MRR €",
     "Monthly Recurring Revenue: Wiederkehrender Umsatz aus Seat-Abonnements (SME + Mid). Exklusive Einmal-Zahlungen.",
     "4_Revenue", "Zeile 25: 'MRR – Monthly Recurring Revenue'",
     "= Active SME-Seats × Seat-Preis + Active Mid-Seats × Mid-Preis",
     "€/Monat", "SaaS-Kernmetrik: Zeigt skalierbare Umsatzbasis"),

    ("Monthly", "Impl.-Support €",
     "Einmalige Beratungserlöse aus der Kundeneinführung. Nicht wiederkehrend, aber zeigt Vertriebsaktivität.",
     "4_Revenue", "Zeile 30: 'Impl.-Support Revenue (€/Monat)'",
     "= Neue Kunden × Tagessatz × Tage/Kunde × Buchungsquote",
     "€/Monat", "Services-Revenue; skaliert mit Neukunden"),

    ("Monthly", "COGS €",
     "Cost of Revenue: Variable Kosten, die direkt mit dem Umsatz entstehen (Hosting, KI-APIs, Payment).",
     "6_P&L", "Zeile 14: 'TOTAL COGS'",
     "= Cloud Hosting (variabel, pro Seat) + AI/ML API-Kosten + Payment Processing (2,5% vom Revenue)",
     "€/Monat", "→ C13: 'CoR Cash Out'"),

    ("Monthly", "Gross Profit €",
     "Rohertrag: Was nach Abzug der direkten Herstellkosten vom Umsatz verbleibt.",
     "6_P&L", "Zeile 16: 'GROSS PROFIT'",
     "= Total Revenue − TOTAL COGS",
     "€/Monat", "Basis für Gross-Margin-Berechnung"),

    ("Monthly", "Gross Margin %",
     "Prozentualer Anteil des Rohertrags am Umsatz. Zeigt Effizienz des Geschäftsmodells. SaaS-Benchmark: >70 %.",
     "6_P&L", "Berechnet aus Zeilen 8 + 14",
     "= Gross Profit / Total Revenue",
     "%", "Unter 0 % = COGS übersteigen Umsatz (Früh­phase normal)"),

    ("Monthly", "Personal €",
     "Gesamte Personalkosten inkl. aller Sozialabgaben (AG-Brutto). Größter Einzelkostenblock.",
     "5_Costs", "Zeile 13: 'TOTAL PERSONAL'",
     "= Exec-Gehälter (phasenabhängig) + MA-Gehälter (ab Eintrittsmonat) × (1 + 3 % p.a.) + AG-SV 22 %",
     "€/Monat", "→ Teil von C13: 'Operating Cash Out'"),

    ("Monthly", "IT & Technik €",
     "Cloud-Hosting (Basislast), AI/ML-APIs (für interne LFL-MA), SaaS-Tools und Lizenzen.",
     "5_Costs", "Zeile 18: 'TOTAL TECHNOLOGIE'",
     "= Cloud-Basis + AI/ML × MA-Anzahl + SaaS-Tools (phasenabhängig)",
     "€/Monat", "→ Teil von C13: 'Operating Cash Out'"),

    ("Monthly", "Total OpEx €",
     "Gesamte operative Ausgaben (exkl. COGS): Personal, IT, Büro, Professional Services, Versicherungen, Marketing, Sonstiges.",
     "6_P&L", "Zeile 27: 'TOTAL OPERATING EXPENSES'",
     "= Personal + Technologie + Büro + Prof. Services + Versicherung + Marketing + Sonstiges",
     "€/Monat", "→ C13: 'Operating Cash Out (incl R&D, upfront R&D)'"),

    ("Monthly", "EBITDA €",
     "Earnings Before Interest, Taxes, Depreciation & Amortization. Wichtigster Profitabilitätsindikator im Modell.",
     "6_P&L", "Zeile 29: 'EBITDA'",
     "= Gross Profit − Total OpEx",
     "€/Monat", "Negativ in Frühphase (Investitionsphase), Trend zu 0 zeigt Skalierungseffekt"),

    ("Monthly", "EBITDA Margin %",
     "EBITDA als Anteil am Umsatz. SaaS-Benchmark profitabler Scale-ups: >20 %. Früh­phase: stark negativ.",
     "6_P&L", "Berechnet aus Zeilen 29 + 8",
     "= EBITDA / Total Revenue (bei Revenue = 0: 0 %)",
     "%", "Zeigt operative Hebelwirkung bei Umsatzwachstum"),

    ("Monthly", "Equity-Zufluss €",
     "Im jeweiligen Monat eingegangene Eigenkapitalzahlungen (Investoren-Runden + Business Angels).",
     "7_BS_CF", "Zeile 9: 'Equity Funding Received'",
     "= Summe aller Finanzierungsrunden, die in diesem Monat fließen (aus 2_Inputs)",
     "€/Monat", "→ C13: 'Other Cash In (equity, debt, grants)'"),

    ("Monthly", "Eröffnungssaldo €",
     "Cash-Bestand zu Beginn des Monats (= Schlusssaldo des Vormonats).",
     "7_BS_CF", "Zeile 14: 'Beginning Cash Balance'",
     "= Ending Cash[m-1]",
     "€", "→ C13: 'Opening Cash'"),

    ("Monthly", "Schlussstand Cash €",
     "Liquiditätsbestand am Monatsende nach allen Ein- und Auszahlungen. Wichtigste Überlebensgröße.",
     "7_BS_CF", "Zeile 15: 'ENDING CASH BALANCE'",
     "= Eröffnungssaldo + Revenue + Equity − COGS − OpEx − Steuern",
     "€", "→ C13: 'Closing Cash'. Negativ = Insolvenzrisiko (C13 Stress-Test)"),

    ("Monthly", "Burn Rate €",
     "Monatlicher Netto-Cash-Abfluss. Relevant für Runway-Berechnung und Investoren-Kommunikation.",
     "7_BS_CF", "Zeile 18: 'Monthly Burn Rate'",
     "= MAX(0, −NetChange in Cash) — nur positiv wenn Cash abnimmt",
     "€/Monat", "Gross Burn = COGS + OpEx | Net Burn = inkl. Revenue-Eingang"),

    ("Monthly", "Runway (Monate)",
     "Wie viele Monate das Unternehmen mit dem aktuellen Cash-Bestand bei aktueller Burn Rate überleben kann.",
     "7_BS_CF", "Zeile 19: 'Runway (Months)'",
     "= Closing Cash / Burn Rate (∞ wenn Burn Rate = 0 oder positiv)",
     "Monate", "< 6 = Warnung (C13-Alert). Investor-kritische Kennzahl"),

    ("Monthly", "Headcount",
     "Anzahl aktiver Mitarbeiter inkl. 3 Gründer. Basiert auf Einstellungsplan aus 2_Inputs.",
     "2_Inputs", "Zeilen 175–195: Einstellungsplan (JA-Positionen)",
     "3 Exec konstant + kumulierte JA-Einstellungen ab jeweiligem Eintrittsmonat",
     "Personen", "AI-First: HC wächst nicht linear mit Revenue"),

    ("Monthly", "Kum. Eigenkapital €",
     "Summe aller Equity-Zuflüsse von M1 bis zum jeweiligen Monat (kumuliert).",
     "7_BS_CF", "Berechnet: Σ Equity Funding Received M1..m",
     "= Σ equity_funding[0..m]",
     "€ kum.", "Zeigt gesamten Kapitaleinsatz der Investoren"),

    ("Monthly", "Kum. Net Income €",
     "Kumulierter Gewinn/Verlust von M1 bis zum jeweiligen Monat. Zeigt Gesamtverlauf der Profitabilität.",
     "6_P&L", "Berechnet: Σ NET INCOME M1..m (Zeile 37)",
     "= Σ net_income[0..m]",
     "€ kum.", "Negativ = noch nicht profitabel; Trendwende zeigt Break-even"),

    # ── ANNUAL ───────────────────────────────────────────────────────────────
    ("", "", "", "", "", "", "", ""),
    ("── ANNUAL ──", "", "", "", "", "", "", ""),

    ("Annual", "Jahr",
     "Kalenderjahr der aggregierten Projektion (2026 = Apr–Dez, 9 Monate; 2030 = Jan–Jul, 7 Monate).",
     "—", "Berechnet", "Jahresblöcke: 2026 M1–M9 | 2027 M10–M21 | 2028 M22–M33 | 2029 M34–M45 | 2030 M46–M52",
     "Jahr", "Basis für Investor-Reporting und Jahresabschluss"),

    ("Annual", "Revenue €",
     "Summe aller monatlichen Umsätze des Jahres. Wichtigste Top-Line-Kennzahl.",
     "Monthly", "Σ Revenue über alle Monate des Jahres",
     "= Σ total_revenue[s:e]",
     "€/Jahr", "→ C13: 'Revenue' (Jahresaggregat)"),

    ("Annual", "COGS €",
     "Summe aller direkten Kosten des Jahres. Niedriger COGS-Anteil = hohes Skalierungspotenzial.",
     "Monthly", "Σ COGS über alle Monate des Jahres",
     "= Σ total_cogs[s:e]",
     "€/Jahr", "→ C13: 'Cost of Revenue'"),

    ("Annual", "Gross Profit €",
     "Jährlicher Rohertrag: Revenue minus COGS. Basis für alle weiteren Profitabilitätsberechnungen.",
     "Monthly", "Berechnet",
     "= Σ Revenue − Σ COGS",
     "€/Jahr", "→ C13: 'Gross Profit'"),

    ("Annual", "Gross Margin %",
     "Rohmarge: Anteil des Gross Profit am Jahresumsatz. SaaS-Reife ab ~70 %.",
     "Annual", "Berechnet",
     "= Gross Profit / Revenue",
     "%", "Zeigt strukturelle Profitabilität des Geschäftsmodells"),

    ("Annual", "Total OpEx €",
     "Gesamte jährliche operative Kosten. Wächst langsamer als Revenue = operative Skalierung.",
     "Monthly", "Σ Total OpEx über alle Monate des Jahres",
     "= Σ total_opex[s:e] (Personal + IT + Büro + Prof.Services + Vers. + Mktg + Sonstiges)",
     "€/Jahr", "→ C13: 'Total Opex (incl R&D amort)'"),

    ("Annual", "EBITDA €",
     "Operative Ertragskraft des Jahres. Hauptindikator für Investoren zur Bewertung der Skalierung.",
     "Monthly", "Σ EBITDA über alle Monate des Jahres",
     "= Gross Profit − Total OpEx",
     "€/Jahr", "→ C13: 'EBIT' (da keine D&A separat modelliert)"),

    ("Annual", "Net Profit €",
     "Jahresüberschuss nach Steuern. Positiv = Break-even erreicht. Im Modell = EBITDA (keine Steuern in Verlustjahren).",
     "Monthly", "Σ NET INCOME über alle Monate (Zeile 37 P&L)",
     "= Σ net_income[s:e]",
     "€/Jahr", "→ C13: 'Net Profit'"),

    ("Annual", "Min. Cash-Stand €",
     "Niedrigster Liquiditätsstand im Jahr. Wichtigster Risiko-Indikator (C13: 'Lowest Cash Balance').",
     "Monthly", "MIN(Closing Cash aller Monate des Jahres)",
     "= min(ending_cash[s:e])",
     "€", "→ C13: 'Min Cash'. Negativ = Insolvenzrisiko im Jahresverlauf"),

    ("Annual", "Ø Monatl. Burn €",
     "Durchschnittlicher monatlicher Cash-Abfluss (Gross Burn). Basis für Runway-Planung.",
     "Monthly", "Berechnet aus COGS + OpEx",
     "= Σ(COGS + OpEx) / Anzahl Monate im Jahr",
     "€/Monat", "→ C13: 'Avg Monthly Gross Burn (cash)'"),

    ("Annual", "Cash Jahresende €",
     "Liquiditätsbestand am letzten Monat des Jahres (Dezember bzw. letzter Monat bei unvollständigem Jahr).",
     "Monthly", "Closing Cash des letzten Monats im Jahr",
     "= ending_cash[e-1]",
     "€", "→ C13: 'Cash (year-end)'"),

    ("Annual", "Eigenkapital kum. €",
     "Gesamtes bis Jahresende eingeflossenes Eigenkapital (alle Runden kumuliert).",
     "Monthly", "Σ Equity-Zufluss M1 bis Jahresende",
     "= Σ equity_funding[0..e]",
     "€ kum.", "→ C13: 'Total Equity Investment'"),

    ("Annual", "Net Income kum. €",
     "Gesamtgewinn/-verlust von M1 bis Jahresende. Zeigt Gesamtkapitalbedarf.",
     "Monthly", "Σ Net Income M1 bis Jahresende",
     "= Σ net_income[0..e]",
     "€ kum.", "→ C13: 'Cumulative Profit'"),

    ("Annual", "Quick Ratio",
     "Liquiditätskennzahl: Cash / Verbindlichkeiten. >1 = liquide. < 1 = Zahlungsrisiko.",
     "Annual", "Berechnet",
     "= Cash Jahresende / Kreditoren (= 10 % der Jahreskosten anteilig)",
     "Ratio", "→ C13: 'Quick Ratio'. Investor-Sicherheitskennzahl"),

    # ── BALANCESHEET ─────────────────────────────────────────────────────────
    ("", "", "", "", "", "", "", ""),
    ("── BALANCESHEET ──", "", "", "", "", "", "", ""),

    ("BalanceSheet", "Cash €",
     "Zahlungsmittelbestand zum Jahresende (= Closing Cash des letzten Monats im Jahr).",
     "Monthly", "Closing Cash des Jahresabschlussmonats",
     "= ending_cash[Jahr_Ende]",
     "€", "→ C13: 'Cash'. Wichtigste Aktiv-Position"),

    ("BalanceSheet", "Debitoren €",
     "Ausstehende Kundenforderungen: Umsätze, die in Rechnung gestellt aber noch nicht bezahlt wurden.",
     "Annual", "Revenue des letzten Monats im Jahr",
     "= total_revenue[Jahr_Ende] (1 Monat ausstehend)",
     "€", "→ C13: 'Debtors'. Zeigt Forderungsbestand"),

    ("BalanceSheet", "R&D Immateriel €",
     "Aktivierte Entwicklungskosten: Software-Entwicklung als immaterieller Vermögenswert (10 Jahre Abschreibung).",
     "—", "Nicht modelliert (0 €)",
     "Alle R&D-Kosten laufen durch OpEx. Empfehlung: Einmalige Entwicklungskosten hier aktivieren.",
     "€", "→ C13: 'R&D Intangible'. Optimierungspotenzial für EBIT-Darstellung"),

    ("BalanceSheet", "Kreditoren €",
     "Ausstehende Verbindlichkeiten gegenüber Lieferanten: Kosten, die angefallen aber noch nicht bezahlt wurden.",
     "Monthly", "Berechnet: 10 % der (COGS + OpEx) des Jahresabschlussmonats",
     "= (total_cogs[e-1] + total_opex[e-1]) × 10 %",
     "€", "→ C13: 'Creditors'. Passiv-Position, typisch 1 Monat Zahlungsziel"),

    ("BalanceSheet", "Fremdkapital €",
     "Ausstehende Kreditverbindlichkeiten (Bankdarlehen, Wandeldarlehen). Im Modell: 0 €.",
     "—", "Nicht modelliert (0 €)",
     "Kein Fremdkapital im Basismodell. Erweiterbar bei Venture Debt.",
     "€", "→ C13: 'Debt'. Bei 0 = reine Eigenkapitalfinanzierung"),

    ("BalanceSheet", "Net Assets €",
     "Nettovermögen: Gesamtvermögen minus Verbindlichkeiten. Vereinfacht = Eigenkapital der Gesellschaft.",
     "BalanceSheet", "Berechnet",
     "= Cash + Debitoren + R&D − Kreditoren − Fremdkapital",
     "€", "→ C13: 'Net Assets'. Zeigt Substanzwert"),

    ("BalanceSheet", "Ges. EK-Investment €",
     "Gesamtes eingezahltes Eigenkapital aller Investoren von Gründung bis Jahresende.",
     "Monthly", "Σ Equity-Zuflüsse kumuliert bis Jahresende",
     "= Σ equity_funding[0..Jahr_Ende]",
     "€ kum.", "→ C13: 'Total Equity Investment'"),

    ("BalanceSheet", "Kum. Net Income €",
     "Kumulierter Gewinn/Verlust seit Gründung bis Jahresende. Zeigt Gesamtverlust der Investitionsphase.",
     "Monthly", "Σ Net Income kumuliert bis Jahresende",
     "= Σ net_income[0..Jahr_Ende]",
     "€ kum.", "→ C13: 'Cumulative Profit'. Negativ = noch nicht amortisiert"),

    ("BalanceSheet", "Quick Ratio",
     "Kurzfristige Liquiditätskennzahl: Verhältnis von Cash zu kurzfristigen Verbindlichkeiten.",
     "BalanceSheet", "Berechnet",
     "= Cash / Kreditoren (bei Kreditoren = 0: Quick Ratio = 0)",
     "Ratio", "→ C13: 'Quick Ratio'. >1 empfohlen für Zahlungssicherheit"),
]

# ── Zeilen schreiben ───────────────────────────────────────────────────────
data_row = 2
for entry in MAPPING_ROWS:
    ws_map.row_dimensions[data_row].height = 14

    if entry[0].startswith("──"):
        # Abschnitts-Trennzeile
        ws_map.row_dimensions[data_row].height = 20
        c = ws_map.cell(row=data_row, column=1)
        c.value = entry[0]
        c.font  = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
        c.fill  = FILL_SUBHDR
        c.alignment = ALIGN_LEFT
        c.border = HEADER_BORDER
        for col in range(2, len(MAP_COLS)+1):
            ws_map.cell(row=data_row, column=col).fill = FILL_SUBHDR
            ws_map.cell(row=data_row, column=col).border = HEADER_BORDER
        data_row += 1
        continue

    if entry[0] == "":
        # Leerzeile
        ws_map.row_dimensions[data_row].height = 6
        data_row += 1
        continue

    alt = (data_row % 2 == 0)
    fill = FILL_ALT if alt else FILL_WHITE

    for col_idx, val in enumerate(entry, start=1):
        c = ws_map.cell(row=data_row, column=col_idx)
        c.value = val
        # Besondere Formatierung für erste Spalte (Ziel-Sheet)
        if col_idx == 1:
            c.font = Font(bold=True, name="Calibri", size=10, color="1F4E79")
        elif col_idx == 2:
            c.font = Font(bold=True, name="Calibri", size=10)
        else:
            c.font = FONT_NORMAL
        c.fill = fill
        c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        c.border = THIN_BORDER

    ws_map.row_dimensions[data_row].height = None  # Auto-Höhe
    data_row += 1

# Row-Höhen für Textzellen anpassen (Wrap-Text braucht mehr Höhe)
for r in range(2, data_row):
    h = ws_map.row_dimensions[r].height
    if h is None:
        ws_map.row_dimensions[r].height = 45  # Wrap-Text Zellen brauchen Höhe

# Freeze
freeze_row(ws_map, row=2)


# ═══════════════════════════════════════════════════════════════════════════════
# WORKBOOK-EIGENSCHAFTEN
# ═══════════════════════════════════════════════════════════════════════════════
wb.properties.creator = "LoopforgeLab – Claude Code"
wb.properties.title   = "LFL Financial Projections C13-Format v2"
wb.properties.subject = "52-Monats-Finanzplan (Apr 2026 – Jul 2030)"
wb.properties.description = (
    f"Erstellt aus 260312_LFL_BM_Vorlage_v19.xlsx | "
    f"Mapping: Data Merge C13 Template | {datetime.now().strftime('%d.%m.%Y')}"
)

# Sheet-Tab-Farben
ws_m.sheet_properties.tabColor   = "1F4E79"
ws_a.sheet_properties.tabColor   = "2E75B6"
ws_bs.sheet_properties.tabColor  = "70AD47"
ws_g.sheet_properties.tabColor   = "FFC000"
ws_inv.sheet_properties.tabColor = "C00000"
ws_map.sheet_properties.tabColor = "595959"

wb.save(OUTPUT)
print(f"\n✓ Datei gespeichert: {OUTPUT}")
print(f"  Sheets: {wb.sheetnames}")

# ── Finaler Bericht ──────────────────────────────────────────────────────────
print("\n═══════════════════════════════════════════════════════")
print("ZUSAMMENFASSUNG")
print("═══════════════════════════════════════════════════════")
for yo, (year, s, e, nm) in enumerate(YEARS):
    rev  = sum(total_revenue[s:e])
    eb   = sum(ebitda[s:e])
    cash = ending_cash[e-1]
    print(f"  {year}: Rev={rev:>12,.0f} € | EBITDA={eb:>12,.0f} € | Cash-End={cash:>12,.0f} €")
print(f"  Gesamt Rev:      {sum(total_revenue):>12,.0f} €")
print(f"  Gesamt Equity:   {sum(equity_funding):>12,.0f} €")
print(f"  Kum. Net Income: {cum_ni[51]:>12,.0f} €")
print(f"  Max Cash:        {max(ending_cash):>12,.0f} €")
print(f"  Min Cash:        {min(ending_cash):>12,.0f} €")
neg = [i+1 for i, c in enumerate(ending_cash) if c < 0]
if neg: print(f"  ⚠  Negative Cash in Monat(en): {neg}")
print("═══════════════════════════════════════════════════════")
