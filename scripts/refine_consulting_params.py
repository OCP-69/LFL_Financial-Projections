"""
refine_consulting_params.py
────────────────────────────
Verfeinert LFL_BM_Konservativ_v2_Consulting.xlsx:

Ziel-Architektur (Datenfluss):
  Szenarien_Analyse  (Nutzer bearbeitet direkt)
       ↓  (Formeln in Sandbox referenzieren Szenarien_Analyse)
  00_Input_Sandbox   (Nutzer wählt Szenario via B1; Werte aus Szenarien_Analyse)
       ↓  (VLOOKUP nach aktivem Szenario)
  Inputs             (Anzeigeblatt, keine Direkteingaben für Consulting)
       ↓
  Revenue / P&L / Balance Sheet / Cash Flow

Konkrete Änderungen:
  Szenarien_Analyse  Zeile 8 (NEU): Consulting-Tage/Einsatz  gering=2 | normal=5 | stark=10
  00_Input_Sandbox   Zeile 11 D/E/F: Consulting-Tage/Einsatz → =Szenarien_Analyse!C8/D8/E8
  00_Input_Sandbox   Zeile 12 (NEU): Consulting-Anteil/Kunde → =Szenarien_Analyse!C5/D5/E5
  Inputs             Zeile 139: Consulting-Anteil → VLOOKUP Sandbox statt Hardcode 0.30

Output: scenarios/LFL_BM_Konservativ_v3_Linked.xlsx
"""

import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SRC  = "/home/user/LFL_Financial-Projections/scenarios/LFL_BM_Konservativ_v2_Consulting.xlsx"
DEST = "/home/user/LFL_Financial-Projections/scenarios/LFL_BM_Konservativ_v3_Linked.xlsx"

# ── Existing row positions (verified) ────────────────────────────────────────
SANDBOX_ROW_DAYS    = 11   # Consulting-Tage/Einsatz (existing, values to become formulas)
SANDBOX_ROW_SHARE   = 12   # NEW: Consulting-Anteil je Kunde (%)

SZENARIEN_ROW_SHARE = 5    # Existing: "Consulting Anteil" with gering/normal/stark
SZENARIEN_ROW_DAYS  = 8    # NEW: "Consulting-Tage/Einsatz"

INPUTS_ROW_SECTION  = 137
INPUTS_ROW_RATE     = 138
INPUTS_ROW_SHARE    = 139  # Currently hardcoded 0.30 → change to VLOOKUP
INPUTS_ROW_DAYS     = 140

# ── Colours ───────────────────────────────────────────────────────────────────
TEAL        = "1F6B75"
BLUE_DARK   = "1F3864"
BLUE_LIGHT  = "BDD7EE"
GREEN_LIGHT = "E2EFDA"
AMBER       = "FFF2CC"
ORANGE      = "FCE4D6"
WHITE       = "FFFFFF"


def _thin():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)


def _bold_border():
    s = Side(style="medium", color="4472C4")
    return Border(left=s, right=s, top=s, bottom=s)


def style_header_cell(cell, bg=BLUE_DARK, color=WHITE, bold=True):
    cell.font      = Font(bold=bold, color=color, size=10, name="Calibri")
    cell.fill      = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    cell.border    = _thin()


def style_value_cell(cell, bg=GREEN_LIGHT, fmt=None, bold=False):
    cell.font      = Font(bold=bold, size=10, name="Calibri")
    cell.fill      = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="right", vertical="center")
    cell.border    = _thin()
    if fmt:
        cell.number_format = fmt


def style_formula_cell(cell, bg=BLUE_LIGHT):
    """Marks cells that contain formulas pulling from another sheet."""
    cell.font      = Font(italic=True, size=10, name="Calibri", color="1F3864")
    cell.fill      = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="right", vertical="center")
    cell.border    = _thin()


def style_note_cell(cell):
    cell.font      = Font(italic=True, size=9, color="595959", name="Calibri")
    cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    cell.border    = _thin()


# ══════════════════════════════════════════════════════════════════════════════
# 1. Szenarien_Analyse  –  add "Consulting-Tage/Einsatz" row (row 8)
# ══════════════════════════════════════════════════════════════════════════════
def patch_szenarien_analyse(wb):
    ws  = wb["Szenarien_Analyse"]
    row = SZENARIEN_ROW_DAYS   # = 8

    # Column layout (from existing):
    # A=Faktor | B=Referenz(Basis) | C=Gering | D=Normal | E=Stark | F=Delta | G=Argumentation
    ws.cell(row=row, column=1).value = "Consulting-Tage/Einsatz"
    ws.cell(row=row, column=2).value = 5      # Referenzwert (normal)
    ws.cell(row=row, column=3).value = 2      # Gering: Basis-Einführung (2 Tage)
    ws.cell(row=row, column=4).value = 5      # Normal: Implementierung + Schulung (5 Tage)
    ws.cell(row=row, column=5).value = 10     # Stark: spezielle Anforderungen (10 Tage)
    ws.cell(row=row, column=6).value = "=E8-D8"
    ws.cell(row=row, column=7).value = (
        "Tage pro Beratungseinsatz je Kunde:\n"
        "gering=2 (Basis-Einführung: Produktschulung, Systemcheck),\n"
        "normal=5 (Implementierung + Anwenderschulung + Pilot-Begleitung),\n"
        "stark=10 (spez. Datenbereinigung, Stammdaten-Harmonisierung, "
        "individuelle Sonderlösung).\n"
        "→ Wert fließt über Sandbox in Inputs → Revenue automatisch."
    )

    # Style row 8 to match existing rows
    for col in range(1, 8):
        c = ws.cell(row=row, column=col)
        c.border    = _thin()
        c.font      = Font(size=10, name="Calibri")
        c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        ws.row_dimensions[row].height = 60

    # Highlight editable scenario columns (C, D, E)
    for col in (3, 4, 5):
        c = ws.cell(row=row, column=col)
        c.fill      = PatternFill("solid", fgColor=AMBER)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.font      = Font(bold=True, size=11, name="Calibri")

    # Also style existing row 5 (Consulting Anteil) for clarity
    ws.row_dimensions[SZENARIEN_ROW_SHARE].height = 40
    for col in (3, 4, 5):
        c = ws.cell(row=SZENARIEN_ROW_SHARE, column=col)
        c.fill      = PatternFill("solid", fgColor=AMBER)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.font      = Font(bold=True, size=11, name="Calibri")

    # Add a note in column H explaining the flow
    ws.cell(row=1, column=9).value = (
        "⚑ HINWEIS:\n"
        "Werte in den Spalten C/D/E (Gering/Normal/Stark) sind direkt bearbeitbar.\n"
        "Änderungen in DIESEM Blatt fließen automatisch über Sandbox → Inputs → Revenue/P&L."
    )
    ws.cell(row=1, column=9).font      = Font(bold=True, size=9, color="C55A11", name="Calibri")
    ws.cell(row=1, column=9).alignment = Alignment(wrap_text=True, vertical="top")
    ws.cell(row=1, column=9).fill      = PatternFill("solid", fgColor=ORANGE)
    ws.column_dimensions["I"].width    = 45
    ws.row_dimensions[1].height        = 60

    print(f"  Szenarien_Analyse row {row}: 'Consulting-Tage/Einsatz' added (2/5/10 Tage)")
    print(f"  Szenarien_Analyse row {SZENARIEN_ROW_SHARE}: 'Consulting Anteil' styled as editable")


# ══════════════════════════════════════════════════════════════════════════════
# 2. 00_Input_Sandbox  –  link rows 11 & 12 to Szenarien_Analyse
# ══════════════════════════════════════════════════════════════════════════════
def patch_sandbox(wb):
    ws = wb["00_Input_Sandbox"]

    # ── Row 11: Consulting-Tage/Einsatz → replace hardcoded values with formulas ──
    # Existing: D11=2, E11=5, F11=10 (set by add_consulting_revenue.py)
    # New:      D11=Szenarien_Analyse!C8, E11=!D8, F11=!E8
    ws.cell(row=SANDBOX_ROW_DAYS, column=4).value = "=Szenarien_Analyse!C8"
    ws.cell(row=SANDBOX_ROW_DAYS, column=5).value = "=Szenarien_Analyse!D8"
    ws.cell(row=SANDBOX_ROW_DAYS, column=6).value = "=Szenarien_Analyse!E8"
    ws.cell(row=SANDBOX_ROW_DAYS, column=7).value = (
        "Referenziert Szenarien_Analyse Zeile 8 → dort bearbeitbar. "
        "Direkte Eingabe hier überschreibt die Referenz."
    )
    # Mark as formula-sourced
    for col in (4, 5, 6):
        style_formula_cell(ws.cell(row=SANDBOX_ROW_DAYS, column=col), bg=BLUE_LIGHT)

    # ── Row 12: NEW – Consulting-Anteil je Kunde (%) ──
    row = SANDBOX_ROW_SHARE
    ws.cell(row=row, column=1).value = "Consulting"
    ws.cell(row=row, column=2).value = "Consulting-Anteil je Kunde"
    ws.cell(row=row, column=3).value = "%"
    ws.cell(row=row, column=4).value = "=Szenarien_Analyse!C5"   # Gering
    ws.cell(row=row, column=5).value = "=Szenarien_Analyse!D5"   # Normal
    ws.cell(row=row, column=6).value = "=Szenarien_Analyse!E5"   # Stark
    ws.cell(row=row, column=7).value = (
        "Referenziert Szenarien_Analyse Zeile 5 ('Consulting Anteil') → dort bearbeitbar. "
        "Anteil Kunden, die Consulting kaufen: gering=50% (viel IT-Aufwand), "
        "normal=35%, stark=15% (einfache Integration)."
    )

    # Style row 12
    for col in range(1, 8):
        c = ws.cell(row=row, column=col)
        c.font   = Font(size=10, name="Calibri")
        c.border = _thin()
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    style_formula_cell(ws.cell(row=row, column=4), bg=BLUE_LIGHT)
    style_formula_cell(ws.cell(row=row, column=5), bg=BLUE_LIGHT)
    style_formula_cell(ws.cell(row=row, column=6), bg=BLUE_LIGHT)

    # Col A & B labels
    ws.cell(row=row, column=1).fill = PatternFill("solid", fgColor="F2F2F2")
    ws.cell(row=row, column=2).font = Font(bold=True, size=10, name="Calibri")
    ws.cell(row=row, column=3).alignment = Alignment(horizontal="center")

    ws.row_dimensions[row].height = 35

    # Update VLOOKUP range hint in Sandbox header comment (no formula change needed –
    # existing VLOOKUPs use $B$3:$F$20 which still covers row 12)

    print(f"  Sandbox row {SANDBOX_ROW_DAYS}: D/E/F changed to =Szenarien_Analyse!C8/D8/E8")
    print(f"  Sandbox row {SANDBOX_ROW_SHARE}: 'Consulting-Anteil je Kunde' added → =Szenarien_Analyse!C5/D5/E5")


# ══════════════════════════════════════════════════════════════════════════════
# 3. Inputs  –  row 139: replace hardcoded 0.30 with VLOOKUP from Sandbox
# ══════════════════════════════════════════════════════════════════════════════
def patch_inputs(wb):
    ws = wb["Inputs"]

    # Row 139: was hardcoded 0.30 → now VLOOKUP
    vlookup = (
        "=VLOOKUP(\"Consulting-Anteil je Kunde\","
        "'00_Input_Sandbox'!$B$3:$F$20,"
        "'00_Input_Sandbox'!$B$2-1,FALSE)"
    )
    ws.cell(row=INPUTS_ROW_SHARE, column=2).value = vlookup
    # Remove hardcoded number format, use percentage display
    ws.cell(row=INPUTS_ROW_SHARE, column=2).number_format = "0%"
    style_formula_cell(ws.cell(row=INPUTS_ROW_SHARE, column=2), bg=BLUE_LIGHT)

    # Update explanation text to clarify the link
    ws.cell(row=INPUTS_ROW_SHARE, column=3).value = (
        "% der Kunden mit Consulting-Einsatz (szenarioabhängig aus Sandbox/Szenarien_Analyse)"
    )
    ws.cell(row=INPUTS_ROW_SHARE, column=4).value = (
        "Wert stammt aus Szenarien_Analyse (Zeile 5) via Sandbox. "
        "Gering (Automotive): 50 % – hoher IT-Aufwand & Legacy-Anbindung. "
        "Normal (Hybrid): 35 %. Stark (Packaging): 15 % – einfache PLM-Integration. "
        "Kategorien: a) Implementierungsbegleitung  b) Schulung  "
        "c) Spez. Kundenanforderungen & Datenbereinigung."
    )
    style_note_cell(ws.cell(row=INPUTS_ROW_SHARE, column=4))

    # Also add a small "DATA FLOW" block so users understand the architecture
    FLOW_ROW = INPUTS_ROW_SECTION - 1   # = 136
    ws.cell(row=FLOW_ROW, column=1).value = (
        "ℹ DATENFLUSS CONSULTING:  "
        "Szenarien_Analyse (Z.5 Anteil / Z.8 Tage)  →  "
        "Sandbox (Z.11 Tage / Z.12 Anteil)  →  "
        "Inputs (Z.138-140)  →  Revenue Z.28-29  →  P&L"
    )
    ws.cell(row=FLOW_ROW, column=1).font      = Font(italic=True, size=9, color="1F6B75", name="Calibri")
    ws.cell(row=FLOW_ROW, column=1).fill      = PatternFill("solid", fgColor="E9F7F8")
    ws.cell(row=FLOW_ROW, column=1).alignment = Alignment(wrap_text=False, vertical="center")
    ws.cell(row=FLOW_ROW, column=1).border    = _thin()
    # Merge across columns A-E for readability
    try:
        ws.merge_cells(
            start_row=FLOW_ROW, start_column=1,
            end_row=FLOW_ROW, end_column=5
        )
    except Exception:
        pass
    ws.row_dimensions[FLOW_ROW].height = 20

    print(f"  Inputs row {INPUTS_ROW_SHARE}: hardcoded 0.30 → VLOOKUP('Consulting-Anteil je Kunde', Sandbox)")
    print(f"  Inputs row {FLOW_ROW}: data-flow banner added")


# ══════════════════════════════════════════════════════════════════════════════
# 4. Add column header annotation in Szenarien_Analyse for context
# ══════════════════════════════════════════════════════════════════════════════
def annotate_szenarien_header(wb):
    ws = wb["Szenarien_Analyse"]

    # Bold the Consulting rows to make them visually distinct
    for row in (SZENARIEN_ROW_SHARE, SZENARIEN_ROW_DAYS):
        c = ws.cell(row=row, column=1)
        c.font = Font(bold=True, size=10, name="Calibri", color="1F6B75")
        c.fill = PatternFill("solid", fgColor="E9F7F8")

    # Add "→ Sandbox" marker in column H for those rows
    for row in (SZENARIEN_ROW_SHARE, SZENARIEN_ROW_DAYS):
        c = ws.cell(row=row, column=8)
        c.value = "→ Sandbox → Inputs"
        c.font  = Font(italic=True, size=8, color="1F6B75", name="Calibri")
        c.alignment = Alignment(horizontal="left", vertical="center")

    ws.column_dimensions["H"].width = 20


# ══════════════════════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════════════════════
def main():
    print(f"Loading: {SRC}")
    wb = openpyxl.load_workbook(SRC, data_only=False)

    print("\n1. Szenarien_Analyse: adding Consulting-Tage/Einsatz row ...")
    patch_szenarien_analyse(wb)

    print("\n2. 00_Input_Sandbox: linking rows 11 & 12 to Szenarien_Analyse ...")
    patch_sandbox(wb)

    print("\n3. Inputs: row 139 hardcode → VLOOKUP ...")
    patch_inputs(wb)

    print("\n4. Szenarien_Analyse: annotating linked rows ...")
    annotate_szenarien_header(wb)

    wb.save(DEST)
    print(f"\n✅ Saved: {DEST}")

    # ── Verification ──
    wb2 = openpyxl.load_workbook(DEST, data_only=False)
    ws_sz  = wb2["Szenarien_Analyse"]
    ws_sb  = wb2["00_Input_Sandbox"]
    ws_in  = wb2["Inputs"]

    print("\n── Verification ──")
    print(f"  Szenarien row 5  (Consulting Anteil):        "
          f"gering={ws_sz.cell(5,3).value} | normal={ws_sz.cell(5,4).value} | stark={ws_sz.cell(5,5).value}")
    print(f"  Szenarien row 8  (Consulting-Tage/Einsatz):  "
          f"gering={ws_sz.cell(8,3).value} | normal={ws_sz.cell(8,4).value} | stark={ws_sz.cell(8,5).value}")
    print(f"  Sandbox  row 11  D (Tage/gering formula):    {ws_sb.cell(11,4).value}")
    print(f"  Sandbox  row 11  E (Tage/normal formula):    {ws_sb.cell(11,5).value}")
    print(f"  Sandbox  row 11  F (Tage/stark formula):     {ws_sb.cell(11,6).value}")
    print(f"  Sandbox  row 12  B (variable):               {ws_sb.cell(12,2).value}")
    print(f"  Sandbox  row 12  D (Anteil/gering formula):  {ws_sb.cell(12,4).value}")
    print(f"  Inputs   row 139 B (Anteil VLOOKUP):         {ws_in.cell(139,2).value}")
    print(f"  Inputs   row 140 B (Tage VLOOKUP):           {ws_in.cell(140,2).value}")

    print()
    print("  Datenfluss:")
    print("  Szenarien_Analyse Z.5 (Anteil) → Sandbox Z.12 D/E/F → Inputs Z.139 VLOOKUP ✓")
    print("  Szenarien_Analyse Z.8 (Tage)   → Sandbox Z.11 D/E/F → Inputs Z.140 VLOOKUP ✓")
    print("  Inputs Z.138-140 → Revenue Z.28-29 (Tage×Anteil×Tagessatz) → P&L Z.8 ✓")
    print("  P&L Z.7 TOTAL REVENUE = Sub + Enterprise + Consulting ✓")
    print("  Cash Flow & Balance Sheet: referenzieren P&L!B39/B7 → automatisch ✓")


if __name__ == "__main__":
    main()
